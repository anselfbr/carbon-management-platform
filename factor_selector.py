from __future__ import annotations

import re
import shutil
import tempfile
import zipfile
import time
import xml.etree.ElementTree as ET
import math
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook, Workbook

ACTIVITY_SHEET_NAME = "Input Sheet Activity Data"
ACTIVITY_SHEET_ALIASES = ["Input Sheet Activity Data", "Raw Material Activity Data", "Activity Data"]
RAW_MATERIAL_SHEET_ALIASES = ["Input Sheet Raw Material", "Raw Material", "Raw Materials"]
DATA_START_ROW = 3
M3_MAX_UPLOAD_TOTAL_ROWS = 50000
M3_MAX_UPLOAD_DATA_ROWS = max(1, M3_MAX_UPLOAD_TOTAL_ROWS - (DATA_START_ROW - 1))
CCL_SHEET_NAME = "02.料號CCL分類表"
LCIA_SHEET_NAME = "LCIA"

FACTOR_SELECTOR_VERSION = "CMP_MODULE3A_DYNAMIC_TEMPLATE_HEADERS_V9_20260721"


def _norm(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\n", " ").replace("\r", " ")
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", text)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def _find_header_row(ws, aliases: Iterable[str], max_scan_rows: int = 30) -> int:
    alias_keys = {_norm(a) for a in aliases}
    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = {_norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        if values & alias_keys:
            return row
    return 1


def _find_col(ws, aliases: list[str], header_rows: int = DATA_START_ROW - 1, required: bool = True) -> int | None:
    alias_keys = [_norm(a) for a in aliases if str(a or "").strip()]
    rows = list(range(1, max(1, header_rows) + 1))
    if 2 in rows:
        rows = [2] + [r for r in rows if r != 2]
    for row in rows:
        for col in range(1, ws.max_column + 1):
            if _norm(ws.cell(row, col).value) in alias_keys:
                return col
    if required:
        raise ValueError(f"找不到欄位：{', '.join(aliases)}")
    return None


def _find_col_in_header_row(ws, header_row: int, aliases: list[str], required: bool = True) -> int | None:
    alias_keys = [_norm(a) for a in aliases if str(a or "").strip()]
    for col in range(1, ws.max_column + 1):
        if _norm(ws.cell(header_row, col).value) in alias_keys:
            return col
    if required:
        raise ValueError(f"找不到欄位：{', '.join(aliases)}")
    return None




def _normalize_material_key(value: Any) -> str:
    text = _text(value).strip().upper()
    if text.endswith(".0"):
        text = text[:-2]
    text = text.replace("\t", "").replace("\n", "").replace("\r", "")
    return text.strip()


def _emit_progress(
    callback: Callable[..., None] | None,
    progress: int,
    step: str,
    remaining_seconds: int | None = None,
    processed_rows: int | None = None,
    total_rows: int | None = None,
) -> None:
    if not callback:
        return
    try:
        callback(progress, step, remaining_seconds, processed_rows=processed_rows, total_rows=total_rows)
    except TypeError:
        try:
            callback(progress, step, remaining_seconds)
        except TypeError:
            callback(progress, step)


def _safe_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text or text.lower() in {"nan", "none"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _read_ccl_mapping(ccl_path: str | Path, progress_callback: Callable[..., None] | None = None) -> dict[str, Dict[str, Any]]:
    _emit_progress(progress_callback, 10, "讀取 CCL 係數組配表", 30)
    wb = load_workbook(ccl_path, read_only=True, data_only=True)
    try:
        sheet_name = CCL_SHEET_NAME if CCL_SHEET_NAME in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        header_row = _find_header_row(ws, ["Material"])

        # CCL 對照表正式固定欄位：Material / CCL Item / 碳係數 / 係數單位
        material_col = _find_col_in_header_row(ws, header_row, ["Material"])
        ccl_item_col = _find_col_in_header_row(ws, header_row, ["CCL Item"])
        factor_name_col = _find_col_in_header_row(ws, header_row, ["係數名稱"], required=False)
        factor_col = _find_col_in_header_row(ws, header_row, ["碳係數"])
        unit_col = _find_col_in_header_row(ws, header_row, ["係數單位"], required=False)

        mapping: dict[str, Dict[str, Any]] = {}
        total_rows = max(1, ws.max_row - header_row)
        start_time = time.perf_counter()
        value_rows = ws.iter_rows(
            min_row=header_row + 1,
            max_row=ws.max_row,
            values_only=True,
        )
        for idx, values in enumerate(value_rows, start=1):
            material = _text(values[material_col - 1] if len(values) >= material_col else None)
            if not material:
                continue
            key = _normalize_material_key(material)
            if not key:
                continue
            factor_value = values[factor_col - 1] if len(values) >= factor_col else None
            unit_value = values[unit_col - 1] if unit_col and len(values) >= unit_col else ""
            safe_factor = _safe_number(factor_value)
            mapping[key] = {
                "material": material,
                "ccl_item": _text(values[ccl_item_col - 1] if len(values) >= ccl_item_col else None),
                "factor_name": _text(values[factor_name_col - 1] if factor_name_col and len(values) >= factor_name_col else ""),
                "emission_factor": safe_factor if safe_factor is not None else factor_value,
                "unit": _text(unit_value),
            }
            if idx == 1 or idx % 2000 == 0:
                elapsed = max(0.001, time.perf_counter() - start_time)
                rate = idx / elapsed
                remaining = int(max(1, (total_rows - idx) / rate + 12)) if rate > 0 else 30
                _emit_progress(progress_callback, 10 + int(min(20, idx / total_rows * 20)), "建立 CCL Material → Factor 對應索引", remaining)
        _emit_progress(progress_callback, 32, f"CCL 索引建立完成，共 {len(mapping):,} 筆", 25)
        return mapping
    finally:
        wb.close()

def _first_header_rows(ws, header_row_count: int = DATA_START_ROW - 1) -> list[list[Any]]:
    rows: list[list[Any]] = []
    max_col = max(1, int(getattr(ws, "max_column", 1) or 1))
    for values in ws.iter_rows(min_row=1, max_row=header_row_count, values_only=True):
        row = list(values or [])
        if len(row) < max_col:
            row.extend([None] * (max_col - len(row)))
        rows.append(row)
    while len(rows) < header_row_count:
        rows.append([None] * max_col)
    return rows


def _find_col_from_header_rows(header_rows: list[list[Any]], aliases: list[str], required: bool = True) -> int | None:
    alias_keys = {_norm(a) for a in aliases if str(a or "").strip()}
    # Bulk templates use row 2 as the user-visible header row. Prefer it over
    # row 1 because row 1 can contain internal/helper keys in hidden columns.
    row_order = [1, 0] + list(range(2, min(len(header_rows), DATA_START_ROW - 1)))
    for row_idx in row_order:
        if row_idx >= len(header_rows):
            continue
        for col_idx, value in enumerate(header_rows[row_idx], start=1):
            if _norm(value) in alias_keys:
                return col_idx
    if required:
        raise ValueError(f"找不到欄位：{', '.join(aliases)}")
    return None


def _ensure_output_col(header_rows: list[list[Any]], aliases: list[str], preferred_header: str) -> int:
    col = _find_col_from_header_rows(header_rows, aliases, required=False)
    if col:
        return col
    max_len = max((len(r) for r in header_rows), default=0)
    new_col = max_len + 1
    for row in header_rows:
        if len(row) < new_col:
            row.extend([None] * (new_col - len(row)))
    if not header_rows:
        header_rows.append([None] * new_col)
    header_rows[0][new_col - 1] = preferred_header
    return new_col


def _pad_row(row: list[Any], width: int) -> list[Any]:
    if len(row) < width:
        row.extend([None] * (width - len(row)))
    elif len(row) > width:
        row = row[:width]
    return row


def _copy_non_activity_sheet_streaming(src_ws, dst_wb: Workbook) -> int:
    dst_ws = dst_wb.create_sheet(title=src_ws.title)
    rows = 0
    for values in src_ws.iter_rows(values_only=True):
        dst_ws.append(list(values or []))
        rows += 1
    return rows




# ---------------------------------------------------------------------------
# Module 3 final export: keep M2B/M2C/M3 as lightweight intermediate files,
# then apply the original Raw Material Bulk Template only at the final output.
# This avoids carrying full Excel template structures through every large-data
# step while still producing a third-party-uploadable workbook at M3.
# ---------------------------------------------------------------------------
_XML_INVALID_TEXT_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

_ACTIVITY_FINAL_FACTOR_COLUMNS: list[tuple[str, list[str], str]] = [
    ("factor_name", ["Factor Name", "factor_name", "CCL Item", "係數名稱"], "Factor Name"),
    ("emission_factor", ["Emission Factor", "emission_factor", "Carbon Factor", "碳係數"], "Emission Factor"),
    ("factor_source", ["Factor Source", "factor_source", "Emission Factor Source", "係數來源"], "Factor Source"),
    ("factor_comment", ["Factor Comment", "factor_comment", "Emission Factor Comment", "係數備註"], "Factor Comment"),
    ("country_area", ["Country/Area", "Country Area", "country_area", "Country", "Area", "國家地區"], "Country/Area"),
    ("enabled_date", ["Enabled Date", "activation_date", "Effective Date", "啟用日期"], "Enabled Date"),
    ("data_quality", ["Data Quality", "data_quality", "資料品質"], "Data Quality"),
]

_ACTIVITY_SOURCE_ALIASES: dict[str, list[str]] = {
    "raw_name": ["Raw Material Name", "raw_material_name"],
    "raw_code": ["Raw Material Code", "raw_material_code", "Material"],
    "doc_start": ["Doc. Start Date", "doc_start_date", "Document Start Date"],
    "doc_end": ["Doc. End Date", "doc_end_date", "Document End Date"],
    "document_type": ["Document Type", "document_type"],
    "document_number": ["Document Number", "document_number"],
    "usage": ["Usage", "Activity Data", "activity_data", "usage"],
    "unit": ["Activity Data Unit", "activity_data_unit", "Unit"],
    "net_weight": ["Net weight", "Net Weight", "net_weight"],
    "gross_weight": ["Gross weight", "Gross Weight", "gross_weight"],
    "weight_unit": ["Weight Unit (optional)", "Weight Unit", "weight_unit"],
    "data_source": ["Data Source", "data_source"],
    "data_source_other": ["Data Source Other", "Data Source other", "data_source_other"],
    "supplier_name": ["Supplier Name (optional)", "Supplier Name", "supplier_name", "supplier_name_resolved"],
    "transport_origin": ["Transportation Origin", "transportation_origin"],
    "transport_destination": ["Transportation Destination", "transportation_destination"],
    "calculate_transportation_emissions": [
        "Calculate Transportation Emissions (Required)",
        "Calculate Transportation Emissions",
        "calculate_transportation_emissions",
    ],
    "target_product": ["Product Name", "allocated_target_product_service", "Target Product", "target_product"],
    "comment": ["Comment", "comment"],
    "material_group": ["Material Group", "Material group", "material_group"],
}

_RAW_SOURCE_ALIASES: dict[str, list[str]] = {
    "raw_name": ["Raw Material Name", "raw_material_name"],
    "raw_code": ["Raw Material Code", "raw_material_code"],
    "description": ["Raw Material Description (Optional)", "Raw Material Description", "raw_material_description", "description"],
}


_ACTIVITY_SOURCE_DEFAULT_COLS: dict[str, int] = {
    "raw_name": 1,
    "raw_code": 2,
    "doc_start": 3,
    "doc_end": 4,
    "document_type": 5,
    "document_number": 6,
    "usage": 7,
    "unit": 8,
    "net_weight": 9,
    "gross_weight": 10,
    "weight_unit": 11,
    "data_source": 12,
    "data_source_other": 13,
    "supplier_name": 14,
    "transport_origin": 15,
    "transport_destination": 16,
    "target_product": 17,
    "comment": 18,
    "material_group": 19,
}

_RAW_SOURCE_DEFAULT_COLS: dict[str, int] = {
    "raw_name": 1,
    "raw_code": 2,
    "description": 6,
}


def _xml_clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, datetime):
        text = value.strftime("%Y/%m/%d")
    elif isinstance(value, date):
        text = value.strftime("%Y/%m/%d")
    else:
        text = str(value)
    return _XML_INVALID_TEXT_RE.sub("", text)


def _xlsx_col_letter(col_idx: int) -> str:
    result = ""
    value = int(col_idx)
    while value > 0:
        value, rem = divmod(value - 1, 26)
        result = chr(65 + rem) + result
    return result or "A"


def _xlsx_cell_ref(row_idx: int, col_idx: int) -> str:
    return f"{_xlsx_col_letter(col_idx)}{int(row_idx)}"


_NO_FORMULA_CACHE = object()


class _FormulaCachedValue:
    """Internal marker: keep the template formula and attach its cached result."""

    __slots__ = ("value",)

    def __init__(self, value: Any):
        self.value = value


def _xml_escape_value(value: Any) -> bytes:
    text = _xml_clean_text(value)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .encode("utf-8")
    )


def _manual_cell_xml(
    row_idx: int,
    col_idx: int,
    value: Any,
    style_id: str | int | None = None,
    formula_xml: bytes | None = None,
    formula_cache_value: Any = _NO_FORMULA_CACHE,
    emit_empty_style: bool = False,
) -> bytes:
    ref = _xlsx_cell_ref(row_idx, col_idx)
    style_attr = b""
    if style_id is not None and str(style_id).strip() != "":
        style_attr = f' s="{str(style_id).strip()}"'.encode("utf-8")
    if formula_xml:
        if formula_cache_value is _NO_FORMULA_CACHE:
            return b'<c r="' + ref.encode("utf-8") + b'"' + style_attr + b'>' + formula_xml + b'</c>'
        cache = formula_cache_value
        if isinstance(cache, bool):
            return (
                b'<c r="' + ref.encode("utf-8") + b'"' + style_attr + b' t="b">'
                + formula_xml + f'<v>{1 if cache else 0}</v></c>'.encode("utf-8")
            )
        if isinstance(cache, (int, float)) and not isinstance(cache, bool):
            try:
                number = float(cache)
                if math.isfinite(number):
                    raw = str(int(number)) if number.is_integer() else repr(number)
                    return (
                        b'<c r="' + ref.encode("utf-8") + b'"' + style_attr + b'>'
                        + formula_xml + b'<v>' + raw.encode("utf-8") + b'</v></c>'
                    )
            except Exception:
                pass
        # Formula results in AB~AH are lookup keys/names.  Store them as a
        # formula string result (t="str") while the cell number format remains
        # the original/General style.  This is not Excel Text number format.
        return (
            b'<c r="' + ref.encode("utf-8") + b'"' + style_attr + b' t="str">'
            + formula_xml + b'<v>' + _xml_escape_value(cache) + b'</v></c>'
        )
    if value is None or (isinstance(value, str) and value == ""):
        if emit_empty_style and style_attr:
            return b'<c r="' + ref.encode("utf-8") + b'"' + style_attr + b'/>'
        return b""
    if isinstance(value, bool):
        return b'<c r="' + ref.encode("utf-8") + b'"' + style_attr + f' t="b"><v>{1 if value else 0}</v></c>'.encode("utf-8")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            number = float(value)
            if math.isfinite(number):
                raw = str(int(number)) if number.is_integer() else repr(number)
                return b'<c r="' + ref.encode("utf-8") + b'"' + style_attr + b'><v>' + raw.encode("utf-8") + b'</v></c>'
        except Exception:
            pass
    text = _xml_clean_text(value)
    if text == "":
        if emit_empty_style and style_attr:
            return b'<c r="' + ref.encode("utf-8") + b'"' + style_attr + b'/>'
        return b""
    escaped = (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )
    return b'<c r="' + ref.encode("utf-8") + b'"' + style_attr + b' t="inlineStr"><is><t xml:space="preserve">' + escaped.encode("utf-8") + b'</t></is></c>'


def _rewrite_formula_row_refs(formula_xml: bytes | None, source_row: int, target_row: int) -> bytes | None:
    if not formula_xml or source_row == target_row:
        return formula_xml
    # Adjust simple A3 / $A$3 style references in row-level template formulas.
    src = str(int(source_row)).encode("ascii")
    dst = str(int(target_row)).encode("ascii")
    pattern = rb'(?<![A-Za-z0-9_])((?:\$?[A-Z]{1,3})\$?)' + src + rb'(?![0-9])'
    return re.sub(pattern, lambda m: m.group(1) + dst, formula_xml)


def _coerce_general_numeric_value(value: Any) -> Any:
    """Convert numeric-looking helper text to a real Excel number when safe."""
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text:
        return value
    compact = text.replace(",", "")
    if not re.fullmatch(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?", compact):
        return value
    unsigned = compact.lstrip("+-")
    integer_part = unsigned.split(".", 1)[0].split("e", 1)[0].split("E", 1)[0]
    # Preserve code-like values such as 00123.
    if len(integer_part) > 1 and integer_part.startswith("0") and not unsigned.startswith("0."):
        return value
    try:
        number = float(compact)
    except Exception:
        return value
    if not math.isfinite(number):
        return value
    return int(number) if number.is_integer() and "e" not in compact.lower() and "." not in compact else number


def _manual_row_xml(
    row_idx: int,
    values: list[Any],
    width: int | None = None,
    style_by_col: dict[int, str] | None = None,
    formula_by_col: dict[int, bytes] | None = None,
    template_row_attrs: bytes | None = None,
    template_formula_row: int = DATA_START_ROW,
    emit_empty_styles: bool = False,
    general_numeric_cols: set[int] | None = None,
) -> bytes:
    width = int(width or len(values) or 1)
    style_by_col = style_by_col or {}
    formula_by_col = formula_by_col or {}
    general_numeric_cols = general_numeric_cols or set()
    cells = []
    for col_idx in range(1, width + 1):
        value = values[col_idx - 1] if col_idx <= len(values) else None
        formula_cache_value = _NO_FORMULA_CACHE
        if isinstance(value, _FormulaCachedValue):
            formula_cache_value = value.value
            value = None
        if col_idx in general_numeric_cols and formula_cache_value is _NO_FORMULA_CACHE:
            value = _coerce_general_numeric_value(value)
        formula_xml = None
        if col_idx in formula_by_col and (
            formula_cache_value is not _NO_FORMULA_CACHE
            or value is None
            or (isinstance(value, str) and value == "")
        ):
            formula_xml = _rewrite_formula_row_refs(formula_by_col.get(col_idx), template_formula_row, row_idx)
        if formula_cache_value is not _NO_FORMULA_CACHE and not formula_xml:
            # Defensive fallback for an unexpected template without a formula.
            value = formula_cache_value
            formula_cache_value = _NO_FORMULA_CACHE
        cell = _manual_cell_xml(
            row_idx,
            col_idx,
            value,
            style_id=style_by_col.get(col_idx),
            formula_xml=formula_xml,
            formula_cache_value=formula_cache_value,
            emit_empty_style=emit_empty_styles,
        )
        if cell:
            cells.append(cell)
    attrs = b''
    if template_row_attrs:
        attrs = b' ' + template_row_attrs.strip()
        # Remove r/spans from copied attrs; those are regenerated for the target row.
        attrs = re.sub(rb'\s+r="[^"]*"', b'', attrs)
        attrs = re.sub(rb"\s+r='[^']*'", b'', attrs)
        attrs = re.sub(rb'\s+spans="[^"]*"', b'', attrs)
        attrs = re.sub(rb"\s+spans='[^']*'", b'', attrs)
    return b'<row r="%d" spans="1:%d"%s>%s</row>' % (int(row_idx), int(width), attrs, b"".join(cells))

def _max_header_width(header_rows: list[list[Any]]) -> int:
    return max((len(r) for r in header_rows), default=0)


def _pad_header_rows(header_rows: list[list[Any]], width: int) -> list[list[Any]]:
    out = [list(r) for r in header_rows]
    while len(out) < DATA_START_ROW - 1:
        out.append([])
    for row in out:
        if len(row) < width:
            row.extend([None] * (width - len(row)))
    return out


def _header_labels(header_rows: list[list[Any]], col_idx: int) -> list[str]:
    labels = []
    for row in header_rows:
        if col_idx - 1 < len(row):
            text = _text(row[col_idx - 1])
            if text:
                labels.append(text)
    return labels


def _visible_bulk_header_row_missing(header_rows: list[list[Any]]) -> bool:
    if len(header_rows) < 2:
        return True
    return sum(1 for value in header_rows[1] if _text(value)) < 3


def _find_col_in_header_values(
    header_rows: list[list[Any]],
    aliases: list[str],
    required: bool = False,
    fallback_col: int | None = None,
    prefer_fallback_when_row2_missing: bool = False,
) -> int | None:
    alias_keys = {_norm(a) for a in aliases if _text(a)}
    width = _max_header_width(header_rows)
    row2_missing = _visible_bulk_header_row_missing(header_rows)
    if prefer_fallback_when_row2_missing and row2_missing and fallback_col:
        # Legacy M2B/M2C files omitted row 2 but still contained helper keys in
        # hidden columns. Prefer official visible columns such as H/K/L/N.
        return int(fallback_col)
    row_order = [1, 0] + list(range(2, len(header_rows)))
    for row_idx in row_order:
        if row_idx >= len(header_rows):
            continue
        for col_idx in range(1, width + 1):
            if col_idx - 1 < len(header_rows[row_idx]) and _norm(header_rows[row_idx][col_idx - 1]) in alias_keys:
                return col_idx
    if required:
        raise ValueError(f"找不到欄位：{', '.join(aliases)}")
    return int(fallback_col) if fallback_col else None


def _ensure_header_column(header_rows: list[list[Any]], aliases: list[str], preferred_header: str, group_header: str = "Factor") -> int:
    existing = _find_col_in_header_values(header_rows, aliases, required=False)
    if existing:
        return existing
    width = _max_header_width(header_rows) + 1
    while len(header_rows) < 2:
        header_rows.append([])
    for row in header_rows:
        if len(row) < width:
            row.extend([None] * (width - len(row)))
    header_rows[0][width - 1] = group_header
    header_rows[1][width - 1] = preferred_header
    return width


def _row_value(values: tuple[Any, ...] | list[Any], col_idx: int | None) -> Any:
    if not col_idx:
        return None
    idx = int(col_idx) - 1
    return values[idx] if 0 <= idx < len(values) else None


def _copy_matching_value_to_target(target_row: list[Any], target_col: int | None, source_values: tuple[Any, ...], source_cols: dict[str, int | None], key: str) -> None:
    if not target_col:
        return
    src_col = source_cols.get(key)
    if src_col:
        while len(target_row) < target_col:
            target_row.append(None)
        target_row[target_col - 1] = _row_value(source_values, src_col)



# M3A supports only the revised third-party Raw Material Bulk Template, but it
# locates both the visible transportation flag and hidden helper formulas by
# their header keys rather than fixed Excel column letters. This prevents future
# template insertions/reordering from breaking M3A.
_M3_DOCUMENT_TYPE_HELPER_KEY = "document_type"
_M3_TRANSPORT_CALC_FIELD_ALIASES = tuple(_ACTIVITY_SOURCE_ALIASES["calculate_transportation_emissions"])
_M3_TRANSPORT_CALC_DEFAULT_VALUE = "Yes"
_M3_ACTIVITY_HELPER_FIELDS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("document_type", ("document_type",)),
    ("activity_data_unit", ("activity_data_unit",)),
    ("weight_unit", ("weight_unit",)),
    ("data_source", ("data_source",)),
    ("supplier_name_resolved", ("supplier_name_resolved",)),
    ("supplier_code_resolved", ("supplier_code_resolved",)),
    ("country_area", ("country_area",)),
)


def _find_internal_header_col(target_activity_headers: list[list[Any]], aliases: Iterable[str]) -> int | None:
    """Find a hidden/helper field by its internal header key, preferring row 1."""
    alias_keys = {_norm(alias) for alias in aliases if _text(alias)}
    if not alias_keys:
        return None
    width = _max_header_width(target_activity_headers)
    row_order = [0] + list(range(1, len(target_activity_headers)))
    for row_idx in row_order:
        if row_idx >= len(target_activity_headers):
            continue
        row = target_activity_headers[row_idx]
        for col_idx in range(1, width + 1):
            if col_idx - 1 < len(row) and _norm(row[col_idx - 1]) in alias_keys:
                return col_idx
    return None


def _activity_helper_layout(target_activity_headers: list[list[Any]], width: int) -> tuple[tuple[int, ...], int]:
    """Locate revised-template fields dynamically from their table headers."""
    transport_col = _find_col_in_header_values(
        target_activity_headers,
        list(_M3_TRANSPORT_CALC_FIELD_ALIASES),
        required=False,
    )
    if not transport_col:
        raise ValueError(
            "正式 Raw Material Bulk Template 版本不符：缺少欄位 "
            "Calculate Transportation Emissions (Required)。"
        )

    helper_cols: list[int] = []
    missing_headers: list[str] = []
    for helper_key, aliases in _M3_ACTIVITY_HELPER_FIELDS:
        col_idx = _find_internal_header_col(target_activity_headers, aliases)
        if not col_idx:
            missing_headers.append(helper_key)
        else:
            helper_cols.append(int(col_idx))

    if missing_headers:
        raise ValueError(
            "正式 Raw Material Bulk Template 版本不符：缺少隱藏公式欄位表頭："
            + ", ".join(missing_headers)
            + "。"
        )
    if len(set(helper_cols)) != len(helper_cols):
        raise ValueError("正式 Raw Material Bulk Template 版本不符：隱藏公式欄位表頭有重複對應。")
    if int(transport_col) in set(helper_cols):
        raise ValueError(
            "正式 Raw Material Bulk Template 版本不符："
            "Calculate Transportation Emissions (Required) 與隱藏公式欄位發生欄位衝突。"
        )
    if helper_cols and int(width or 0) < max(helper_cols):
        raise ValueError("正式 Raw Material Bulk Template 版本不符：隱藏公式欄位超出工作表範圍。")
    return tuple(helper_cols), helper_cols[0]


def _activity_helper_formula_columns_to_preserve(tpl_activity_ws, target_activity_headers: list[list[Any]], width: int) -> set[int]:
    """Require and preserve the seven helper formulas located by header keys."""
    helper_cols, _document_type_idx = _activity_helper_layout(
        target_activity_headers,
        int(width or _max_header_width(target_activity_headers) or 1),
    )
    missing: list[str] = []
    for col_idx in helper_cols:
        try:
            value = tpl_activity_ws.cell(DATA_START_ROW, int(col_idx)).value
        except Exception:
            value = None
        if not (isinstance(value, str) and value.startswith("=")):
            missing.append(_xlsx_col_letter(int(col_idx)))
    if missing:
        raise ValueError(
            "正式 Raw Material Bulk Template 版本不符："
            "依表頭定位的隱藏欄位必須保留公式，缺少公式欄位："
            + ", ".join(missing)
            + "。"
        )
    return set(int(col) for col in helper_cols)


def _activity_helper_cache_specs(
    template_wb,
    target_activity_cols: dict[str, int | None],
    factor_cols: dict[str, int],
    helper_cols: Iterable[int],
) -> dict[int, tuple[int, dict[str, Any], dict[str, Any]]]:
    """Map dynamically located helper formulas to their visible source cells."""
    dropdown_name = _first_existing_name(template_wb.sheetnames, ["Dropdown Values", "Dropdown Value", "Dropdown"])
    if not dropdown_name:
        raise ValueError("正式 Raw Material Bulk Template 缺少 Dropdown Values 分頁，無法建立隱藏公式快取值。")
    ws = template_wb[dropdown_name]
    helper_cols = tuple(int(col) for col in helper_cols)
    if len(helper_cols) != 7:
        raise ValueError(f"正式 Raw Material Bulk Template 隱藏公式欄位數量應為 7，目前偵測為 {len(helper_cols)}。")
    raw_specs = {
        helper_cols[0]: (target_activity_cols.get("document_type"), 1, 2),
        helper_cols[1]: (target_activity_cols.get("unit"), 3, 4),
        helper_cols[2]: (target_activity_cols.get("weight_unit"), 5, 6),
        helper_cols[3]: (target_activity_cols.get("data_source"), 7, 8),
        helper_cols[4]: (target_activity_cols.get("supplier_name"), 15, 23),
        helper_cols[5]: (target_activity_cols.get("supplier_name"), 15, 24),
        helper_cols[6]: (factor_cols.get("country_area"), 22, 25),
    }
    maps: dict[int, tuple[int, dict[str, Any], dict[str, Any], int, int]] = {}
    for helper_col, (source_col, display_col, key_col) in raw_specs.items():
        if source_col:
            maps[int(helper_col)] = (int(source_col), {}, {}, int(display_col), int(key_col))

    # read_only worksheets are optimized for sequential iter_rows. Calling
    # ws.cell() thousands of times reparses the XML and is prohibitively slow.
    for values in ws.iter_rows(min_row=2, max_col=25, values_only=True):
        for helper_col, (_source_col, exact, normalized, display_col, key_col) in maps.items():
            display = _text(values[display_col - 1] if display_col <= len(values) else None)
            if not display:
                continue
            key = values[key_col - 1] if key_col <= len(values) else None
            exact.setdefault(display, key)
            normalized.setdefault(_norm(display), key)

    return {
        helper_col: (source_col, exact, normalized)
        for helper_col, (source_col, exact, normalized, _display_col, _key_col) in maps.items()
    }


def _apply_preserved_formula_cache_cells(
    row_values: list[Any],
    formula_cols: set[int],
    cache_specs: dict[int, tuple[int, dict[str, Any], dict[str, Any]]],
    stats: dict[str, Any] | None = None,
) -> None:
    """Keep helper formulas and attach formula-result caches for third-party parsers."""
    if not formula_cols:
        return
    if stats is not None:
        stats["rows_processed"] = int(stats.get("rows_processed", 0)) + 1
    for col_idx in sorted(int(c) for c in formula_cols):
        while len(row_values) < col_idx:
            row_values.append(None)
        existing = row_values[col_idx - 1]
        cache_value: Any = ""
        source_display = ""
        spec = cache_specs.get(col_idx)
        if spec:
            source_col, exact, normalized = spec
            source_raw = row_values[source_col - 1] if 1 <= source_col <= len(row_values) else None
            source_display = _text(source_raw)
            if source_display:
                if source_display in exact:
                    cache_value = exact[source_display]
                else:
                    cache_value = normalized.get(_norm(source_display), "")
        # If an upstream intermediate already carries a non-formula helper value,
        # retain it as a safe fallback when the official dropdown has no match.
        if (cache_value is None or _text(cache_value) == "") and existing is not None:
            existing_text = _text(existing)
            if existing_text and not existing_text.startswith("="):
                cache_value = existing
        if stats is not None:
            col_name = _xlsx_col_letter(col_idx)
            if source_display and (cache_value is None or _text(cache_value) == ""):
                misses = stats.setdefault("misses_by_column", {})
                misses[col_name] = int(misses.get(col_name, 0)) + 1
            elif cache_value is not None and _text(cache_value) != "":
                hits = stats.setdefault("hits_by_column", {})
                hits[col_name] = int(hits.get(col_name, 0)) + 1
        row_values[col_idx - 1] = _FormulaCachedValue(cache_value if cache_value is not None else "")


def _force_full_calc_on_load(data: bytes) -> bytes:
    """Ask Excel to recalculate preserved formulas when the workbook opens."""
    calc_attrs = b' calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"'
    if b"<calcPr" not in data:
        return data.replace(b"</workbook>", b"<calcPr" + calc_attrs + b"/></workbook>", 1)

    def repl(match: re.Match[bytes]) -> bytes:
        tag = match.group(0)
        for attr in (b"calcMode", b"fullCalcOnLoad", b"forceFullCalc"):
            tag = re.sub(attr + rb'="[^"]*"', b"", tag)
        if tag.endswith(b"/>"):
            return tag[:-2].rstrip() + calc_attrs + b"/>"
        return tag[:-1].rstrip() + calc_attrs + b">"

    return re.sub(rb"<calcPr\b[^>]*/?>", repl, data, count=1)


def _sheet_paths_by_name_from_xlsx(path: str | Path) -> dict[str, str]:
    ns_main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    ns_rel = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    ns_pkg_rel = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    with zipfile.ZipFile(path) as zf:
        workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rels = {}
        for rel in rel_root.findall(f"{ns_pkg_rel}Relationship"):
            rid = rel.attrib.get("Id")
            target = rel.attrib.get("Target", "")
            if rid:
                if target.startswith("/xl/"):
                    rels[rid] = target.lstrip("/")
                elif target.startswith("xl/"):
                    rels[rid] = target
                else:
                    rels[rid] = "xl/" + target.lstrip("/")
        sheet_paths: dict[str, str] = {}
        sheets_el = workbook_root.find(f"{ns_main}sheets")
        if sheets_el is not None:
            for sheet in sheets_el.findall(f"{ns_main}sheet"):
                name = sheet.attrib.get("name", "")
                rid = sheet.attrib.get(f"{ns_rel}id")
                if name and rid in rels:
                    sheet_paths[name] = rels[rid]
        return sheet_paths


def _split_sheet_xml(sheet_xml: bytes) -> tuple[bytes, bytes, bytes, bytes]:
    m = re.search(rb"<sheetData\b[^>]*>", sheet_xml)
    if not m:
        raise ValueError("Template worksheet XML 缺少 sheetData。")
    end_m = re.search(rb"</sheetData>", sheet_xml)
    if not end_m:
        # Handle self-closing sheetData, uncommon but valid.
        self_m = re.search(rb"<sheetData\b[^>]*/>", sheet_xml)
        if not self_m:
            raise ValueError("Template worksheet XML 缺少 /sheetData。")
        prefix = sheet_xml[: self_m.start()] + b"<sheetData>"
        inside = b""
        suffix = b"</sheetData>" + sheet_xml[self_m.end():]
        return prefix, inside, suffix, sheet_xml
    prefix = sheet_xml[: m.end()]
    inside = sheet_xml[m.end(): end_m.start()]
    suffix = sheet_xml[end_m.start():]
    return prefix, inside, suffix, sheet_xml


def _extract_raw_header_row(inside_sheetdata: bytes, row_idx: int) -> bytes | None:
    pattern = rb"<row\b(?=[^>]*\br=['\"]?" + str(int(row_idx)).encode() + rb"['\"]?)[^>]*>.*?</row>"
    m = re.search(pattern, inside_sheetdata, flags=re.DOTALL)
    return m.group(0) if m else None


def _append_header_cells(row_xml: bytes | None, row_idx: int, start_col: int, values: list[Any], final_width: int) -> bytes:
    if row_xml is None:
        row_xml = f'<row r="{int(row_idx)}"></row>'.encode("utf-8")
    row_xml = re.sub(rb'spans="[^"]*"', f'spans="1:{int(final_width)}"'.encode("utf-8"), row_xml)
    if not values:
        return row_xml
    if row_xml.rstrip().endswith(b"/>"):
        row_xml = row_xml.rstrip()[:-2] + b"></row>"
    insert_at = row_xml.rfind(b"</row>")
    if insert_at < 0:
        return _manual_row_xml(row_idx, [None] * (start_col - 1) + values, final_width)
    extra = b"".join(_manual_cell_xml(row_idx, start_col + offset, value) for offset, value in enumerate(values))
    return row_xml[:insert_at] + extra + row_xml[insert_at:]


def _update_dimension_in_prefix(prefix: bytes, max_row: int, max_col: int) -> bytes:
    ref = f'A1:{_xlsx_col_letter(max(1, int(max_col)))}{max(1, int(max_row))}'.encode("utf-8")
    if re.search(rb"<dimension\b[^>]*/>", prefix):
        return re.sub(rb"<dimension\b[^>]*/>", b'<dimension ref="' + ref + b'"/>', prefix, count=1)
    return prefix


def _remove_calc_chain_content_types(data: bytes) -> bytes:
    return re.sub(rb'<Override\b[^>]*PartName="/xl/calcChain\.xml"[^>]*/>', b"", data)


def _remove_calc_chain_rels(data: bytes) -> bytes:
    return re.sub(rb'<Relationship\b[^>]*(?:calcChain|calcchain)[^>]*/>', b"", data)


def _spool_sheet_rows_xml(
    rows_iter,
    temp_path: Path,
    start_row: int,
    width: int,
    progress_callback=None,
    progress_base: int = 0,
    progress_span: int = 0,
    progress_step: str = "",
    style_by_col: dict[int, str] | None = None,
    formula_by_col: dict[int, bytes] | None = None,
    template_row_attrs: bytes | None = None,
    template_formula_row: int = DATA_START_ROW,
    emit_empty_styles: bool = False,
    general_numeric_cols: set[int] | None = None,
) -> tuple[int, int]:
    count = 0
    max_width = int(width or 1)
    start_time = time.perf_counter()
    with temp_path.open("wb") as fh:
        for count, values in enumerate(rows_iter, start=1):
            row_idx = start_row + count - 1
            row_values = list(values or [])
            if len(row_values) > max_width:
                max_width = len(row_values)
            fh.write(_manual_row_xml(
                row_idx,
                row_values,
                max_width,
                style_by_col=style_by_col,
                formula_by_col=formula_by_col,
                template_row_attrs=template_row_attrs,
                template_formula_row=template_formula_row,
                emit_empty_styles=emit_empty_styles,
                general_numeric_cols=general_numeric_cols,
            ))
            if progress_callback and progress_span and (count == 1 or count % 5000 == 0):
                elapsed = max(0.001, time.perf_counter() - start_time)
                rate = count / elapsed
                remaining = int(max(1, 30 if rate <= 0 else (100000 - count) / rate))
                _emit_progress(progress_callback, min(95, progress_base + min(progress_span, int(progress_span * min(1, count / max(count, 1))))), progress_step, remaining, count, None)
    return count, max_width

def _first_existing_name(names: Iterable[str], aliases: Iterable[str]) -> str | None:
    exact = {str(n): str(n) for n in names}
    for alias in aliases:
        if alias in exact:
            return exact[alias]
    norm_lookup = {_norm(n): str(n) for n in names}
    for alias in aliases:
        found = norm_lookup.get(_norm(alias))
        if found:
            return found
    return None


def _style_format_profile(styles_xml: bytes | None) -> tuple[set[str], str]:
    """Return text-formatted style IDs and an existing General style ID.

    The third-party parser treats AB~AH helper cells formatted as Excel Text as
    literal text, even when the cells contain formulas or numeric-looking
    values.  We keep the official template style whenever it is not Text.  For
    Text styles, we reuse an existing General style from the same workbook so
    the workbook package and style table remain untouched.
    """
    if not styles_xml:
        return set(), "0"
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    try:
        root = ET.fromstring(styles_xml)
    except Exception:
        return set(), "0"

    custom_formats: dict[int, str] = {}
    for node in root.findall("m:numFmts/m:numFmt", ns):
        try:
            custom_formats[int(node.attrib.get("numFmtId", "-1"))] = str(node.attrib.get("formatCode", ""))
        except Exception:
            continue

    text_style_ids: set[str] = set()
    general_style_id: str | None = None
    xfs = root.findall("m:cellXfs/m:xf", ns)
    for idx, xf in enumerate(xfs):
        try:
            num_fmt_id = int(xf.attrib.get("numFmtId", "0") or 0)
        except Exception:
            num_fmt_id = 0
        if num_fmt_id == 0 and general_style_id is None:
            general_style_id = str(idx)
        format_code = custom_formats.get(num_fmt_id, "")
        normalized_code = re.sub(r"\s+", "", format_code).strip('"').lower()
        if num_fmt_id == 49 or normalized_code == "@":
            text_style_ids.add(str(idx))
    return text_style_ids, general_style_id or "0"


def _normalize_activity_helper_styles(
    style_by_col: dict[int, str],
    styles_xml: bytes | None,
    helper_cols: Iterable[int],
) -> tuple[dict[int, str], dict[str, Any]]:
    """Keep helper template styles, replacing only Excel Text with General."""
    normalized = dict(style_by_col or {})
    text_style_ids, general_style_id = _style_format_profile(styles_xml)
    replaced: list[str] = []
    for col_idx in tuple(int(col) for col in helper_cols):
        style_id = str(normalized.get(int(col_idx), "") or "")
        if style_id in text_style_ids:
            normalized[int(col_idx)] = general_style_id
            replaced.append(_xlsx_col_letter(int(col_idx)))
    return normalized, {
        "helper_general_style_id": general_style_id,
        "helper_text_style_ids": sorted(text_style_ids),
        "helper_text_styles_replaced": replaced,
    }


def _template_row_format(inside_sheetdata: bytes, row_idx: int, width: int) -> tuple[dict[int, str], dict[int, bytes], bytes | None]:
    row_xml = _extract_raw_header_row(inside_sheetdata, row_idx)
    if not row_xml:
        return {}, {}, None
    open_m = re.match(rb'<row\b([^>]*)>', row_xml, flags=re.DOTALL)
    row_attrs = open_m.group(1) if open_m else None
    style_by_col: dict[int, str] = {}
    formula_by_col: dict[int, bytes] = {}
    # Match self-closing cells first. Otherwise an empty AA cell such as
    # <c r="AA3"/> can be mistaken for the opening tag of the following AB
    # formula cell, moving AB's formula into AA and corrupting the column layout.
    for cell_m in re.finditer(rb'<c\b([^>]*)/>|<c\b([^>]*)>(.*?)</c>', row_xml, flags=re.DOTALL):
        attrs = cell_m.group(1) or cell_m.group(2) or b''
        body = cell_m.group(3) or b''
        ref_m = re.search(rb'\br="([A-Z]+)\d+"', attrs)
        if not ref_m:
            continue
        col_letters = ref_m.group(1).decode("ascii")
        col_idx = 0
        for ch in col_letters:
            col_idx = col_idx * 26 + (ord(ch) - ord("A") + 1)
        s_m = re.search(rb'\bs="([^"]*)"', attrs)
        if s_m:
            style_by_col[col_idx] = s_m.group(1).decode("utf-8", "ignore")
        f_m = re.search(rb'(<f\b[^>]*>.*?</f>)', body, flags=re.DOTALL)
        if f_m:
            formula_by_col[col_idx] = f_m.group(1)
    if style_by_col:
        last_style = style_by_col[max(style_by_col)]
        for col in range(1, int(width or 1) + 1):
            style_by_col.setdefault(col, style_by_col.get(col - 1, last_style))
    return style_by_col, formula_by_col, row_attrs


def _extend_data_validation_sqref(data: bytes, max_row: int, max_col: int) -> bytes:
    max_row = max(DATA_START_ROW, int(max_row or DATA_START_ROW))
    def repl(match: re.Match[bytes]) -> bytes:
        raw = match.group(1).decode("utf-8", "ignore")
        parts = raw.split()
        out_parts = []
        for part in parts:
            m = re.match(r'^(\$?[A-Z]{1,3})\$?3:(\$?[A-Z]{1,3})\$?\d+$', part)
            if m:
                end_col = m.group(2).replace('$', '')
                out_parts.append(f'{m.group(1)}3:{end_col}{max_row}')
                continue
            single_m = re.match(r'^(\$?[A-Z]{1,3})\$?3$', part)
            if single_m:
                col = single_m.group(1).replace('$', '')
                out_parts.append(f'{single_m.group(1)}3:{col}{max_row}')
                continue
            out_parts.append(part)
        return b'sqref="' + ' '.join(out_parts).encode("utf-8") + b'"'
    return re.sub(rb'sqref="([^"]+)"', repl, data)


def _extend_autofilter_ref(data: bytes, header_row: int, max_col: int) -> bytes:
    if b'<autoFilter' not in data:
        return data
    ref = f'A{int(header_row)}:{_xlsx_col_letter(max(1, int(max_col)))}{int(header_row)}'.encode("utf-8")
    return re.sub(rb'<autoFilter\b[^>]*/>', b'<autoFilter ref="' + ref + b'"/>', data, count=1)


def _write_sheet_part_from_template(zout: zipfile.ZipFile, arcname: str, template_xml: bytes, header_rows: list[list[Any]], data_xml_path: Path, data_count: int, final_width: int) -> None:
    prefix, inside, suffix, _ = _split_sheet_xml(template_xml)
    max_row = DATA_START_ROW - 1 + int(data_count)
    prefix = _update_dimension_in_prefix(prefix, max_row, int(final_width))
    suffix = _extend_data_validation_sqref(suffix, max_row, final_width)
    suffix = _extend_autofilter_ref(suffix, DATA_START_ROW - 1, final_width)
    raw_row_1 = _extract_raw_header_row(inside, 1)
    raw_row_2 = _extract_raw_header_row(inside, 2)
    original_width = 0
    for row_xml in (raw_row_1, raw_row_2):
        if row_xml:
            refs = re.findall(rb'<c\b[^>]*\br="([A-Z]+)\d+"', row_xml)
            for ref in refs:
                value = 0
                for ch in ref.decode("ascii"):
                    value = value * 26 + (ord(ch) - ord("A") + 1)
                original_width = max(original_width, value)
    if original_width <= 0:
        original_width = min(_max_header_width(header_rows), final_width)
    header_rows = _pad_header_rows(header_rows, final_width)
    append_start = original_width + 1
    row1_append = header_rows[0][append_start - 1: final_width] if final_width >= append_start else []
    row2_append = header_rows[1][append_start - 1: final_width] if len(header_rows) > 1 and final_width >= append_start else []
    row1_xml = _append_header_cells(raw_row_1, 1, append_start, row1_append, final_width)
    row2_xml = _append_header_cells(raw_row_2, 2, append_start, row2_append, final_width)
    with zout.open(arcname, "w") as out:
        out.write(prefix)
        out.write(row1_xml)
        out.write(row2_xml)
        if data_xml_path.exists():
            with data_xml_path.open("rb") as src:
                shutil.copyfileobj(src, out, length=1024 * 1024)
        out.write(suffix)

def _write_template_applied_workbook(
    template_path: str | Path,
    output_path: str | Path,
    activity_header_rows: list[list[Any]],
    activity_rows_iter,
    raw_header_rows: list[list[Any]],
    raw_rows_iter,
    activity_width: int,
    raw_width: int,
) -> dict[str, int]:
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    sheet_paths = _sheet_paths_by_name_from_xlsx(template_path)
    activity_template_name = _first_existing_name(sheet_paths.keys(), ACTIVITY_SHEET_ALIASES)
    raw_template_name = _first_existing_name(sheet_paths.keys(), RAW_MATERIAL_SHEET_ALIASES)
    activity_sheet_path = sheet_paths.get(activity_template_name or "")
    raw_sheet_path = sheet_paths.get(raw_template_name or "")
    if not activity_sheet_path or not raw_sheet_path:
        raise ValueError("Raw Material Bulk Template 缺少 Activity Data 或 Raw Material 分頁。")

    with tempfile.TemporaryDirectory(prefix="cmp_m3_template_apply_") as tmp:
        tmpdir = Path(tmp)
        activity_xml_tmp = tmpdir / "activity_rows.xml"
        raw_xml_tmp = tmpdir / "raw_rows.xml"
        with zipfile.ZipFile(template_path, "r") as zpeek:
            activity_template_xml_peek = zpeek.read(activity_sheet_path)
            raw_template_xml_peek = zpeek.read(raw_sheet_path)
            styles_xml_peek = zpeek.read("xl/styles.xml") if "xl/styles.xml" in zpeek.namelist() else None
        _, activity_inside, _, _ = _split_sheet_xml(activity_template_xml_peek)
        _, raw_inside, _, _ = _split_sheet_xml(raw_template_xml_peek)
        activity_style_by_col, activity_formula_by_col, activity_row_attrs = _template_row_format(activity_inside, DATA_START_ROW, activity_width)
        helper_cols, document_type_helper_col = _activity_helper_layout(activity_header_rows, activity_width)
        activity_style_by_col, helper_style_summary = _normalize_activity_helper_styles(
            activity_style_by_col,
            styles_xml_peek,
            helper_cols,
        )
        missing_formula_cols = [col for col in helper_cols if col not in activity_formula_by_col]
        if missing_formula_cols:
            raise ValueError(
                "正式 Raw Material Bulk Template 版本不符："
                "依表頭定位的隱藏欄位必須包含完整公式，缺少："
                + ", ".join(_xlsx_col_letter(col) for col in missing_formula_cols)
            )
        raw_style_by_col, raw_formula_by_col, raw_row_attrs = _template_row_format(raw_inside, DATA_START_ROW, raw_width)
        activity_count, activity_actual_width = _spool_sheet_rows_xml(
            activity_rows_iter,
            activity_xml_tmp,
            DATA_START_ROW,
            activity_width,
            style_by_col=activity_style_by_col,
            formula_by_col=activity_formula_by_col,
            template_row_attrs=activity_row_attrs,
            emit_empty_styles=False,
            general_numeric_cols=set(helper_cols),
        )
        raw_count, raw_actual_width = _spool_sheet_rows_xml(
            raw_rows_iter,
            raw_xml_tmp,
            DATA_START_ROW,
            raw_width,
            style_by_col=raw_style_by_col,
            formula_by_col=raw_formula_by_col,
            template_row_attrs=raw_row_attrs,
            emit_empty_styles=False,
        )
        final_activity_width = max(int(activity_width), int(activity_actual_width))
        final_raw_width = max(int(raw_width), int(raw_actual_width))

        with zipfile.ZipFile(template_path, "r") as zin, zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as zout:
            for item in zin.infolist():
                name = item.filename
                if item.is_dir():
                    zout.writestr(item, b"")
                    continue
                if name == "xl/calcChain.xml":
                    continue
                if name in {activity_sheet_path, raw_sheet_path}:
                    continue
                data = zin.read(name)
                if name == "[Content_Types].xml":
                    data = _remove_calc_chain_content_types(data)
                elif name == "xl/_rels/workbook.xml.rels":
                    data = _remove_calc_chain_rels(data)
                elif name == "xl/workbook.xml":
                    data = _force_full_calc_on_load(data)
                zout.writestr(item, data)
            activity_template_xml = zin.read(activity_sheet_path)
            raw_template_xml = zin.read(raw_sheet_path)
            _write_sheet_part_from_template(zout, activity_sheet_path, activity_template_xml, activity_header_rows, activity_xml_tmp, activity_count, final_activity_width)
            _write_sheet_part_from_template(zout, raw_sheet_path, raw_template_xml, raw_header_rows, raw_xml_tmp, raw_count, final_raw_width)
    return {
        "activity_rows": int(activity_count),
        "raw_material_rows": int(raw_count),
        "activity_helper_text_styles_replaced": list(helper_style_summary.get("helper_text_styles_replaced", [])),
        "activity_helper_general_style_id": helper_style_summary.get("helper_general_style_id", "0"),
        # Backward-compatible diagnostic keys retained for callers that already read them.
        "ab_ah_text_styles_replaced": list(helper_style_summary.get("helper_text_styles_replaced", [])),
        "ab_ah_general_style_id": helper_style_summary.get("helper_general_style_id", "0"),
    }


def _build_source_col_map(
    header_rows: list[list[Any]],
    aliases: dict[str, list[str]],
    defaults: dict[str, int] | None = None,
) -> dict[str, int | None]:
    defaults = defaults or {}
    return {
        key: _find_col_in_header_values(
            header_rows,
            names,
            required=False,
            fallback_col=defaults.get(key),
            prefer_fallback_when_row2_missing=bool(defaults),
        )
        for key, names in aliases.items()
    }


def _build_exact_source_lookup(header_rows: list[list[Any]]) -> dict[str, int]:
    lookup: dict[str, int] = {}
    width = _max_header_width(header_rows)
    for col_idx in range(1, width + 1):
        for label in _header_labels(header_rows, col_idx):
            key = _norm(label)
            if key and key not in lookup:
                lookup[key] = col_idx
    return lookup


def _apply_ccl_factors_to_raw_material_bulk_final_template(
    raw_material_bulk_path: str | Path,
    ccl_mapping_path: str | Path,
    output_path: str | Path,
    raw_material_template_path: str | Path,
    progress_callback: Callable[..., None] | None = None,
    ccl_map: Dict[str, Dict[str, Any]] | None = None,
) -> Dict[str, Any]:
    """Apply CCL factors and split final upload workbooks into <=50,000 total rows.

    Third-party upload platform limits each workbook to 50,000 rows. Because the
    Raw Material Bulk template uses two header rows, each output part contains at
    most 49,998 Activity Data rows. For split outputs this function returns a ZIP
    containing one or more template-applied XLSX files.
    """
    perf_start = time.perf_counter()
    if ccl_map is None:
        ccl_map = _read_ccl_mapping(ccl_mapping_path, progress_callback=progress_callback)

    raw_material_bulk_path = Path(raw_material_bulk_path)
    raw_material_template_path = Path(raw_material_template_path)
    output_path = Path(output_path)
    if not raw_material_template_path.exists():
        raise FileNotFoundError(f"找不到 M3 最終套版用 Raw Material Bulk Template：{raw_material_template_path}")

    _emit_progress(progress_callback, 34, "讀取 M3 輕量中繼檔與最終 Bulk Template", 60, 0, None)
    src_wb = load_workbook(raw_material_bulk_path, read_only=True, data_only=False)
    tpl_wb = load_workbook(raw_material_template_path, read_only=True, data_only=False)
    matched = 0
    unmatched = 0
    written_rows = 0
    material_rows = 0
    part_summaries: list[dict[str, Any]] = []
    generated_paths: list[Path] = []
    actual_output_path = output_path
    try:
        src_activity_name = _first_existing_name(src_wb.sheetnames, ACTIVITY_SHEET_ALIASES)
        tpl_activity_name = _first_existing_name(tpl_wb.sheetnames, ACTIVITY_SHEET_ALIASES)
        tpl_raw_name = _first_existing_name(tpl_wb.sheetnames, RAW_MATERIAL_SHEET_ALIASES)
        src_raw_name = _first_existing_name(src_wb.sheetnames, RAW_MATERIAL_SHEET_ALIASES)
        if not src_activity_name:
            raise ValueError("M3 輕量中繼檔找不到 Activity Data 分頁。")
        if not tpl_activity_name:
            raise ValueError("Raw Material Bulk Template 找不到 Activity Data 分頁。")
        if not tpl_raw_name:
            raise ValueError("Raw Material Bulk Template 找不到 Raw Material 分頁。")
        src_activity_ws = src_wb[src_activity_name]
        src_raw_ws = src_wb[src_raw_name] if src_raw_name else (src_wb[src_wb.sheetnames[1]] if len(src_wb.sheetnames) > 1 else None)
        tpl_activity_ws = tpl_wb[tpl_activity_name]
        tpl_raw_ws = tpl_wb[tpl_raw_name]

        source_activity_headers = _first_header_rows(src_activity_ws, DATA_START_ROW - 1)
        source_raw_headers = _first_header_rows(src_raw_ws, DATA_START_ROW - 1) if src_raw_ws is not None else [[], []]
        target_activity_headers = _first_header_rows(tpl_activity_ws, DATA_START_ROW - 1)
        target_raw_headers = _first_header_rows(tpl_raw_ws, DATA_START_ROW - 1)

        # Add M3 factor columns only if the third-party template does not already contain them.
        factor_cols: dict[str, int] = {}
        for key, aliases, preferred in _ACTIVITY_FINAL_FACTOR_COLUMNS:
            factor_cols[key] = _ensure_header_column(target_activity_headers, aliases, preferred, "Factor")

        activity_width = _max_header_width(target_activity_headers)
        raw_width = _max_header_width(target_raw_headers)
        target_activity_cols = {key: _find_col_in_header_values(target_activity_headers, aliases, required=False) for key, aliases in _ACTIVITY_SOURCE_ALIASES.items()}
        target_raw_cols = {key: _find_col_in_header_values(target_raw_headers, aliases, required=False) for key, aliases in _RAW_SOURCE_ALIASES.items()}
        activity_helper_formula_cols_to_preserve = _activity_helper_formula_columns_to_preserve(
            tpl_activity_ws,
            target_activity_headers,
            activity_width,
        )
        source_activity_cols = _build_source_col_map(source_activity_headers, _ACTIVITY_SOURCE_ALIASES, _ACTIVITY_SOURCE_DEFAULT_COLS)
        source_raw_cols = _build_source_col_map(source_raw_headers, _RAW_SOURCE_ALIASES, _RAW_SOURCE_DEFAULT_COLS)
        source_exact = _build_exact_source_lookup(source_activity_headers)
        raw_exact = _build_exact_source_lookup(source_raw_headers)
        activity_helper_cols, _document_type_helper_col = _activity_helper_layout(
            target_activity_headers,
            activity_width,
        )
        activity_helper_cache_specs = _activity_helper_cache_specs(
            tpl_wb,
            target_activity_cols,
            factor_cols,
            activity_helper_cols,
        )
        activity_helper_cache_stats: dict[str, Any] = {
            "rows_processed": 0,
            "hits_by_column": {},
            "misses_by_column": {},
        }

        total_activity_rows = max(0, int(src_activity_ws.max_row or 0) - DATA_START_ROW + 1)
        max_data_rows_per_file = int(M3_MAX_UPLOAD_DATA_ROWS)
        if max_data_rows_per_file < 1:
            raise ValueError("M3 切檔列數設定錯誤：每檔資料列數必須大於 0。")

        # Build a compact Raw Material lookup once. Each split workbook receives
        # only the Raw Material master rows referenced by its Activity Data part.
        raw_lookup: dict[str, list[Any]] = {}
        raw_order: list[str] = []
        if src_raw_ws is not None:
            for values in src_raw_ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
                raw_key_text = _text(_row_value(values, source_raw_cols.get("raw_code")) or _row_value(values, source_raw_cols.get("raw_name")))
                if not raw_key_text:
                    continue
                out = [None] * raw_width
                for tgt_col in range(1, raw_width + 1):
                    for label in _header_labels(target_raw_headers, tgt_col):
                        src_col = raw_exact.get(_norm(label))
                        if src_col:
                            out[tgt_col - 1] = _row_value(values, src_col)
                            break
                for key, tgt_col in target_raw_cols.items():
                    src_col = source_raw_cols.get(key)
                    if tgt_col and src_col:
                        out[tgt_col - 1] = _row_value(values, src_col)
                normalized = _normalize_material_key(raw_key_text)
                if normalized and normalized not in raw_lookup:
                    raw_lookup[normalized] = out
                    raw_order.append(normalized)

        t0 = time.perf_counter()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        base_name = output_path.stem
        if base_name.lower().endswith(".xlsx"):
            base_name = Path(base_name).stem

        def make_part_name(part_idx: int, total_parts: int | None = None) -> str:
            if total_parts and total_parts > 1:
                digits = max(3, len(str(total_parts)))
                return f"{base_name}_part{part_idx:0{digits}d}_of_{total_parts:0{digits}d}.xlsx"
            if output_path.suffix.lower() == ".xlsx":
                return output_path.name
            return f"{base_name}_part{part_idx:03d}.xlsx"

        def write_part(part_idx: int, activity_chunk: list[list[Any]], material_keys: list[str], part_path: Path) -> dict[str, Any]:
            raw_chunk = [raw_lookup[key] for key in material_keys if key in raw_lookup]
            _emit_progress(
                progress_callback,
                min(94, 86 + int(min(8, written_rows / max(1, total_activity_rows) * 8))),
                f"套用正式 Bulk Template 並輸出切檔 {part_idx}",
                35,
                written_rows,
                total_activity_rows,
            )
            written_summary = _write_template_applied_workbook(
                raw_material_template_path,
                part_path,
                target_activity_headers,
                iter(activity_chunk),
                target_raw_headers,
                iter(raw_chunk),
                activity_width,
                raw_width,
            )
            return {
                "part_index": int(part_idx),
                "activity_rows": int(written_summary.get("activity_rows", len(activity_chunk))),
                "raw_material_rows": int(written_summary.get("raw_material_rows", len(raw_chunk))),
                "total_excel_rows": int(written_summary.get("activity_rows", len(activity_chunk))) + (DATA_START_ROW - 1),
            }

        with tempfile.TemporaryDirectory(prefix="cmp_m3_split_upload_") as tmp:
            tmpdir = Path(tmp)
            activity_chunk: list[list[Any]] = []
            material_keys_for_chunk: list[str] = []
            material_keys_seen: set[str] = set()
            part_idx = 0

            start = time.perf_counter()
            for idx, values in enumerate(src_activity_ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=1):
                material = _text(_row_value(values, source_activity_cols.get("raw_code")) or _row_value(values, source_activity_cols.get("raw_name")))
                if not material:
                    continue
                material_rows += 1
                normalized_material = _normalize_material_key(material)
                out = [None] * activity_width
                # First copy columns with identical labels where possible.
                for tgt_col in range(1, activity_width + 1):
                    copied = False
                    for label in _header_labels(target_activity_headers, tgt_col):
                        src_col = source_exact.get(_norm(label))
                        if src_col:
                            out[tgt_col - 1] = _row_value(values, src_col)
                            copied = True
                            break
                    if copied:
                        continue
                for key, tgt_col in target_activity_cols.items():
                    _copy_matching_value_to_target(out, tgt_col, values, source_activity_cols, key)
                if target_activity_cols.get("document_type") and not out[target_activity_cols["document_type"] - 1]:
                    out[target_activity_cols["document_type"] - 1] = "Bill of Materials (BOM)"
                if target_activity_cols.get("data_source") and not out[target_activity_cols["data_source"] - 1]:
                    out[target_activity_cols["data_source"] - 1] = "SAP"
                if target_activity_cols.get("unit") and not out[target_activity_cols["unit"] - 1]:
                    out[target_activity_cols["unit"] - 1] = _row_value(values, source_activity_cols.get("unit")) or "PC"
                transport_calc_col = target_activity_cols.get("calculate_transportation_emissions")
                if transport_calc_col:
                    out[int(transport_calc_col) - 1] = _M3_TRANSPORT_CALC_DEFAULT_VALUE
                # Country/Area is populated only when the CCL factor is matched.
                # Clear any upstream/source value first so unmatched rows remain blank.
                out[factor_cols["country_area"] - 1] = None
                item = ccl_map.get(normalized_material)
                if item:
                    out[factor_cols["country_area"] - 1] = "TBD"
                    out[factor_cols["factor_name"] - 1] = item.get("factor_name") or item.get("ccl_item") or ""
                    out[factor_cols["emission_factor"] - 1] = item.get("emission_factor")
                    out[factor_cols["factor_source"] - 1] = "Ecoinvent"
                    out[factor_cols["factor_comment"] - 1] = "CCLibrary"
                    start_date = _row_value(values, source_activity_cols.get("doc_start"))
                    out[factor_cols["enabled_date"] - 1] = start_date or out[target_activity_cols["doc_start"] - 1] if target_activity_cols.get("doc_start") else start_date
                    out[factor_cols["data_quality"] - 1] = "SECONDARY"
                    matched += 1
                else:
                    unmatched += 1

                _apply_preserved_formula_cache_cells(
                    out,
                    activity_helper_formula_cols_to_preserve,
                    activity_helper_cache_specs,
                    activity_helper_cache_stats,
                )
                activity_chunk.append(out)
                written_rows += 1
                if normalized_material and normalized_material not in material_keys_seen:
                    material_keys_seen.add(normalized_material)
                    material_keys_for_chunk.append(normalized_material)

                if progress_callback and (written_rows == 1 or written_rows % 5000 == 0):
                    elapsed = max(0.001, time.perf_counter() - start)
                    rate = written_rows / elapsed
                    remaining = int(max(1, (total_activity_rows - written_rows) / rate + 25)) if rate > 0 and total_activity_rows else 60
                    progress = 40 + int(min(45, written_rows / max(1, total_activity_rows) * 45))
                    _emit_progress(progress_callback, progress, "M3 係數對應並準備 5 萬列切檔", remaining, written_rows, total_activity_rows)

                if len(activity_chunk) >= max_data_rows_per_file:
                    part_idx += 1
                    part_path = tmpdir / f"part_{part_idx:06d}.xlsx"
                    summary = write_part(part_idx, activity_chunk, material_keys_for_chunk, part_path)
                    generated_paths.append(part_path)
                    part_summaries.append(summary)
                    activity_chunk = []
                    material_keys_for_chunk = []
                    material_keys_seen = set()

            if activity_chunk or not generated_paths:
                part_idx += 1
                part_path = tmpdir / f"part_{part_idx:06d}.xlsx"
                summary = write_part(part_idx, activity_chunk, material_keys_for_chunk, part_path)
                generated_paths.append(part_path)
                part_summaries.append(summary)

            total_parts = len(generated_paths)
            force_zip = output_path.suffix.lower() == ".zip" or total_parts > 1
            if force_zip:
                actual_output_path = output_path if output_path.suffix.lower() == ".zip" else output_path.with_suffix(".zip")
                with zipfile.ZipFile(actual_output_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as zout:
                    for idx, part_path in enumerate(generated_paths, start=1):
                        arcname = make_part_name(idx, total_parts)
                        zout.write(part_path, arcname=arcname)
                        part_summaries[idx - 1]["output_filename"] = arcname
                        try:
                            part_summaries[idx - 1]["file_size_bytes"] = part_path.stat().st_size
                            part_summaries[idx - 1]["file_size_mb"] = round(part_path.stat().st_size / 1024 / 1024, 2)
                        except OSError:
                            pass
            else:
                actual_output_path = output_path
                shutil.copyfile(generated_paths[0], actual_output_path)
                part_summaries[0]["output_filename"] = actual_output_path.name
                try:
                    part_summaries[0]["file_size_bytes"] = actual_output_path.stat().st_size
                    part_summaries[0]["file_size_mb"] = round(actual_output_path.stat().st_size / 1024 / 1024, 2)
                except OSError:
                    pass
        split_write_seconds = time.perf_counter() - t0
    finally:
        try:
            src_wb.close()
        finally:
            tpl_wb.close()

    total_time = time.perf_counter() - perf_start
    _emit_progress(progress_callback, 100, "CCL 係數對應完成，已依 5 萬列限制切檔", 0, written_rows, max(total_activity_rows, written_rows))
    output_file_size_bytes = None
    output_file_size_mb = None
    try:
        output_file_size_bytes = actual_output_path.stat().st_size
        output_file_size_mb = round(output_file_size_bytes / 1024 / 1024, 2)
    except OSError:
        pass
    return {
        "output_filename": actual_output_path.name,
        "download_url": f"/download/{actual_output_path.name}",
        "ccl_mapping_rows": len(ccl_map),
        "matched_rows": int(matched),
        "unmatched_rows": int(unmatched),
        "written_rows": int(written_rows),
        "total_rows": int(material_rows),
        "activity_rows": int(written_rows),
        "raw_material_rows": int(sum(part.get("raw_material_rows", 0) for part in part_summaries)),
        "performance_seconds": {"total": round(total_time, 3), "split_write": round(split_write_seconds, 3)},
        "factor_selector_version": FACTOR_SELECTOR_VERSION,
        "large_dataset_mode": True,
        "template_strategy": "M3 final output split into third-party-uploadable workbooks; each part preserves original Raw Material Bulk Template headers/sheets",
        "compact_template_write": True,
        "empty_styled_cells_omitted": True,
        "activity_helper_formula_cache": activity_helper_cache_stats,
        "activity_helper_formula_cache_columns": [_xlsx_col_letter(col) for col in activity_helper_cols],
        "calculate_transportation_emissions_column": target_activity_cols.get("calculate_transportation_emissions"),
        "calculate_transportation_emissions_value": _M3_TRANSPORT_CALC_DEFAULT_VALUE,
        "final_template_filename": raw_material_template_path.name,
        "split_enabled": True,
        "split_reason": "third-party upload row limit",
        "max_upload_total_rows_per_file": int(M3_MAX_UPLOAD_TOTAL_ROWS),
        "max_activity_data_rows_per_file": int(M3_MAX_UPLOAD_DATA_ROWS),
        "split_file_count": int(len(part_summaries)),
        "split_files": part_summaries,
        "output_file_size_bytes": output_file_size_bytes,
        "output_file_size_mb": output_file_size_mb,
    }

def apply_ccl_factors_to_raw_material_bulk(
    raw_material_bulk_path: str | Path,
    ccl_mapping_path: str | Path,
    output_path: str | Path,
    progress_callback: Callable[..., None] | None = None,
    ccl_map: Dict[str, Dict[str, Any]] | None = None,
    raw_material_template_path: str | Path | None = None,
) -> Dict[str, Any]:
    """Fill Module 3 CCL factor fields into a Module 2 raw-material bulk workbook.

    Large Dataset Mode avoids copying/loading the full workbook template in normal
    openpyxl mode. It reads the input workbook in read-only mode and writes a new
    workbook in write-only mode, preserving the formal sheet names and required
    raw-material bulk columns while appending missing factor columns when M2B/M2C
    produced a lightweight bulk file.
    """
    perf_start = time.perf_counter()
    perf: dict[str, float] = {}
    raw_material_bulk_path = Path(raw_material_bulk_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if raw_material_template_path:
        return _apply_ccl_factors_to_raw_material_bulk_final_template(
            raw_material_bulk_path=raw_material_bulk_path,
            ccl_mapping_path=ccl_mapping_path,
            output_path=output_path,
            raw_material_template_path=raw_material_template_path,
            progress_callback=progress_callback,
            ccl_map=ccl_map,
        )

    t0 = time.perf_counter()
    if ccl_map is None:
        ccl_map = _read_ccl_mapping(ccl_mapping_path, progress_callback=progress_callback)
    perf["read_ccl_and_build_dict"] = time.perf_counter() - t0

    t0 = time.perf_counter()
    _emit_progress(progress_callback, 34, "以低記憶體模式開啟 raw material bulk", 60, 0, None)
    src_wb = load_workbook(raw_material_bulk_path, read_only=True, data_only=False)
    perf["open_readonly_workbook"] = time.perf_counter() - t0
    if ACTIVITY_SHEET_NAME not in src_wb.sheetnames:
        src_wb.close()
        raise ValueError(f"找不到分頁：{ACTIVITY_SHEET_NAME}")

    out_wb = Workbook(write_only=True)

    matched = 0
    unmatched = 0
    written_rows = 0
    non_empty_material_rows = 0
    copied_other_rows = 0
    total_activity_rows = 0

    try:
        src_ws = src_wb[ACTIVITY_SHEET_NAME]
        total_activity_rows = max(0, int(src_ws.max_row or 0) - DATA_START_ROW + 1)
        header_rows = _first_header_rows(src_ws, DATA_START_ROW - 1)

        cols = {
            "material": _find_col_from_header_rows(header_rows, ["Raw Material Code", "Raw Material Number", "Material", "Material Number", "原物料代碼", "料號"]),
            "doc_start": _find_col_from_header_rows(header_rows, ["Doc. Start Date", "Document Start Date", "開始日期"], required=False),
            "factor_name": _ensure_output_col(header_rows, ["Factor Name", "Emission Factor Name", "係數名稱"], "Factor Name"),
            "emission_factor": _ensure_output_col(header_rows, ["Emission Factor", "Carbon Factor", "碳係數"], "Emission Factor"),
            "factor_source": _ensure_output_col(header_rows, ["Factor Source", "Emission Factor Source", "係數來源"], "Factor Source"),
            "factor_comment": _ensure_output_col(header_rows, ["Factor Comment", "Emission Factor Comment", "係數備註"], "Factor Comment"),
            "country": _ensure_output_col(header_rows, ["Country/Area", "Country Area", "Country", "Area", "國家地區"], "Country/Area"),
            "enabled_date": _ensure_output_col(header_rows, ["Enabled Date", "Effective Date", "啟用日期"], "Enabled Date"),
            "data_quality": _ensure_output_col(header_rows, ["Data Quality", "資料品質"], "Data Quality"),
        }
        output_width = max(max(len(r) for r in header_rows), max(c for c in cols.values() if c))

        _emit_progress(progress_callback, 38, "建立正式 Raw Material Bulk 欄位結構", 45, 0, total_activity_rows)
        out_ws = out_wb.create_sheet(title=ACTIVITY_SHEET_NAME)
        for hrow in header_rows:
            out_ws.append(_pad_row(list(hrow), output_width))

        t0 = time.perf_counter()
        for idx, values in enumerate(src_ws.iter_rows(min_row=DATA_START_ROW, max_row=src_ws.max_row, values_only=True), start=1):
            row_values = _pad_row(list(values or []), output_width)
            material = _text(row_values[cols["material"] - 1] if len(row_values) >= cols["material"] else None)
            if material:
                non_empty_material_rows += 1
                # Country/Area is populated only when the CCL factor is matched.
                # Clear any upstream/source value first so unmatched rows remain blank.
                row_values[cols["country"] - 1] = None
                item = ccl_map.get(_normalize_material_key(material))
                if item:
                    row_values[cols["country"] - 1] = "TBD"
                    row_values[cols["factor_name"] - 1] = item.get("factor_name") or item.get("ccl_item") or ""
                    row_values[cols["emission_factor"] - 1] = item.get("emission_factor")
                    row_values[cols["factor_source"] - 1] = "Ecoinvent"
                    row_values[cols["factor_comment"] - 1] = "CCLibrary"
                    row_values[cols["enabled_date"] - 1] = row_values[cols["doc_start"] - 1] if cols.get("doc_start") else None
                    row_values[cols["data_quality"] - 1] = "SECONDARY"
                    matched += 1
                    written_rows += 1
                else:
                    unmatched += 1
            out_ws.append(row_values)
            if idx == 1 or idx % 1000 == 0:
                elapsed = max(0.001, time.perf_counter() - t0)
                rate = idx / elapsed
                remaining = int(max(1, (total_activity_rows - idx) / rate + 15)) if rate > 0 and total_activity_rows else 60
                progress = 40 + int(min(48, idx / max(1, total_activity_rows) * 48))
                _emit_progress(
                    progress_callback,
                    progress,
                    "比對原物料並串流寫入 CCL 係數欄位",
                    remaining,
                    idx,
                    total_activity_rows,
                )
        perf["stream_map_and_write_activity"] = time.perf_counter() - t0

        t0 = time.perf_counter()
        _emit_progress(progress_callback, 89, "複製 Raw Material Bulk 其他分頁", 20, total_activity_rows, total_activity_rows)
        for sheet_name in src_wb.sheetnames:
            if sheet_name == ACTIVITY_SHEET_NAME:
                continue
            copied_other_rows += _copy_non_activity_sheet_streaming(src_wb[sheet_name], out_wb)
        perf["copy_other_sheets"] = time.perf_counter() - t0

        t0 = time.perf_counter()
        _emit_progress(progress_callback, 94, "儲存已填入係數的正式 Bulk 檔", 15, total_activity_rows, total_activity_rows)
        out_wb.save(output_path)
        perf["save_writeonly_workbook"] = time.perf_counter() - t0
    finally:
        src_wb.close()

    total_time = time.perf_counter() - perf_start
    perf["total"] = total_time
    _emit_progress(progress_callback, 100, "CCL 係數對應完成", 0, total_activity_rows, total_activity_rows)

    return {
        "output_filename": output_path.name,
        "download_url": f"/download/{output_path.name}",
        "ccl_mapping_rows": len(ccl_map),
        "matched_rows": matched,
        "unmatched_rows": unmatched,
        "written_rows": written_rows,
        "total_rows": non_empty_material_rows,
        "activity_rows": total_activity_rows,
        "copied_other_sheet_rows": copied_other_rows,
        "performance_seconds": {k: round(v, 3) for k, v in perf.items()},
        "factor_selector_version": FACTOR_SELECTOR_VERSION,
        "large_dataset_mode": True,
        "template_strategy": "write_only workbook; preserve formal sheet names/headers; append missing factor columns; no full-template load",
    }


def _short_m3_bulk_output_base_name(original_name: str) -> str:
    """Build a concise, readable M3A output basename from an M2B/M2C workbook.

    Example:
    supplier_mapped_raw_material_activity_data_bulk_中國石碣廠-IPS_3a7a714e15_e17fd5fe2b.xlsx
    -> raw_material_bulk_中國石碣廠-IPS_e17fd5fe

    The first upstream token is removed and the final job token is shortened to
    eight characters. The site / BU label remains intact for user recognition.
    """
    stem = Path(original_name).stem
    lowered = stem.lower()
    prefixes = (
        "factor_filled_supplier_mapped_raw_material_activity_data_bulk_",
        "factor_filled_supplier_mapped_raw_material_bulk_",
        "supplier_mapped_raw_material_activity_data_bulk_",
        "supplier_mapped_raw_material_bulk_",
        "factor_filled_raw_material_activity_data_bulk_",
        "raw_material_activity_data_bulk_",
        "factor_filled_raw_material_bulk_",
        "raw_material_bulk_",
    )
    remainder = stem
    for prefix in prefixes:
        if lowered.startswith(prefix):
            remainder = stem[len(prefix):]
            break

    # M2C names normally end with two hexadecimal tokens:
    # <site-BU>_<module2b-token>_<module2c-token>. Keep only the final token.
    two_tokens = re.match(r"^(?P<label>.+)_(?P<upstream>[0-9a-fA-F]{8,})_(?P<job>[0-9a-fA-F]{8,})$", remainder)
    if two_tokens:
        label = two_tokens.group("label").strip(" _-")
        short_token = two_tokens.group("job")[:8].lower()
        return f"raw_material_bulk_{label}_{short_token}"

    # M2B input may contain only one token. Keep a shortened form of that token.
    one_token = re.match(r"^(?P<label>.+)_(?P<job>[0-9a-fA-F]{8,})$", remainder)
    if one_token:
        label = one_token.group("label").strip(" _-")
        short_token = one_token.group("job")[:8].lower()
        return f"raw_material_bulk_{label}_{short_token}"

    cleaned = remainder.strip(" _-") or "output"
    return f"raw_material_bulk_{cleaned}"


def _is_raw_material_bulk_zip_member(filename: str) -> bool:
    """Return True only for Raw Material Bulk workbooks inside Module 2 ZIP packages.

    Module 2C packages may also contain supplier_bulk_create workbooks. Those are
    not Raw Material Activity Data bulk files and must not be filled with factors.
    """
    name = Path(filename).name.lower()
    if name.startswith("~$") or not name.endswith((".xlsx", ".xlsm", ".xls")):
        return False
    if "supplier_bulk" in name or "supplier_create" in name or "supplier-bulk" in name:
        return False
    raw_tokens = (
        "raw_material",
        "raw-material",
        "raw materials",
        "raw_materials",
        "activity_data_bulk",
        "activity data bulk",
    )
    return any(token in name for token in raw_tokens)


def apply_ccl_factors_to_raw_material_bulk_package(
    raw_material_bulk_path: str | Path,
    ccl_mapping_path: str | Path,
    output_path: str | Path,
    progress_callback: Callable[..., None] | None = None,
    raw_material_template_path: str | Path | None = None,
) -> Dict[str, Any]:
    """Apply CCL factors to a Module 2 raw-material bulk Excel or ZIP package.

    ZIP input is processed one workbook at a time. Each workbook is streamed from
    the extracted file to a write-only XLSX, then immediately added to the output
    ZIP. Output files are not accumulated in memory.
    """
    raw_material_bulk_path = Path(raw_material_bulk_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if raw_material_bulk_path.suffix.lower() != ".zip":
        return apply_ccl_factors_to_raw_material_bulk(
            raw_material_bulk_path,
            ccl_mapping_path,
            output_path,
            progress_callback=progress_callback,
            raw_material_template_path=raw_material_template_path,
        )

    _emit_progress(progress_callback, 2, "讀取 Module 2 ZIP 內原物料 Bulk 檔案", 60, 0, None)
    with zipfile.ZipFile(raw_material_bulk_path, "r") as zin:
        all_excel_members = [
            info for info in zin.infolist()
            if not info.is_dir()
            and not Path(info.filename).name.startswith("~$")
            and Path(info.filename).suffix.lower() in {".xlsx", ".xlsm", ".xls"}
        ]
        excel_members = [info for info in all_excel_members if _is_raw_material_bulk_zip_member(info.filename)]
        skipped_excel_members = [Path(info.filename).name for info in all_excel_members if info not in excel_members]
        if not excel_members:
            raise ValueError("Module 2 ZIP 內找不到原物料 Bulk Excel 檔案；已排除 supplier_bulk_create 等非原物料 Bulk 檔。")

        ccl_map = _read_ccl_mapping(ccl_mapping_path, progress_callback=progress_callback)
        totals = {
            "ccl_mapping_rows": len(ccl_map),
            "matched_rows": 0,
            "unmatched_rows": 0,
            "written_rows": 0,
            "total_rows": 0,
            "activity_rows": 0,
        }
        processed_files: list[dict[str, Any]] = []

        with tempfile.TemporaryDirectory(prefix="cmp_module3_zip_") as tmpdir:
            tmpdir_path = Path(tmpdir)
            total_files = len(excel_members)
            with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as zout:
                for file_idx, info in enumerate(excel_members, start=1):
                    original_name = Path(info.filename).name
                    input_file = tmpdir_path / f"input_{file_idx}_{original_name}"
                    concise_base_name = _short_m3_bulk_output_base_name(original_name)
                    filled_name = f"{concise_base_name}.xlsx"
                    # Use a temporary ZIP per input workbook so large single workbooks
                    # can be split into multiple XLSX parts, then flatten those parts
                    # into the final M3 package. The temporary ZIP basename controls
                    # the final split workbook names inside the delivered ZIP.
                    filled_file = tmpdir_path / (f"{concise_base_name}.zip" if raw_material_template_path else f"{concise_base_name}.xlsx")
                    with zin.open(info, "r") as src, input_file.open("wb") as dst:
                        shutil.copyfileobj(src, dst, length=1024 * 1024)

                    base_pct = 10 + int((file_idx - 1) / max(1, total_files) * 80)
                    _emit_progress(progress_callback, base_pct, f"處理 ZIP 內第 {file_idx}/{total_files} 個 Bulk：{original_name}", 60, 0, None)

                    def nested_progress(p: int, step: str, remaining_seconds: int | None = None, *, processed_rows: int | None = None, total_rows: int | None = None) -> None:
                        # Map each file's internal 34-100% progress into its share of 10-90% total progress.
                        file_span = 80 / max(1, total_files)
                        normalized = max(0, min(1, (int(p) - 34) / 66)) if p >= 34 else 0
                        package_progress = int(10 + ((file_idx - 1) + normalized) * file_span)
                        display_step = f"{step}: {original_name}"
                        _emit_progress(progress_callback, package_progress, display_step, remaining_seconds, processed_rows, total_rows)

                    summary = apply_ccl_factors_to_raw_material_bulk(
                        input_file,
                        ccl_mapping_path,
                        filled_file,
                        progress_callback=nested_progress,
                        ccl_map=ccl_map,
                        raw_material_template_path=raw_material_template_path,
                    )
                    flattened_outputs: list[str] = []
                    if raw_material_template_path and filled_file.exists() and filled_file.suffix.lower() == ".zip" and zipfile.is_zipfile(filled_file):
                        with zipfile.ZipFile(filled_file, "r") as nested_zip:
                            for nested_info in nested_zip.infolist():
                                if nested_info.is_dir():
                                    continue
                                nested_name = Path(nested_info.filename).name
                                arcname = nested_name
                                with nested_zip.open(nested_info, "r") as src:
                                    zout.writestr(arcname, src.read())
                                flattened_outputs.append(arcname)
                    elif filled_file.exists():
                        zout.write(filled_file, arcname=filled_name)
                        flattened_outputs.append(filled_name)

                    processed_files.append({
                        "filename": original_name,
                        "output_filename": flattened_outputs[0] if len(flattened_outputs) == 1 else "",
                        "output_files": flattened_outputs,
                        "split_file_count": summary.get("split_file_count", len(flattened_outputs) or 1),
                        "split_files": summary.get("split_files", []),
                        "matched_rows": summary.get("matched_rows", 0),
                        "unmatched_rows": summary.get("unmatched_rows", 0),
                        "written_rows": summary.get("written_rows", 0),
                        "total_rows": summary.get("total_rows", 0),
                        "activity_rows": summary.get("activity_rows", 0),
                    })
                    for key in ["matched_rows", "unmatched_rows", "written_rows", "total_rows", "activity_rows"]:
                        totals[key] += int(summary.get(key, 0) or 0)

    _emit_progress(progress_callback, 100, "ZIP 內全部 Bulk 係數對應完成", 0, totals.get("activity_rows", 0), totals.get("activity_rows", 0))
    return {
        "output_filename": output_path.name,
        "download_url": f"/download/{output_path.name}",
        "input_package_filename": raw_material_bulk_path.name,
        "processed_file_count": len(processed_files),
        "processed_files": processed_files,
        "split_enabled": True,
        "split_reason": "third-party upload row limit",
        "max_upload_total_rows_per_file": int(M3_MAX_UPLOAD_TOTAL_ROWS),
        "max_activity_data_rows_per_file": int(M3_MAX_UPLOAD_DATA_ROWS),
        "split_file_count": int(sum(int(item.get("split_file_count", 0) or 0) for item in processed_files)),
        "skipped_non_raw_material_bulk_files": skipped_excel_members,
        "raw_material_bulk_file_filter": "include raw_material/activity_data_bulk workbooks; exclude supplier_bulk_create workbooks",
        "factor_selector_version": FACTOR_SELECTOR_VERSION,
        "large_dataset_mode": True,
        "template_strategy": "M2B/M2C/M3 lightweight intermediates; final M3 applies original Raw Material Bulk Template via compact streaming OpenXML package writer" if raw_material_template_path else "write_only workbook; formal bulk columns; no full-template load",
        "compact_template_write": bool(raw_material_template_path),
        "final_template_filename": Path(raw_material_template_path).name if raw_material_template_path else "",
        **totals,
    }


def _read_factor_map_from_filled_bulk(
    factor_filled_bulk_path: str | Path,
    progress_callback: Callable[..., None] | None = None,
) -> dict[str, Dict[str, Any]]:
    """Build a compact Material -> factor map from an already completed M3 file.

    M3A uses this map to apply a newly downloaded official Raw Material template
    without re-reading the original CCL mapping file or re-running M3 matching.
    """
    factor_filled_bulk_path = Path(factor_filled_bulk_path)
    _emit_progress(progress_callback, 8, "讀取 M3 係數對應完成檔", 30, 0, None)
    wb = load_workbook(factor_filled_bulk_path, read_only=True, data_only=False)
    try:
        activity_name = _first_existing_name(wb.sheetnames, ACTIVITY_SHEET_ALIASES)
        if not activity_name:
            raise ValueError("M3 係數對應完成檔找不到 Activity Data 分頁。")
        ws = wb[activity_name]
        headers = _first_header_rows(ws, DATA_START_ROW - 1)
        material_col = _find_col_from_header_rows(
            headers,
            ["Raw Material Code", "Raw Material Number", "Material", "Material Number", "原物料代碼", "料號"],
        )
        factor_name_col = _find_col_from_header_rows(headers, ["Factor Name", "Emission Factor Name", "CCL Item", "係數名稱"], required=False)
        factor_col = _find_col_from_header_rows(headers, ["Emission Factor", "Carbon Factor", "碳係數"], required=False)
        mapping: dict[str, Dict[str, Any]] = {}
        total_rows = max(0, int(ws.max_row or 0) - DATA_START_ROW + 1)
        for idx, values in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=1):
            material = _text(_row_value(values, material_col))
            if not material:
                continue
            factor_name = _text(_row_value(values, factor_name_col)) if factor_name_col else ""
            factor_value = _row_value(values, factor_col) if factor_col else None
            if not factor_name and factor_value in (None, ""):
                continue
            normalized = _normalize_material_key(material)
            if normalized and normalized not in mapping:
                mapping[normalized] = {
                    "material": material,
                    "ccl_item": factor_name,
                    "factor_name": factor_name,
                    "emission_factor": factor_value,
                    "unit": "",
                }
            if idx == 1 or idx % 5000 == 0:
                progress = 8 + int(min(12, idx / max(1, total_rows) * 12))
                _emit_progress(progress_callback, progress, "建立 M3 已完成係數索引", 25, idx, total_rows)
        return mapping
    finally:
        wb.close()


def apply_final_template_to_factor_filled_bulk(
    factor_filled_bulk_path: str | Path,
    raw_material_template_path: str | Path,
    output_path: str | Path,
    progress_callback: Callable[..., None] | None = None,
) -> Dict[str, Any]:
    """Apply an official third-party template to one M3 factor-filled workbook."""
    factor_filled_bulk_path = Path(factor_filled_bulk_path)
    ccl_map = _read_factor_map_from_filled_bulk(factor_filled_bulk_path, progress_callback=progress_callback)
    summary = _apply_ccl_factors_to_raw_material_bulk_final_template(
        raw_material_bulk_path=factor_filled_bulk_path,
        ccl_mapping_path=factor_filled_bulk_path,  # unused because ccl_map is supplied
        output_path=output_path,
        raw_material_template_path=raw_material_template_path,
        progress_callback=progress_callback,
        ccl_map=ccl_map,
    )
    summary["m3a_source_filename"] = factor_filled_bulk_path.name
    summary["m3a_reused_completed_factor_mapping"] = True
    return summary


def apply_final_template_to_factor_filled_package(
    factor_filled_bulk_path: str | Path,
    raw_material_template_path: str | Path,
    output_path: str | Path,
    progress_callback: Callable[..., None] | None = None,
) -> Dict[str, Any]:
    """Apply the user-uploaded official template to the latest M3 output package.

    This is the M3A stage. It preserves the official template's dropdown values,
    validations, formulas and hidden sheets, while reusing M3's completed factor
    mapping instead of asking the user to upload the CCL file again.
    """
    factor_filled_bulk_path = Path(factor_filled_bulk_path)
    raw_material_template_path = Path(raw_material_template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if not raw_material_template_path.exists():
        raise FileNotFoundError(f"找不到正式原物料批次檔 Template：{raw_material_template_path}")

    if factor_filled_bulk_path.suffix.lower() != ".zip":
        return apply_final_template_to_factor_filled_bulk(
            factor_filled_bulk_path,
            raw_material_template_path,
            output_path,
            progress_callback=progress_callback,
        )

    _emit_progress(progress_callback, 2, "讀取 M3 係數對應完成 ZIP", 60, 0, None)
    with zipfile.ZipFile(factor_filled_bulk_path, "r") as zin:
        all_excel_members = [
            info for info in zin.infolist()
            if not info.is_dir()
            and not Path(info.filename).name.startswith("~$")
            and Path(info.filename).suffix.lower() in {".xlsx", ".xlsm", ".xls"}
        ]
        excel_members = [info for info in all_excel_members if _is_raw_material_bulk_zip_member(info.filename)]
        skipped_excel_members = [Path(info.filename).name for info in all_excel_members if info not in excel_members]
        if not excel_members:
            raise ValueError("M3 ZIP 內找不到已填入係數的原物料 Bulk Excel 檔案。")

        totals = {
            "matched_rows": 0,
            "unmatched_rows": 0,
            "written_rows": 0,
            "total_rows": 0,
            "activity_rows": 0,
            "raw_material_rows": 0,
        }
        processed_files: list[dict[str, Any]] = []
        with tempfile.TemporaryDirectory(prefix="cmp_module3a_template_") as tmpdir:
            tmpdir_path = Path(tmpdir)
            total_files = len(excel_members)
            with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as zout:
                for file_idx, info in enumerate(excel_members, start=1):
                    original_name = Path(info.filename).name
                    input_file = tmpdir_path / f"input_{file_idx}_{original_name}"
                    concise_base_name = _short_m3_bulk_output_base_name(original_name)
                    filled_file = tmpdir_path / f"{concise_base_name}.zip"
                    with zin.open(info, "r") as src, input_file.open("wb") as dst:
                        shutil.copyfileobj(src, dst, length=1024 * 1024)

                    base_pct = 5 + int((file_idx - 1) / max(1, total_files) * 88)
                    _emit_progress(progress_callback, base_pct, f"M3A 套用第 {file_idx}/{total_files} 個正式 Template：{original_name}", 60, 0, None)

                    def nested_progress(p: int, step: str, remaining_seconds: int | None = None, *, processed_rows: int | None = None, total_rows: int | None = None) -> None:
                        file_span = 88 / max(1, total_files)
                        normalized = max(0, min(1, int(p) / 100))
                        package_progress = int(5 + ((file_idx - 1) + normalized) * file_span)
                        _emit_progress(progress_callback, package_progress, f"{step}: {original_name}", remaining_seconds, processed_rows, total_rows)

                    summary = apply_final_template_to_factor_filled_bulk(
                        input_file,
                        raw_material_template_path,
                        filled_file,
                        progress_callback=nested_progress,
                    )
                    flattened_outputs: list[str] = []
                    if filled_file.exists() and zipfile.is_zipfile(filled_file):
                        with zipfile.ZipFile(filled_file, "r") as nested_zip:
                            for nested_info in nested_zip.infolist():
                                if nested_info.is_dir():
                                    continue
                                arcname = Path(nested_info.filename).name
                                with nested_zip.open(nested_info, "r") as src:
                                    zout.writestr(arcname, src.read())
                                flattened_outputs.append(arcname)
                    elif filled_file.exists():
                        arcname = f"{concise_base_name}.xlsx"
                        zout.write(filled_file, arcname=arcname)
                        flattened_outputs.append(arcname)

                    processed_files.append({
                        "filename": original_name,
                        "output_filename": flattened_outputs[0] if len(flattened_outputs) == 1 else "",
                        "output_files": flattened_outputs,
                        "split_file_count": summary.get("split_file_count", len(flattened_outputs) or 1),
                        "split_files": summary.get("split_files", []),
                        "matched_rows": summary.get("matched_rows", 0),
                        "unmatched_rows": summary.get("unmatched_rows", 0),
                        "written_rows": summary.get("written_rows", 0),
                        "total_rows": summary.get("total_rows", 0),
                        "activity_rows": summary.get("activity_rows", 0),
                        "raw_material_rows": summary.get("raw_material_rows", 0),
                    })
                    for key in totals:
                        totals[key] += int(summary.get(key, 0) or 0)

    _emit_progress(progress_callback, 100, "M3A 正式原物料批次檔套版完成", 0, totals.get("activity_rows", 0), totals.get("activity_rows", 0))
    return {
        "output_filename": output_path.name,
        "download_url": f"/download/{output_path.name}",
        "input_package_filename": factor_filled_bulk_path.name,
        "final_template_filename": raw_material_template_path.name,
        "processed_file_count": len(processed_files),
        "processed_files": processed_files,
        "split_enabled": True,
        "split_reason": "third-party upload row limit",
        "max_upload_total_rows_per_file": int(M3_MAX_UPLOAD_TOTAL_ROWS),
        "max_activity_data_rows_per_file": int(M3_MAX_UPLOAD_DATA_ROWS),
        "split_file_count": int(sum(int(item.get("split_file_count", 0) or 0) for item in processed_files)),
        "skipped_non_raw_material_bulk_files": skipped_excel_members,
        "factor_selector_version": FACTOR_SELECTOR_VERSION,
        "large_dataset_mode": True,
        "template_strategy": "M3A applies user-uploaded official template to completed M3 factor-filled intermediates",
        "compact_template_write": True,
        "m3a_reused_completed_factor_mapping": True,
        **totals,
    }

def _resolve_lcia_target_column(ws) -> int:
    """Find IPCC 2021 + climate change total excl. biogenic CO2 + GWP100 without fixed column letters."""
    candidates: list[tuple[int, int]] = []
    for col in range(1, ws.max_column + 1):
        method = _text(ws.cell(1, col).value).lower()
        category = _text(ws.cell(2, col).value).lower()
        indicator = _text(ws.cell(3, col).value).lower()
        if "ipcc 2021" not in method:
            continue
        if "climate change: total (excl. biogenic co2)" not in category:
            continue
        if "global warming potential (gwp100)" not in indicator:
            continue
        score = 0
        if "no lt" not in category and "no lt" not in method:
            score += 3
        if "incl. slcfs" not in category:
            score += 2
        candidates.append((score, col))
    if not candidates:
        raise ValueError("LCIA 檔案找不到 IPCC 2021 / climate change: total (excl. biogenic CO2) / GWP100 欄位")
    return sorted(candidates, reverse=True)[0][1]



def _resolve_lcia_metadata_columns(ws) -> dict[str, int]:
    """Resolve LCIA metadata columns by row-4 headers instead of fixed column letters."""
    aliases = {
        "activity_name": ["Activity Name"],
        "reference_product_name": ["Reference Product Name", "Reference Product", "Product Name"],
        "geography": ["Geography", "Geographical representativeness"],
        "reference_product_unit": ["Reference Product Unit", "Unit"],
    }
    resolved: dict[str, int] = {}
    header_row = 4
    for key, names in aliases.items():
        alias_keys = {_norm(name) for name in names}
        for col in range(1, ws.max_column + 1):
            if _norm(ws.cell(header_row, col).value) in alias_keys:
                resolved[key] = col
                break
        if key not in resolved and key == "reference_product_name":
            resolved[key] = 4  # ecoinvent LCIA common position; used only as a fallback for keyword search
        elif key not in resolved:
            raise ValueError(f"LCIA 檔案找不到欄位：{', '.join(names)}")
    return resolved

def _process_type_mode(process_type: str | None) -> str:
    """Normalize process type filter.

    Business rule:
    - production_with_transport: Activity Name must start with ``market for``.
    - production_only: all non-``market for`` activities are treated as production only.
    - all: no process-type filtering.
    """
    value = str(process_type or "all").strip().lower()
    if value in {"market_for", "market", "production_with_transport"}:
        return "market_for"
    if value in {"production", "production_only"}:
        return "production_only"
    return "all"


def _matches_process_type(activity_name: str, process_type: str | None) -> bool:
    activity = str(activity_name or "").strip().lower()
    mode = _process_type_mode(process_type)
    if mode == "market_for":
        return activity.startswith("market for")
    if mode == "production_only":
        return not activity.startswith("market for")
    return True



def _format_emission_factor_unit(unit: Any) -> str:
    raw_unit = _text(unit).strip()
    if not raw_unit:
        return ""
    normalized = raw_unit.lower().replace(" ", "")
    if normalized.startswith("kgco2e/"):
        return raw_unit
    return f"kgCO2e / {raw_unit}"


_SOURCE_DISPLAY_NAMES = {
    "APOS": "Ecoinvent 3.12 APOS",
    "Cut-off": "Ecoinvent 3.12 Cut-off",
}

_LCIA_CACHE: dict[str, Dict[str, Any]] = {}


def _source_display_name(source: str) -> str:
    return _SOURCE_DISPLAY_NAMES.get(source, source)


def _load_lcia_cache(path: str | Path, source: str) -> Dict[str, Any]:
    """Load LCIA rows once into memory so repeated keyword searches do not re-read Excel."""
    path = Path(path)
    cache_key = str(path.resolve())
    stat = path.stat()
    cached = _LCIA_CACHE.get(cache_key)
    if cached and cached.get("mtime") == stat.st_mtime and cached.get("size") == stat.st_size:
        return cached

    wb = load_workbook(path, read_only=True, data_only=True)
    if LCIA_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"{path.name} 找不到分頁：{LCIA_SHEET_NAME}")
    ws = wb[LCIA_SHEET_NAME]
    value_col = _resolve_lcia_target_column(ws)
    meta_cols = _resolve_lcia_metadata_columns(ws)

    rows: list[Dict[str, Any]] = []
    geographies: set[str] = set()
    display_source = _source_display_name(source)
    for values in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        activity_name = _text(values[meta_cols["activity_name"] - 1] if len(values) >= meta_cols["activity_name"] else "")
        row_geography = _text(values[meta_cols["geography"] - 1] if len(values) >= meta_cols["geography"] else "")
        reference_product_name = _text(values[meta_cols["reference_product_name"] - 1] if len(values) >= meta_cols["reference_product_name"] else "")
        ref_unit = _text(values[meta_cols["reference_product_unit"] - 1] if len(values) >= meta_cols["reference_product_unit"] else "")
        factor_value = values[value_col - 1] if len(values) >= value_col else None
        if row_geography:
            geographies.add(row_geography)
        # Keyword search is intentionally limited to Activity Name and Reference Product Name only.
        activity_searchable = activity_name.lower()
        reference_searchable = reference_product_name.lower()
        searchable = f"{activity_searchable} {reference_searchable}"
        rows.append({
            "source": display_source,
            "source_key": source,
            "activity_name": activity_name,
            "geography": row_geography,
            "emission_factor": factor_value,
            "reference_product_unit": ref_unit,
            "emission_factor_unit": _format_emission_factor_unit(ref_unit),
            "reference_product_name": reference_product_name,
            "ipcc2021_gwp100": factor_value,
            "indicator": "IPCC 2021 | climate change: total (excl. biogenic CO2) | global warming potential (GWP100)",
            "_activity_lower": activity_name.lower(),
            "_activity_searchable": activity_searchable,
            "_reference_searchable": reference_searchable,
            "_searchable": searchable,
        })

    cached = {
        "mtime": stat.st_mtime,
        "size": stat.st_size,
        "source": source,
        "rows": rows,
        "geographies": geographies,
    }
    _LCIA_CACHE[cache_key] = cached
    return cached



def _keyword_matches_activity_or_reference(keyword: str, searchable: str) -> bool:
    """Match keyword against Activity Name and Reference Product Name.

    English / alphanumeric keywords are matched as complete words or complete
    phrases, so a query such as "tin" will not match "coating" merely because
    the letters appear inside another word. Non-English keywords continue to use
    a direct containment check because word boundaries are less reliable.
    """
    key = str(keyword or "").strip().lower()
    text = str(searchable or "").lower()
    if not key:
        return True
    if re.fullmatch(r"[a-z0-9][a-z0-9\s\-_/.,()+]*", key, flags=re.I):
        # Treat separators and punctuation as boundaries, but do not match inside
        # another English/alphanumeric token.
        escaped = re.escape(key)
        escaped = escaped.replace(r"\ ", r"\s+")
        pattern = rf"(?<![a-z0-9]){escaped}(?![a-z0-9])"
        return re.search(pattern, text, flags=re.I) is not None
    return key in text

def _search_lcia_file(
    path: str | Path,
    keyword: str,
    source: str,
    limit: int,
    geography: str | None = "all",
    process_type: str | None = "all",
    activity_name_keyword: str | None = "",
    reference_product_keyword: str | None = "",
) -> tuple[list[Dict[str, Any]], int]:
    cache = _load_lcia_cache(path, source)
    results: list[Dict[str, Any]] = []
    total_count = 0
    key = keyword.lower().strip()
    activity_key = str(activity_name_keyword or "").lower().strip()
    reference_key = str(reference_product_keyword or "").lower().strip()
    geography_key = str(geography or "all").strip()
    for row in cache["rows"]:
        if geography_key.lower() != "all" and row.get("geography") != geography_key:
            continue
        if not _matches_process_type(row.get("_activity_lower", ""), process_type):
            continue
        if activity_key and not _keyword_matches_activity_or_reference(activity_key, row.get("_activity_searchable", "")):
            continue
        if reference_key and not _keyword_matches_activity_or_reference(reference_key, row.get("_reference_searchable", "")):
            continue
        if not activity_key and not reference_key and not _keyword_matches_activity_or_reference(key, row.get("_searchable", "")):
            continue
        total_count += 1
        if len(results) < limit:
            clean_row = {k: v for k, v in row.items() if not k.startswith("_") and k != "source_key"}
            results.append(clean_row)
    return results, total_count



def preload_factor_libraries(apos_path: str | Path | None, cutoff_path: str | Path | None) -> Dict[str, Any]:
    """Preload APOS and Cut-off LCIA workbooks into memory at application startup.

    This keeps the first user search from paying the Excel read cost. Missing
    files are skipped so local development can still start without the databases.
    """
    loaded: list[str] = []
    skipped: list[str] = []
    errors: list[dict[str, str]] = []
    for path, source in ((apos_path, "APOS"), (cutoff_path, "Cut-off")):
        if not path:
            skipped.append(source)
            continue
        p = Path(path)
        if not p.exists():
            skipped.append(source)
            continue
        try:
            _load_lcia_cache(p, source)
            loaded.append(source)
        except Exception as exc:  # keep startup robust
            errors.append({"source": source, "message": str(exc)})
    return {
        "loaded": loaded,
        "skipped": skipped,
        "errors": errors,
        "factor_selector_version": FACTOR_SELECTOR_VERSION,
    }

def collect_factor_library_geographies(*paths: str | Path | None) -> list[str]:
    """Return the common geography filters shown in the UI.

    The database can contain many Geography values, but the factor library UI is
    intentionally limited to the common review filters requested for CMP:
    GLO, RoW, and RER.
    """
    return ["GLO", "RoW", "RER"]

def search_factor_library(
    keyword: str,
    apos_path: str | Path | None,
    cutoff_path: str | Path | None,
    limit: int = 10,
    source: str = "all",
    geography: str = "all",
    process_type: str = "all",
    page: int = 1,
    page_size: int = 10,
    activity_name_keyword: str | None = "",
    reference_product_keyword: str | None = "",
) -> Dict[str, Any]:
    keyword = str(keyword or "").strip()
    activity_name_keyword = str(activity_name_keyword or "").strip()
    reference_product_keyword = str(reference_product_keyword or "").strip()
    if len(activity_name_keyword) < 2 and len(reference_product_keyword) < 2 and len(keyword) < 2:
        raise ValueError("請至少在 Activity Name 或 Reference Product Name 輸入 2 個字元")
    page_size = int(page_size or limit or 10)
    if page_size not in {10, 20, 50}:
        page_size = 10
    page = max(1, int(page or 1))
    # Load enough rows to slice the requested page after APOS -> Cut-off priority ordering.
    fetch_limit = page * page_size
    selected_source = str(source or "all").strip().lower()
    ordered_results: list[Dict[str, Any]] = []
    total_count = 0
    if selected_source in {"all", "apos"} and apos_path and Path(apos_path).exists():
        apos_results, apos_count = _search_lcia_file(apos_path, keyword, "APOS", fetch_limit, geography, process_type, activity_name_keyword, reference_product_keyword)
        ordered_results.extend(apos_results)
        total_count += apos_count
    remaining_fetch = max(0, fetch_limit - len(ordered_results))
    if remaining_fetch > 0 and selected_source in {"all", "cut-off", "cutoff", "cut off"} and cutoff_path and Path(cutoff_path).exists():
        cutoff_results, cutoff_count = _search_lcia_file(cutoff_path, keyword, "Cut-off", remaining_fetch, geography, process_type, activity_name_keyword, reference_product_keyword)
        ordered_results.extend(cutoff_results)
        total_count += cutoff_count
    elif selected_source in {"all", "cut-off", "cutoff", "cut off"} and cutoff_path and Path(cutoff_path).exists():
        # Count Cut-off matches even when the requested page is fully occupied by APOS results.
        _, cutoff_count = _search_lcia_file(cutoff_path, keyword, "Cut-off", 0, geography, process_type, activity_name_keyword, reference_product_keyword)
        total_count += cutoff_count
    start = (page - 1) * page_size
    end = start + page_size
    results = ordered_results[start:end]
    total_pages = max(1, (total_count + page_size - 1) // page_size) if total_count else 0
    return {
        "keyword": keyword,
        "count": len(results),
        "total_count": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "priority": "APOS first, then Cut-off",
        "filters": {
            "source": source or "all",
            "geography": geography or "all",
            "process_type": process_type or "all",
            "activity_name_keyword": activity_name_keyword,
            "reference_product_keyword": reference_product_keyword,
        },
        "results": results,
        "factor_selector_version": FACTOR_SELECTOR_VERSION,
    }


# =========================================================
# Module 3 Factor Library Search - SQLite index implementation
# Version 1: keep existing UI/API, replace Excel cache search with SQLite FTS
# =========================================================
import sqlite3
from contextlib import closing

FACTOR_SELECTOR_VERSION = "CMP_MODULE3A_DYNAMIC_TEMPLATE_HEADERS_V9_20260721"
FACTOR_DB_FILENAME = "factors.db"
FACTOR_DB_SCHEMA_VERSION = "20260704_v1"


def _resolve_factor_excel_path(path: str | Path | None, source: str) -> Path | None:
    """Resolve factor workbook path, including ZIP-safe filenames such as (#U9867#U554f)."""
    if path:
        p = Path(path)
        if p.exists():
            return p
        base = p.parent if p.parent.exists() else Path(__file__).resolve().parent / "data" / "factor_library"
    else:
        base = Path(__file__).resolve().parent / "data" / "factor_library"
    if not base.exists():
        return None
    source_key = str(source or "").lower()
    patterns = ["*.xlsx", "*.xlsm", "*.xls"]
    for pattern in patterns:
        for candidate in sorted(base.glob(pattern)):
            name = candidate.name.lower()
            if source_key == "apos" and "apos" in name and "lcia" in name:
                return candidate
            if source_key in {"cut-off", "cutoff", "cut off"} and ("cut-off" in name or "cutoff" in name) and "lcia" in name:
                return candidate
    return None


def _factor_db_path(*paths: str | Path | None) -> Path:
    for path in paths:
        if path:
            parent = Path(path).parent
            if parent.exists():
                return parent / FACTOR_DB_FILENAME
    return Path(__file__).resolve().parent / "data" / "factor_library" / FACTOR_DB_FILENAME


def _excel_signature(path: Path | None) -> str:
    if not path or not path.exists():
        return "missing"
    stat = path.stat()
    return f"{path.name}|{stat.st_size}|{int(stat.st_mtime)}"


def _connect_factor_db(db_path: Path) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA temp_store=MEMORY")
    return conn


def _init_factor_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS factor_meta (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS factor_library (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT NOT NULL,
            source_key TEXT NOT NULL,
            activity_name TEXT,
            geography TEXT,
            emission_factor REAL,
            emission_factor_text TEXT,
            reference_product_unit TEXT,
            emission_factor_unit TEXT,
            reference_product_name TEXT,
            ipcc2021_gwp100 REAL,
            ipcc2021_gwp100_text TEXT,
            indicator TEXT
        );
        CREATE VIRTUAL TABLE IF NOT EXISTS factor_library_fts USING fts5(
            activity_name,
            reference_product_name,
            content='factor_library',
            content_rowid='id'
        );
        CREATE INDEX IF NOT EXISTS idx_factor_source_geo ON factor_library(source_key, geography);
        CREATE INDEX IF NOT EXISTS idx_factor_activity_lower ON factor_library(activity_name);
        """
    )
    conn.commit()


def _get_meta(conn: sqlite3.Connection, key: str) -> str | None:
    row = conn.execute("SELECT value FROM factor_meta WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else None


def _set_meta(conn: sqlite3.Connection, key: str, value: str) -> None:
    conn.execute(
        "INSERT INTO factor_meta(key, value) VALUES(?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (key, value),
    )


def _coerce_float_or_none(value: Any) -> float | None:
    number = _safe_number(value)
    return float(number) if number is not None else None


def _insert_lcia_rows_to_db(conn: sqlite3.Connection, path: Path, source: str) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    if LCIA_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"{path.name} 找不到分頁：{LCIA_SHEET_NAME}")
    ws = wb[LCIA_SHEET_NAME]
    value_col = _resolve_lcia_target_column(ws)
    meta_cols = _resolve_lcia_metadata_columns(ws)
    display_source = _source_display_name(source)
    source_key = "apos" if source.lower() == "apos" else "cut-off"
    indicator = "IPCC 2021 | climate change: total (excl. biogenic CO2) | global warming potential (GWP100)"

    batch: list[tuple[Any, ...]] = []
    inserted = 0
    for values in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        activity_name = _text(values[meta_cols["activity_name"] - 1] if len(values) >= meta_cols["activity_name"] else "")
        reference_product_name = _text(values[meta_cols["reference_product_name"] - 1] if len(values) >= meta_cols["reference_product_name"] else "")
        if not activity_name and not reference_product_name:
            continue
        row_geography = _text(values[meta_cols["geography"] - 1] if len(values) >= meta_cols["geography"] else "")
        ref_unit = _text(values[meta_cols["reference_product_unit"] - 1] if len(values) >= meta_cols["reference_product_unit"] else "")
        raw_factor = values[value_col - 1] if len(values) >= value_col else None
        factor_number = _coerce_float_or_none(raw_factor)
        factor_text = _text(raw_factor)
        batch.append((
            display_source,
            source_key,
            activity_name,
            row_geography,
            factor_number,
            factor_text,
            ref_unit,
            _format_emission_factor_unit(ref_unit),
            reference_product_name,
            factor_number,
            factor_text,
            indicator,
        ))
        if len(batch) >= 1000:
            conn.executemany(
                """
                INSERT INTO factor_library(
                    source, source_key, activity_name, geography, emission_factor, emission_factor_text,
                    reference_product_unit, emission_factor_unit, reference_product_name,
                    ipcc2021_gwp100, ipcc2021_gwp100_text, indicator
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                batch,
            )
            inserted += len(batch)
            batch.clear()
    if batch:
        conn.executemany(
            """
            INSERT INTO factor_library(
                source, source_key, activity_name, geography, emission_factor, emission_factor_text,
                reference_product_unit, emission_factor_unit, reference_product_name,
                ipcc2021_gwp100, ipcc2021_gwp100_text, indicator
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            batch,
        )
        inserted += len(batch)
    wb.close()
    return inserted


def ensure_factor_database(apos_path: str | Path | None, cutoff_path: str | Path | None, force_rebuild: bool = False) -> Dict[str, Any]:
    """Build or reuse SQLite factor index. Excel is read only when the index is missing/outdated."""
    apos = _resolve_factor_excel_path(apos_path, "APOS")
    cutoff = _resolve_factor_excel_path(cutoff_path, "Cut-off")
    db_path = _factor_db_path(apos or apos_path, cutoff or cutoff_path)
    signature = f"schema={FACTOR_DB_SCHEMA_VERSION};apos={_excel_signature(apos)};cutoff={_excel_signature(cutoff)}"

    with closing(_connect_factor_db(db_path)) as conn:
        _init_factor_db(conn)
        existing_signature = _get_meta(conn, "signature")
        if not force_rebuild and existing_signature == signature:
            total = conn.execute("SELECT COUNT(*) AS c FROM factor_library").fetchone()["c"]
            return {
                "db_path": str(db_path),
                "rebuilt": False,
                "rows": int(total),
                "apos_path": str(apos) if apos else "",
                "cutoff_path": str(cutoff) if cutoff else "",
                "factor_selector_version": FACTOR_SELECTOR_VERSION,
            }

        conn.execute("DELETE FROM factor_library_fts")
        conn.execute("DELETE FROM factor_library")
        conn.execute("DELETE FROM factor_meta")
        rows_by_source: dict[str, int] = {}
        if apos and apos.exists():
            rows_by_source["APOS"] = _insert_lcia_rows_to_db(conn, apos, "APOS")
        if cutoff and cutoff.exists():
            rows_by_source["Cut-off"] = _insert_lcia_rows_to_db(conn, cutoff, "Cut-off")
        conn.execute(
            "INSERT INTO factor_library_fts(rowid, activity_name, reference_product_name) "
            "SELECT id, activity_name, reference_product_name FROM factor_library"
        )
        _set_meta(conn, "schema_version", FACTOR_DB_SCHEMA_VERSION)
        _set_meta(conn, "signature", signature)
        _set_meta(conn, "apos_path", str(apos) if apos else "")
        _set_meta(conn, "cutoff_path", str(cutoff) if cutoff else "")
        conn.commit()
        total = conn.execute("SELECT COUNT(*) AS c FROM factor_library").fetchone()["c"]
        return {
            "db_path": str(db_path),
            "rebuilt": True,
            "rows": int(total),
            "rows_by_source": rows_by_source,
            "apos_path": str(apos) if apos else "",
            "cutoff_path": str(cutoff) if cutoff else "",
            "factor_selector_version": FACTOR_SELECTOR_VERSION,
        }


def preload_factor_libraries(apos_path: str | Path | None, cutoff_path: str | Path | None) -> Dict[str, Any]:
    """Compatibility name used by main.py startup. Builds SQLite index without keeping Excel rows in memory."""
    return ensure_factor_database(apos_path, cutoff_path, force_rebuild=False)


def collect_factor_library_geographies(*paths: str | Path | None) -> list[str]:
    return ["GLO", "RoW", "RER"]


def _fts_term_query(text: str) -> str:
    """Convert user input into a conservative FTS5 AND query."""
    raw = str(text or "").strip().lower()
    if not raw:
        return ""
    tokens = re.findall(r"[a-z0-9\u4e00-\u9fff]{2,}", raw, flags=re.I)
    return " ".join(tokens[:8])


def _sqlite_source_keys(source: str) -> list[str]:
    selected = str(source or "all").strip().lower()
    if selected == "apos":
        return ["apos"]
    if selected in {"cut-off", "cutoff", "cut off"}:
        return ["cut-off"]
    return ["apos", "cut-off"]


def _build_sql_conditions(
    source_keys: list[str],
    geography: str,
    process_type: str,
    keyword: str,
    activity_name_keyword: str,
    reference_product_keyword: str,
) -> tuple[str, list[Any], str, list[Any]]:
    where = []
    params: list[Any] = []
    fts_terms: list[str] = []

    placeholders = ",".join("?" for _ in source_keys)
    where.append(f"fl.source_key IN ({placeholders})")
    params.extend(source_keys)

    geography_key = str(geography or "all").strip()
    if geography_key.lower() != "all":
        where.append("fl.geography = ?")
        params.append(geography_key)

    mode = _process_type_mode(process_type)
    if mode == "market_for":
        where.append("LOWER(fl.activity_name) LIKE 'market for%'")
    elif mode == "production_only":
        # Updated business rule: production-only means Activity Name contains production.
        where.append("LOWER(fl.activity_name) LIKE '%production%'")

    activity_query = _fts_term_query(activity_name_keyword)
    reference_query = _fts_term_query(reference_product_keyword)
    keyword_query = _fts_term_query(keyword)
    if activity_query:
        fts_terms.append(f"activity_name : ({activity_query})")
    if reference_query:
        fts_terms.append(f"reference_product_name : ({reference_query})")
    if not activity_query and not reference_query and keyword_query:
        fts_terms.append(keyword_query)

    join = ""
    if fts_terms:
        join = "JOIN factor_library_fts fts ON fts.rowid = fl.id"
        where.append("factor_library_fts MATCH ?")
        params.append(" ".join(fts_terms))

    return " AND ".join(where), params, join, params


def _row_to_result(row: sqlite3.Row) -> Dict[str, Any]:
    factor_value = row["emission_factor"]
    if factor_value is None:
        factor_value = row["emission_factor_text"]
    return {
        "source": row["source"],
        "activity_name": row["activity_name"] or "",
        "geography": row["geography"] or "",
        "emission_factor": factor_value,
        "reference_product_unit": row["reference_product_unit"] or "",
        "emission_factor_unit": row["emission_factor_unit"] or "",
        "reference_product_name": row["reference_product_name"] or "",
        "ipcc2021_gwp100": row["ipcc2021_gwp100"] if row["ipcc2021_gwp100"] is not None else row["ipcc2021_gwp100_text"],
        "indicator": row["indicator"] or "",
    }


def search_factor_library(
    keyword: str,
    apos_path: str | Path | None,
    cutoff_path: str | Path | None,
    limit: int = 10,
    source: str = "all",
    geography: str = "all",
    process_type: str = "all",
    page: int = 1,
    page_size: int = 10,
    activity_name_keyword: str | None = "",
    reference_product_keyword: str | None = "",
) -> Dict[str, Any]:
    keyword = str(keyword or "").strip()
    activity_name_keyword = str(activity_name_keyword or "").strip()
    reference_product_keyword = str(reference_product_keyword or "").strip()
    if len(activity_name_keyword) < 2 and len(reference_product_keyword) < 2 and len(keyword) < 2:
        raise ValueError("請至少在關鍵字查詢或名稱查詢輸入 2 個字元")

    page_size = int(page_size or limit or 10)
    if page_size not in {10, 20, 50}:
        page_size = 10
    page = max(1, int(page or 1))
    offset = (page - 1) * page_size

    db_summary = ensure_factor_database(apos_path, cutoff_path, force_rebuild=False)
    db_path = Path(db_summary["db_path"])
    source_keys = _sqlite_source_keys(source)
    where_sql, params, join_sql, _ = _build_sql_conditions(
        source_keys,
        geography,
        process_type,
        keyword,
        activity_name_keyword,
        reference_product_keyword,
    )

    with closing(_connect_factor_db(db_path)) as conn:
        count_sql = f"SELECT COUNT(*) AS c FROM factor_library fl {join_sql} WHERE {where_sql}"
        total_count = int(conn.execute(count_sql, params).fetchone()["c"])
        rows = conn.execute(
            f"""
            SELECT fl.*
            FROM factor_library fl
            {join_sql}
            WHERE {where_sql}
            ORDER BY CASE fl.source_key WHEN 'apos' THEN 0 ELSE 1 END, fl.id
            LIMIT ? OFFSET ?
            """,
            params + [page_size, offset],
        ).fetchall()

    results = [_row_to_result(row) for row in rows]
    total_pages = max(1, (total_count + page_size - 1) // page_size) if total_count else 0
    return {
        "keyword": keyword,
        "count": len(results),
        "total_count": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "priority": "SQLite index: APOS first, then Cut-off",
        "filters": {
            "source": source or "all",
            "geography": geography or "all",
            "process_type": process_type or "all",
            "activity_name_keyword": activity_name_keyword,
            "reference_product_keyword": reference_product_keyword,
        },
        "results": results,
        "factor_db": {
            "rows": db_summary.get("rows"),
            "rebuilt": db_summary.get("rebuilt"),
        },
        "factor_selector_version": FACTOR_SELECTOR_VERSION,
    }
