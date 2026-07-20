from __future__ import annotations

import shutil
import tempfile
import re
import zipfile
import xml.etree.ElementTree as ET
from xml.sax.saxutils import escape as _xml_escape
from datetime import date
from pathlib import Path
from typing import Any, Dict, Callable

import pandas as pd
from openpyxl import load_workbook


SOURCE_SHEET_NAME = "Plant_Material年度產量"
ACTIVITY_SHEET_NAME = "Input Sheet Activity Data"
PRODUCTS_SHEET_NAME = "Input Sheet Products"

# Bulk template row 1 = system field key
# Bulk template row 2 = display header
# Data starts from row 3
DATA_START_ROW = 3

BULK_FORMATTER_VERSION = "CMP_BULK_V8_4_M1B_OPENXML_CELL_ORDER_FIX"


def _sanitize_filename(value: Any) -> str:
    text = str(value or "Unknown").strip()
    text = re.sub(r"[\\/:*?\"<>|]+", "_", text)
    text = re.sub(r"\s+", "_", text)
    return text[:80] or "Unknown"



def _normalize_col(value: Any) -> str:
    return str(value or "").strip().replace("\n", " ").replace("\r", " ")


def _find_column(df: pd.DataFrame, candidates: list[str]) -> str:
    normalized = {_normalize_col(c).lower(): c for c in df.columns}
    for name in candidates:
        key = _normalize_col(name).lower()
        if key in normalized:
            return normalized[key]
    raise ValueError(f"找不到必要欄位：{', '.join(candidates)}")


def _find_optional_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    try:
        return _find_column(df, candidates)
    except ValueError:
        return None

def _normalize_header(value: Any) -> str:
    return str(value or "").strip().replace("\n", " ").replace("\r", " ").lower()


def _find_excel_column(ws, candidates: list[str], header_rows: tuple[int, ...] = (1, 2)) -> int | None:
    """Find a column in an Excel template by row 1 system key or row 2 display header."""
    candidate_keys = {_normalize_header(c) for c in candidates if str(c or "").strip()}
    for row_idx in header_rows:
        for col_idx in range(1, ws.max_column + 1):
            if _normalize_header(ws.cell(row_idx, col_idx).value) in candidate_keys:
                return col_idx
    return None


def _is_wip(product_type: Any, is_wip: Any = None) -> bool:
    product_type_text = str(product_type or "").strip().upper()
    is_wip_text = str(is_wip or "").strip().upper()
    return product_type_text == "WIP" or is_wip_text in ["Y", "YES", "TRUE", "1"]


def _production_site(product_type: Any) -> str:
    """Rule Master only mode.

    Production Site should come from Step1 output / rule_master.csv.
    Do not infer Production Site from Product Type in Step2.
    """
    return ""


def _as_year(value: Any) -> int:
    if pd.isna(value):
        raise ValueError("Year 欄位有空白值")
    if hasattr(value, "year"):
        return int(value.year)
    return int(str(value).strip()[:4])


def _safe_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()



def _safe_number(value: Any) -> float:
    if pd.isna(value):
        return 0.0
    text = str(value or "").strip()
    if text.upper() in ["", "NAN", "NONE"]:
        return 0.0
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return 0.0


def _normalize_material(value: Any) -> str:
    return str(value or "").strip().upper()


def normalize_working_hour_source(value: Any) -> str:
    text = str(value or "direct").strip().lower().replace("-", "_").replace(" ", "_")
    if text in {"include_semi", "semi", "semi_finished", "include_semifinished", "rollup", "rolled_up", "total"}:
        return "include_semi"
    return "direct"




def _read_working_hour_rollup(working_hour_rollup_path: str | Path | None) -> pd.DataFrame:
    if not working_hour_rollup_path:
        return pd.DataFrame()
    path = Path(working_hour_rollup_path)
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_excel(path, sheet_name="Summary", dtype=object)
    except ValueError:
        return pd.read_excel(path, sheet_name=0, dtype=object)


def _build_total_hour_lookup_from_rollup(working_hour_rollup_path: str | Path | None) -> tuple[dict[tuple[str, str], float], dict[str, float]]:
    rollup = _read_working_hour_rollup(working_hour_rollup_path)
    if rollup.empty:
        return {}, {}
    material_col = _find_optional_column(rollup, ["Material Number", "Material", "Product"])
    site_col = _find_optional_column(rollup, ["Production Site", "production site", "生產廠區", "廠區", "廠別"])
    total_col = _find_optional_column(rollup, ["Total Annual Working Hour", "Total Working Hour", "Total Hour", "Total Hours"])
    if not material_col or not total_col:
        return {}, {}
    work = rollup.copy()
    work["_material_key"] = work[material_col].apply(_normalize_material)
    work["_site_key"] = work[site_col].apply(lambda x: str(x or "").strip().upper()) if site_col else ""
    work["_total_hour"] = work[total_col].apply(_safe_number)
    by_site: dict[tuple[str, str], float] = {}
    for _, r in work.groupby(["_site_key", "_material_key"], dropna=False, as_index=False)["_total_hour"].sum().iterrows():
        site = str(r["_site_key"] or "").strip()
        material = str(r["_material_key"] or "").strip()
        if material:
            by_site[(site, material)] = float(r["_total_hour"] or 0)
    by_mat: dict[str, float] = {}
    for _, r in work.groupby(["_material_key"], dropna=False, as_index=False)["_total_hour"].sum().iterrows():
        material = str(r["_material_key"] or "").strip()
        if material:
            by_mat[material] = float(r["_total_hour"] or 0)
    return by_site, by_mat


def _build_structural_wip_set_from_rollup(working_hour_rollup_path: str | Path | None) -> set[str]:
    """Read BOM-structural semi-finished materials from an M2A roll-up workbook.

    This is intentionally used only when M1B Working Hour Source is
    ``include_semi``.  Direct Working Hour mode remains independent of M2A.

    Primary source:
      - sheet ``Semi Hour per PC`` / column ``Semi Material``

    Backward-compatible source:
      - sheet ``Roll-up Detail`` / column ``Semi Material``

    An empty, header-only sheet means the enterprise has no structural WIP and
    therefore returns an empty set.
    """
    if not working_hour_rollup_path:
        return set()
    path = Path(working_hour_rollup_path)
    if not path.exists():
        return set()

    last_error: Exception | None = None
    for sheet_name in ("Semi Hour per PC", "Roll-up Detail"):
        try:
            frame = pd.read_excel(path, sheet_name=sheet_name, dtype=object)
        except ValueError as exc:
            last_error = exc
            continue
        if frame.empty and len(frame.columns) == 0:
            continue
        material_col = _find_optional_column(frame, ["Semi Material", "Semi-finished Material", "WIP Material"])
        if not material_col:
            continue
        return {
            key
            for key in frame[material_col].map(_normalize_material).tolist()
            if key
        }

    # Current M2A output always includes one of the structural-WIP sheets.
    # Raising here prevents silently treating every M1A material as a finished
    # product when Include Semi-finished Working Hour was explicitly selected.
    detail = f" ({last_error})" if last_error else ""
    raise ValueError(
        "Module 2A working_hour_rollup 找不到『Semi Hour per PC』或『Roll-up Detail』的 Semi Material 欄位，"
        f"無法依標準 BOM 排除結構性半品{detail}"
    )


def _read_latest_bom_structure(bom_structure_path: str | Path | None) -> pd.DataFrame:
    if not bom_structure_path:
        return pd.DataFrame()
    path = Path(bom_structure_path)
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_excel(path, sheet_name="BOM Structure", dtype=object)
    except ValueError:
        return pd.read_excel(path, sheet_name=0, dtype=object)


def _build_semi_hour_per_product(step1_df: pd.DataFrame, material_col: str, qty_col: str, direct_hour_col: str | None, bom_structure_path: str | Path | None) -> dict[str, float]:
    if not direct_hour_col:
        return {}
    bom_df = _read_latest_bom_structure(bom_structure_path)
    if bom_df.empty:
        return {}
    target_col = _find_optional_column(bom_df, ["Target Product", "target_product"])
    component_col = _find_optional_column(bom_df, ["Component", "component"])
    accumulated_qty_col = _find_optional_column(bom_df, ["Accumulated Quantity", "usage", "Quantity"])
    semi_col = _find_optional_column(bom_df, ["Is Semi-finished", "Is Semi", "semi_finished"])
    if not target_col or not component_col or not accumulated_qty_col:
        return {}
    annual = step1_df.copy()
    annual["_material_key"] = annual[material_col].apply(_normalize_material)
    annual["_annual_qty"] = annual[qty_col].apply(_safe_number)
    annual["_direct_hours"] = annual[direct_hour_col].apply(_safe_number)
    grouped = annual.groupby("_material_key", dropna=False, as_index=False).agg({"_annual_qty":"sum", "_direct_hours":"sum"})
    hour_per_pc, qty_by_material = {}, {}
    for _, r in grouped.iterrows():
        material = str(r["_material_key"] or "").strip()
        qty = float(r["_annual_qty"] or 0)
        hours = float(r["_direct_hours"] or 0)
        if material and qty != 0:
            hour_per_pc[material] = hours / qty
            qty_by_material[material] = qty
    if not hour_per_pc:
        return {}
    bom = bom_df.copy()
    bom["_target_key"] = bom[target_col].apply(_normalize_material)
    bom["_component_key"] = bom[component_col].apply(_normalize_material)
    bom["_accumulated_qty"] = bom[accumulated_qty_col].apply(_safe_number)
    if semi_col:
        bom = bom[bom[semi_col].astype(str).str.strip().str.upper().isin(["Y", "YES", "TRUE", "1"])].copy()
    else:
        bom = bom[bom["_component_key"].isin(hour_per_pc.keys())].copy()
    semi_per_pc_by_target: dict[str, float] = {}
    for _, r in bom.iterrows():
        target = str(r["_target_key"] or "").strip()
        component = str(r["_component_key"] or "").strip()
        if not target or not component:
            continue
        contribution = float(r["_accumulated_qty"] or 0) * float(hour_per_pc.get(component, 0) or 0)
        if contribution:
            semi_per_pc_by_target[target] = semi_per_pc_by_target.get(target, 0.0) + contribution
    return {target: per_pc * float(qty_by_material.get(target, 0) or 0) for target, per_pc in semi_per_pc_by_target.items() if float(qty_by_material.get(target, 0) or 0)}


def _clear_target_cells(ws, start_row: int, columns: list[int]) -> None:
    """
    只清除指定欄位的儲存格內容，不重建 sheet、不刪除列欄、不改格式。
    這樣可以保留原始 bulk template 的格式、資料驗證、凍結窗格、欄寬、隱藏分頁等設定。
    """
    max_row = ws.max_row
    for row_idx in range(start_row, max_row + 1):
        for col_idx in columns:
            ws.cell(row_idx, col_idx).value = None





def _template_header_columns(template_path: Path, sheet_name: str) -> dict[str, int]:
    paths = _xlsx_sheet_paths(template_path)
    sheet_path = paths.get(sheet_name)
    if not sheet_path:
        return {}
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with zipfile.ZipFile(template_path, "r") as zf:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall(f"{ns}si"):
                shared.append("".join(t.text or "" for t in si.iter(f"{ns}t")))
        sheet = ET.fromstring(zf.read(sheet_path))
        result: dict[str, int] = {}
        data = sheet.find(f"{ns}sheetData")
        if data is None:
            return result
        for row in data.findall(f"{ns}row")[:2]:
            for cell in row.findall(f"{ns}c"):
                ref = cell.attrib.get("r", "")
                typ = cell.attrib.get("t")
                val = ""
                v = cell.find(f"{ns}v")
                inline = cell.find(f"{ns}is")
                if typ == "s" and v is not None and v.text is not None:
                    try: val = shared[int(v.text)]
                    except (ValueError, IndexError): val = ""
                elif typ == "inlineStr" and inline is not None:
                    val = "".join(t.text or "" for t in inline.iter(f"{ns}t"))
                elif v is not None:
                    val = v.text or ""
                if val:
                    result[_normalize_header(val)] = _cell_col_index(ref)
        return result


def _find_header_col_map(header_map: dict[str, int], candidates: list[str]) -> int | None:
    for candidate in candidates:
        found = header_map.get(_normalize_header(candidate))
        if found:
            return found
    return None


def _xlsx_sheet_paths(path: str | Path) -> dict[str, str]:
    ns_main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    ns_rel = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    ns_pkg = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    with zipfile.ZipFile(path, "r") as zf:
        wb = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rels = {}
        for rel in rels_root.findall(f"{ns_pkg}Relationship"):
            rid = rel.attrib.get("Id")
            target = rel.attrib.get("Target", "")
            if not rid:
                continue
            if target.startswith("/xl/"):
                rels[rid] = target.lstrip("/")
            elif target.startswith("xl/"):
                rels[rid] = target
            else:
                rels[rid] = "xl/" + target.lstrip("/")
        result = {}
        sheets = wb.find(f"{ns_main}sheets")
        if sheets is not None:
            for sh in sheets.findall(f"{ns_main}sheet"):
                name = sh.attrib.get("name", "")
                rid = sh.attrib.get(f"{ns_rel}id")
                if name and rid in rels:
                    result[name] = rels[rid]
        return result


def _cell_col_index(ref: str) -> int:
    value = 0
    for ch in re.match(r"[A-Z]+", ref or "").group(0):
        value = value * 26 + ord(ch) - 64
    return value


def _col_letter(idx: int) -> str:
    out = ""
    n = int(idx)
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


def _replace_cell_xml(row_xml: bytes, row_idx: int, col_idx: int, kind: str, value: Any = None, formula_cache: str | None = None) -> bytes:
    ref = f"{_col_letter(col_idx)}{row_idx}"
    pattern = re.compile(rb'<c\b[^>]*?\br="' + re.escape(ref.encode()) + rb'"[^>]*?(?:/>|>.*?</c>)', re.DOTALL)
    m = pattern.search(row_xml)
    old = m.group(0) if m else b""
    style_m = re.search(rb'\bs="([^"]+)"', old)
    style = (b' s="' + style_m.group(1) + b'"') if style_m else b""
    if kind == "blank":
        new = b'<c r="' + ref.encode() + b'"' + style + b'/>'
    elif kind == "number":
        new = b'<c r="' + ref.encode() + b'"' + style + b'><v>' + str(value).encode() + b'</v></c>'
    elif kind == "text":
        txt = _xml_escape(str(value or ""), {'"':'&quot;'})
        new = b'<c r="' + ref.encode() + b'"' + style + b' t="inlineStr"><is><t xml:space="preserve">' + txt.encode('utf-8') + b'</t></is></c>'
    elif kind == "formula_cache":
        f_m = re.search(rb'(<f\b[^>]*>.*?</f>)', old, re.DOTALL)
        formula = f_m.group(1) if f_m else b""
        cache = _xml_escape(str(formula_cache or ""))
        new = b'<c r="' + ref.encode() + b'"' + style + b' t="str">' + formula + b'<v>' + cache.encode('utf-8') + b'</v></c>'
    elif kind == "formula_blank":
        f_m = re.search(rb'(<f\b[^>]*>.*?</f>)', old, re.DOTALL)
        formula = f_m.group(1) if f_m else b""
        new = b'<c r="' + ref.encode() + b'"' + style + b' t="str">' + formula + b'<v></v></c>'
    else:
        raise ValueError(kind)
    if m:
        return row_xml[:m.start()] + new + row_xml[m.end():]
    pos = row_xml.rfind(b'</row>')
    return row_xml[:pos] + new + row_xml[pos:] if pos >= 0 else row_xml


def _sort_row_cells_by_column(row_xml: bytes) -> bytes:
    """Ensure worksheet cells are serialized in ascending column order.

    Excel validates the order of <c> nodes inside each <row>. Third-party importers
    often resolve cells only by their r= coordinate and therefore tolerate unsorted
    cells, but desktop Excel repairs/removes those records.
    """
    open_m = re.match(rb'(<row\b[^>]*>)', row_xml, re.DOTALL)
    close_pos = row_xml.rfind(b'</row>')
    if not open_m or close_pos < 0:
        return row_xml
    body = row_xml[open_m.end():close_pos]
    cell_pattern = re.compile(rb'<c\b[^>]*?\br="([A-Z]+)\d+"[^>]*?(?:/>|>.*?</c>)', re.DOTALL)
    matches = list(cell_pattern.finditer(body))
    if len(matches) < 2:
        return row_xml
    # Worksheet rows in the official template contain only cell nodes plus whitespace.
    # Keep leading/trailing whitespace and serialize all cells in schema-valid order.
    prefix = body[:matches[0].start()]
    suffix = body[matches[-1].end():]
    cells = [( _cell_col_index(m.group(1).decode('ascii')), m.group(0)) for m in matches]
    cells.sort(key=lambda item: item[0])
    new_body = prefix + b''.join(cell for _, cell in cells) + suffix
    return row_xml[:open_m.end()] + new_body + row_xml[close_pos:]


def _rewrite_product_activity_sheet(template_xml: bytes, rows: list[dict[str, Any]], sheet_kind: str, labor_col: int | None, labor_unit_col: int | None) -> bytes:
    start_m = re.search(rb'<sheetData\b[^>]*>', template_xml)
    end_m = re.search(rb'</sheetData>', template_xml)
    if not start_m or not end_m:
        raise ValueError("Product Activity Template worksheet XML 缺少 sheetData")
    inside = template_xml[start_m.end():end_m.start()]
    row_pattern = re.compile(rb'<row\b[^>]*\br="(\d+)"[^>]*>.*?</row>', re.DOTALL)
    out=[]
    data_count=len(rows)
    max_template_row=2
    for m in row_pattern.finditer(inside):
        row_idx=int(m.group(1)); max_template_row=max(max_template_row,row_idx)
        row_xml=m.group(0)
        if row_idx < DATA_START_ROW:
            out.append(row_xml); continue
        item = rows[row_idx-DATA_START_ROW] if row_idx-DATA_START_ROW < data_count else None
        # Official templates already contain blank preallocated rows and formulas.
        # Leave unused rows byte-for-byte unchanged for speed and compatibility.
        if item is None:
            out.append(row_xml)
            continue
        if sheet_kind == "activity":
            visible=[1,2,3,4,5,6,7,8]
            if labor_col: visible.append(labor_col)
            if labor_unit_col: visible.append(labor_unit_col)
            if item is None:
                for c in visible: row_xml=_replace_cell_xml(row_xml,row_idx,c,"blank")
                for c in (20,21,22,23,24): row_xml=_replace_cell_xml(row_xml,row_idx,c,"formula_blank")
            else:
                year=int(item['year'])
                # Excel 1900 date system serials; 1899-12-30 matches openpyxl/Excel.
                start_serial=(date(year,1,1)-date(1899,12,30)).days
                end_serial=(date(year,12,31)-date(1899,12,30)).days
                vals={1:("text",item['product_name']),2:("number",start_serial),3:("number",end_serial),4:("text","Target Product"),5:("text",item['production_site']),6:("number",item['qty']),7:("text","SAP"),8:("blank",None)}
                if labor_col: vals[labor_col]=(("number",item['labor_hours']) if item['labor_hours'] is not None else ("blank",None))
                if labor_unit_col: vals[labor_unit_col]=(("text","hours") if item['labor_hours'] is not None else ("blank",None))
                for c,(k,v) in vals.items(): row_xml=_replace_cell_xml(row_xml,row_idx,c,k,v)
                for c,cache in {20:"FINISHED_PRODUCT",21:"SAP",22:"",23:"HOURS",24:""}.items():
                    row_xml=_replace_cell_xml(row_xml,row_idx,c,"formula_cache",formula_cache=cache)
        else:
            visible=[1,3,4,6]
            if item is None:
                for c in visible: row_xml=_replace_cell_xml(row_xml,row_idx,c,"blank")
                for c in range(22,29): row_xml=_replace_cell_xml(row_xml,row_idx,c,"formula_blank")
            else:
                for c,k,v in [(1,"text",item['product_name']),(3,"text",item['product_description']),(4,"text","Cradle-to-Gate"),(6,"text","PC")]:
                    row_xml=_replace_cell_xml(row_xml,row_idx,c,k,v)
                for c,cache in {22:"CRADLE_TO_GATE",23:"PC",24:"",25:"",26:"",27:"",28:""}.items():
                    row_xml=_replace_cell_xml(row_xml,row_idx,c,"formula_cache",formula_cache=cache)
        row_xml = _sort_row_cells_by_column(row_xml)
        out.append(row_xml)
    if data_count > max(0,max_template_row-DATA_START_ROW+1):
        raise ValueError(f"Product Activity Template 預留列數不足：需要 {data_count} 筆，Template 僅支援 {max_template_row-DATA_START_ROW+1} 筆")
    new_inside=b''.join(out)
    return template_xml[:start_m.end()] + new_inside + template_xml[end_m.start():]




def _remove_calc_chain_content_types(data: bytes) -> bytes:
    return re.sub(rb'<Override\b[^>]*PartName="/xl/calcChain\.xml"[^>]*/>', b"", data)


def _remove_calc_chain_rels(data: bytes) -> bytes:
    return re.sub(rb'<Relationship\b[^>]*(?:calcChain|calcchain)[^>]*/>', b"", data)


def _force_full_calc_on_load(data: bytes) -> bytes:
    """移除舊計算鏈後，要求 Excel 開啟時重新計算保留公式。"""
    calc_attrs = b' calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"'
    if b"<calcPr" not in data:
        return data.replace(b"</workbook>", b"<calcPr" + calc_attrs + b"/></workbook>", 1)

    def repl(match: re.Match[bytes]) -> bytes:
        tag = match.group(0)
        for attr in (b"calcMode", b"fullCalcOnLoad", b"forceFullCalc"):
            tag = re.sub(attr + rb'="[^"]*"', b"", tag)
        if tag.endswith(b"/>"):
            return tag[:-2].rstrip() + calc_attrs + b"/>"
        return tag[:-1].rstrip() + calc_attrs + b">"

    return re.sub(rb"<calcPr\b[^>]*/?>", repl, data, count=1)

def _write_product_activity_openxml(template_path: Path, output_path: Path, rows: list[dict[str, Any]], labor_col: int | None, labor_unit_col: int | None) -> None:
    paths=_xlsx_sheet_paths(template_path)
    activity_path=paths.get(ACTIVITY_SHEET_NAME); products_path=paths.get(PRODUCTS_SHEET_NAME)
    if not activity_path or not products_path:
        raise ValueError("Product Activity Bulk Template 缺少必要分頁")
    with zipfile.ZipFile(template_path,'r') as zin, zipfile.ZipFile(output_path,'w',compression=zipfile.ZIP_DEFLATED,compresslevel=6) as zout:
        for item in zin.infolist():
            name = item.filename
            if item.is_dir():
                zout.writestr(item,b'')
                continue
            # Products!A3 原本可能是動態陣列公式；M1B 改寫為實際值後，
            # Template 舊 calcChain 會與工作表公式狀態不一致，Excel 可能判定檔案損壞。
            if name == 'xl/calcChain.xml':
                continue
            data=zin.read(name)
            if name==activity_path:
                data=_rewrite_product_activity_sheet(data,rows,'activity',labor_col,labor_unit_col)
            elif name==products_path:
                data=_rewrite_product_activity_sheet(data,rows,'products',labor_col,labor_unit_col)
            elif name == '[Content_Types].xml':
                data = _remove_calc_chain_content_types(data)
            elif name == 'xl/_rels/workbook.xml.rels':
                data = _remove_calc_chain_rels(data)
            elif name == 'xl/workbook.xml':
                data = _force_full_calc_on_load(data)
            zout.writestr(item,data)


def generate_product_activity_bulk_file(
    step1_output_path: str | Path,
    bulk_template_path: str | Path,
    output_path: str | Path,
    working_hour_source: str = "direct",
    bom_structure_path: str | Path | None = None,
    working_hour_rollup_path: str | Path | None = None,
    progress_callback: Callable[..., None] | None = None,
) -> Dict[str, Any]:
    """
    方案 1：直接複製原始 Bulk Template，再只覆蓋指定分頁的指定儲存格內容。

    不重新建立 Workbook。
    不新增 / 刪除分頁。
    不破壞原始 template 的樣式、欄寬、資料驗證、隱藏分頁與公式。
    原本 bulk template 的下拉選單會保留；前提是 template 的 Data Validation
    原本就涵蓋寫入的列數範圍。
    """

    step1_output_path = Path(step1_output_path)
    bulk_template_path = Path(bulk_template_path)
    output_path = Path(output_path)

    def report(step: str, processed: int = 0, total: int = 0, progress: int = 0, **extra: Any) -> None:
        if progress_callback is None:
            return
        progress_callback(
            step,
            processed=int(processed or 0),
            total=int(total or 0),
            progress=max(0, min(100, int(progress or 0))),
            **extra,
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)

    df = pd.read_excel(step1_output_path, sheet_name=SOURCE_SHEET_NAME, dtype=object)
    source_total_rows = int(len(df))
    report("讀取產品活動來源資料", source_total_rows, source_total_rows, 12)

    # Production Site is optional in old Step1 outputs.
    # If present, Step2 writes it into Activity Data and can split files by site.
    production_site_col = _find_optional_column(
        df,
        ["Production Site", "production site", "生產廠區", "廠區", "廠別"],
    )

    year_col = _find_column(df, ["Year"])
    material_col = _find_column(df, ["Material Number"])
    material_desc_col = _find_optional_column(df, ["Material description", "Material Description", "產品描述", "品名"])
    product_type_col = _find_column(df, ["產品類型", "Product Type"])
    qty_col = _find_column(df, ["年度生產量", "Annual Quantity", "Delivered quantity"])
    labor_hours_col = _find_optional_column(
        df,
        ["年度總工時", "Total working hours", "Selected Hours", "Total Hours", "Working Hours"],
    )

    working_hour_source = normalize_working_hour_source(working_hour_source)
    semi_hour_by_material: dict[str, float] = {}
    rollup_by_site_material: dict[tuple[str, str], float] = {}
    rollup_by_material: dict[str, float] = {}
    structural_wip_materials: set[str] = set()
    if working_hour_source == "include_semi":
        if working_hour_rollup_path is not None and Path(working_hour_rollup_path).exists():
            rollup_by_site_material, rollup_by_material = _build_total_hour_lookup_from_rollup(working_hour_rollup_path)
            structural_wip_materials = _build_structural_wip_set_from_rollup(working_hour_rollup_path)
        elif bom_structure_path is not None and Path(bom_structure_path).exists():
            semi_hour_by_material = _build_semi_hour_per_product(
                step1_df=df,
                material_col=material_col,
                qty_col=qty_col,
                direct_hour_col=labor_hours_col,
                bom_structure_path=bom_structure_path,
            )
        else:
            raise ValueError("No Working Hour Roll-up result found. Please complete Module 2 → BOM Expansion with Step 1 Output first, then return to Step 2 to generate Product Activity Data Bulk.")

    is_wip_col = None
    for candidate in ["Is_WIP", "Is WIP", "WIP"]:
        try:
            is_wip_col = _find_column(df, [candidate])
            break
        except ValueError:
            pass

    # 直接從 OpenXML 表頭取得欄位位置；不載入或重新儲存整本 Workbook。
    activity_header_map = _template_header_columns(bulk_template_path, ACTIVITY_SHEET_NAME)
    product_header_map = _template_header_columns(bulk_template_path, PRODUCTS_SHEET_NAME)
    if not activity_header_map or not product_header_map:
        raise ValueError("Product Activity Bulk Template 缺少必要分頁或表頭")
    activity_labor_hours_col = _find_header_col_map(activity_header_map, [
        "Working Hour (optional)", "Working Hours (optional)", "Working Hour", "Working Hours",
        "年度總工時", "Total working hours", "Total Hours", "Labor Hours", "生產工時", "工時", "Hours"
    ])
    activity_labor_hours_unit_col = _find_header_col_map(activity_header_map, [
        "Working Hours Unit (optional)", "Working Hour Unit (optional)", "Working Hours Unit", "Working Hour Unit",
        "工時單位", "生產工時單位", "Hours Unit", "Hour Unit"
    ])

    excluded_wip_rows = 0
    excluded_structural_wip_rows = 0
    skipped_blank_rows = 0
    excluded_zero_labor_hour_rows = 0
    source_valid_rows = 0
    duplicated_product_rows_merged = 0

    # Step2 consolidation:
    # Third-party bulk upload validates Product Name uniqueness. Therefore,
    # before writing to the template, consolidate rows by finished product
    # Material Number / Product Name. Quantity and working hours are summed.
    consolidated_rows: dict[str, Dict[str, Any]] = {}

    for source_index, (_, row) in enumerate(df.iterrows(), start=1):
        if source_index == source_total_rows or source_index % 1000 == 0:
            consolidation_progress = 12 + int((source_index / max(1, source_total_rows)) * 28)
            report("整理產品活動來源資料", source_index, source_total_rows, consolidation_progress)
        raw_product_name = row.get(material_col)
        if pd.isna(raw_product_name) or str(raw_product_name).strip() == "":
            skipped_blank_rows += 1
            continue

        product_name = str(raw_product_name).strip()
        product_key = _normalize_material(product_name)
        if not product_key:
            skipped_blank_rows += 1
            continue

        product_type = row.get(product_type_col)
        is_wip = row.get(is_wip_col) if is_wip_col else None

        # M1A 明確判定的 WIP 一律不寫入 Bulk。
        if _is_wip(product_type, is_wip):
            excluded_wip_rows += 1
            continue

        # 只有選擇「包含半品工時」時，才使用 M2A working_hour_rollup
        # 的 Semi Material 清單，依標準 BOM 結構排除中間半品。
        # Direct Working Hour 模式不讀取 Roll-up，也不套用本條件。
        if working_hour_source == "include_semi" and product_key in structural_wip_materials:
            excluded_structural_wip_rows += 1
            continue

        year = _as_year(row.get(year_col))
        qty = _safe_number(row.get(qty_col))
        product_description = _safe_text(row.get(material_desc_col)) if material_desc_col else ""
        production_site = _safe_text(row.get(production_site_col)) if production_site_col else ""

        direct_labor_hours = 0.0
        if labor_hours_col:
            raw_labor_hours = pd.to_numeric(row.get(labor_hours_col), errors="coerce")
            direct_labor_hours = 0 if pd.isna(raw_labor_hours) else float(raw_labor_hours)

        if product_key not in consolidated_rows:
            consolidated_rows[product_key] = {
                "product_name": product_name,
                "year": year,
                "qty": 0.0,
                "product_description": product_description,
                "production_site": production_site,
                "direct_labor_hours": 0.0,
                "source_row_count": 0,
            }
        else:
            duplicated_product_rows_merged += 1

        item = consolidated_rows[product_key]
        item["qty"] = float(item.get("qty", 0) or 0) + qty
        item["direct_labor_hours"] = float(item.get("direct_labor_hours", 0) or 0) + direct_labor_hours
        item["source_row_count"] = int(item.get("source_row_count", 0) or 0) + 1

        # Keep the first non-empty descriptive fields.
        if not item.get("product_description") and product_description:
            item["product_description"] = product_description
        if not item.get("production_site") and production_site:
            item["production_site"] = production_site
        if not item.get("year") and year:
            item["year"] = year

        source_valid_rows += 1

    rows_to_write: list[Dict[str, Any]] = []
    consolidated_total = int(len(consolidated_rows))
    for consolidated_index, (material_key, item) in enumerate(consolidated_rows.items(), start=1):
        if consolidated_index == consolidated_total or consolidated_index % 1000 == 0:
            validation_progress = 40 + int((consolidated_index / max(1, consolidated_total)) * 12)
            report("檢查產品活動資料", consolidated_index, consolidated_total, validation_progress)
        production_site = str(item.get("production_site") or "").strip()
        direct_labor_hours = float(item.get("direct_labor_hours", 0) or 0)
        labor_hours = None

        if labor_hours_col:
            if working_hour_source == "include_semi":
                site_key = str(production_site or "").strip().upper()
                if rollup_by_site_material or rollup_by_material:
                    if (site_key, material_key) in rollup_by_site_material:
                        labor_hours = float(rollup_by_site_material.get((site_key, material_key), 0) or 0)
                    elif material_key in rollup_by_material:
                        labor_hours = float(rollup_by_material.get(material_key, 0) or 0)
                    else:
                        labor_hours = direct_labor_hours
                else:
                    semi_labor_hours = float(semi_hour_by_material.get(material_key, 0) or 0)
                    labor_hours = direct_labor_hours + semi_labor_hours
            else:
                labor_hours = direct_labor_hours

            # 年度總工時等於 0 的產品不寫入 Bulk。
            # 先合併再排除，避免同產品多筆有正負/空值時誤判。
            if labor_hours == 0:
                excluded_zero_labor_hour_rows += int(item.get("source_row_count", 1) or 1)
                continue

        rows_to_write.append({
            "product_name": str(item.get("product_name") or "").strip(),
            "year": int(item.get("year") or date.today().year),
            "qty": float(item.get("qty", 0) or 0),
            "product_description": str(item.get("product_description") or "").strip(),
            "production_site": production_site,
            "labor_hours": labor_hours,
        })

    rows_total = int(len(rows_to_write))
    report("寫入 Product Activity Bulk", 0, rows_total, 55)
    _write_product_activity_openxml(
        bulk_template_path, output_path, rows_to_write,
        activity_labor_hours_col, activity_labor_hours_unit_col,
    )
    report("Product Activity Bulk 已完成", rows_total, rows_total, 100)

    return {
        "source_rows": int(len(df)),
        "activity_rows": int(rows_total),
        "product_rows": int(rows_total),
        "excluded_wip_rows": int(excluded_wip_rows),
        "excluded_structural_wip_rows": int(excluded_structural_wip_rows),
        "structural_wip_filter_applied": bool(working_hour_source == "include_semi"),
        "structural_wip_material_count": int(len(structural_wip_materials)),
        "structural_wip_materials_sample": sorted(structural_wip_materials)[:50],
        "skipped_blank_rows": int(skipped_blank_rows),
        "excluded_zero_labor_hour_rows": int(excluded_zero_labor_hour_rows),
        "step2_product_name_consolidation_enabled": True,
        "valid_rows_before_consolidation": int(source_valid_rows),
        "unique_product_names_after_consolidation": int(rows_total),
        "duplicated_product_rows_merged": int(duplicated_product_rows_merged),
        "output_filename": output_path.name,
        "template_copy_mode": True,
        "openxml_template_preserve_mode": True,
        "openpyxl_workbook_save_used": False,
        "product_description_from_material_description": True,
        "labor_hours_from_step1": bool(labor_hours_col),
        "labor_hours_written_to_bulk": bool(labor_hours_col and activity_labor_hours_col),
        "labor_hours_template_column": int(activity_labor_hours_col) if activity_labor_hours_col else None,
        "labor_hours_unit_written_to_bulk": bool(labor_hours_col and activity_labor_hours_unit_col),
        "labor_hours_unit_template_column": int(activity_labor_hours_unit_col) if activity_labor_hours_unit_col else None,
        "working_hour_source": working_hour_source,
        "working_hour_rollup_used": bool(working_hour_rollup_path and Path(working_hour_rollup_path).exists()),
        "semi_finished_working_hour_products": int(len(structural_wip_materials) or len(semi_hour_by_material)),
        "bulk_formatter_version": BULK_FORMATTER_VERSION,
    }



def generate_product_activity_bulk_files_by_site(
    step1_output_path: str | Path,
    bulk_template_path: str | Path,
    output_dir: str | Path,
    token: str | None = None,
    working_hour_source: str = "direct",
    bom_structure_path: str | Path | None = None,
    working_hour_rollup_path: str | Path | None = None,
) -> Dict[str, Any]:
    """Generate one or multiple bulk files according to Production Site.

    If Step1 output contains multiple Production Site values, this function generates
    one xlsx per site. Files are not zipped; the API returns multiple download links.
    """
    step1_output_path = Path(step1_output_path)
    bulk_template_path = Path(bulk_template_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    token = token or "bulk"

    df = pd.read_excel(step1_output_path, sheet_name=SOURCE_SHEET_NAME, dtype=object)
    production_site_col = _find_optional_column(df, ["Production Site", "production site", "生產廠區", "廠區", "廠別"])

    if production_site_col is None:
        output_path = output_dir / f"product_activity_data_bulk_create_{token}.xlsx"
        summary = generate_product_activity_bulk_file(step1_output_path, bulk_template_path, output_path, working_hour_source=working_hour_source, bom_structure_path=bom_structure_path, working_hour_rollup_path=working_hour_rollup_path)
        return {
            "split_by_production_site": False,
            "production_site_count": 1,
            "files": [{
                "production_site": "ALL",
                "filename": output_path.name,
                "download_url": f"/download/{output_path.name}",
                "summary": summary,
            }],
            **summary,
        }

    df[production_site_col] = df[production_site_col].fillna("").astype(str).str.strip()
    sites = [s for s in sorted(df[production_site_col].unique()) if s]

    if len(sites) <= 1:
        site = sites[0] if sites else "ALL"
        output_path = output_dir / f"product_activity_data_bulk_create_{_sanitize_filename(site)}_{token}.xlsx"
        summary = generate_product_activity_bulk_file(step1_output_path, bulk_template_path, output_path, working_hour_source=working_hour_source, bom_structure_path=bom_structure_path, working_hour_rollup_path=working_hour_rollup_path)
        return {
            "split_by_production_site": False,
            "production_site_count": len(sites) or 1,
            "files": [{
                "production_site": site,
                "filename": output_path.name,
                "download_url": f"/download/{output_path.name}",
                "summary": summary,
            }],
            **summary,
        }

    files: list[Dict[str, Any]] = []
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)
        for site in sites:
            site_df = df[df[production_site_col] == site].copy()
            safe_site = _sanitize_filename(site)
            temp_step1 = tmpdir_path / f"step1_{safe_site}.xlsx"
            output_path = output_dir / f"product_activity_data_bulk_create_{safe_site}_{token}.xlsx"

            with pd.ExcelWriter(temp_step1, engine="openpyxl") as writer:
                site_df.to_excel(writer, index=False, sheet_name=SOURCE_SHEET_NAME)

            summary = generate_product_activity_bulk_file(temp_step1, bulk_template_path, output_path, working_hour_source=working_hour_source, bom_structure_path=bom_structure_path, working_hour_rollup_path=working_hour_rollup_path)
            files.append({
                "production_site": site,
                "filename": output_path.name,
                "download_url": f"/download/{output_path.name}",
                "summary": summary,
            })

    total_activity_rows = sum(int(f["summary"].get("activity_rows", 0)) for f in files)
    total_product_rows = sum(int(f["summary"].get("product_rows", 0)) for f in files)
    total_excluded_wip = sum(int(f["summary"].get("excluded_wip_rows", 0)) for f in files)
    total_excluded_structural_wip = sum(int(f["summary"].get("excluded_structural_wip_rows", 0)) for f in files)

    return {
        "split_by_production_site": True,
        "production_site_count": len(files),
        "activity_rows": total_activity_rows,
        "product_rows": total_product_rows,
        "excluded_wip_rows": total_excluded_wip,
        "excluded_structural_wip_rows": total_excluded_structural_wip,
        "structural_wip_filter_applied": normalize_working_hour_source(working_hour_source) == "include_semi",
        "files": files,
    }



def generate_product_activity_bulk_files_by_site_zip(
    step1_output_path: str | Path,
    bulk_template_path: str | Path,
    output_dir: str | Path,
    token: str | None = None,
    working_hour_source: str = "direct",
    bom_structure_path: str | Path | None = None,
    working_hour_rollup_path: str | Path | None = None,
    progress_callback: Callable[..., None] | None = None,
) -> Dict[str, Any]:
    """Generate Bulk files by Production Site and package them into one ZIP."""
    import zipfile

    step1_output_path = Path(step1_output_path)
    bulk_template_path = Path(bulk_template_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    token = token or "bulk"

    def report(step: str, processed: int = 0, total: int = 0, progress: int = 0, **extra: Any) -> None:
        if progress_callback is None:
            return
        progress_callback(
            step,
            processed=int(processed or 0),
            total=int(total or 0),
            progress=max(0, min(100, int(progress or 0))),
            **extra,
        )

    report("讀取產品活動來源檔", 0, 0, 2)
    df = pd.read_excel(step1_output_path, sheet_name=SOURCE_SHEET_NAME, dtype=object)
    source_total_rows = int(len(df))
    report("讀取產品活動來源檔", source_total_rows, source_total_rows, 8)
    production_site_col = _find_optional_column(
        df,
        ["Production Site", "production site", "生產廠區", "廠區", "廠別"],
    )

    generated_files: list[Dict[str, Any]] = []

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)

        if production_site_col is None:
            output_path = tmpdir_path / f"product_activity_data_bulk_create_ALL_{token}.xlsx"
            summary = generate_product_activity_bulk_file(
                step1_output_path,
                bulk_template_path,
                output_path,
                working_hour_source=working_hour_source,
                bom_structure_path=bom_structure_path,
                working_hour_rollup_path=working_hour_rollup_path,
                progress_callback=lambda step, processed=0, total=0, progress=0, **extra: report(
                    step, processed, total, 8 + int(progress * 0.82), production_site="ALL", **extra
                ),
            )
            generated_files.append({
                "production_site": "ALL",
                "filename": output_path.name,
                "path": output_path,
                "summary": summary,
            })
        else:
            df[production_site_col] = df[production_site_col].fillna("").astype(str).str.strip()
            sites = [s for s in sorted(df[production_site_col].unique()) if s]

            if not sites:
                sites = ["ALL"]
                df[production_site_col] = "ALL"

            site_count = max(1, len(sites))
            for site_index, site in enumerate(sites):
                site_df = df[df[production_site_col] == site].copy()
                safe_site = _sanitize_filename(site)
                temp_step1 = tmpdir_path / f"step1_{safe_site}_{token}.xlsx"
                output_path = tmpdir_path / f"product_activity_data_bulk_create_{safe_site}_{token}.xlsx"

                with pd.ExcelWriter(temp_step1, engine="openpyxl") as writer:
                    site_df.to_excel(writer, index=False, sheet_name=SOURCE_SHEET_NAME)

                stage_start = 8 + int((site_index / site_count) * 82)
                stage_span = 82 / site_count
                summary = generate_product_activity_bulk_file(
                    temp_step1,
                    bulk_template_path,
                    output_path,
                    working_hour_source=working_hour_source,
                    bom_structure_path=bom_structure_path,
                    working_hour_rollup_path=working_hour_rollup_path,
                    progress_callback=lambda step, processed=0, total=0, progress=0, _site=site, _start=stage_start, _span=stage_span, **extra: report(
                        f"{_site}｜{step}", processed, total, int(_start + (progress / 100) * _span), production_site=_site, **extra
                    ),
                )
                generated_files.append({
                    "production_site": site,
                    "filename": output_path.name,
                    "path": output_path,
                    "summary": summary,
                })

        zip_name = f"product_activity_data_bulk_by_production_site_{token}.zip"
        zip_path = output_dir / zip_name

        report("封裝 Product Activity Bulk ZIP", 0, len(generated_files), 92)
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
            for file_index, item in enumerate(generated_files, start=1):
                z.write(item["path"], item["filename"])
                report("封裝 Product Activity Bulk ZIP", file_index, len(generated_files), 92 + int((file_index / max(1, len(generated_files))) * 7))

    files_summary = [
        {
            "production_site": item["production_site"],
            "filename": item["filename"],
            "activity_rows": int(item["summary"].get("activity_rows", 0)),
            "product_rows": int(item["summary"].get("product_rows", 0)),
            "excluded_wip_rows": int(item["summary"].get("excluded_wip_rows", 0)),
            "excluded_structural_wip_rows": int(item["summary"].get("excluded_structural_wip_rows", 0)),
            "structural_wip_material_count": int(item["summary"].get("structural_wip_material_count", 0)),
        }
        for item in generated_files
    ]

    total_activity_rows = sum(item["activity_rows"] for item in files_summary)
    report("Product Activity Bulk ZIP 已完成", total_activity_rows, total_activity_rows, 100)

    return {
        "split_by_production_site": len(generated_files) > 1,
        "production_site_count": len(generated_files),
        "activity_rows": total_activity_rows,
        "product_rows": sum(item["product_rows"] for item in files_summary),
        "excluded_wip_rows": sum(item["excluded_wip_rows"] for item in files_summary),
        "excluded_structural_wip_rows": sum(item["excluded_structural_wip_rows"] for item in files_summary),
        "structural_wip_filter_applied": normalize_working_hour_source(working_hour_source) == "include_semi",
        "structural_wip_material_count": max((item["structural_wip_material_count"] for item in files_summary), default=0),
        "files": files_summary,
        "output_filename": zip_name,
        "download_url": f"/download/{zip_name}",
        "zip_output": True,
        "working_hour_source": normalize_working_hour_source(working_hour_source),
        "bom_structure_used": bool(bom_structure_path and Path(bom_structure_path).exists()),
        "working_hour_rollup_used": bool(working_hour_rollup_path and Path(working_hour_rollup_path).exists()),
    }
