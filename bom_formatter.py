from __future__ import annotations

import csv
import math
import os
import re
import shutil
import tempfile
import time
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict

from template_header_resolver import aliases as _profile_aliases
from dropdown_value_resolver import DropdownValuesCache

import pandas as pd
from openpyxl import Workbook, load_workbook
try:
    import xlsxwriter
except Exception:  # pragma: no cover
    xlsxwriter = None


ACTIVITY_SHEET_NAME = "Input Sheet Activity Data"
RAW_MATERIAL_SHEET_NAME = "Input Sheet Raw Material"
DATA_START_ROW = 3
BOM_FORMATTER_VERSION = "DIP_V29_M2C_STREAMING_OPENXML_20260724"
M2_MAX_ACTIVITY_DATA_ROWS = 50000
M2_ACTIVITY_HELPER_FORMULA_COLS = tuple(range(28, 36))  # AB~AI
M2C_XLSX_COMPRESSION_LEVEL = max(0, min(9, int(os.getenv("M2C_XLSX_COMPRESSION_LEVEL", "1") or 1)))
M2C_COPY_BUFFER_BYTES = max(1024 * 1024, int(os.getenv("M2C_COPY_BUFFER_BYTES", str(4 * 1024 * 1024)) or (4 * 1024 * 1024)))


DEFAULT_MAPPING = {
    "material_col": "Material",
    "parent_col": "Parent Node",
    "component_col": "Component",
    "qty_col": "CS03 Qty",
    "unit_col": "CS03 UoM",
    "description_col": "Component Description",
    "material_group_col": "Material group",
    "valid_from_col": "BOM Valid From",
    "altitem_group_col": "Altitem group",
    "usage_probability_col": "Usage probability%",
    "net_weight_col": "Net weight",
    "gross_weight_col": "Gross weight",
    "weight_uom_col": "Weight UoM",
}


def _normalize_col(value: Any) -> str:
    return str(value or "").strip().replace("\n", " ").replace("\r", " ")


def _resolve_mapping(mapping: dict[str, str | None] | None) -> dict[str, str]:
    mapping = mapping or {}
    resolved = {}
    for key, default_value in DEFAULT_MAPPING.items():
        value = str(mapping.get(key) or "").strip()
        resolved[key] = value if value else default_value
    return resolved


def _find_column(df: pd.DataFrame, column_name: str) -> str:
    target = _normalize_col(column_name).lower()
    normalized = {_normalize_col(c).lower(): c for c in df.columns}
    if target in normalized:
        return normalized[target]
    raise ValueError(f"找不到 BOM 欄位：{column_name}")


def _find_optional_column(df: pd.DataFrame, column_name: str) -> str | None:
    try:
        return _find_column(df, column_name)
    except ValueError:
        return None


def _safe_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _safe_number(value: Any) -> float:
    if pd.isna(value):
        return 0.0
    text = str(value).strip()
    if text.upper() in ["", "NAN", "NONE"]:
        return 0.0
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return 0.0


def _normalize_altitem_group(value: Any) -> str:
    """Normalize SAP Altitem group values for grouping alternative materials.

    Excel sometimes reads group values as 1.0 instead of 1.  Empty, zero,
    and non-numeric values are treated as no alternative group.
    """
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.upper() in ["", "NAN", "NONE"]:
        return ""
    try:
        number = float(text.replace(",", ""))
    except ValueError:
        return text
    if number == 0:
        return ""
    if number.is_integer():
        return str(int(number))
    return str(number)


def _usage_probability_ratio(value: Any) -> float | None:
    """Convert Usage probability% to a multiplier.

    Supports both SAP-style percent values (50 -> 0.5) and ratio values
    already stored as decimals (0.5 -> 0.5). Blank/invalid values return
    None so normal BOM quantity remains unchanged.
    """
    if pd.isna(value):
        return None
    text = str(value).strip().replace("%", "")
    if text.upper() in ["", "NAN", "NONE"]:
        return None
    try:
        number = float(text.replace(",", ""))
    except ValueError:
        return None
    if number < 0:
        return None
    return number / 100.0 if number > 1 else number


def _apply_altitem_usage_probability(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Apply alternative-item usage probability to CS03 quantity.

    Rule: when Altitem group has value, the row is an SAP alternative item,
    so effective quantity = CS03 Qty × Usage probability%. Blank Altitem group
    keeps original CS03 Qty.
    """
    work = df.copy()
    if "_altitem_group" not in work.columns:
        work["_altitem_group"] = ""
    if "_usage_probability_ratio" not in work.columns:
        work["_usage_probability_ratio"] = None

    alt_mask = work["_altitem_group"].astype(str).str.strip() != ""
    if not alt_mask.any():
        work["_qty_original"] = work["_qty"]
        work["_qty_adjusted_by_altitem"] = False
        return work, {
            "altitem_rows": 0,
            "altitem_groups": 0,
            "altitem_adjusted_rows": 0,
            "altitem_probability_missing_rows": 0,
        }

    group_keys = ["_parent", "_altitem_group"]
    has_probability = work["_usage_probability_ratio"].apply(lambda x: x is not None)

    # CMP official rule:
    # If Altitem group has a value, CS03 Qty must be converted to effective qty
    # by multiplying Usage probability%. Do not require duplicated rows in the
    # same group, because SAP may export only the selected/available alternative
    # item row while the probability still represents real usage.
    apply_mask = alt_mask & has_probability

    work["_qty_original"] = work["_qty"]
    work.loc[apply_mask, "_qty"] = work.loc[apply_mask, "_qty"] * work.loc[apply_mask, "_usage_probability_ratio"].astype(float)
    work["_qty_adjusted_by_altitem"] = apply_mask

    alt_groups = (
        work.loc[alt_mask, group_keys]
        .drop_duplicates()
        .shape[0]
    )
    missing_probability_rows = int((alt_mask & ~has_probability).sum())
    return work, {
        "altitem_rows": int(alt_mask.sum()),
        "altitem_groups": int(alt_groups),
        "altitem_adjusted_rows": int(apply_mask.sum()),
        "altitem_probability_missing_rows": missing_probability_rows,
        "altitem_rule": "Effective CS03 Qty = CS03 Qty × Usage probability% when Altitem group has value; blank Altitem group keeps original CS03 Qty.",
    }


_XML_NUMERIC_CHAR_REF_RE = re.compile(rb"&#(?:x([0-9A-Fa-f]+)|([0-9]+));")
_XML_INVALID_ASCII_CONTROL_RE = re.compile(rb"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def _is_valid_xml_char_number(codepoint: int) -> bool:
    """Return True if the code point is valid in XML 1.0."""
    return (
        codepoint in (0x09, 0x0A, 0x0D)
        or 0x20 <= codepoint <= 0xD7FF
        or 0xE000 <= codepoint <= 0xFFFD
        or 0x10000 <= codepoint <= 0x10FFFF
    )


def _clean_invalid_xml_bytes(data: bytes) -> bytes:
    """Remove invalid XML 1.0 character references/control bytes from XLSX XML parts.

    Some SAP-exported Excel files contain strings such as ``&#11;`` or
    ``&#x0B;`` inside sharedStrings.xml. openpyxl rejects those XML files with
    "reference to invalid character number". Removing only invalid XML
    characters keeps the workbook readable without changing normal BOM values.
    """

    def replace_numeric_ref(match: re.Match[bytes]) -> bytes:
        raw_hex, raw_dec = match.groups()
        try:
            codepoint = int(raw_hex, 16) if raw_hex is not None else int(raw_dec, 10)
        except Exception:
            return b""
        return match.group(0) if _is_valid_xml_char_number(codepoint) else b""

    data = _XML_NUMERIC_CHAR_REF_RE.sub(replace_numeric_ref, data)
    data = _XML_INVALID_ASCII_CONTROL_RE.sub(b"", data)
    return data


def _repair_xlsx_invalid_xml(source_path: str | Path, repaired_path: str | Path) -> Path:
    """Create a temporary XLSX copy with invalid XML characters removed."""
    source_path = Path(source_path)
    repaired_path = Path(repaired_path)
    with zipfile.ZipFile(source_path, "r") as zin, zipfile.ZipFile(repaired_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.lower().endswith((".xml", ".rels")):
                data = _clean_invalid_xml_bytes(data)
            zout.writestr(item, data)
    return repaired_path


def _read_excel_first_sheet(path: str | Path) -> pd.DataFrame:
    """Read the first sheet from a sanitized XLSX copy.

    SAP/exported BOM workbooks can contain invalid XML numeric character
    references such as ``&#11;`` in sharedStrings.xml. openpyxl fails before
    pandas can build the DataFrame. Therefore we always sanitize XML parts into
    a temporary workbook first, then read that repaired workbook.
    """
    path = Path(path)
    with tempfile.TemporaryDirectory() as tmp_dir:
        repaired_path = Path(tmp_dir) / f"repaired_{path.name}"
        try:
            _repair_xlsx_invalid_xml(path, repaired_path)
            return pd.read_excel(repaired_path, sheet_name=0, dtype=object)
        except Exception as repaired_exc:
            # Fall back to direct read only if the repair/copy itself was the issue.
            # If direct read also fails, return the clearer repaired-read error.
            try:
                return pd.read_excel(path, sheet_name=0, dtype=object)
            except Exception:
                raise ValueError(f"Excel XML 修復後仍無法讀取：{repaired_exc}") from repaired_exc


def _date_from_value(value: Any) -> date:
    if pd.isna(value) or value in [None, ""]:
        return date(datetime.now().year, 1, 1)
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return date(datetime.now().year, 1, 1)
    return parsed.date()


def _year_start(d: date) -> date:
    return date(d.year, 1, 1)

def _year_end(d: date) -> date:
    return date(d.year, 12, 31)


def _clear_target_cells(ws, start_row: int, columns: list[int]) -> None:
    max_row = max(ws.max_row, start_row)
    for row_idx in range(start_row, max_row + 1):
        for col_idx in columns:
            ws.cell(row_idx, col_idx).value = None

def _normalize_template_header(value: Any) -> str:
    """Normalize a bulk-template header for resilient column matching."""
    text = str(value or "").strip().lower()
    text = text.replace("\n", " ").replace("\r", " ")
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", text)


def _find_template_columns(ws, aliases: list[str], header_rows: int = DATA_START_ROW - 1) -> list[int]:
    """Find worksheet columns by exact header text in the uploaded bulk template.

    V14.2 rule: Module 2 output follows the user's uploaded template instead
    of a historical fixed column order. Matching is exact after normalization to
    avoid accidental matches such as Raw Material Name -> Raw Material Code.
    """
    alias_keys = [_normalize_template_header(a) for a in aliases if str(a or "").strip()]
    if not alias_keys:
        return []

    found: list[int] = []
    max_header_row = max(1, int(header_rows or 1))
    row_order = list(range(1, max_header_row + 1))
    if 2 in row_order:
        row_order = [2] + [r for r in row_order if r != 2]

    for row in row_order:
        for col in range(1, ws.max_column + 1):
            header_key = _normalize_template_header(ws.cell(row, col).value)
            if header_key and header_key in alias_keys and col not in found:
                found.append(col)
    return found


def _find_template_column(ws, aliases: list[str], fallback_col: int | None = None) -> int:
    cols = _find_template_columns(ws, aliases)
    if cols:
        return int(cols[0])
    if fallback_col is not None:
        return int(fallback_col)
    raise ValueError(f"Bulk template 缺少必要欄位：{', '.join(aliases)}")


def _find_template_optional_column(ws, aliases: list[str]) -> int | None:
    cols = _find_template_columns(ws, aliases)
    return int(cols[0]) if cols else None


def _write_template_value(ws, row_idx: int, col_idx: int | None, value: Any) -> None:
    if col_idx:
        ws.cell(row_idx, int(col_idx)).value = value


def _clear_template_columns(ws, start_row: int, columns: list[int]) -> None:
    unique_columns = sorted({int(c) for c in columns if c})
    if unique_columns:
        _clear_target_cells(ws, start_row, unique_columns)

RAW_MATERIAL_NAME_ALIASES = _profile_aliases("raw_material_activity", "raw_name", ["原料名稱"])
RAW_MATERIAL_CODE_ALIASES = _profile_aliases("raw_material_activity", "raw_code", ["Raw Material ID", "Raw Material Number", "原料代碼"])
RAW_MATERIAL_DESC_ALIASES = ["raw_material_description", "Raw Material Description (Optional)", "Raw Material Description", "Description", "原物料描述", "品名"]
DOC_START_DATE_ALIASES = _profile_aliases("raw_material_activity", "doc_start", ["Document Start Date"])
DOC_END_DATE_ALIASES = _profile_aliases("raw_material_activity", "doc_end", ["Document End Date"])
DOCUMENT_TYPE_ALIASES = _profile_aliases("raw_material_activity", "document_type", ["文件類型"])
DOCUMENT_NUMBER_ALIASES = _profile_aliases("raw_material_activity", "document_number", ["Document Number", "文件號碼"])
USAGE_ALIASES = _profile_aliases("raw_material_activity", "usage")
ACTIVITY_DATA_UNIT_ALIASES = _profile_aliases("raw_material_activity", "unit", ["單位"])
DATA_SOURCE_ALIASES = _profile_aliases("raw_material_activity", "data_source")
DATA_SOURCE_OTHER_ALIASES = _profile_aliases("raw_material_activity", "data_source_other", ["其他資料來源"])
CALCULATE_TRANSPORTATION_EMISSIONS_ALIASES = _profile_aliases(
    "raw_material_activity",
    "calculate_transportation_emissions",
    [
        "Calculate Transportation Emissions",
        "calculate_transportation_emissions",
        "is_transportation_emission_calculated",
    ],
)
TRANSPORT_ORIGIN_ALIASES = _profile_aliases("raw_material_activity", "transport_origin")
TRANSPORT_DESTINATION_ALIASES = _profile_aliases("raw_material_activity", "transport_destination", ["運輸終點"])
SUPPLIER_NAME_ALIASES = _profile_aliases("raw_material_activity", "supplier_name")
PRODUCT_LINK_ALIASES = _profile_aliases("raw_material_activity", "target_product", ["Target Product", "Product Code", "Product Name", "產品代碼", "產品名稱"])
COMMENT_ALIASES = _profile_aliases("raw_material_activity", "comment", ["Comment"])
MATERIAL_GROUP_ALIASES = _profile_aliases("raw_material_activity", "material_group", ["Material group"])
NET_WEIGHT_ALIASES = ["net_weight", "Net Weight (optional)", "Net Weight", "Net weight", "淨重"]
GROSS_WEIGHT_ALIASES = ["gross_weight", "Gross Weight (optional)", "Gross Weight", "Gross weight", "毛重"]
WEIGHT_UNIT_ALIASES = ["weight_unit", "Weight Unit (optional)", "Weight Unit", "Weight UoM", "Weight UOM", "重量單位"]


def _sheet_has_dropdown_label(wb, dropdown_column_name: str, label: str) -> bool:
    """Check whether a visible dropdown label exists in the template.

    The user's bulk template has visible input columns and hidden key/helper
    columns. For Document Type, the visible value is "Bill of Materials (BOM)"
    while the helper formula converts it to key "BOM". Writing the key into
    the visible column causes the helper formula to fail.
    """
    if "Dropdown Values" not in wb.sheetnames:
        return False
    ws = wb["Dropdown Values"]
    target_header = _normalize_template_header(dropdown_column_name)
    target_label = str(label or "").strip()
    if not target_header or not target_label:
        return False

    candidate_cols: list[int] = []
    for col in range(1, ws.max_column + 1):
        if _normalize_template_header(ws.cell(1, col).value) == target_header:
            candidate_cols.append(col)
    for col in candidate_cols:
        for row in range(2, min(ws.max_row, 20000) + 1):
            if str(ws.cell(row, col).value or "").strip() == target_label:
                return True
    return False


def _dropdown_display_map(wb, display_col: int, key_col: int) -> dict[str, str]:
    result: dict[str, str] = {}
    if "Dropdown Values" not in getattr(wb, "sheetnames", []):
        return result
    ws = wb["Dropdown Values"]
    start_col = min(int(display_col), int(key_col))
    end_col = max(int(display_col), int(key_col))
    display_offset = int(display_col) - start_col
    key_offset = int(key_col) - start_col
    max_row = min(int(getattr(ws, "max_row", 0) or 0), 20000)
    for values in ws.iter_rows(min_row=2, max_row=max_row, min_col=start_col, max_col=end_col, values_only=True):
        display = str(values[display_offset] or "").strip() if len(values) > display_offset else ""
        key = str(values[key_offset] or "").strip() if len(values) > key_offset else ""
        if display and key:
            result[key.upper()] = display
    return result


def _localized_dropdown_display(value: Any, mapping: dict[str, str], default: str = "") -> str:
    text = str(value or "").strip()
    if not text:
        text = str(default or "").strip()
    if not text:
        return ""
    return mapping.get(text.upper(), text)


def _document_type_for_template(wb, dropdown_cache: DropdownValuesCache | None = None) -> str:
    """Return the template-localized visible value whose internal key is BOM."""
    cache = dropdown_cache or DropdownValuesCache.from_workbook(wb)
    return cache.display("document_type", "BOM", "Bill of Materials (BOM)")

def _transport_calculation_yes_for_template(wb, dropdown_cache: DropdownValuesCache | None = None) -> str:
    """Return the localized visible dropdown value whose internal key is YES."""
    cache = dropdown_cache or DropdownValuesCache.from_workbook(wb)
    return cache.display("calculate_transportation_emissions", "YES", "Yes")

def _read_bom(bom_path: str | Path, mapping: dict[str, str | None] | None = None) -> tuple[pd.DataFrame, dict[str, str]]:
    df = _read_excel_first_sheet(bom_path)
    m = _resolve_mapping(mapping)

    material_col = _find_optional_column(df, m.get("material_col", "Material"))
    parent_col = _find_column(df, m["parent_col"])
    component_col = _find_column(df, m["component_col"])
    qty_col = _find_column(df, m["qty_col"])
    unit_col = _find_column(df, m["unit_col"])
    description_col = _find_optional_column(df, m["description_col"])
    material_group_col = _find_optional_column(df, m["material_group_col"])
    valid_from_col = _find_optional_column(df, m["valid_from_col"])
    altitem_group_col = _find_optional_column(df, m["altitem_group_col"])
    usage_probability_col = _find_optional_column(df, m["usage_probability_col"])
    net_weight_col = _find_optional_column(df, m.get("net_weight_col", "Net weight"))
    gross_weight_col = _find_optional_column(df, m.get("gross_weight_col", "Gross weight"))
    weight_uom_col = _find_optional_column(df, m.get("weight_uom_col", "Weight UoM"))

    df = df.copy()
    df["_bom_material"] = df[material_col].apply(_safe_text) if material_col else ""
    df["_parent"] = df[parent_col].apply(_safe_text)
    df["_component"] = df[component_col].apply(_safe_text)
    df["_qty"] = df[qty_col].apply(_safe_number)
    df["_uom"] = df[unit_col].apply(_safe_text)
    df["_description"] = df[description_col].apply(_safe_text) if description_col else ""
    df["_material_group"] = df[material_group_col].apply(_safe_text) if material_group_col else ""
    df["_valid_from"] = df[valid_from_col].apply(_date_from_value) if valid_from_col else date(datetime.now().year, 1, 1)
    df["_altitem_group"] = df[altitem_group_col].apply(_normalize_altitem_group) if altitem_group_col else ""
    df["_usage_probability_ratio"] = df[usage_probability_col].apply(_usage_probability_ratio) if usage_probability_col else None
    df["_net_weight"] = df[net_weight_col].apply(_safe_number) if net_weight_col else ""
    df["_gross_weight"] = df[gross_weight_col].apply(_safe_number) if gross_weight_col else ""
    df["_weight_uom"] = df[weight_uom_col].apply(_safe_text) if weight_uom_col else ""

    df = df[(df["_parent"] != "") & (df["_component"] != "")].copy()
    df, altitem_summary = _apply_altitem_usage_probability(df)

    used_columns = {
        "material_col": material_col or "",
        "parent_col": parent_col,
        "component_col": component_col,
        "qty_col": qty_col,
        "unit_col": unit_col,
        "description_col": description_col or "",
        "material_group_col": material_group_col or "",
        "valid_from_col": valid_from_col or "",
        "altitem_group_col": altitem_group_col or "",
        "usage_probability_col": usage_probability_col or "",
        "net_weight_col": net_weight_col or "",
        "gross_weight_col": gross_weight_col or "",
        "weight_uom_col": weight_uom_col or "",
        **altitem_summary,
    }
    return df, used_columns




def _as_bom_path_list(bom_path: str | Path | list[str | Path] | tuple[str | Path, ...]) -> list[Path]:
    """Normalize single or multiple BOM paths while preserving backward compatibility."""
    if isinstance(bom_path, (list, tuple)):
        return [Path(p) for p in bom_path]
    return [Path(bom_path)]


def _bom_version_display(value: Any) -> str:
    """Return a stable display value for BOM Valid From in diagnostics."""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return _safe_text(value)


def _bom_definition_signature(group: pd.DataFrame) -> tuple[tuple[Any, ...], ...]:
    """Build a deterministic calculation-content signature for one BOM version.

    Material + BOM Valid From identifies a BOM version.  The signature compares
    every calculation-relevant detail row inside that version so duplicated
    exports can be removed safely without deleting different components from
    the same BOM.
    """
    records: list[tuple[Any, ...]] = []
    for _, values in group.iterrows():
        probability = values.get("_usage_probability_ratio", None)
        if probability is None or (isinstance(probability, float) and pd.isna(probability)):
            probability_value: Any = None
        else:
            probability_value = round(float(probability), 12)
        records.append((
            _safe_text(values.get("_parent")),
            _safe_text(values.get("_component")),
            round(float(values.get("_qty", 0.0) or 0.0), 12),
            _safe_text(values.get("_uom")),
            _safe_text(values.get("_altitem_group")),
            probability_value,
        ))
    # Duplicate detail rows inside one source do not change the BOM definition.
    unique_records = set(records)
    return tuple(sorted(unique_records, key=lambda item: tuple("" if v is None else str(v) for v in item)))


def _deduplicate_bom_versions(merged: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Deduplicate complete BOM versions across files and reject conflicts.

    Version key: Material + BOM Valid From.
    - Same key and identical full detail signature: keep one complete version.
    - Same key but different detail signature: stop instead of merging or adding.
    """
    if merged is None or merged.empty:
        return merged.copy(), {
            "bom_version_duplicate_groups_removed": 0,
            "bom_version_duplicate_rows_removed": 0,
            "bom_version_conflicts": 0,
        }

    work = merged.copy()
    if "_source_file" not in work.columns:
        work["_source_file"] = ""

    keep_mask = pd.Series(True, index=work.index)
    duplicate_groups_removed = 0
    duplicate_rows_removed = 0
    conflict_messages: list[str] = []

    valid_material_mask = work["_bom_material"].astype(str).str.strip() != ""
    version_work = work.loc[valid_material_mask]
    for (material, valid_from), version_group in version_work.groupby(["_bom_material", "_valid_from"], dropna=False, sort=False):
        source_groups = list(version_group.groupby("_source_file", dropna=False, sort=False))
        if len(source_groups) <= 1:
            continue

        signatures: list[tuple[str, tuple[tuple[Any, ...], ...], pd.Index]] = []
        for source_file, source_group in source_groups:
            signatures.append((_safe_text(source_file), _bom_definition_signature(source_group), source_group.index))

        first_source, first_signature, _first_index = signatures[0]
        mismatched_sources = [source for source, signature, _idx in signatures[1:] if signature != first_signature]
        if mismatched_sources:
            files = [first_source] + [source for source, _signature, _idx in signatures[1:]]
            conflict_messages.append(
                f"Material={_safe_text(material)}、BOM Valid From={_bom_version_display(valid_from)}；來源檔案：{', '.join(files)}"
            )
            continue

        # All complete definitions are identical. Keep the first source and
        # remove every row belonging to duplicate source files for this version.
        for _source, _signature, duplicate_index in signatures[1:]:
            keep_mask.loc[duplicate_index] = False
            duplicate_groups_removed += 1
            duplicate_rows_removed += int(len(duplicate_index))

    if conflict_messages:
        preview = "；".join(conflict_messages[:10])
        remaining = len(conflict_messages) - 10
        if remaining > 0:
            preview += f"；另有 {remaining} 組衝突"
        raise ValueError(
            "偵測到相同 Material + BOM Valid From，但完整 BOM 明細不一致。"
            "系統已停止，避免將不同 BOM 合併或數量相加：" + preview
        )

    result = work.loc[keep_mask].copy().reset_index(drop=True)
    return result, {
        "bom_version_duplicate_groups_removed": int(duplicate_groups_removed),
        "bom_version_duplicate_rows_removed": int(duplicate_rows_removed),
        "bom_version_conflicts": 0,
        "bom_version_key_rule": "Material + BOM Valid From identifies one BOM version; identical full definitions are deduplicated, inconsistent definitions stop processing.",
    }


def _read_boms(
    bom_paths: str | Path | list[str | Path] | tuple[str | Path, ...],
    mapping: dict[str, str | None] | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Read and merge one or multiple standard BOM Excel files.

    Complete BOM versions are identified by Material + BOM Valid From.  When
    the same version is uploaded more than once, its complete calculation
    content must be identical before a duplicate copy is removed.  Conflicting
    definitions are rejected instead of silently merged.
    """
    paths = _as_bom_path_list(bom_paths)
    if not paths:
        raise ValueError("請至少上傳一個 Standard BOM Excel 檔案")

    frames: list[pd.DataFrame] = []
    used_columns: dict[str, str] | None = None
    source_rows: list[dict[str, Any]] = []
    errors: list[str] = []

    for path in paths:
        try:
            df, cols = _read_bom(path, mapping=mapping)
            part = df.copy()
            part["_source_file"] = path.name
            frames.append(part)
            used_columns = used_columns or cols
            source_rows.append({"filename": path.name, "rows": int(len(part))})
        except Exception as exc:
            errors.append(f"{path.name}: {exc}")

    if errors:
        raise ValueError("；".join(errors))
    if not frames:
        raise ValueError("沒有可處理的 BOM 資料")

    merged = pd.concat(frames, ignore_index=True)
    before_version_dedup = int(len(merged))
    merged, version_summary = _deduplicate_bom_versions(merged)

    # Remove duplicated detail rows inside the retained complete version.  This
    # is a second safety layer for accidental duplicate rows in one source file.
    before_row_dedup = int(len(merged))
    dedup_subset = [
        "_bom_material", "_parent", "_component", "_qty", "_uom",
        "_description", "_material_group", "_valid_from", "_altitem_group",
        "_usage_probability_ratio", "_net_weight", "_gross_weight", "_weight_uom",
    ]
    merged = merged.drop_duplicates(subset=dedup_subset, keep="first").reset_index(drop=True)
    after_dedup = int(len(merged))

    used = dict(used_columns or {})
    used["bom_files"] = int(len(paths))
    used["bom_rows_before_dedup"] = before_version_dedup
    used["bom_rows_after_version_dedup"] = before_row_dedup
    used["bom_rows_after_dedup"] = after_dedup
    used["bom_duplicate_rows_removed"] = before_version_dedup - after_dedup
    used["bom_source_files"] = source_rows
    used.update(version_summary)
    return merged, used

def _exclude_zero_usage_rows(exploded: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Exclude true zero-usage raw material rows for Raw Material Bulk output.

    Rule: rows whose final calculated usage equals exactly 0 are not exported
    to raw_material_activity_data_bulk. Non-zero decimal usages are preserved.
    """
    if exploded is None or exploded.empty or "usage" not in exploded.columns:
        return exploded.copy() if isinstance(exploded, pd.DataFrame) else pd.DataFrame(), 0

    work = exploded.copy()
    usage_numeric = pd.to_numeric(work["usage"], errors="coerce").fillna(0.0)
    keep_mask = usage_numeric != 0
    excluded_rows = int((~keep_mask).sum())
    if excluded_rows:
        work = work.loc[keep_mask].copy().reset_index(drop=True)
    return work, excluded_rows


def _exclude_blank_material_group_rows(exploded: pd.DataFrame) -> tuple[pd.DataFrame, int, list[str], list[str]]:
    """Exclude final BOM leaves whose Standard BOM Material group is blank.

    M2A must retain the complete expanded BOM for traceability. This filter is
    intentionally applied only when preparing M2B Raw Material Bulk output.
    Blank includes None/NaN, empty text, half-width spaces and full-width spaces.
    Text values such as 0, TBC and TBD remain valid Material group values.
    """
    if exploded is None or exploded.empty:
        return exploded.copy() if isinstance(exploded, pd.DataFrame) else pd.DataFrame(), 0, [], []

    work = exploded.copy()
    if "material_group" not in work.columns:
        work["material_group"] = ""

    material_group_text = work["material_group"].apply(_safe_text)
    keep_mask = material_group_text != ""
    excluded_rows = int((~keep_mask).sum())
    excluded_raw_materials = sorted({
        _safe_text(value)
        for value in work.loc[~keep_mask, "raw_material"].tolist()
        if _safe_text(value)
    }) if "raw_material" in work.columns else []
    excluded_targets = sorted({
        _safe_text(value)
        for value in work.loc[~keep_mask, "target_product"].tolist()
        if _safe_text(value)
    }) if "target_product" in work.columns else []

    if excluded_rows:
        work = work.loc[keep_mask].copy().reset_index(drop=True)
    return work, excluded_rows, excluded_raw_materials, excluded_targets



def _exclude_zero_total_working_hour_target_rows(
    exploded: pd.DataFrame,
    total_hour_by_material: dict[str, float] | None,
) -> tuple[pd.DataFrame, int]:
    """Exclude raw material rows whose target product has zero total working hours.

    Rule: if a Target Product is found in the Step1/BOM working-hour roll-up map
    and its Total Annual Working Hour equals 0, all raw material activity rows
    under that Target Product are excluded from raw_material_activity_data_bulk.
    Products not found in the map are kept to avoid silently dropping rows when
    Step1 data is incomplete.
    """
    if exploded is None or exploded.empty or "target_product" not in exploded.columns:
        return exploded.copy() if isinstance(exploded, pd.DataFrame) else pd.DataFrame(), 0
    if not total_hour_by_material:
        return exploded.copy(), 0

    work = exploded.copy()
    target_keys = work["target_product"].apply(_normalize_material_key)
    known_zero_targets = {
        str(k).strip().upper()
        for k, v in total_hour_by_material.items()
        if str(k or "").strip() and float(v or 0.0) == 0.0
    }
    drop_mask = target_keys.isin(known_zero_targets)
    excluded_rows = int(drop_mask.sum())
    if excluded_rows:
        work = work.loc[~drop_mask].copy().reset_index(drop=True)
    return work, excluded_rows


def _calculate_total_working_hour_by_target(
    step1_output_path: str | Path,
    bom_df: pd.DataFrame,
) -> tuple[dict[str, float], Dict[str, Any]]:
    """Calculate Target Product total annual working hours including semi-finished roll-up.

    Total Annual Working Hour = direct annual working hour of the target product
    + annual working-hour contribution from semi-finished components in the BOM.
    This mirrors generate_working_hour_rollup_file but returns a lightweight map
    for filtering Raw Material Bulk rows.
    """
    step1_output_path = Path(step1_output_path)
    try:
        step1_df = pd.read_excel(step1_output_path, sheet_name=STEP1_SOURCE_SHEET_NAME, dtype=object)
    except Exception:
        step1_df = pd.read_excel(step1_output_path, sheet_name=0, dtype=object)

    material_col = _find_step1_column(step1_df, ["Material Number", "Material", "Product Material Number"])
    qty_col = _find_step1_optional_column(step1_df, ["年度生產量", "Annual Quantity", "Delivered quantity"])
    hour_col = _find_step1_optional_column(step1_df, ["年度總工時", "Total working hours", "Selected Hours", "Total Hours", "Working Hours"])
    if not hour_col:
        return {}, {
            "working_hour_filter_applied": False,
            "working_hour_filter_reason": "Step1 Output 找不到年度總工時欄位，未套用 Total Annual Working Hour = 0 過濾。",
            "zero_total_working_hour_targets": 0,
        }

    work = step1_df.copy()
    work["_material_key"] = work[material_col].apply(_normalize_material_key)
    work["_annual_qty"] = work[qty_col].apply(_safe_number) if qty_col else 0.0
    work["_direct_hour"] = work[hour_col].apply(_safe_number)
    work = work[work["_material_key"] != ""].copy()

    material_totals = work.groupby(["_material_key"], dropna=False, as_index=False).agg({"_annual_qty": "sum", "_direct_hour": "sum"})
    qty_by_material: dict[str, float] = {}
    direct_by_material: dict[str, float] = {}
    hour_per_pc_by_material: dict[str, float] = {}
    for _, r in material_totals.iterrows():
        material = str(r["_material_key"] or "").strip().upper()
        qty = float(r["_annual_qty"] or 0.0)
        hours = float(r["_direct_hour"] or 0.0)
        qty_by_material[material] = qty
        direct_by_material[material] = hours
        hour_per_pc_by_material[material] = hours / qty if qty else 0.0

    structure, _structure_summary = _explode_bom_structure(bom_df)
    semi_by_target: dict[str, float] = {}
    if not structure.empty:
        b = structure.copy()
        b["_target_key"] = b["Target Product"].apply(_normalize_material_key)
        b["_component_key"] = b["Component"].apply(_normalize_material_key)
        b["_accumulated_qty"] = b["Accumulated Quantity"].apply(_safe_number)
        b = b[b["Is Semi-finished"].astype(str).str.strip().str.upper().isin(["Y", "YES", "TRUE", "1"])].copy()
        for _, edge in b.iterrows():
            target = str(edge["_target_key"] or "").strip().upper()
            semi = str(edge["_component_key"] or "").strip().upper()
            target_qty = float(qty_by_material.get(target, 0.0) or 0.0)
            acc_qty = float(edge["_accumulated_qty"] or 0.0)
            semi_hr_pc = float(hour_per_pc_by_material.get(semi, 0.0) or 0.0)
            contribution = target_qty * acc_qty * semi_hr_pc
            if contribution:
                semi_by_target[target] = semi_by_target.get(target, 0.0) + contribution

    total_by_target: dict[str, float] = {}
    for material in sorted(set(direct_by_material) | set(semi_by_target)):
        total_by_target[material] = float(direct_by_material.get(material, 0.0) or 0.0) + float(semi_by_target.get(material, 0.0) or 0.0)

    zero_targets = [m for m, h in total_by_target.items() if float(h or 0.0) == 0.0]
    return total_by_target, {
        "working_hour_filter_applied": True,
        "working_hour_filter_rule": "Exclude all Raw Material Activity rows when Target Product Total Annual Working Hour, including semi-finished roll-up, equals 0.",
        "working_hour_mapped_targets": int(len(total_by_target)),
        "zero_total_working_hour_targets": int(len(zero_targets)),
    }

def _bom_scope_key(material: Any, valid_from: Any) -> tuple[str, Any]:
    return (_safe_text(material), valid_from)


def _build_bom_scope_context(df: pd.DataFrame) -> dict[str, Any]:
    """Build version-aware local graphs and cross-Material lookup tables."""
    work = df.copy()
    if "_bom_material" not in work.columns:
        work["_bom_material"] = ""
    work["_bom_material"] = work["_bom_material"].apply(_safe_text)

    has_material = work["_bom_material"].astype(str).str.strip().any()
    if has_material:
        grouped = list(work.groupby(["_bom_material", "_valid_from"], dropna=False, sort=False))
    else:
        grouped = [(('', None), work)]

    scopes: dict[tuple[str, Any], dict[str, Any]] = {}
    material_index: dict[str, list[tuple[str, Any]]] = defaultdict(list)

    for raw_key, scoped_df in grouped:
        if has_material:
            material_value, valid_from = raw_key
        else:
            material_value, valid_from = "", None
        material = _safe_text(material_value)
        key = _bom_scope_key(material, valid_from)
        parent_set = set(scoped_df["_parent"].dropna().astype(str))
        component_set = set(scoped_df["_component"].dropna().astype(str))
        if material and material in parent_set:
            roots = [material]
        else:
            roots = sorted(parent_set - component_set) or sorted(parent_set)

        children: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for _, row in scoped_df.iterrows():
            parent = _safe_text(row.get("_parent", ""))
            child = {
                "source_material": material,
                "scope_valid_from": valid_from,
                "parent": parent,
                "component": _safe_text(row.get("_component", "")),
                "qty": float(row.get("_qty", 0.0) or 0.0),
                "qty_original": row.get("_qty_original", row.get("_qty", 0.0)),
                "qty_adjusted_by_altitem": bool(row.get("_qty_adjusted_by_altitem", False)),
                "altitem_group": row.get("_altitem_group", ""),
                "usage_probability_ratio": row.get("_usage_probability_ratio", None),
                "uom": _safe_text(row.get("_uom", "")),
                "description": _safe_text(row.get("_description", "")),
                "material_group": _safe_text(row.get("_material_group", "")),
                "valid_from": row.get("_valid_from", valid_from),
                "net_weight": row.get("_net_weight", ""),
                "gross_weight": row.get("_gross_weight", ""),
                "weight_uom": row.get("_weight_uom", ""),
                "source_file": row.get("_source_file", ""),
            }
            children[parent].append(child)

        scopes[key] = {
            "key": key,
            "material": material,
            "valid_from": valid_from,
            "roots": roots,
            "children": children,
            "parent_set": parent_set,
            "component_set": component_set,
        }
        if material:
            material_index[material].append(key)

    return {"work": work, "scopes": scopes, "material_index": material_index, "has_material": bool(has_material)}


def _select_cross_material_scope(
    component: str,
    preferred_valid_from: Any,
    material_index: dict[str, list[tuple[str, Any]]],
    scopes: dict[tuple[str, Any], dict[str, Any]],
) -> tuple[str, Any] | None:
    """Select the component's own BOM version for cross-Material expansion."""
    candidates = list(material_index.get(component, []))
    # A valid cross-Material BOM must actually have the component as an
    # expandable parent node.  A Material value alone is not enough.
    candidates = [key for key in candidates if component in scopes[key]["children"]]
    if not candidates:
        return None

    exact = [key for key in candidates if key[1] == preferred_valid_from]
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        return exact[0]
    if len(candidates) == 1:
        return candidates[0]

    candidate_dates = ", ".join(_bom_version_display(key[1]) for key in candidates)
    raise ValueError(
        f"半成品 {component} 找到多個 BOM Valid From（{candidate_dates}），"
        f"但沒有與上階 {_bom_version_display(preferred_valid_from)} 完全相同的唯一版本，無法安全展開。"
    )


def _explode_bom(df: pd.DataFrame) -> tuple[pd.DataFrame, Dict[str, Any]]:
    """Explode local and cross-Material semi-finished BOMs by BOM version.

    Local expansion has priority.  When a Component has no local children but
    exists as another Material's BOM root, the matching BOM Valid From version
    is used and its final raw materials are rolled back to the original target.
    """
    output_rows: list[dict[str, Any]] = []
    cycle_count = 0
    product_count = 0
    cross_material_expansions = 0
    semi_finished_total: set[str] = set()

    empty_trace_columns = [
        "target_product", "target_valid_from", "source_material", "raw_material", "usage", "unit", "description", "material_group", "valid_from", "level",
        "immediate_parent", "trace_path", "parent_accumulated_qty", "qty_this_level_effective",
        "qty_this_level_original", "qty_adjusted_by_altitem", "altitem_group", "usage_probability_ratio", "usage_per_path", "source_file",
    ]
    if df is None or df.empty:
        trace_detail = pd.DataFrame(columns=empty_trace_columns)
        exploded = pd.DataFrame(columns=["target_product", "target_valid_from", "raw_material", "usage", "unit", "description", "material_group", "valid_from", "level"])
        exploded.attrs["trace_detail"] = trace_detail
        return exploded, {
            "products": 0, "semi_finished": 0, "raw_materials": 0,
            "activity_rows": 0, "max_level": 0, "cycles_skipped": 0,
            "cross_material_expansions": 0,
        }

    context = _build_bom_scope_context(df)
    scopes = context["scopes"]
    material_index = context["material_index"]

    for scope_key, root_scope in scopes.items():
        material = root_scope["material"]
        target_valid_from = root_scope["valid_from"]
        roots = root_scope["roots"]
        product_count += len(roots)

        for root in roots:
            target_product = material or root
            # scope_key, current parent, accumulated qty, level, path nodes,
            # path scope keys.  Both node and scope tracking protect cycles.
            stack: list[tuple[tuple[str, Any], str, float, int, list[str], list[tuple[str, Any]]]] = [
                (scope_key, root, 1.0, 0, [root], [scope_key])
            ]

            while stack:
                current_scope_key, current_parent, accumulated_qty, level, path, scope_path = stack.pop()
                current_scope = scopes[current_scope_key]
                current_children = current_scope["children"]

                for child in current_children.get(current_parent, []):
                    component = child["component"]
                    qty = child["qty"]
                    next_qty = accumulated_qty * qty
                    next_level = level + 1

                    if component in path:
                        cycle_count += 1
                        continue

                    next_scope_key: tuple[str, Any] | None = None
                    # Local Parent Node definition always wins.
                    if component in current_children:
                        next_scope_key = current_scope_key
                    elif context["has_material"]:
                        next_scope_key = _select_cross_material_scope(
                            component=component,
                            preferred_valid_from=child.get("valid_from", current_scope["valid_from"]),
                            material_index=material_index,
                            scopes=scopes,
                        )

                    if next_scope_key is not None:
                        if next_scope_key in scope_path and next_scope_key != current_scope_key:
                            cycle_count += 1
                            continue
                        semi_finished_total.add(component)
                        if next_scope_key != current_scope_key:
                            cross_material_expansions += 1
                        stack.append((
                            next_scope_key,
                            component,
                            next_qty,
                            next_level,
                            path + [component],
                            scope_path + ([next_scope_key] if next_scope_key != current_scope_key else []),
                        ))
                        continue

                    output_rows.append({
                        "target_product": target_product,
                        "target_valid_from": target_valid_from,
                        "source_material": current_scope["material"],
                        "raw_material": component,
                        "usage": next_qty,
                        "unit": child["uom"],
                        "description": child["description"],
                        "material_group": child["material_group"],
                        "net_weight": child.get("net_weight", ""),
                        "gross_weight": child.get("gross_weight", ""),
                        "weight_uom": child.get("weight_uom", ""),
                        "valid_from": child["valid_from"],
                        "level": next_level,
                        "immediate_parent": current_parent,
                        "trace_path": " > ".join(path + [component]),
                        "parent_accumulated_qty": accumulated_qty,
                        "qty_this_level_effective": qty,
                        "qty_this_level_original": child.get("qty_original", qty),
                        "qty_adjusted_by_altitem": child.get("qty_adjusted_by_altitem", False),
                        "altitem_group": child.get("altitem_group", ""),
                        "usage_probability_ratio": child.get("usage_probability_ratio", None),
                        "usage_per_path": next_qty,
                        "source_file": child.get("source_file", ""),
                    })

    trace_detail = pd.DataFrame(output_rows)
    if trace_detail.empty:
        trace_detail = pd.DataFrame(columns=empty_trace_columns)

    exploded = trace_detail.copy()
    if exploded.empty:
        exploded = pd.DataFrame(columns=[
            "target_product", "target_valid_from", "raw_material", "usage", "unit", "description", "material_group", "valid_from", "level"
        ])
    else:
        exploded = (
            exploded.groupby(["target_product", "target_valid_from", "raw_material", "unit"], dropna=False, as_index=False)
            .agg({
                "usage": "sum",
                "description": "first",
                "material_group": "first",
                "net_weight": "first",
                "gross_weight": "first",
                "weight_uom": "first",
                "valid_from": "first",
                "level": "max",
            })
            .sort_values(["target_product", "target_valid_from", "raw_material"])
            .reset_index(drop=True)
        )

    exploded.attrs["trace_detail"] = trace_detail
    summary = {
        "products": int(product_count),
        "semi_finished": int(len(semi_finished_total)),
        "raw_materials": int(exploded["raw_material"].nunique()) if not exploded.empty else 0,
        "activity_rows": int(len(exploded)),
        "max_level": int(exploded["level"].max()) if not exploded.empty else 0,
        "cycles_skipped": int(cycle_count),
        "cross_material_expansions": int(cross_material_expansions),
        "bom_scope_rule": "Expand within Material first; if no local children exist, match Component to another Material BOM using BOM Valid From and roll its raw materials back to the original target product.",
    }
    return exploded, summary

def _write_raw_material_bulk_from_exploded(
    exploded: pd.DataFrame,
    raw_material_template_path: str | Path,
    output_path: str | Path,
) -> Dict[str, Any]:
    """Write Raw Material Bulk workbook from an already exploded BOM DataFrame.

    This keeps Module 2 template-driven: data is written by header name, not
    fixed Excel column positions. It is reused by the all-site export and the
    Production Site split export.

    V14.9: true zero-usage raw material rows are excluded from Bulk output;
    non-zero decimal usages are preserved.
    """
    exploded, zero_usage_rows_excluded = _exclude_zero_usage_rows(exploded)

    raw_material_template_path = Path(raw_material_template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(raw_material_template_path, output_path)

    wb = load_workbook(output_path)

    if ACTIVITY_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"找不到 raw material bulk 分頁：{ACTIVITY_SHEET_NAME}")
    if RAW_MATERIAL_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"找不到 raw material bulk 分頁：{RAW_MATERIAL_SHEET_NAME}")

    activity_ws = wb[ACTIVITY_SHEET_NAME]
    raw_ws = wb[RAW_MATERIAL_SHEET_NAME]

    activity_cols = {
        "raw_name": _find_template_column(activity_ws, RAW_MATERIAL_NAME_ALIASES, 1),
        "raw_code": _find_template_column(activity_ws, RAW_MATERIAL_CODE_ALIASES, 2),
        "start_date": _find_template_column(activity_ws, DOC_START_DATE_ALIASES, 3),
        "end_date": _find_template_column(activity_ws, DOC_END_DATE_ALIASES, 4),
        "document_type": _find_template_column(activity_ws, DOCUMENT_TYPE_ALIASES, 5),
        "document_number": _find_template_column(activity_ws, DOCUMENT_NUMBER_ALIASES, 6),
        "usage": _find_template_column(activity_ws, USAGE_ALIASES, 7),
        "unit": _find_template_column(activity_ws, ACTIVITY_DATA_UNIT_ALIASES, 8),
        "data_source": _find_template_column(activity_ws, DATA_SOURCE_ALIASES, 12),
        "data_source_other": _find_template_column(activity_ws, DATA_SOURCE_OTHER_ALIASES, 13),
        "transport_origin": _find_template_column(activity_ws, TRANSPORT_ORIGIN_ALIASES, 15),
        "transport_destination": _find_template_column(activity_ws, TRANSPORT_DESTINATION_ALIASES, 16),
        "supplier_name": _find_template_column(activity_ws, SUPPLIER_NAME_ALIASES, 14),
        "target_product": _find_template_column(activity_ws, PRODUCT_LINK_ALIASES, 17),
        "comment": _find_template_column(activity_ws, COMMENT_ALIASES, 18),
        "material_group": _find_template_column(activity_ws, MATERIAL_GROUP_ALIASES, 19),
        "net_weight": _find_template_optional_column(activity_ws, NET_WEIGHT_ALIASES),
        "gross_weight": _find_template_optional_column(activity_ws, GROSS_WEIGHT_ALIASES),
        "weight_unit": _find_template_optional_column(activity_ws, WEIGHT_UNIT_ALIASES),
    }
    raw_cols = {
        "raw_name": _find_template_column(raw_ws, RAW_MATERIAL_NAME_ALIASES, 1),
        "raw_code": _find_template_column(raw_ws, RAW_MATERIAL_CODE_ALIASES, 2),
        "description": _find_template_column(raw_ws, RAW_MATERIAL_DESC_ALIASES, 6),
    }

    dropdown_cache = DropdownValuesCache.from_workbook(wb)
    document_type_value = _document_type_for_template(wb, dropdown_cache)
    activity_unit_display_map = dropdown_cache.input_to_display_map("activity_data_unit")
    weight_unit_display_map = dropdown_cache.input_to_display_map("weight_unit")
    data_source_display_map = dropdown_cache.input_to_display_map("data_source")

    _clear_template_columns(activity_ws, DATA_START_ROW, list(activity_cols.values()))
    _clear_template_columns(raw_ws, DATA_START_ROW, list(raw_cols.values()))

    row_idx = DATA_START_ROW
    for _, r in exploded.iterrows():
        valid_from = r["valid_from"]
        if not isinstance(valid_from, date):
            valid_from = _date_from_value(valid_from)

        raw_material = r["raw_material"]
        target_product = r["target_product"]
        usage_value = float(r["usage"]) if not pd.isna(r["usage"]) else 0

        _write_template_value(activity_ws, row_idx, activity_cols["raw_name"], raw_material)
        _write_template_value(activity_ws, row_idx, activity_cols["raw_code"], raw_material)
        _write_template_value(activity_ws, row_idx, activity_cols["start_date"], _year_start(valid_from))
        _write_template_value(activity_ws, row_idx, activity_cols["end_date"], _year_end(valid_from))
        _write_template_value(activity_ws, row_idx, activity_cols["document_type"], document_type_value)
        _write_template_value(activity_ws, row_idx, activity_cols["document_number"], "")
        _write_template_value(activity_ws, row_idx, activity_cols["usage"], usage_value)
        _write_template_value(activity_ws, row_idx, activity_cols["unit"], _localized_dropdown_display(r["unit"], activity_unit_display_map))
        _write_template_value(activity_ws, row_idx, activity_cols["data_source"], _localized_dropdown_display("SAP", data_source_display_map, "SAP"))
        _write_template_value(activity_ws, row_idx, activity_cols["data_source_other"], "")
        _write_template_value(activity_ws, row_idx, activity_cols["transport_origin"], "")
        _write_template_value(activity_ws, row_idx, activity_cols["transport_destination"], r.get("transport_destination", ""))
        _write_template_value(activity_ws, row_idx, activity_cols["supplier_name"], "")
        _write_template_value(activity_ws, row_idx, activity_cols["target_product"], target_product)
        _write_template_value(activity_ws, row_idx, activity_cols["comment"], "")
        _write_template_value(activity_ws, row_idx, activity_cols["material_group"], r["material_group"])
        _write_template_value(activity_ws, row_idx, activity_cols.get("net_weight"), r.get("net_weight", ""))
        _write_template_value(activity_ws, row_idx, activity_cols.get("gross_weight"), r.get("gross_weight", ""))
        _write_template_value(activity_ws, row_idx, activity_cols.get("weight_unit"), _localized_dropdown_display(r.get("weight_uom", ""), weight_unit_display_map))

        activity_ws.cell(row_idx, activity_cols["start_date"]).number_format = "yyyy/mm/dd"
        activity_ws.cell(row_idx, activity_cols["end_date"]).number_format = "yyyy/mm/dd"
        row_idx += 1

    raw_unique = (
        exploded.sort_values(["raw_material"])
        .drop_duplicates(subset=["raw_material"])
        [["raw_material", "description"]]
        if not exploded.empty else pd.DataFrame(columns=["raw_material", "description"])
    )

    row_idx = DATA_START_ROW
    for _, r in raw_unique.iterrows():
        raw_material = r["raw_material"]
        description = r["description"]

        _write_template_value(raw_ws, row_idx, raw_cols["raw_name"], raw_material)
        _write_template_value(raw_ws, row_idx, raw_cols["raw_code"], raw_material)
        _write_template_value(raw_ws, row_idx, raw_cols["description"], description)
        row_idx += 1

    wb.save(output_path)

    return {
        "output_filename": output_path.name,
        "activity_template_columns": activity_cols,
        "raw_material_template_columns": raw_cols,
        "activity_rows": int(len(exploded)),
        "raw_materials": int(exploded["raw_material"].nunique()) if not exploded.empty else 0,
        "zero_usage_rows_excluded": int(zero_usage_rows_excluded),
    }


def generate_raw_material_bulk_file(
    bom_path: str | Path | list[str | Path] | tuple[str | Path, ...],
    raw_material_template_path: str | Path,
    output_path: str | Path,
    mapping: dict[str, str | None] | None = None,
) -> Dict[str, Any]:
    """Generate one Raw Material Bulk workbook for all target products."""
    output_path = Path(output_path)

    bom_df, used_columns = _read_boms(bom_path, mapping=mapping)
    exploded, summary = _explode_bom(bom_df)
    exploded, zero_usage_rows_excluded = _exclude_zero_usage_rows(exploded)
    summary["zero_usage_rows_excluded"] = int(zero_usage_rows_excluded)
    summary["activity_rows"] = int(len(exploded))
    summary["raw_materials"] = int(exploded["raw_material"].nunique()) if not exploded.empty else 0

    write_summary = _write_raw_material_bulk_from_exploded(
        exploded=exploded,
        raw_material_template_path=raw_material_template_path,
        output_path=output_path,
    )
    summary.update(write_summary)

    summary["used_columns"] = used_columns
    summary["bom_files"] = int(used_columns.get("bom_files", 1)) if isinstance(used_columns, dict) else 1
    summary["bom_rows_before_dedup"] = int(used_columns.get("bom_rows_before_dedup", 0)) if isinstance(used_columns, dict) else 0
    summary["bom_rows_after_dedup"] = int(used_columns.get("bom_rows_after_dedup", 0)) if isinstance(used_columns, dict) else 0
    summary["bom_duplicate_rows_removed"] = int(used_columns.get("bom_duplicate_rows_removed", 0)) if isinstance(used_columns, dict) else 0
    return summary


def _sanitize_filename_part(value: Any, fallback: str = "Unassigned") -> str:
    text = str(value or "").strip()
    if not text:
        text = fallback
    text = re.sub(r"[\\/:*?\"<>|]+", "_", text)
    text = re.sub(r"\s+", "_", text)
    return text[:80] or fallback


def _read_step1_product_master_maps(step1_output_path: str | Path) -> tuple[dict[str, str], Dict[str, Any]]:
    """Read Step1 output and return product master-data maps.

    Module 2 uses Step1 output only for product-level master data:
    - Material Number / Target Product -> Production Site for split export.

    Material Group is a raw-material attribute and must remain sourced from
    the Standard BOM Material group column. Step1 output must not overwrite it.
    """
    step1_output_path = Path(step1_output_path)
    try:
        df = pd.read_excel(step1_output_path, sheet_name="Plant_Material年度產量", dtype=object)
    except Exception:
        df = pd.read_excel(step1_output_path, sheet_name=0, dtype=object)

    material_col = _find_step1_column(df, ["Material Number", "Material", "Product Material Number"])
    site_col = _find_step1_column(df, ["Production Site", "production site", "生產廠區", "廠區", "廠別"])

    work = df.copy()
    work["_material_key"] = work[material_col].apply(_normalize_material_key)
    work["_production_site"] = work[site_col].apply(_safe_text)
    work = work[work["_material_key"] != ""].copy()

    site_map: dict[str, str] = {}
    duplicate_site_conflicts: list[str] = []

    for material, group in work.groupby("_material_key", dropna=False):
        sites = sorted({str(x).strip() for x in group["_production_site"] if str(x).strip()})

        if sites:
            site_map[str(material)] = sites[0]
            if len(sites) > 1:
                duplicate_site_conflicts.append(f"{material}: {', '.join(sites)}")

    return site_map, {
        "step1_rows": int(len(df)),
        "step1_mapped_materials": int(len(site_map)),
        "step1_site_conflicts": duplicate_site_conflicts,
        "material_group_source": "Standard BOM",
        "transportation_destination_source": "Step1 Production Site",
    }


def _read_step1_annual_quantity_map(step1_output_path: str | Path) -> tuple[dict[str, float], Dict[str, Any]]:
    """Read Module 1 Step1 output and return Finished Product -> annual quantity.

    Raw Material Bulk Usage is the annual required amount, not per-PC BOM usage:
        final usage = exploded BOM usage per PC × annual finished-product quantity.
    """
    step1_output_path = Path(step1_output_path)
    try:
        df = pd.read_excel(step1_output_path, sheet_name="Plant_Material年度產量", dtype=object)
    except Exception:
        df = pd.read_excel(step1_output_path, sheet_name=0, dtype=object)

    material_col = _find_step1_column(df, ["Material Number", "Material", "Product Material Number"])
    qty_col = _find_step1_column(df, ["年度生產量", "Annual Quantity", "Delivered quantity"])

    work = df.copy()
    work["_material_key"] = work[material_col].apply(_normalize_material_key)
    work["_annual_qty"] = work[qty_col].apply(_safe_number)
    work = work[work["_material_key"] != ""].copy()

    qty_map: dict[str, float] = {}
    for _, r in work.groupby("_material_key", dropna=False, as_index=False)["_annual_qty"].sum().iterrows():
        material = str(r["_material_key"] or "").strip().upper()
        if material:
            qty_map[material] = float(r["_annual_qty"] or 0.0)

    return qty_map, {
        "annual_quantity_source": "Module 1 Step1 Plant_Material年度產量",
        "annual_quantity_mapped_products": int(len(qty_map)),
        "raw_material_usage_rule": "Final Usage = BOM exploded usage per PC × Module 1 annual finished-product quantity",
    }

def _step1_row_is_wip(product_type: Any = "", is_wip: Any = "") -> bool:
    """Return whether a Step1 product row should be treated as WIP/semi-finished."""
    type_text = _safe_text(product_type).strip().upper()
    wip_text = _safe_text(is_wip).strip().upper()
    if wip_text in {"1", "Y", "YES", "TRUE", "T", "WIP", "是"}:
        return True
    if type_text == "WIP" or "WIP" in type_text or "半品" in type_text:
        return True
    return False


def _read_working_hour_rollup_m2b_product_eligibility_map(
    working_hour_rollup_path: str | Path,
) -> tuple[dict[str, bool], dict[str, str], Dict[str, Any]]:
    """Read M2A Working Hour Roll-up and return whether each product is allowed into M2B.

    The roll-up Summary already contains Direct + Semi-finished working hours.
    M2B must use Total Annual Working Hour from this file so products whose
    direct hour is zero but semi-finished roll-up hour is positive are retained.
    """
    path = Path(working_hour_rollup_path)
    if not path.exists():
        return {}, {}, {
            "m2b_product_filter_applied": False,
            "m2b_product_filter_reason": f"找不到 M2A working_hour_rollup：{path}",
            "m2b_product_filter_source": "M2A working_hour_rollup missing",
            "m2b_excluded_products": 0,
        }

    try:
        df = pd.read_excel(path, sheet_name="Summary", dtype=object)
    except Exception:
        df = pd.read_excel(path, sheet_name=0, dtype=object)

    material_col = _find_step1_column(df, ["Material Number", "Material", "Product Material Number"])
    total_hour_col = _find_step1_optional_column(
        df,
        ["Total Annual Working Hour", "總年度工時", "年度總工時", "Total working hours"],
    )
    type_col = _find_step1_optional_column(df, ["Product Type", "產品類型"])
    wip_col = _find_step1_optional_column(df, ["Is_WIP", "Is WIP", "WIP"])

    if not total_hour_col:
        return {}, {}, {
            "m2b_product_filter_applied": False,
            "m2b_product_filter_reason": "M2A working_hour_rollup 找不到 Total Annual Working Hour 欄位。",
            "m2b_product_filter_source": path.name,
            "m2b_excluded_products": 0,
        }

    work = df.copy()
    work["_material_key"] = work[material_col].apply(_normalize_material_key)
    work["_total_annual_hour"] = work[total_hour_col].apply(_safe_number)
    work["_product_type"] = work[type_col].apply(_safe_text) if type_col else ""
    work["_is_wip_text"] = work[wip_col].apply(_safe_text) if wip_col else ""
    work["_is_wip_bool"] = work.apply(
        lambda r: _step1_row_is_wip(r.get("_product_type", ""), r.get("_is_wip_text", "")),
        axis=1,
    )
    work = work[work["_material_key"] != ""].copy()

    eligibility: dict[str, bool] = {}
    reasons: dict[str, str] = {}
    excluded_wip = 0
    excluded_zero_hour = 0
    eligible_count = 0

    for material, group in work.groupby("_material_key", dropna=False):
        key = str(material or "").strip().upper()
        if not key:
            continue
        non_wip = group[~group["_is_wip_bool"].astype(bool)]
        if non_wip.empty:
            eligibility[key] = False
            reasons[key] = "WIP"
            excluded_wip += 1
            continue
        positive_hour = non_wip[
            pd.to_numeric(non_wip["_total_annual_hour"], errors="coerce").fillna(0.0) > 0.0
        ]
        if positive_hour.empty:
            eligibility[key] = False
            reasons[key] = "Total Annual Working Hour=0/空白/非數字"
            excluded_zero_hour += 1
            continue
        eligibility[key] = True
        eligible_count += 1

    excluded_examples = [
        {"material": material, "reason": reasons.get(material, "")}
        for material in sorted(reasons.keys())[:50]
    ]
    return eligibility, reasons, {
        "m2b_product_filter_applied": True,
        "m2b_product_filter_source": f"M2A working_hour_rollup: {path.name}",
        "m2b_product_filter_rule": (
            "Before Module 2B Raw Material Bulk output, exclude WIP products and target products "
            "whose M2A Total Annual Working Hour (Direct + Semi-finished roll-up) is 0/blank/non-numeric."
        ),
        "m2b_products_checked": int(len(eligibility)),
        "m2b_eligible_products": int(eligible_count),
        "m2b_excluded_products": int(len(reasons)),
        "m2b_excluded_wip_products": int(excluded_wip),
        "m2b_excluded_zero_hour_products": int(excluded_zero_hour),
        "m2b_excluded_product_examples": excluded_examples,
    }


def _read_step1_m2b_product_eligibility_map(step1_output_path: str | Path) -> tuple[dict[str, bool], dict[str, str], Dict[str, Any]]:
    """Read Step1 output and return whether each product is allowed into M2B.

    M2B product-entry filter:
    - WIP / semi-finished products are excluded.
    - Finished products whose Module 1A 年度總工時 is 0, blank, or non-numeric are excluded.

    The current M2B large-dataset flow is keyed by target product material.  If
    a material appears more than once, it is kept only when at least one non-WIP
    Step1 row has 年度總工時 > 0.
    """
    step1_output_path = Path(step1_output_path)
    try:
        df = pd.read_excel(step1_output_path, sheet_name=STEP1_SOURCE_SHEET_NAME, dtype=object)
    except Exception:
        df = pd.read_excel(step1_output_path, sheet_name=0, dtype=object)

    material_col = _find_step1_column(df, ["Material Number", "Material", "Product Material Number"])
    hour_col = _find_step1_optional_column(df, ["年度總工時", "Total working hours", "Selected Hours", "Total Hours", "Working Hours"])
    type_col = _find_step1_optional_column(df, ["產品類型", "Product Type"])
    wip_col = _find_step1_optional_column(df, ["Is_WIP", "Is WIP", "WIP"])

    if not hour_col:
        return {}, {}, {
            "m2b_product_filter_applied": False,
            "m2b_product_filter_reason": "Module 1A 年度產品產量與分類結果找不到年度總工時欄位，未套用成品年度總工時=0排除。",
            "m2b_product_filter_rule": "Exclude WIP products and finished products whose Module 1A 年度總工時 is 0/blank/non-numeric before Module 2B Raw Material Bulk output.",
            "m2b_excluded_products": 0,
        }

    work = df.copy()
    work["_material_key"] = work[material_col].apply(_normalize_material_key)
    work["_annual_total_hour"] = work[hour_col].apply(_safe_number)
    work["_product_type"] = work[type_col].apply(_safe_text) if type_col else ""
    work["_is_wip_text"] = work[wip_col].apply(_safe_text) if wip_col else ""
    work["_is_wip_bool"] = work.apply(lambda r: _step1_row_is_wip(r.get("_product_type", ""), r.get("_is_wip_text", "")), axis=1)
    work = work[work["_material_key"] != ""].copy()

    eligibility: dict[str, bool] = {}
    reasons: dict[str, str] = {}
    excluded_wip = 0
    excluded_zero_hour = 0
    eligible_count = 0

    for material, group in work.groupby("_material_key", dropna=False):
        key = str(material or "").strip().upper()
        if not key:
            continue
        non_wip = group[~group["_is_wip_bool"].astype(bool)]
        if non_wip.empty:
            eligibility[key] = False
            reasons[key] = "WIP"
            excluded_wip += 1
            continue
        positive_hour = non_wip[pd.to_numeric(non_wip["_annual_total_hour"], errors="coerce").fillna(0.0) > 0.0]
        if positive_hour.empty:
            eligibility[key] = False
            reasons[key] = "年度總工時=0/空白/非數字"
            excluded_zero_hour += 1
            continue
        eligibility[key] = True
        eligible_count += 1

    excluded_examples = [
        {"material": material, "reason": reasons.get(material, "")}
        for material in sorted(reasons.keys())[:50]
    ]
    return eligibility, reasons, {
        "m2b_product_filter_applied": True,
        "m2b_product_filter_source": "Module 1A Plant_Material年度產量",
        "m2b_product_filter_rule": "Before Module 2B Raw Material Bulk output, exclude target products that are WIP or whose Module 1A 年度總工時 is 0/blank/non-numeric.",
        "m2b_products_checked": int(len(eligibility)),
        "m2b_eligible_products": int(eligible_count),
        "m2b_excluded_products": int(len(reasons)),
        "m2b_excluded_wip_products": int(excluded_wip),
        "m2b_excluded_zero_hour_products": int(excluded_zero_hour),
        "m2b_excluded_product_examples": excluded_examples,
    }



def _apply_annual_quantity_to_exploded_usage(
    exploded: pd.DataFrame,
    annual_qty_map: dict[str, float] | None,
) -> tuple[pd.DataFrame, Dict[str, Any]]:
    """Convert per-PC BOM usage into annual raw-material requirement."""
    if exploded is None or exploded.empty or "target_product" not in exploded.columns or "usage" not in exploded.columns:
        return exploded.copy() if isinstance(exploded, pd.DataFrame) else pd.DataFrame(), {
            "annual_quantity_applied": False,
            "annual_quantity_missing_rows": 0,
        }

    annual_qty_map = annual_qty_map or {}
    work = exploded.copy()
    target_keys = work["target_product"].apply(_normalize_material_key)
    annual_qty = target_keys.map(annual_qty_map)
    found_mask = annual_qty.notna()

    work["usage_per_pc"] = pd.to_numeric(work["usage"], errors="coerce").fillna(0.0)
    work["annual_finished_product_qty"] = annual_qty.where(found_mask, 1.0).astype(float)
    work.loc[found_mask, "usage"] = work.loc[found_mask, "usage_per_pc"] * work.loc[found_mask, "annual_finished_product_qty"]

    missing_targets = sorted(set(target_keys.loc[~found_mask].astype(str))) if (~found_mask).any() else []
    return work, {
        "annual_quantity_applied": True,
        "annual_quantity_matched_rows": int(found_mask.sum()),
        "annual_quantity_missing_rows": int((~found_mask).sum()),
        "annual_quantity_missing_targets": missing_targets[:50],
        "usage_per_pc_column_added": True,
        "annual_finished_product_qty_column_added": True,
    }




def _write_bom_trace_detail_file(
    trace_detail: pd.DataFrame,
    annual_qty_map: dict[str, float] | None,
    output_path: str | Path,
) -> Dict[str, Any]:
    """Write ungrouped BOM trace rows for debugging unexpected usage totals.

    This file shows every raw-material path before final groupby aggregation,
    so a value such as 1.3 can be checked as 0.3 + 1.0 from separate paths.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    annual_qty_map = annual_qty_map or {}
    work = trace_detail.copy() if isinstance(trace_detail, pd.DataFrame) else pd.DataFrame()
    if work.empty:
        work = pd.DataFrame(columns=[
            "target_product", "raw_material", "immediate_parent", "trace_path",
            "usage_per_path", "annual_finished_product_qty", "final_usage_per_path"
        ])
    else:
        target_keys = work["target_product"].apply(_normalize_material_key)
        annual_qty = target_keys.map(annual_qty_map)
        found_mask = annual_qty.notna()
        work["annual_finished_product_qty"] = annual_qty.where(found_mask, 1.0).astype(float)
        work["final_usage_per_path"] = pd.to_numeric(work.get("usage_per_path", work.get("usage", 0.0)), errors="coerce").fillna(0.0) * work["annual_finished_product_qty"]
        work["annual_qty_found"] = found_mask

    preferred_cols = [
        "target_product", "source_material", "raw_material", "unit", "immediate_parent", "level", "trace_path",
        "parent_accumulated_qty", "qty_this_level_original", "qty_this_level_effective",
        "qty_adjusted_by_altitem", "altitem_group", "usage_probability_ratio",
        "usage_per_path", "annual_finished_product_qty", "final_usage_per_path",
        "description", "material_group", "valid_from", "annual_qty_found", "source_file",
    ]
    ordered_cols = [c for c in preferred_cols if c in work.columns] + [c for c in work.columns if c not in preferred_cols]
    work = work[ordered_cols]

    summary = pd.DataFrame()
    if not work.empty and {"target_product", "raw_material", "unit", "final_usage_per_path"}.issubset(work.columns):
        summary = (
            work.groupby(["target_product", "raw_material", "unit"], dropna=False, as_index=False)
            .agg({
                "final_usage_per_path": "sum",
                "usage_per_path": "sum" if "usage_per_path" in work.columns else "first",
                "trace_path": "count" if "trace_path" in work.columns else "first",
            })
            .rename(columns={"final_usage_per_path": "final_usage_total", "usage_per_path": "usage_per_pc_total", "trace_path": "path_count"})
            .sort_values(["target_product", "raw_material"])
            .reset_index(drop=True)
        )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        work.to_excel(writer, index=False, sheet_name="Trace Detail")
        summary.to_excel(writer, index=False, sheet_name="Grouped Summary")
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            for col in sheet.columns:
                max_len = 12
                letter = col[0].column_letter
                for cell in col[:500]:
                    max_len = max(max_len, len(str(cell.value or "")) + 2)
                sheet.column_dimensions[letter].width = min(max_len, 60)

    return {
        "bom_trace_filename": output_path.name,
        "bom_trace_download_url": f"/download/{output_path.name}",
        "bom_trace_rows": int(len(work)),
        "bom_trace_grouped_rows": int(len(summary)),
    }

def generate_raw_material_bulk_files_by_site_zip(
    bom_path: str | Path | list[str | Path] | tuple[str | Path, ...],
    raw_material_template_path: str | Path,
    output_dir: str | Path,
    token: str,
    step1_output_path: str | Path,
    mapping: dict[str, str | None] | None = None,
) -> Dict[str, Any]:
    """Generate one Raw Material Bulk workbook per Production Site and ZIP them.

    Split source:
    - BOM explosion gives Target Product -> Raw Material usage.
    - Step1 output gives Material Number / Target Product -> Production Site.
    - Rows whose target product is not found in Step1 are exported under
      "Unassigned" so they are visible instead of silently dropped.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    bom_df, used_columns = _read_boms(bom_path, mapping=mapping)
    exploded, base_summary = _explode_bom(bom_df)
    exploded, zero_usage_rows_excluded = _exclude_zero_usage_rows(exploded)

    annual_qty_map, annual_qty_source_summary = _read_step1_annual_quantity_map(step1_output_path)
    # Production mode: do not generate bom_trace_detail_*.xlsx.
    # That file was only for BOM debugging and must not be included in the Module 2 ZIP,
    # because Module 3 consumes every Excel in the package as raw-material bulk input.
    trace_summary: dict[str, Any] = {}
    exploded, annual_usage_summary = _apply_annual_quantity_to_exploded_usage(exploded, annual_qty_map)
    exploded, zero_annual_usage_rows_excluded = _exclude_zero_usage_rows(exploded)

    total_hour_by_target, working_hour_summary = _calculate_total_working_hour_by_target(
        step1_output_path=step1_output_path,
        bom_df=bom_df,
    )
    exploded, zero_total_working_hour_rows_excluded = _exclude_zero_total_working_hour_target_rows(
        exploded=exploded,
        total_hour_by_material=total_hour_by_target,
    )

    base_summary["zero_usage_rows_excluded"] = int(zero_usage_rows_excluded)
    base_summary["zero_annual_usage_rows_excluded"] = int(zero_annual_usage_rows_excluded)
    base_summary["zero_total_working_hour_rows_excluded"] = int(zero_total_working_hour_rows_excluded)
    base_summary.update(annual_qty_source_summary)
    base_summary.update(annual_usage_summary)
    base_summary.update(trace_summary)
    base_summary["activity_rows"] = int(len(exploded))
    base_summary["raw_materials"] = int(exploded["raw_material"].nunique()) if not exploded.empty else 0
    base_summary.update(working_hour_summary)

    site_map, step1_summary = _read_step1_product_master_maps(step1_output_path)

    work = exploded.copy()
    if work.empty:
        work["_production_site"] = ""
    else:
        work["_target_key"] = work["target_product"].apply(_normalize_material_key)
        work["_production_site"] = work["_target_key"].map(site_map).fillna("Unassigned")
        work["_production_site"] = work["_production_site"].apply(lambda x: str(x or "").strip() or "Unassigned")
        # V14.6: Transportation Destination in Raw Material Bulk follows Step1 Production Site.
        # Step1 Output is the product master-data source; Standard BOM only provides material structure and usage.
        work["transport_destination"] = work["_production_site"]

        # V14.7: Material Group is a raw-material attribute from Standard BOM.
        # Do not overwrite it with Step1 product classification fields.
        if "material_group" not in work.columns:
            work["material_group"] = ""

    site_values = sorted({str(x).strip() or "Unassigned" for x in work["_production_site"].tolist()}) if not work.empty else ["Unassigned"]

    generated_files: list[dict[str, Any]] = []
    zip_filename = f"raw_material_activity_data_bulk_by_site_{token}.zip"
    zip_path = output_dir / zip_filename

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for site in site_values:
            site_df = work[work["_production_site"] == site].copy()
            site_df = site_df.drop(columns=["_target_key", "_production_site"], errors="ignore")
            safe_site = _sanitize_filename_part(site)
            file_path = output_dir / f"raw_material_activity_data_bulk_{safe_site}_{token}.xlsx"

            write_summary = _write_raw_material_bulk_from_exploded(
                exploded=site_df,
                raw_material_template_path=raw_material_template_path,
                output_path=file_path,
            )
            zf.write(file_path, arcname=file_path.name)
            generated_files.append({
                "production_site": site,
                "filename": file_path.name,
                "activity_rows": int(write_summary.get("activity_rows", 0)),
                "raw_materials": int(write_summary.get("raw_materials", 0)),
                "supplier_usage_split_source_rows": int(write_summary.get("supplier_usage_split_source_rows", 0)),
                "supplier_usage_split_output_rows": int(write_summary.get("supplier_usage_split_output_rows", 0)),
                "supplier_usage_total_before_split": float(write_summary.get("supplier_usage_total_before_split", 0.0) or 0.0),
                "supplier_usage_total_after_split": float(write_summary.get("supplier_usage_total_after_split", 0.0) or 0.0),
            })

    unassigned_rows = int((work["_production_site"] == "Unassigned").sum()) if not work.empty else 0

    summary = dict(base_summary)
    summary.update({
        "output_filename": zip_filename,
        "download_url": f"/download/{zip_filename}",
        "split_by_production_site": True,
        "production_site_files": generated_files,
        "production_site_count": int(len(site_values)),
        "unassigned_rows": unassigned_rows,
        "used_columns": used_columns,
        "bom_files": int(used_columns.get("bom_files", 1)) if isinstance(used_columns, dict) else 1,
        "bom_rows_before_dedup": int(used_columns.get("bom_rows_before_dedup", 0)) if isinstance(used_columns, dict) else 0,
        "bom_rows_after_dedup": int(used_columns.get("bom_rows_after_dedup", 0)) if isinstance(used_columns, dict) else 0,
        "bom_duplicate_rows_removed": int(used_columns.get("bom_duplicate_rows_removed", 0)) if isinstance(used_columns, dict) else 0,
    })
    summary.update(step1_summary)
    return summary


def _explode_bom_structure(df: pd.DataFrame) -> tuple[pd.DataFrame, Dict[str, Any]]:
    """Create version-aware structure including cross-Material semi-finished edges."""
    rows: list[dict[str, Any]] = []
    cycle_count = 0
    product_count = 0
    cross_material_expansions = 0
    semi_finished_total: set[str] = set()
    columns = [
        "Target Product", "Target BOM Valid From", "Source Material", "Source BOM Valid From",
        "Parent Material", "Component", "Quantity Per Parent", "Accumulated Quantity",
        "Unit", "Component Description", "Material Group", "Valid From", "Level",
        "Is Semi-finished", "Expansion Type", "Source File",
    ]

    if df is None or df.empty:
        structure = pd.DataFrame(columns=columns)
        return structure, {
            "products": 0, "semi_finished": 0, "structure_rows": 0,
            "max_level": 0, "cycles_skipped": 0, "cross_material_expansions": 0,
        }

    context = _build_bom_scope_context(df)
    scopes = context["scopes"]
    material_index = context["material_index"]

    for scope_key, root_scope in scopes.items():
        material = root_scope["material"]
        target_valid_from = root_scope["valid_from"]
        roots = root_scope["roots"]
        product_count += len(roots)

        for root in roots:
            target_product = material or root
            stack: list[tuple[tuple[str, Any], str, float, int, list[str], list[tuple[str, Any]]]] = [
                (scope_key, root, 1.0, 0, [root], [scope_key])
            ]
            while stack:
                current_scope_key, current_parent, accumulated_qty, level, path, scope_path = stack.pop()
                current_scope = scopes[current_scope_key]
                current_children = current_scope["children"]

                for child in current_children.get(current_parent, []):
                    component = child["component"]
                    next_qty = accumulated_qty * child["qty"]
                    next_level = level + 1
                    if component in path:
                        cycle_count += 1
                        continue

                    next_scope_key: tuple[str, Any] | None = None
                    expansion_type = "Raw material"
                    if component in current_children:
                        next_scope_key = current_scope_key
                        expansion_type = "Local Material"
                    elif context["has_material"]:
                        next_scope_key = _select_cross_material_scope(
                            component=component,
                            preferred_valid_from=child.get("valid_from", current_scope["valid_from"]),
                            material_index=material_index,
                            scopes=scopes,
                        )
                        if next_scope_key is not None:
                            expansion_type = "Cross Material"

                    is_semi = next_scope_key is not None
                    if is_semi:
                        semi_finished_total.add(component)
                        if next_scope_key != current_scope_key:
                            cross_material_expansions += 1

                    rows.append({
                        "Target Product": target_product,
                        "Target BOM Valid From": target_valid_from,
                        "Source Material": current_scope["material"],
                        "Source BOM Valid From": current_scope["valid_from"],
                        "Parent Material": current_parent,
                        "Component": component,
                        "Quantity Per Parent": child["qty"],
                        "Accumulated Quantity": next_qty,
                        "Unit": child["uom"],
                        "Component Description": child["description"],
                        "Material Group": child["material_group"],
                        "Valid From": child["valid_from"],
                        "Level": next_level,
                        "Is Semi-finished": "Y" if is_semi else "N",
                        "Expansion Type": expansion_type,
                        "Source File": child.get("source_file", ""),
                    })

                    if not is_semi:
                        continue
                    if next_scope_key in scope_path and next_scope_key != current_scope_key:
                        cycle_count += 1
                        continue
                    stack.append((
                        next_scope_key,
                        component,
                        next_qty,
                        next_level,
                        path + [component],
                        scope_path + ([next_scope_key] if next_scope_key != current_scope_key else []),
                    ))

    structure = pd.DataFrame(rows, columns=columns)
    summary = {
        "products": int(product_count),
        "semi_finished": int(len(semi_finished_total)),
        "structure_rows": int(len(structure)),
        "max_level": int(structure["Level"].max()) if not structure.empty else 0,
        "cycles_skipped": int(cycle_count),
        "cross_material_expansions": int(cross_material_expansions),
    }
    return structure, summary

def export_bom_structure_file(
    bom_path: str | Path | list[str | Path] | tuple[str | Path, ...],
    output_path: str | Path,
    mapping: dict[str, str | None] | None = None,
) -> Dict[str, Any]:
    """Export latest normalized BOM structure for Step 2 semi-finished working-hour roll-up."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    bom_df, used_columns = _read_boms(bom_path, mapping=mapping)
    structure, summary = _explode_bom_structure(bom_df)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        structure.to_excel(writer, index=False, sheet_name="BOM Structure")
        ws = writer.book["BOM Structure"]
        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_len = 12
            letter = col[0].column_letter
            for cell in col[:1000]:
                max_len = max(max_len, len(str(cell.value or "")) + 2)
            ws.column_dimensions[letter].width = min(max_len, 45)
    summary["output_filename"] = output_path.name
    summary["used_columns"] = used_columns
    summary["bom_files"] = int(used_columns.get("bom_files", 1)) if isinstance(used_columns, dict) else 1
    summary["bom_rows_before_dedup"] = int(used_columns.get("bom_rows_before_dedup", 0)) if isinstance(used_columns, dict) else 0
    summary["bom_rows_after_dedup"] = int(used_columns.get("bom_rows_after_dedup", 0)) if isinstance(used_columns, dict) else 0
    summary["bom_duplicate_rows_removed"] = int(used_columns.get("bom_duplicate_rows_removed", 0)) if isinstance(used_columns, dict) else 0
    return summary


# =========================================================
# V12 Working Hour Roll-up
# Step1 Output + BOM Structure -> working_hour_rollup.xlsx
# =========================================================

STEP1_SOURCE_SHEET_NAME = "Plant_Material年度產量"


def _find_step1_column(df: pd.DataFrame, candidates: list[str]) -> str:
    normalized = {_normalize_col(c).lower(): c for c in df.columns}
    for name in candidates:
        key = _normalize_col(name).lower()
        if key in normalized:
            return normalized[key]
    raise ValueError(f"找不到 Step1 欄位：{', '.join(candidates)}")


def _find_step1_optional_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    try:
        return _find_step1_column(df, candidates)
    except ValueError:
        return None


def _normalize_material_key(value: Any) -> str:
    return str(value or "").strip().upper()


def _read_bom_structure_file(path: str | Path) -> pd.DataFrame:
    path = Path(path)
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_excel(path, sheet_name="BOM Structure", dtype=object)
    except ValueError:
        return pd.read_excel(path, sheet_name=0, dtype=object)


def generate_working_hour_rollup_file(
    step1_output_path: str | Path,
    bom_structure_path: str | Path,
    output_path: str | Path,
) -> Dict[str, Any]:
    """Generate auditable working-hour roll-up workbook.

    Summary.Total Annual Working Hour is the source used by Step2 when
    Include Semi-finished Working Hour is selected.
    """
    step1_output_path = Path(step1_output_path)
    bom_structure_path = Path(bom_structure_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    step1_df = pd.read_excel(step1_output_path, sheet_name=STEP1_SOURCE_SHEET_NAME, dtype=object)
    bom_df = _read_bom_structure_file(bom_structure_path)
    if bom_df.empty:
        raise ValueError("BOM Structure is empty. Please complete Module 2 → BOM Expansion first.")

    material_col = _find_step1_column(step1_df, ["Material Number"])
    qty_col = _find_step1_column(step1_df, ["年度生產量", "Annual Quantity", "Delivered quantity"])
    hour_col = _find_step1_optional_column(step1_df, ["年度總工時", "Total working hours", "Selected Hours", "Total Hours", "Working Hours"])
    if not hour_col:
        raise ValueError("Step1 Output 找不到年度總工時欄位，無法產生 working_hour_rollup.xlsx")

    plant_col = _find_step1_optional_column(step1_df, ["Plant"])
    site_col = _find_step1_optional_column(step1_df, ["Production Site", "production site", "生產廠區", "廠區", "廠別"])
    type_col = _find_step1_optional_column(step1_df, ["產品類型", "Product Type"])
    wip_col = _find_step1_optional_column(step1_df, ["Is_WIP", "Is WIP", "WIP"])

    work = step1_df.copy()
    work["_material_key"] = work[material_col].apply(_normalize_material_key)
    work["_annual_qty"] = work[qty_col].apply(_safe_number)
    work["_direct_hour"] = work[hour_col].apply(_safe_number)
    work["_plant"] = work[plant_col].apply(_safe_text) if plant_col else ""
    work["_production_site"] = work[site_col].apply(_safe_text) if site_col else ""
    work["_product_type"] = work[type_col].apply(_safe_text) if type_col else ""
    work["_is_wip"] = work[wip_col].apply(_safe_text) if wip_col else ""

    summary_base = work.groupby(
        ["_material_key", "_plant", "_production_site", "_product_type", "_is_wip"],
        dropna=False,
        as_index=False,
    ).agg({"_annual_qty": "sum", "_direct_hour": "sum"})

    parent_values = set(bom_df["Parent Material"].dropna().astype(str).str.strip().str.upper()) if "Parent Material" in bom_df.columns else set()
    component_values = set(bom_df["Component"].dropna().astype(str).str.strip().str.upper()) if "Component" in bom_df.columns else set()
    semi_materials = parent_values.intersection(component_values)

    material_totals = work.groupby(["_material_key"], dropna=False, as_index=False).agg({"_annual_qty": "sum", "_direct_hour": "sum"})
    qty_by_material = {}
    direct_by_material = {}
    hour_per_pc_by_material = {}
    for _, r in material_totals.iterrows():
        material = str(r["_material_key"] or "").strip()
        qty = float(r["_annual_qty"] or 0)
        hours = float(r["_direct_hour"] or 0)
        qty_by_material[material] = qty
        direct_by_material[material] = hours
        hour_per_pc_by_material[material] = hours / qty if qty else 0.0

    semi_hour_per_pc_rows = [{
        "Semi Material": m,
        "Semi Annual Qty": qty_by_material.get(m, 0.0),
        "Semi Direct Annual Working Hour": direct_by_material.get(m, 0.0),
        "Semi Direct Hour per PC": hour_per_pc_by_material.get(m, 0.0),
    } for m in sorted(semi_materials)]
    semi_hour_per_pc_df = pd.DataFrame(semi_hour_per_pc_rows)

    target_col = _find_step1_optional_column(bom_df, ["Target Product", "target_product"])
    parent_col = _find_step1_optional_column(bom_df, ["Parent Material", "parent_material"])
    comp_col = _find_step1_optional_column(bom_df, ["Component", "component"])
    acc_col = _find_step1_optional_column(bom_df, ["Accumulated Quantity", "usage", "Quantity"])
    level_col = _find_step1_optional_column(bom_df, ["Level", "level"])
    semi_flag_col = _find_step1_optional_column(bom_df, ["Is Semi-finished", "Is Semi", "semi_finished"])
    if not target_col or not comp_col or not acc_col:
        raise ValueError("BOM Structure 缺少 Target Product、Component 或 Accumulated Quantity 欄位")

    b = bom_df.copy()
    b["_target_key"] = b[target_col].apply(_normalize_material_key)
    b["_component_key"] = b[comp_col].apply(_normalize_material_key)
    b["_accumulated_qty"] = b[acc_col].apply(_safe_number)
    if semi_flag_col:
        b = b[b[semi_flag_col].astype(str).str.strip().str.upper().isin(["Y", "YES", "TRUE", "1"])].copy()
    else:
        b = b[b["_component_key"].isin(semi_materials)].copy()

    detail_rows = []
    semi_by_key: dict[tuple[str, str, str], float] = {}
    for _, target_row in summary_base.iterrows():
        target = str(target_row["_material_key"] or "").strip()
        plant = _safe_text(target_row["_plant"])
        site = _safe_text(target_row["_production_site"])
        target_qty = float(target_row["_annual_qty"] or 0)
        for _, edge in b[b["_target_key"] == target].iterrows():
            semi = str(edge["_component_key"] or "").strip()
            acc_qty = float(edge["_accumulated_qty"] or 0)
            semi_hr_pc = float(hour_per_pc_by_material.get(semi, 0.0) or 0.0)
            contrib_pc = acc_qty * semi_hr_pc
            contrib_annual = target_qty * contrib_pc
            if contrib_annual:
                key = (target, plant, site)
                semi_by_key[key] = semi_by_key.get(key, 0.0) + contrib_annual
            detail_rows.append({
                "Target Product": target,
                "Plant": plant,
                "Production Site": site,
                "Target Annual Qty": target_qty,
                "Parent Material": edge.get(parent_col, "") if parent_col else "",
                "Semi Material": semi,
                "BOM Accumulated Qty": acc_qty,
                "Semi Direct Hour per PC": semi_hr_pc,
                "Semi Hour Contribution per PC": contrib_pc,
                "Semi Annual Working Hour Contribution": contrib_annual,
                "Level": edge.get(level_col, "") if level_col else "",
            })

    detail_df = pd.DataFrame(detail_rows)
    if detail_df.empty:
        detail_df = pd.DataFrame(columns=["Target Product", "Plant", "Production Site", "Target Annual Qty", "Parent Material", "Semi Material", "BOM Accumulated Qty", "Semi Direct Hour per PC", "Semi Hour Contribution per PC", "Semi Annual Working Hour Contribution", "Level"])

    summary_rows = []
    for _, r in summary_base.iterrows():
        material = str(r["_material_key"] or "").strip()
        plant = _safe_text(r["_plant"])
        site = _safe_text(r["_production_site"])
        qty = float(r["_annual_qty"] or 0)
        direct = float(r["_direct_hour"] or 0)
        semi = float(semi_by_key.get((material, plant, site), 0.0) or 0.0)
        total = direct + semi
        summary_rows.append({
            "Material Number": material,
            "Plant": plant,
            "Production Site": site,
            "Product Type": _safe_text(r["_product_type"]),
            "Is_WIP": _safe_text(r["_is_wip"]),
            "Annual Qty": qty,
            "Direct Annual Working Hour": direct,
            "Semi Annual Working Hour": semi,
            "Total Annual Working Hour": total,
            "Direct Hour per PC": direct / qty if qty else 0.0,
            "Semi Hour per PC": semi / qty if qty else 0.0,
            "Total Hour per PC": total / qty if qty else 0.0,
        })
    summary_df = pd.DataFrame(summary_rows).sort_values(["Plant", "Production Site", "Material Number"]).reset_index(drop=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        detail_df.to_excel(writer, index=False, sheet_name="Roll-up Detail")
        semi_hour_per_pc_df.to_excel(writer, index=False, sheet_name="Semi Hour per PC")
        bom_df.to_excel(writer, index=False, sheet_name="BOM Structure")
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            for col in sheet.columns:
                max_len = 12
                letter = col[0].column_letter
                for cell in col[:1000]:
                    max_len = max(max_len, len(str(cell.value or "")) + 2)
                sheet.column_dimensions[letter].width = min(max_len, 48)

    return {
        "output_filename": output_path.name,
        "summary_rows": int(len(summary_df)),
        "detail_rows": int(len(detail_df)),
        "semi_materials": int(len(semi_hour_per_pc_df)),
        "total_direct_hours": float(summary_df["Direct Annual Working Hour"].sum()) if not summary_df.empty else 0.0,
        "total_semi_hours": float(summary_df["Semi Annual Working Hour"].sum()) if not summary_df.empty else 0.0,
        "total_hours": float(summary_df["Total Annual Working Hour"].sum()) if not summary_df.empty else 0.0,
    }



def generate_working_hour_rollup_file_from_standard_bom(
    step1_output_path: str | Path,
    bom_path: str | Path | list[str | Path] | tuple[str | Path, ...],
    output_path: str | Path,
    mapping: dict[str, str | None] | None = None,
    progress_callback=None,
) -> Dict[str, Any]:
    """Generate M1 Step2 working-hour roll-up directly from Standard BOM.

    This replaces the previous M2A intermediate path:
      Standard BOM -> huge BOM Structure workbook -> read huge workbook -> roll-up.

    The old path created an openpyxl workbook containing hundreds of thousands of
    BOM Structure rows and could exceed Render memory.  This function keeps the
    Standard BOM graph in memory, streams the audit workbook with write_only=True,
    and only creates the Summary sheet required by M1 Step2 plus lightweight
    supporting sheets.
    """
    step1_output_path = Path(step1_output_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if progress_callback:
        progress_callback(step="Reading Step1 output for working-hour roll-up", processed=0, total=0, progress=88)

    step1_df = pd.read_excel(step1_output_path, sheet_name=STEP1_SOURCE_SHEET_NAME, dtype=object)

    material_col = _find_step1_column(step1_df, ["Material Number"])
    qty_col = _find_step1_column(step1_df, ["年度生產量", "Annual Quantity", "Delivered quantity"])
    hour_col = _find_step1_optional_column(step1_df, ["年度總工時", "Total working hours", "Selected Hours", "Total Hours", "Working Hours"])
    if not hour_col:
        raise ValueError("Step1 Output 找不到年度總工時欄位，無法產生 working_hour_rollup.xlsx")

    plant_col = _find_step1_optional_column(step1_df, ["Plant"])
    site_col = _find_step1_optional_column(step1_df, ["Production Site", "production site", "生產廠區", "廠區", "廠別"])
    type_col = _find_step1_optional_column(step1_df, ["產品類型", "Product Type"])
    wip_col = _find_step1_optional_column(step1_df, ["Is_WIP", "Is WIP", "WIP"])

    work = step1_df.copy()
    work["_material_key"] = work[material_col].apply(_normalize_material_key)
    work["_annual_qty"] = work[qty_col].apply(_safe_number)
    work["_direct_hour"] = work[hour_col].apply(_safe_number)
    work["_plant"] = work[plant_col].apply(_safe_text) if plant_col else ""
    work["_production_site"] = work[site_col].apply(_safe_text) if site_col else ""
    work["_product_type"] = work[type_col].apply(_safe_text) if type_col else ""
    work["_is_wip"] = work[wip_col].apply(_safe_text) if wip_col else ""
    work = work[work["_material_key"] != ""].copy()

    summary_base = work.groupby(
        ["_material_key", "_plant", "_production_site", "_product_type", "_is_wip"],
        dropna=False,
        as_index=False,
    ).agg({"_annual_qty": "sum", "_direct_hour": "sum"})

    material_totals = work.groupby(["_material_key"], dropna=False, as_index=False).agg({"_annual_qty": "sum", "_direct_hour": "sum"})
    qty_by_material: dict[str, float] = {}
    direct_by_material: dict[str, float] = {}
    hour_per_pc_by_material: dict[str, float] = {}
    for _, r in material_totals.iterrows():
        material = str(r["_material_key"] or "").strip()
        qty = float(r["_annual_qty"] or 0.0)
        hours = float(r["_direct_hour"] or 0.0)
        if material:
            qty_by_material[material] = qty
            direct_by_material[material] = hours
            hour_per_pc_by_material[material] = hours / qty if qty else 0.0

    target_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for _, r in summary_base.iterrows():
        target = str(r["_material_key"] or "").strip()
        if not target:
            continue
        target_groups[target].append({
            "plant": _safe_text(r["_plant"]),
            "site": _safe_text(r["_production_site"]),
            "product_type": _safe_text(r["_product_type"]),
            "is_wip": _safe_text(r["_is_wip"]),
            "annual_qty": float(r["_annual_qty"] or 0.0),
            "direct_hour": float(r["_direct_hour"] or 0.0),
        })

    if progress_callback:
        progress_callback(step="Reading Standard BOM for working-hour roll-up", processed=0, total=0, progress=89)
    bom_df, used_columns = _read_boms(bom_path, mapping=mapping)
    if bom_df.empty:
        raise ValueError("Standard BOM is empty. Please complete Module 2A with valid BOM files.")

    wb = Workbook(write_only=True)
    summary_ws = wb.create_sheet("Summary")
    detail_ws = wb.create_sheet("Roll-up Detail")
    semi_ws = wb.create_sheet("Semi Hour per PC")
    metadata_ws = wb.create_sheet("Metadata")

    summary_header = [
        "Material Number", "Plant", "Production Site", "Product Type", "Is_WIP",
        "Annual Qty", "Direct Annual Working Hour", "Semi Annual Working Hour", "Total Annual Working Hour",
        "Direct Hour per PC", "Semi Hour per PC", "Total Hour per PC",
    ]
    detail_header = [
        "Target Product", "Plant", "Production Site", "Target Annual Qty", "Parent Material", "Semi Material",
        "BOM Accumulated Qty", "Semi Direct Hour per PC", "Semi Hour Contribution per PC",
        "Semi Annual Working Hour Contribution", "Level",
    ]
    semi_header = ["Semi Material", "Semi Annual Qty", "Semi Direct Annual Working Hour", "Semi Direct Hour per PC"]
    summary_ws.append(summary_header)
    detail_ws.append(detail_header)
    semi_ws.append(semi_header)
    metadata_ws.append(["Key", "Value"])
    metadata_ws.append(["generator", "generate_working_hour_rollup_file_from_standard_bom"])
    metadata_ws.append(["source_rule", "M2A streams Working Hour Roll-up directly from Standard BOM; no large BOM Structure workbook is exported."])
    metadata_ws.append(["step1_source", step1_output_path.name])
    metadata_ws.append(["bom_files", int(used_columns.get("bom_files", 1)) if isinstance(used_columns, dict) else 1])
    metadata_ws.append(["bom_rows_after_dedup", int(used_columns.get("bom_rows_after_dedup", len(bom_df))) if isinstance(used_columns, dict) else int(len(bom_df))])

    semi_by_key: dict[tuple[str, str, str], float] = {}
    semi_materials: set[str] = set()
    detail_rows = 0
    cycle_count = 0
    products_seen: set[str] = set()
    max_level = 0
    scoped_count = 0

    context = _build_bom_scope_context(bom_df)
    scopes = context["scopes"]
    material_index = context["material_index"]
    total_scopes = max(len(scopes), 1)
    cross_material_expansions = 0

    for scoped_count, (scope_key, root_scope) in enumerate(scopes.items(), start=1):
        material = root_scope["material"]
        roots = root_scope["roots"]

        for root in roots:
            target_product = material or root
            target_key = _normalize_material_key(target_product)
            if target_key not in target_groups:
                continue
            products_seen.add(target_key)

            stack: list[tuple[tuple[str, Any], str, float, int, tuple[str, ...], tuple[tuple[str, Any], ...]]] = [
                (scope_key, root, 1.0, 0, (root,), (scope_key,))
            ]
            while stack:
                current_scope_key, current_parent, accumulated_qty, level, path, scope_path = stack.pop()
                current_scope = scopes[current_scope_key]
                current_children = current_scope["children"]

                for child in current_children.get(current_parent, []):
                    component = child["component"]
                    if component in path:
                        cycle_count += 1
                        continue

                    next_qty = accumulated_qty * float(child.get("qty") or 0.0)
                    next_level = level + 1
                    max_level = max(max_level, next_level)

                    next_scope_key: tuple[str, Any] | None = None
                    if component in current_children:
                        next_scope_key = current_scope_key
                    elif context["has_material"]:
                        next_scope_key = _select_cross_material_scope(
                            component=component,
                            preferred_valid_from=child.get("valid_from", current_scope["valid_from"]),
                            material_index=material_index,
                            scopes=scopes,
                        )

                    if next_scope_key is None:
                        continue
                    if next_scope_key in scope_path and next_scope_key != current_scope_key:
                        cycle_count += 1
                        continue

                    semi_key = _normalize_material_key(component)
                    if semi_key:
                        semi_materials.add(semi_key)
                    if next_scope_key != current_scope_key:
                        cross_material_expansions += 1

                    semi_hr_pc = float(hour_per_pc_by_material.get(semi_key, 0.0) or 0.0)
                    for tg in target_groups.get(target_key, []):
                        plant = _safe_text(tg.get("plant"))
                        site = _safe_text(tg.get("site"))
                        target_qty = float(tg.get("annual_qty") or 0.0)
                        contrib_pc = next_qty * semi_hr_pc
                        contrib_annual = target_qty * contrib_pc
                        if contrib_annual:
                            key = (target_key, plant, site)
                            semi_by_key[key] = semi_by_key.get(key, 0.0) + contrib_annual
                        detail_ws.append([
                            target_key, plant, site, target_qty, current_parent, semi_key,
                            next_qty, semi_hr_pc, contrib_pc, contrib_annual, next_level,
                        ])
                        detail_rows += 1

                    stack.append((
                        next_scope_key,
                        component,
                        next_qty,
                        next_level,
                        path + (component,),
                        scope_path + ((next_scope_key,) if next_scope_key != current_scope_key else ()),
                    ))

        if progress_callback and (scoped_count % 25 == 0 or scoped_count == total_scopes):
            progress = 89 + int((scoped_count / total_scopes) * 7)
            progress_callback(
                step="Generating Working Hour Roll-up (streaming)",
                processed=int(detail_rows),
                total=int(len(summary_base)),
                progress=min(96, progress),
            )

    summary_rows_count = 0
    total_direct = 0.0
    total_semi = 0.0
    summary_sorted = summary_base.sort_values(["_plant", "_production_site", "_material_key"]).reset_index(drop=True)
    for _, r in summary_sorted.iterrows():
        material = str(r["_material_key"] or "").strip()
        plant = _safe_text(r["_plant"])
        site = _safe_text(r["_production_site"])
        qty = float(r["_annual_qty"] or 0.0)
        direct = float(r["_direct_hour"] or 0.0)
        semi = float(semi_by_key.get((material, plant, site), 0.0) or 0.0)
        total = direct + semi
        total_direct += direct
        total_semi += semi
        summary_ws.append([
            material,
            plant,
            site,
            _safe_text(r["_product_type"]),
            _safe_text(r["_is_wip"]),
            qty,
            direct,
            semi,
            total,
            direct / qty if qty else 0.0,
            semi / qty if qty else 0.0,
            total / qty if qty else 0.0,
        ])
        summary_rows_count += 1

    for semi in sorted(semi_materials):
        semi_qty = float(qty_by_material.get(semi, 0.0) or 0.0)
        semi_direct = float(direct_by_material.get(semi, 0.0) or 0.0)
        semi_ws.append([semi, semi_qty, semi_direct, semi_direct / semi_qty if semi_qty else 0.0])

    metadata_ws.append(["summary_rows", summary_rows_count])
    metadata_ws.append(["detail_rows", detail_rows])
    metadata_ws.append(["semi_materials", len(semi_materials)])
    metadata_ws.append(["products_matched_to_bom", len(products_seen)])
    metadata_ws.append(["max_level", max_level])
    metadata_ws.append(["cycles_skipped", cycle_count])
    metadata_ws.append(["cross_material_expansions", cross_material_expansions])
    metadata_ws.append(["created_at", datetime.now().isoformat(timespec="seconds")])

    if progress_callback:
        progress_callback(step="Saving Working Hour Roll-up", processed=summary_rows_count, total=summary_rows_count, progress=97)
    wb.save(output_path)

    return {
        "output_filename": output_path.name,
        "summary_rows": int(summary_rows_count),
        "detail_rows": int(detail_rows),
        "semi_materials": int(len(semi_materials)),
        "products": int(len(products_seen)),
        "max_level": int(max_level),
        "cycles_skipped": int(cycle_count),
        "cross_material_expansions": int(cross_material_expansions),
        "total_direct_hours": float(total_direct),
        "total_semi_hours": float(total_semi),
        "total_hours": float(total_direct + total_semi),
        "used_columns": used_columns,
        "streaming_working_hour_rollup": True,
        "bom_structure_exported": False,
    }

# =========================================================
# Module 2 Supplier Master + Supplier Bulk Export
# Clean implementation based on official baseline.
# =========================================================

SUPPLIER_BULK_SHEET_NAME = "Input Sheet"
SUPPLIER_BULK_TEMPLATE_FILENAME = "supplier_bulk_create_template_v1.xlsx"

SUPPLIER_MATERIAL_ALIASES = [
    "Raw Material Code", "Raw Material Number", "Material", "Material Number", "Component", "Component Number",
    "物料", "料號", "原物料代碼", "原料代碼", "元件料號",
]
SUPPLIER_VENDOR_ALIASES = [
    "Vendor", "Vender", "Vendor Code", "Vendor Number", "Supplier Vendor", "供應商代碼", "供應商編號", "廠商代碼",
]
SUPPLIER_ADDRESS_ALIASES = [
    # Module 2C supplier mapping rule: Transportation Origin must prefer the
    # uploaded supplier file's explicit Address column. Keep the historical
    # supplier-address aliases as fallback for older supplier masters.
    "Address", "地址",
    "Supplier Address", "Supplier Address 1", "Supplier Address1", "Supplier Address Line1",
    "Supplier Address (English)", "Supplier Address (Local)", "Supplier Addr", "Supplier_Address",
    "供應商地址", "廠商地址",
]
SUPPLIER_VENDOR_NAME2_ALIASES = [
    # Preferred display-name source requested for M2C supplier mapping.
    "Vendor Name-2", "Vendor Name 2", "Vendor Name_2", "Vendor Name2", "VendorName2",
    "Vendor Name - 2", "供應商名稱2", "廠商名稱2",
]
SUPPLIER_BULK_NAME_ALIASES = _profile_aliases("supplier_bulk", "supplier_name", ["Supplier Name (optional)", "Supplier Name(optional)"])
SUPPLIER_BULK_CODE_ALIASES = _profile_aliases("supplier_bulk", "supplier_code", ["Supplier Code (optional)", "Vendor", "Vendor Code"])
SUPPLIER_BULK_COUNTRY_ALIASES = _profile_aliases("supplier_bulk", "country_area", ["Country / Area", "Country", "Country Area", "國家"])
SUPPLIER_BULK_ADDRESS_ALIASES = _profile_aliases("supplier_bulk", "supplier_address", ["Supplier Address", "Supplier Address1", "供應商地址"])
SUPPLIER_BULK_UNIT_ALIASES = _profile_aliases("supplier_bulk", "unit_name", ["Unit", "Transportation Destination", "Production Site", "廠區"])
SUPPLIER_PLANT_ALIASES = [
    "Plant", "Plant Code", "Production Plant", "Production Site", "Site", "Factory",
    "工廠", "工廠代碼", "廠別", "廠區", "廠區代碼", "生產廠區",
]

# Plant/Unit Name aliases used only for selecting the uploaded site-specific TBC
# supplier. 2670 can produce both A2 and A9 products, so its TBC record is made
# available to both destinations when the supplier master identifies the site by
# Plant code instead of the full Unit Name.
_PLANT_TO_UNIT_NAMES: dict[str, tuple[str, ...]] = {
    "2670": ("中國常州廠(A9)-IPS", "中國常州廠(A2)-IPS"),
    "3760": ("中國石碣廠-IPS",),
    "3775": ("中國石碣廠-IPS",),
    "4070": ("泰國廠-IPS",),
    "4270": ("越南海防廠-IPS",),
    "429A": ("越南海防廠-IPS",),
    "A9": ("中國常州廠(A9)-IPS",),
    "A2": ("中國常州廠(A2)-IPS",),
}

_PLANT_COUNTRY_AREA: dict[str, str] = {
    "2670": "China",
    "3760": "China",
    "3775": "China",
    "4070": "Thailand",
    "4270": "Viet Nam",
    "429A": "Viet Nam",
    "A9": "China",
    "A2": "China",
}

# Broaden raw material template header aliases while keeping old behavior.
TRANSPORT_ORIGIN_ALIASES = list(dict.fromkeys(TRANSPORT_ORIGIN_ALIASES + [
    "Transportation Origin (optional)", "Transport Origin", "Origin", "運輸起點(optional)",
]))
TRANSPORT_DESTINATION_ALIASES = list(dict.fromkeys(TRANSPORT_DESTINATION_ALIASES + [
    "Transportation Destination (optional)", "Transport Destination", "Destination", "運輸終點(optional)",
]))
SUPPLIER_NAME_ALIASES = list(dict.fromkeys(SUPPLIER_NAME_ALIASES + [
    "Supplier Name(optional)", "Supplier Name Optional", "Supplier Name（optional）",
]))


def _normalize_vendor_code(value: Any) -> str:
    text = _safe_text(value).upper()
    if text.endswith(".0"):
        text = text[:-2]
    text = re.sub(r"\s+", "", text)
    return text


def _site_lookup_keys(value: Any) -> list[str]:
    """Return normalized lookup keys for Plant codes and Unit Name labels."""
    text = _safe_text(value)
    if not text:
        return []

    keys: list[str] = []

    def add(candidate: Any) -> None:
        candidate_text = _safe_text(candidate)
        if not candidate_text:
            return
        for key in (
            _normalize_template_header(candidate_text),
            re.sub(r"\s+", "", candidate_text).upper(),
            _normalize_vendor_code(candidate_text),
        ):
            if key and key not in keys:
                keys.append(key)

    add(text)
    plant_code = _normalize_vendor_code(text)
    for unit_name in _PLANT_TO_UNIT_NAMES.get(plant_code, ()):
        add(unit_name)

    compact = re.sub(r"\s+", "", text).upper()
    alias_matches = {
        "中國常州廠(A9)-IPS": ("A9", "CHANGZHOUA9"),
        "中國常州廠(A2)-IPS": ("A2", "CHANGZHOUA2"),
        "中國石碣廠-IPS": ("廣州", "广州", "石碣", "GUANGZHOU", "SHIJIE"),
        "泰國廠-IPS": ("泰國", "泰国", "THAILAND", "THAI"),
        "越南海防廠-IPS": ("越南", "海防", "VIETNAM", "VIET NAM", "HAIPHONG", "HAI PHONG"),
    }
    for unit_name, aliases in alias_matches.items():
        if any(re.sub(r"\s+", "", alias).upper() in compact for alias in aliases):
            add(unit_name)

    return keys


def _country_area_for_unit_name(unit_name: Any, plant_value: Any = "", uploaded_country: Any = "") -> str:
    """Resolve the Supplier Bulk country using the destination Unit Name first."""
    unit_text = _safe_text(unit_name)
    compact = re.sub(r"\s+", "", unit_text).upper()
    if any(token in compact for token in ("常州", "廣州", "广州", "CHANGZHOU", "GUANGZHOU", "SHIJIE")):
        return "China"
    if any(token in compact for token in ("泰國", "泰国", "THAILAND", "THAI")):
        return "Thailand"
    if any(token in compact for token in ("越南", "海防", "VIETNAM", "HAIPHONG")):
        return "Viet Nam"

    plant_code = _normalize_vendor_code(plant_value)
    if plant_code in _PLANT_COUNTRY_AREA:
        return _PLANT_COUNTRY_AREA[plant_code]

    country = _safe_text(uploaded_country)
    normalized_country = re.sub(r"\s+", " ", country).strip()
    if normalized_country.upper() in {"VIETNAM", "VIET NAM"}:
        return "Viet Nam"
    return normalized_country


def _country_area_zh_for_tbc(unit_name: Any, uploaded_country: Any = "") -> str:
    """Return a Chinese Country/Area label for Supplier Bulk TBC rows only."""
    english_country = _country_area_for_unit_name(unit_name, uploaded_country=uploaded_country)
    normalized = re.sub(r"\s+", " ", _safe_text(english_country)).strip()
    upper = normalized.upper()
    country_map = {
        "CHINA": "中國",
        "PRC": "中國",
        "PEOPLE'S REPUBLIC OF CHINA": "中國",
        "THAILAND": "泰國",
        "VIETNAM": "越南",
        "VIET NAM": "越南",
    }
    if upper in country_map:
        return country_map[upper]

    # Also normalize common Chinese variants if the uploaded supplier master
    # already contains Chinese country names.
    chinese_map = {
        "中国": "中國",
        "中國": "中國",
        "泰国": "泰國",
        "泰國": "泰國",
        "越南": "越南",
    }
    return chinese_map.get(normalized, normalized)


def _supplier_bulk_unit_prefixed_name(unit_name: Any, supplier_name: Any, supplier_code: Any) -> str:
    """Build Supplier Bulk name as ``Unit Name-Supplier Name``.

    This formatting applies only to the separate Supplier Bulk workbook.
    M2C mapped activity data keeps its existing supplier display format.
    """
    unit = _safe_text(unit_name)
    name = _safe_text(supplier_name) or _normalize_vendor_code(supplier_code)
    if unit and name:
        prefix = f"{unit}-"
        # Keep the function idempotent in case an already formatted row is
        # normalized more than once.
        if name.startswith(prefix):
            return name
        return f"{unit}-{name}"
    return name or unit


def _select_uploaded_tbc_supplier_for_destination(
    tbc_supplier_map: dict[str, dict[str, str]] | None,
    transportation_destination: Any,
) -> dict[str, str] | None:
    for key in _site_lookup_keys(transportation_destination):
        if tbc_supplier_map and key in tbc_supplier_map:
            return tbc_supplier_map[key]
    return None


def _format_supplier_display_name(vendor_code: Any, vendor_name: Any) -> str:
    """Format the base supplier display value as ``Supplier Name - Supplier Code``."""
    vendor = _normalize_vendor_code(vendor_code)
    name = _safe_text(vendor_name)
    if vendor and name:
        return f"{name} - {vendor}"
    return name or vendor


def _raw_material_supplier_display_name(transport_destination: Any, vendor_code: Any, vendor_name: Any) -> str:
    """Build Raw Material Bulk Supplier Name as destination-prefixed display text.

    Output format:
        ``Transportation Destination-Supplier Name - Supplier Code``

    The function is idempotent so rows that already carry the destination prefix
    are not prefixed a second time.
    """
    destination = _safe_text(transport_destination)
    base_display = _format_supplier_display_name(vendor_code, vendor_name)
    if destination and base_display:
        prefix = f"{destination}-"
        if base_display.startswith(prefix):
            return base_display
        return f"{prefix}{base_display}"
    return base_display or destination


def _strip_unit_prefix_from_supplier_name(value: Any, unit_name: Any) -> str:
    """Remove one leading ``Unit Name-`` prefix from a supplier display value."""
    text = _safe_text(value)
    unit = _safe_text(unit_name)
    if text and unit:
        prefix = f"{unit}-"
        if text.startswith(prefix):
            return text[len(prefix):]
    return text


def _supplier_bulk_name_only(supplier_master_name: Any, supplier_display_name: Any, vendor_code: Any, unit_name: Any = "") -> str:
    """Return the plain supplier name for Supplier Bulk Create output.

    M2C mapped Raw Material Bulk uses ``Vendor Name - Vendor`` for display and
    supplier mapping/dropdown matching. The
    separate supplier_bulk_create workbook should contain only the supplier name
    in its ``Supplier Name`` column.
    """
    master_name = _safe_text(supplier_master_name)
    if master_name:
        return master_name

    display_name = _strip_unit_prefix_from_supplier_name(supplier_display_name, unit_name)
    vendor = _normalize_vendor_code(vendor_code)
    if not display_name:
        return vendor

    # Fallback for rows that only retain a combined display value. Support both
    # legacy ``Vendor - Vendor Name`` and current ``Vendor Name - Vendor`` forms.
    match = re.match(r"^\s*(.*?)\s+-\s+(.*?)\s*$", display_name)
    if match:
        left_text = _safe_text(match.group(1))
        right_text = _safe_text(match.group(2))
        left_code = _normalize_vendor_code(left_text)
        right_code = _normalize_vendor_code(right_text)
        if vendor and left_code == vendor and right_text:
            return right_text
        if vendor and right_code == vendor and left_text:
            return left_text
    return display_name


def _supplier_header_key(value: Any) -> str:
    return re.sub(r"[^0-9A-Z]+", "", str(value or "").upper())


def _find_any_dataframe_column(df: pd.DataFrame, aliases: list[str]) -> str | None:
    normalized = {_normalize_template_header(c): c for c in df.columns}
    for alias in aliases:
        key = _normalize_template_header(alias)
        if key in normalized:
            return normalized[key]
    return None


def _find_supplier_col_by_rule(df: pd.DataFrame, kind: str) -> str | None:
    keyed = [(_supplier_header_key(c), c) for c in df.columns]
    if kind == "material":
        for key, col in keyed:
            if "MATERIAL" in key and not any(x in key for x in ["GROUP", "DESC", "DESCRIPTION"]):
                return col
        return _find_any_dataframe_column(df, SUPPLIER_MATERIAL_ALIASES)
    if kind == "vendor":
        for key, col in keyed:
            if key.endswith("VENDOR") or key.endswith("VENDER") or "VENDORCODE" in key or "VENDORNUMBER" in key:
                return col
        return _find_any_dataframe_column(df, SUPPLIER_VENDOR_ALIASES)
    if kind == "vendor_name":
        # Prefer Vendor Name-2 for Supplier Name (optional):
        #   <Vendor> + " - " + <Vendor Name-2>
        # If the new field is not present, keep historical compatibility with
        # Vendor Name / Supplier Name / Search Term.
        preferred = _find_any_dataframe_column(df, SUPPLIER_VENDOR_NAME2_ALIASES)
        if preferred:
            return preferred
        for key, col in keyed:
            if key in {"VENDORNAME2", "VENDORNAME02", "VENDORNAME002"}:
                return col
        for key, col in keyed:
            if "VENDORNAME" in key or "SUPPLIERNAME" in key or "SEARCHTERM" in key:
                return col
        return None
    if kind == "country":
        for key, col in keyed:
            if key.endswith("COUNTRY") or "COUNTRYAREA" in key or key == "COUNTRYREGION":
                return col
        return None
    if kind == "city":
        for key, col in keyed:
            if key.endswith("CITY") or key == "CITY":
                return col
        return None
    if kind == "street":
        for key, col in keyed:
            if "STREET" in key or "ADDRESSLINE" in key:
                return col
        return None
    if kind == "incoterms2":
        for key, col in keyed:
            if "INCOTERMS" in key and ("PART2" in key or key.endswith("2")):
                return col
        return None
    if kind == "plant":
        exact = _find_any_dataframe_column(df, SUPPLIER_PLANT_ALIASES)
        if exact:
            return exact
        for key, col in keyed:
            if key in {"PLANT", "PLANTCODE", "PRODUCTIONPLANT", "PRODUCTIONSITE", "SITE", "FACTORY"}:
                return col
        return None
    return None


def _find_supplier_address_column(df: pd.DataFrame) -> str | None:
    # Prefer exact Address for Transportation Origin. This is intentionally
    # before Supplier Address / country-city-street composition because the
    # supplier file may contain multiple address-like fields.
    exact_address = _find_any_dataframe_column(df, ["Address", "地址"])
    if exact_address:
        return exact_address
    exact = _find_any_dataframe_column(df, SUPPLIER_ADDRESS_ALIASES)
    if exact:
        return exact
    for col in df.columns:
        key = _normalize_template_header(col)
        raw_key = _supplier_header_key(col)
        if raw_key == "ADDRESS" or key == "address":
            return col
        if "email" in key or "mail" in key:
            continue
        if "supplier" in key and "address" in key:
            return col
    return None


def _build_supplier_address(row: pd.Series, address_col: str | None, country_col: str | None, city_col: str | None, street_col: str | None, incoterms2_col: str | None) -> str:
    address = _safe_text(row.get(address_col)) if address_col else ""
    if address:
        return address
    parts: list[str] = []
    for col in [country_col, city_col, street_col]:
        text = _safe_text(row.get(col)) if col else ""
        if text and text not in parts:
            parts.append(text)
    if parts:
        return " ".join(parts)
    return _safe_text(row.get(incoterms2_col)) if incoterms2_col else ""


def _supplier_record_score(record: dict[str, str]) -> int:
    score = 0
    if record.get("country_area"):
        score += 1000
    if record.get("supplier_address"):
        score += min(len(record.get("supplier_address") or ""), 500)
    if record.get("supplier_master_name"):
        score += 50
    return score


def _read_supplier_files(
    supplier_paths: list[str | Path] | tuple[str | Path, ...] | None,
) -> tuple[dict[str, list[dict[str, str]]], dict[str, dict[str, str]], Dict[str, Any]]:
    """Read one or many supplier masters and normalize to Material -> suppliers.

    A/B supplier formats are supported. If the same Material+Vendor appears in
    multiple uploaded files, the record with richer address information wins.
    """
    if not supplier_paths:
        return {}, {}, {"supplier_files": 0, "supplier_rows": 0, "supplier_mapped_materials": 0, "supplier_mapped_suppliers": 0, "supplier_site_tbc_records": 0, "supplier_skipped_files": []}

    by_material_vendor: dict[tuple[str, str], dict[str, str]] = {}
    site_tbc_by_key: dict[str, dict[str, str]] = {}
    total_rows = 0
    skipped_files: list[str] = []

    for path in supplier_paths:
        path = Path(path)
        try:
            df = pd.read_excel(path, sheet_name=0, dtype=object)
        except Exception as exc:
            skipped_files.append(f"{path.name}: {exc}")
            continue

        material_col = _find_supplier_col_by_rule(df, "material")
        vendor_col = _find_supplier_col_by_rule(df, "vendor")
        vendor_name_col = _find_supplier_col_by_rule(df, "vendor_name")
        address_col = _find_supplier_address_column(df)
        country_col = _find_supplier_col_by_rule(df, "country")
        city_col = _find_supplier_col_by_rule(df, "city")
        street_col = _find_supplier_col_by_rule(df, "street")
        incoterms2_col = _find_supplier_col_by_rule(df, "incoterms2")
        plant_col = _find_supplier_col_by_rule(df, "plant")

        if not vendor_col or (not material_col and not plant_col):
            skipped_files.append(f"{path.name}: missing vendor and material/plant column")
            continue

        total_rows += int(len(df))
        for _, row in df.iterrows():
            material_key = _normalize_material_key(row.get(material_col)) if material_col else ""
            vendor_code = _normalize_vendor_code(row.get(vendor_col))
            plant_value = _safe_text(row.get(plant_col)) if plant_col else ""
            if not vendor_code:
                continue
            vendor_name = _safe_text(row.get(vendor_name_col)) if vendor_name_col else ""
            country = _safe_text(row.get(country_col)) if country_col else ""
            supplier_address = _build_supplier_address(row, address_col, country_col, city_col, street_col, incoterms2_col)
            candidate = {
                "vendor_code": vendor_code,
                "supplier_code": vendor_code,
                "supplier_master_name": vendor_name,
                "supplier_name": _format_supplier_display_name(vendor_code, vendor_name),
                "country_area": country,
                "supplier_address": supplier_address,
                "transport_origin": supplier_address,
                "plant": plant_value,
                "source_file": path.name,
            }
            if material_key:
                key = (material_key, vendor_code)
                current = by_material_vendor.get(key)
                if current is None or _supplier_record_score(candidate) > _supplier_record_score(current):
                    by_material_vendor[key] = candidate

            # IPS supplier masters may define one generic TBC supplier per Plant
            # without a material number. Preserve it as the preferred fallback for
            # all unmatched raw materials at that Unit Name.
            if vendor_code == "TBC" and plant_value:
                tbc_candidate = dict(candidate)
                tbc_candidate["supplier_code"] = "TBC"
                tbc_candidate["vendor_code"] = "TBC"
                tbc_candidate["supplier_master_name"] = "TBC"
                tbc_candidate["supplier_name"] = "TBC - TBC"
                for site_key in _site_lookup_keys(plant_value):
                    current_tbc = site_tbc_by_key.get(site_key)
                    if current_tbc is None or _supplier_record_score(tbc_candidate) > _supplier_record_score(current_tbc):
                        site_tbc_by_key[site_key] = tbc_candidate

    records: dict[str, list[dict[str, str]]] = {}
    for (material_key, _vendor_code), record in by_material_vendor.items():
        records.setdefault(material_key, []).append(record)
    for material_key in records:
        records[material_key].sort(key=lambda r: r.get("vendor_code", ""))

    unique_tbc_records = {
        (record.get("plant", ""), record.get("supplier_address", ""), record.get("country_area", ""))
        for record in site_tbc_by_key.values()
    }
    return records, site_tbc_by_key, {
        "supplier_files": int(len(supplier_paths)),
        "supplier_rows": int(total_rows),
        "supplier_mapped_materials": int(len(records)),
        "supplier_mapped_suppliers": int(sum(len(v) for v in records.values())),
        "supplier_site_tbc_records": int(len(unique_tbc_records)),
        "supplier_skipped_files": skipped_files,
    }



def _extract_site_tbc_supplier_map_from_raw_template(wb) -> dict[str, dict[str, str]]:
    """Read site-specific TBC supplier rows from raw material template Dropdown Values.

    Expected Supplier Name (optional) pattern:
        <Transportation Destination>_TBC - TBC
    Example:
        中國常州廠(A9)-IPS_TBC - TBC

    The Transportation Origin is read from the same row in Dropdown Values.
    """
    if "Dropdown Values" not in wb.sheetnames:
        return {}

    ws = wb["Dropdown Values"]
    supplier_cols = _find_template_columns(
        ws,
        ["Supplier Name (optional)", "Supplier Name(optional)", "Supplier Name", "supplier_name", "供應商名稱"],
        header_rows=10,
    )
    origin_cols = _find_template_columns(
        ws,
        ["Transportation Origin", "Transportation Origin (optional)", "Transport Origin", "transportation_origin", "運輸起點"],
        header_rows=10,
    )
    if not supplier_cols:
        return {}

    supplier_col = int(supplier_cols[0])
    origin_col = int(origin_cols[0]) if origin_cols else None
    tbc_map: dict[str, dict[str, str]] = {}

    for row_idx in range(2, ws.max_row + 1):
        supplier_name = _safe_text(ws.cell(row_idx, supplier_col).value)
        if not supplier_name or "TBC" not in supplier_name.upper():
            continue

        # Preferred exact expandable pattern: <site>_TBC - TBC
        m = re.match(r"^(.*?)\s*_\s*TBC\s*-\s*TBC\s*$", supplier_name, flags=re.IGNORECASE)
        if not m:
            continue
        site_name = _safe_text(m.group(1))
        if not site_name:
            continue

        origin = _safe_text(ws.cell(row_idx, origin_col).value) if origin_col else ""
        item = {
            "supplier_name": supplier_name,
            "transport_origin": origin,
            "supplier_code": "TBC",
            "supplier_master_name": _supplier_name_from_option(supplier_name),
            "supplier_country_area": "",
            "supplier_address": origin,
        }
        tbc_map[_normalize_template_header(site_name)] = item
        tbc_map[re.sub(r"\s+", "", site_name).upper()] = item

    return tbc_map


def _select_tbc_supplier_for_destination(
    tbc_supplier_map: dict[str, dict[str, str]],
    transportation_destination: Any,
) -> dict[str, str] | None:
    destination = _safe_text(transportation_destination)
    if not destination:
        return None
    keys = [
        _normalize_template_header(destination),
        re.sub(r"\s+", "", destination).upper(),
    ]
    for key in keys:
        if key and key in tbc_supplier_map:
            return tbc_supplier_map[key]
    return None


def _extract_supplier_name_options_from_raw_template(wb) -> list[str]:
    """Read all supplier-name candidates from raw_materials_&_activity_data_template_v1.

    Source of truth: sheet "Dropdown Values", column "Supplier Name (optional)".
    """
    values: list[str] = []
    seen: set[str] = set()

    def add(value: Any) -> None:
        text = _safe_text(value)
        if text and text not in seen:
            seen.add(text)
            values.append(text)

    if "Dropdown Values" not in wb.sheetnames:
        return values
    ws = wb["Dropdown Values"]
    cols = _find_template_columns(
        ws,
        ["Supplier Name (optional)", "Supplier Name(optional)", "Supplier Name", "supplier_name", "供應商名稱"],
        header_rows=10,
    )
    if not cols:
        return values
    col_idx = int(cols[0])
    for row_idx in range(2, ws.max_row + 1):
        add(ws.cell(row_idx, col_idx).value)
    return values


def _supplier_name_from_option(option: Any) -> str:
    text = _safe_text(option)
    if "_" in text:
        tail = text.split("_", 1)[1].strip()
        tail = re.sub(r"\s+-\s+.*$", "", tail).strip()
        return tail or text
    return text


def _select_supplier_name_option(options: list[str], destination: Any, vendor_code: Any) -> str:
    dest = _safe_text(destination)
    dest_compact = re.sub(r"\s+", "", dest).upper()
    vendor = _normalize_vendor_code(vendor_code)
    if not vendor:
        return ""
    candidates: list[str] = []
    for option in options or []:
        text = _safe_text(option)
        compact = re.sub(r"\s+", "", text).upper()
        if vendor not in compact:
            continue
        option_dest = text.split("_", 1)[0].strip() if "_" in text else ""
        option_dest_compact = re.sub(r"\s+", "", option_dest).upper()
        has_dest = bool(dest_compact) and (
            dest_compact == option_dest_compact or dest_compact in compact or (option_dest_compact and option_dest_compact in dest_compact)
        )
        if has_dest:
            candidates.append(text)
    if not candidates:
        return ""
    suffix_matches = [c for c in candidates if re.search(r"-\s*[^-]*" + re.escape(vendor) + r"\s*$", c, flags=re.I)]
    return suffix_matches[0] if suffix_matches else candidates[0]


def _split_usage_evenly_by_supplier_count(value: Any, supplier_count: int) -> float:
    """Return one supplier row's equal share of the original usage.

    Supplier mapping expands one raw-material activity row into one row per
    unique supplier.  The original activity quantity must be conserved rather
    than copied to every expanded row.  A single supplier (including TBC
    fallback) keeps the original quantity unchanged.
    """
    usage = _safe_number(value)
    count = max(int(supplier_count or 0), 1)
    return usage / count


def _apply_supplier_mapping_to_exploded(
    exploded: pd.DataFrame,
    supplier_map: dict[str, list[dict[str, str]]],
    supplier_options: list[str],
    tbc_supplier_map: dict[str, dict[str, str]] | None = None,
) -> tuple[pd.DataFrame, Dict[str, Any]]:
    work = exploded.copy()
    for col in ["supplier_name", "transport_origin", "supplier_code", "supplier_master_name", "supplier_country_area", "supplier_address"]:
        if col not in work.columns:
            work[col] = ""

    matched_source_rows = 0
    expanded_rows = 0
    supplier_name_matched = 0
    supplier_name_missing = 0
    tbc_fallback_rows = 0
    supplier_usage_split_source_rows = 0
    supplier_usage_split_output_rows = 0
    supplier_usage_total_before_split = 0.0
    supplier_usage_total_after_split = 0.0
    output_rows: list[dict[str, Any]] = []

    if work.empty:
        return work, {
            "supplier_matched_rows": 0,
            "supplier_expanded_rows": 0,
            "supplier_name_matched_rows": 0,
            "supplier_name_missing_rows": 0,
            "supplier_dropdown_matched_rows": 0,
            "supplier_dropdown_missing_rows": 0,
            "tbc_fallback_rows": 0,
            "supplier_usage_split_source_rows": 0,
            "supplier_usage_split_output_rows": 0,
            "supplier_usage_total_before_split": 0.0,
            "supplier_usage_total_after_split": 0.0,
        }

    # If a raw material does not match a material-level supplier row, prefer the
    # uploaded Plant-specific TBC supplier. Only use the deterministic system TBC
    # when the supplier master does not define a TBC for that destination.
    for _, row in work.iterrows():
        original = row.to_dict()
        raw_key = _normalize_material_key(row.get("raw_material"))
        destination = _safe_text(row.get("transport_destination"))
        suppliers = supplier_map.get(raw_key) or []
        if not suppliers:
            uploaded_tbc = _select_uploaded_tbc_supplier_for_destination(tbc_supplier_map, destination)
            uploaded_address = ""
            uploaded_country = ""
            uploaded_plant = ""
            if uploaded_tbc:
                uploaded_address = uploaded_tbc.get("supplier_address", "") or uploaded_tbc.get("transport_origin", "")
                uploaded_country = uploaded_tbc.get("country_area", "")
                uploaded_plant = uploaded_tbc.get("plant", "")
            fallback_row = dict(original)
            fallback_row["transport_destination"] = destination
            fallback_row["supplier_name"] = _raw_material_supplier_display_name(destination, "TBC", "TBC")
            fallback_row["transport_origin"] = uploaded_address or "TBC"
            fallback_row["supplier_code"] = "TBC"
            fallback_row["supplier_master_name"] = "TBC"
            fallback_row["supplier_country_area"] = _country_area_for_unit_name(destination, uploaded_plant, uploaded_country)
            fallback_row["supplier_address"] = uploaded_address or "TBC"
            supplier_name_matched += 1
            tbc_fallback_rows += 1
            output_rows.append(fallback_row)
            continue
        matched_source_rows += 1
        supplier_count = len(suppliers)
        usage_per_supplier = _split_usage_evenly_by_supplier_count(original.get("usage"), supplier_count)
        if supplier_count > 1:
            supplier_usage_split_source_rows += 1
            supplier_usage_split_output_rows += supplier_count
            supplier_usage_total_before_split += _safe_number(original.get("usage"))
            supplier_usage_total_after_split += usage_per_supplier * supplier_count
        for info in suppliers:
            new_row = dict(original)
            new_row["usage"] = usage_per_supplier
            # Supplier logic only reads destination. It never clears or overwrites it.
            new_row["transport_destination"] = destination
            supplier_address = info.get("supplier_address", "") or info.get("transport_origin", "")
            supplier_code = info.get("supplier_code", "") or info.get("vendor_code", "")
            if _normalize_vendor_code(supplier_code) == "TBC":
                uploaded_tbc = _select_uploaded_tbc_supplier_for_destination(tbc_supplier_map, destination)
                if uploaded_tbc:
                    supplier_address = supplier_address or uploaded_tbc.get("supplier_address", "") or uploaded_tbc.get("transport_origin", "")
            supplier_master_name = info.get("supplier_master_name", "") or _supplier_name_from_option(info.get("supplier_name", ""))
            supplier_name = _raw_material_supplier_display_name(destination, supplier_code, supplier_master_name)
            if not supplier_name:
                supplier_name = _select_supplier_name_option(supplier_options, destination, supplier_code)
            new_row["transport_origin"] = supplier_address
            new_row["supplier_code"] = supplier_code
            new_row["supplier_master_name"] = info.get("supplier_master_name", "") or _supplier_name_from_option(supplier_name)
            if _normalize_vendor_code(supplier_code) == "TBC":
                new_row["supplier_country_area"] = _country_area_for_unit_name(destination, info.get("plant", ""), info.get("country_area", ""))
                new_row["supplier_master_name"] = "TBC"
                supplier_name = _raw_material_supplier_display_name(destination, "TBC", "TBC")
            else:
                new_row["supplier_country_area"] = info.get("country_area", "")
            new_row["supplier_address"] = supplier_address
            new_row["supplier_name"] = supplier_name
            if supplier_name:
                supplier_name_matched += 1
            else:
                supplier_name_missing += 1
            output_rows.append(new_row)
            expanded_rows += 1

    ordered_columns = list(dict.fromkeys(list(work.columns) + ["supplier_code", "supplier_master_name", "supplier_country_area", "supplier_address"]))
    result = pd.DataFrame(output_rows) if output_rows else work
    for col in ordered_columns:
        if col not in result.columns:
            result[col] = ""
    return result[ordered_columns], {
        "supplier_matched_rows": int(matched_source_rows),
        "supplier_expanded_rows": int(expanded_rows),
        "supplier_name_matched_rows": int(supplier_name_matched),
        "supplier_name_missing_rows": int(supplier_name_missing),
        # Backward-compatible summary keys for the current UI text.
        "supplier_dropdown_matched_rows": int(supplier_name_matched),
        "supplier_dropdown_missing_rows": int(supplier_name_missing),
        "tbc_fallback_rows": int(tbc_fallback_rows),
        "supplier_usage_split_source_rows": int(supplier_usage_split_source_rows),
        "supplier_usage_split_output_rows": int(supplier_usage_split_output_rows),
        "supplier_usage_total_before_split": float(supplier_usage_total_before_split),
        "supplier_usage_total_after_split": float(supplier_usage_total_after_split),
        "supplier_usage_split_rule": "equal_share_original_usage_divided_by_unique_supplier_count",
    }


def _first_text(row: pd.Series, names: list[str]) -> str:
    for name in names:
        if name in row.index:
            text = _safe_text(row.get(name))
            if text:
                return text
    return ""


def _supplier_bulk_row_score(row: tuple[str, str, str, str, str]) -> int:
    """Prefer uploaded, information-rich rows when duplicate TBC rows exist."""
    supplier_name, supplier_code, country_area, supplier_address, unit_name = row
    score = 0
    address = _safe_text(supplier_address)
    country = _safe_text(country_area)
    if address and address.upper() != "TBC":
        score += 10000 + min(len(address), 1000)
    elif address:
        score += 1
    if country:
        score += 1000
    if _safe_text(supplier_name):
        score += 100
    if _safe_text(unit_name):
        score += 10
    if _normalize_vendor_code(supplier_code) == "TBC":
        score += 5
    return score


def _normalize_supplier_bulk_rows(
    supplier_rows: Any,
) -> list[tuple[str, str, str, str, str]]:
    """Normalize Supplier Bulk rows and keep only one TBC per Unit Name."""
    selected: dict[tuple[Any, ...], tuple[str, str, str, str, str]] = {}
    for raw_row in supplier_rows or []:
        if len(raw_row) != 5:
            continue
        supplier_name, supplier_code, country_area, supplier_address, unit_name = raw_row
        code = _normalize_vendor_code(supplier_code)
        name = _safe_text(supplier_name)
        country = _safe_text(country_area)
        address = _safe_text(supplier_address)
        unit = _safe_text(unit_name)
        if not code:
            continue

        if code == "TBC":
            base_name = "TBC"
            country = _country_area_zh_for_tbc(unit, uploaded_country=country)
            normalized_unit = _normalize_template_header(unit) or re.sub(r"\s+", "", unit).upper()
            key: tuple[Any, ...] = ("TBC", normalized_unit)
        else:
            base_name = name
            # Use the unformatted supplier name for duplicate detection. Unit
            # Name is already part of the key, so adding the display prefix
            # here would not improve uniqueness.
            key = (base_name, code, country, address, unit)

        name = _supplier_bulk_unit_prefixed_name(unit, base_name, code)
        candidate = (name, code, country, address, unit)
        current = selected.get(key)
        if current is None or _supplier_bulk_row_score(candidate) > _supplier_bulk_row_score(current):
            selected[key] = candidate

    return sorted(selected.values(), key=lambda row: (row[4], row[1], row[0], row[3]))



def _supplier_col_letter(idx: int) -> str:
    out = ""
    n = int(idx)
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


def _supplier_replace_cell_xml(row_xml: bytes, row_idx: int, col_idx: int, value: Any = None, formula_cache: Any = None) -> bytes:
    ref = f"{_supplier_col_letter(col_idx)}{row_idx}"
    pattern = re.compile(rb'<c\b[^>]*?\br="' + re.escape(ref.encode()) + rb'"[^>]*?(?:/>|>.*?</c>)', re.DOTALL)
    match = pattern.search(row_xml)
    old = match.group(0) if match else b""
    style_match = re.search(rb'\bs="([^"]+)"', old)
    style = (b' s="' + style_match.group(1) + b'"') if style_match else b""
    if formula_cache is not None:
        formula_match = re.search(rb'(<f\b[^>]*>.*?</f>)', old, re.DOTALL)
        formula = formula_match.group(1) if formula_match else b""
        cache = str(formula_cache or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        new = b'<c r="' + ref.encode() + b'"' + style + b' t="str">' + formula + b'<v>' + cache.encode("utf-8") + b'</v></c>'
    elif value in (None, ""):
        new = b'<c r="' + ref.encode() + b'"' + style + b'/>'
    else:
        text = str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
        new = b'<c r="' + ref.encode() + b'"' + style + b' t="inlineStr"><is><t xml:space="preserve">' + text.encode("utf-8") + b'</t></is></c>'
    if match:
        return row_xml[:match.start()] + new + row_xml[match.end():]
    pos = row_xml.rfind(b"</row>")
    return row_xml[:pos] + new + row_xml[pos:] if pos >= 0 else row_xml


def _supplier_sort_row_cells(row_xml: bytes) -> bytes:
    open_match = re.match(rb'(<row\b[^>]*>)', row_xml, re.DOTALL)
    close_pos = row_xml.rfind(b"</row>")
    if not open_match or close_pos < 0:
        return row_xml
    body = row_xml[open_match.end():close_pos]
    pattern = re.compile(rb'<c\b[^>]*?\br="([A-Z]+)\d+"[^>]*?(?:/>|>.*?</c>)', re.DOTALL)
    matches = list(pattern.finditer(body))
    if len(matches) < 2:
        return row_xml
    def idx(letters: bytes) -> int:
        n = 0
        for ch in letters.decode("ascii"):
            n = n * 26 + ord(ch) - 64
        return n
    prefix = body[:matches[0].start()]
    suffix = body[matches[-1].end():]
    cells = sorted(((idx(m.group(1)), m.group(0)) for m in matches), key=lambda x: x[0])
    return row_xml[:open_match.end()] + prefix + b"".join(c for _, c in cells) + suffix + row_xml[close_pos:]


def _supplier_dropdown_country_maps(template_path: str | Path) -> tuple[dict[str, str], dict[str, str]]:
    """Return target-template country display->key and key->display maps."""
    display_to_key: dict[str, str] = {}
    key_to_display: dict[str, str] = {}
    wb = load_workbook(template_path, read_only=True, data_only=True)
    try:
        if "Dropdown Values" not in wb.sheetnames:
            return display_to_key, key_to_display
        ws = wb["Dropdown Values"]
        for display, key in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            display_text = _safe_text(display)
            if not display_text:
                continue
            key_text = _safe_text(key) or display_text
            display_to_key[_normalize_template_header(display_text)] = key_text
            key_to_display.setdefault(key_text.upper(), display_text)
    finally:
        wb.close()
    return display_to_key, key_to_display


def _supplier_unit_aliases(value: Any) -> list[str]:
    """Build normalized aliases for a Supplier Template Unit Name."""
    raw = _safe_text(value)
    if not raw:
        return []
    candidates = [raw]
    # Template values commonly append the business unit, e.g. 中國常州廠(A2)-IPS.
    candidates.append(re.sub(r"\s*[-–—]\s*(?:IPS|AE|PC\s*&\s*CE|PC_CE)\s*$", "", raw, flags=re.I))
    # Source values may contain only the plant code or may use older site wording.
    code_match = re.search(r"\(([^)]+)\)", raw)
    if code_match:
        candidates.append(code_match.group(1))
    aliases: list[str] = []
    for candidate in candidates:
        normalized = _normalize_template_header(candidate)
        if normalized and normalized not in aliases:
            aliases.append(normalized)
    return aliases


def _supplier_dropdown_unit_name_map(template_path: str | Path) -> dict[str, str]:
    """Return normalized Unit Name aliases -> exact dropdown display value."""
    alias_to_display: dict[str, str] = {}
    wb = load_workbook(template_path, read_only=True, data_only=True)
    try:
        if "Dropdown Values" not in wb.sheetnames:
            return alias_to_display
        ws = wb["Dropdown Values"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        unit_col = None
        for idx, value in enumerate(header_row, start=1):
            if _normalize_template_header(value) in {
                _normalize_template_header("Unit Name"),
                _normalize_template_header("單位名稱"),
                _normalize_template_header("廠區"),
            }:
                unit_col = idx
                break
        if unit_col is None:
            return alias_to_display
        for row in ws.iter_rows(min_row=2, min_col=unit_col, max_col=unit_col, values_only=True):
            display = _safe_text(row[0] if row else "")
            if not display:
                continue
            for alias in _supplier_unit_aliases(display):
                alias_to_display.setdefault(alias, display)
    finally:
        wb.close()
    return alias_to_display


def _supplier_localized_unit_name(unit_name: Any, alias_to_display: dict[str, str]) -> str:
    """Resolve source site text to an exact Unit Name allowed by the uploaded template."""
    raw = _safe_text(unit_name)
    for alias in _supplier_unit_aliases(raw):
        display = alias_to_display.get(alias)
        if display:
            return display
    return ""


def _supplier_localized_country(country_area: Any, unit_name: Any, display_to_key: dict[str, str], key_to_display: dict[str, str]) -> tuple[str, str]:
    """Resolve a country to the uploaded Supplier Template language and internal key."""
    original = _safe_text(country_area)
    english = _country_area_for_unit_name(unit_name, uploaded_country=original)
    chinese = _country_area_zh_for_tbc(unit_name, uploaded_country=original)
    candidates = [original, english, chinese]
    key = ""
    for candidate in candidates:
        normalized = _normalize_template_header(candidate)
        if normalized and normalized in display_to_key:
            key = display_to_key[normalized]
            break
    if not key:
        known = {
            "CHINA": "CN", "PRC": "CN", "PEOPLESREPUBLICOFCHINA": "CN", "中國": "CN", "中国": "CN",
            "THAILAND": "TH", "泰國": "TH", "泰国": "TH",
            "VIETNAM": "VN", "VIETNAM": "VN", "VIET NAM": "VN", "越南": "VN",
            "TBD": "TBD",
        }
        for candidate in candidates:
            compact = re.sub(r"[^A-Z0-9一-龥]", "", _safe_text(candidate).upper())
            if compact in known:
                key = known[compact]
                break
    if not key:
        key = original
    display = key_to_display.get(_safe_text(key).upper()) or original or english or chinese
    return display, key


def _write_supplier_bulk_openxml(rows: list[tuple[str, str, str, str, str]], template_path: str | Path, output_path: str | Path, cols: dict[str, int]) -> None:
    template_path = Path(template_path)
    output_path = Path(output_path)
    sheet_paths = _xlsx_sheet_paths_by_name(template_path)
    sheet_path = sheet_paths.get(SUPPLIER_BULK_SHEET_NAME) or next(iter(sheet_paths.values()), None)
    if not sheet_path:
        raise ValueError("Supplier Bulk Template 缺少 Input Sheet")
    country_display_to_key, country_key_to_display = _supplier_dropdown_country_maps(template_path)
    unit_name_map = _supplier_dropdown_unit_name_map(template_path)
    if not unit_name_map:
        raise ValueError("Supplier Bulk Template 的 Dropdown Values 缺少 Unit Name 選項。")
    missing_unit_names = sorted({
        _safe_text(row[4]) for row in rows
        if _safe_text(row[4]) and not _supplier_localized_unit_name(row[4], unit_name_map)
    })
    if missing_unit_names:
        preview = "、".join(missing_unit_names[:10])
        more = f"（另有 {len(missing_unit_names) - 10} 個）" if len(missing_unit_names) > 10 else ""
        raise ValueError(f"Supplier Bulk Template 缺少以下 Unit Name 下拉選項：{preview}{more}")
    hidden_cols = {"country": cols.get("country_hidden", 15), "industry": cols.get("industry_hidden", 16), "tier": cols.get("tier_hidden", 17)}
    with zipfile.ZipFile(template_path, "r") as zin:
        template_xml = zin.read(sheet_path)
        start = re.search(rb'<sheetData\b[^>]*>', template_xml)
        end = re.search(rb'</sheetData>', template_xml)
        if not start or not end:
            raise ValueError("Supplier Bulk Template worksheet XML 缺少 sheetData")
        inside = template_xml[start.end():end.start()]
        row_pattern = re.compile(rb'<row\b[^>]*\br="(\d+)"[^>]*>.*?</row>', re.DOTALL)
        out_rows: list[bytes] = []
        max_template_row = 2
        for match in row_pattern.finditer(inside):
            row_idx = int(match.group(1)); max_template_row = max(max_template_row, row_idx)
            row_xml = match.group(0)
            if row_idx < DATA_START_ROW:
                out_rows.append(row_xml); continue
            offset = row_idx - DATA_START_ROW
            if offset < len(rows):
                supplier_name, supplier_code, country_area, supplier_address, unit_name = rows[offset]
                resolved_unit_name = _supplier_localized_unit_name(unit_name, unit_name_map)
                country_display, country_key = _supplier_localized_country(
                    country_area, resolved_unit_name or unit_name, country_display_to_key, country_key_to_display
                )
                for key, value in (("supplier_name", supplier_name), ("supplier_code", supplier_code), ("country_area", country_display), ("supplier_address", supplier_address), ("unit_name", resolved_unit_name)):
                    row_xml = _supplier_replace_cell_xml(row_xml, row_idx, int(cols[key]), value=value)
                row_xml = _supplier_replace_cell_xml(row_xml, row_idx, int(hidden_cols["country"]), formula_cache=country_key)
                row_xml = _supplier_replace_cell_xml(row_xml, row_idx, int(hidden_cols["industry"]), formula_cache="")
                row_xml = _supplier_replace_cell_xml(row_xml, row_idx, int(hidden_cols["tier"]), formula_cache="")
                row_xml = _supplier_sort_row_cells(row_xml)
            out_rows.append(row_xml)
        if len(rows) > max(0, max_template_row - DATA_START_ROW + 1):
            raise ValueError(f"Supplier Bulk Template 預留列數不足：需要 {len(rows)} 筆")
        new_xml = template_xml[:start.end()] + b"".join(out_rows) + template_xml[end.start():]
        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zout:
            for item in zin.infolist():
                name = item.filename
                if item.is_dir():
                    zout.writestr(item, b""); continue
                if name == "xl/calcChain.xml":
                    continue
                data = new_xml if name == sheet_path else zin.read(name)
                if name == "[Content_Types].xml":
                    data = re.sub(rb'<Override\b[^>]*PartName="/xl/calcChain\.xml"[^>]*/>', b"", data)
                elif name == "xl/_rels/workbook.xml.rels":
                    data = re.sub(rb'<Relationship\b[^>]*(?:calcChain|calcchain)[^>]*/>', b"", data)
                elif name == "xl/workbook.xml":
                    if b"<calcPr" not in data:
                        data = data.replace(b"</workbook>", b'<calcPr calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"/></workbook>', 1)
                zout.writestr(item, data)

def _write_supplier_bulk_create_file(expanded_with_suppliers: pd.DataFrame, supplier_bulk_template_path: str | Path, output_path: str | Path) -> Dict[str, Any]:
    supplier_bulk_template_path = Path(supplier_bulk_template_path)
    output_path = Path(output_path)
    if expanded_with_suppliers is None or expanded_with_suppliers.empty:
        return {"supplier_bulk_rows": 0, "supplier_bulk_filename": "", "supplier_bulk_download_url": ""}
    if not supplier_bulk_template_path.exists():
        return {"supplier_bulk_rows": 0, "supplier_bulk_filename": "", "supplier_bulk_download_url": "", "supplier_bulk_error": f"找不到 Supplier Bulk Template：{supplier_bulk_template_path.name}"}

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_meta = load_workbook(supplier_bulk_template_path, read_only=True, data_only=False)
    ws = wb_meta[SUPPLIER_BULK_SHEET_NAME] if SUPPLIER_BULK_SHEET_NAME in wb_meta.sheetnames else wb_meta[wb_meta.sheetnames[0]]
    cols = {
        "supplier_name": _find_template_column(ws, SUPPLIER_BULK_NAME_ALIASES, 1),
        "supplier_code": _find_template_column(ws, SUPPLIER_BULK_CODE_ALIASES, 2),
        "country_area": _find_template_column(ws, SUPPLIER_BULK_COUNTRY_ALIASES, 3),
        "supplier_address": _find_template_column(ws, SUPPLIER_BULK_ADDRESS_ALIASES, 4),
        "unit_name": _find_template_column(ws, SUPPLIER_BULK_UNIT_ALIASES, 5),
        "country_hidden": _find_template_column(ws, ["country"], 15),
        "industry_hidden": _find_template_column(ws, ["industry_category"], 16),
        "tier_hidden": _find_template_column(ws, ["tier"], 17),
    }
    wb_meta.close()

    raw_rows: list[tuple[str, str, str, str, str]] = []
    for _, row in expanded_with_suppliers.iterrows():
        supplier_code = _normalize_vendor_code(_first_text(row, ["supplier_code", "vendor_code", "Vendor", "Vender", "Supplier Code"]))
        if not supplier_code:
            continue
        unit_name = _first_text(row, ["transport_destination", "Transportation Destination", "transportation_destination", "production_site", "Production Site", "Unit Name"])
        supplier_display_name = _first_text(row, ["supplier_name", "Supplier Name", "Supplier Name (optional)"])
        supplier_master_name = _first_text(row, ["supplier_master_name", "Vendor Name-2", "Vendor Name", "Search Term"])
        supplier_name = _supplier_bulk_name_only(supplier_master_name, supplier_display_name, supplier_code, unit_name)
        country_area = _first_text(row, ["supplier_country_area", "Country/Area", "country_area", "Country"])
        supplier_address = _first_text(row, ["supplier_address", "Supplier Address", "transport_origin", "Transportation Origin"])
        raw_rows.append((supplier_name, supplier_code, country_area, supplier_address, unit_name))

    rows = _normalize_supplier_bulk_rows(raw_rows)

    _write_supplier_bulk_openxml(rows, supplier_bulk_template_path, output_path, cols)
    return {
        "supplier_bulk_rows": int(len(rows)),
        "supplier_bulk_filename": output_path.name,
        "supplier_bulk_download_url": f"/download/{output_path.name}",
        "supplier_bulk_template_columns": cols,
    }


def _write_raw_material_bulk_from_exploded(
    exploded: pd.DataFrame,
    raw_material_template_path: str | Path,
    output_path: str | Path,
    supplier_map: dict[str, list[dict[str, str]]] | None = None,
    tbc_supplier_map: dict[str, dict[str, str]] | None = None,
    return_expanded: bool = False,
) -> Dict[str, Any] | tuple[Dict[str, Any], pd.DataFrame]:
    exploded, zero_usage_rows_excluded = _exclude_zero_usage_rows(exploded)

    raw_material_template_path = Path(raw_material_template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(raw_material_template_path, output_path)

    wb = load_workbook(output_path)
    if ACTIVITY_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"找不到 raw material bulk 分頁：{ACTIVITY_SHEET_NAME}")
    if RAW_MATERIAL_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"找不到 raw material bulk 分頁：{RAW_MATERIAL_SHEET_NAME}")

    activity_ws = wb[ACTIVITY_SHEET_NAME]
    raw_ws = wb[RAW_MATERIAL_SHEET_NAME]
    activity_cols = {
        "raw_name": _find_template_column(activity_ws, RAW_MATERIAL_NAME_ALIASES, 1),
        "raw_code": _find_template_column(activity_ws, RAW_MATERIAL_CODE_ALIASES, 2),
        "start_date": _find_template_column(activity_ws, DOC_START_DATE_ALIASES, 3),
        "end_date": _find_template_column(activity_ws, DOC_END_DATE_ALIASES, 4),
        "document_type": _find_template_column(activity_ws, DOCUMENT_TYPE_ALIASES, 5),
        "document_number": _find_template_column(activity_ws, DOCUMENT_NUMBER_ALIASES, 6),
        "usage": _find_template_column(activity_ws, USAGE_ALIASES, 7),
        "unit": _find_template_column(activity_ws, ACTIVITY_DATA_UNIT_ALIASES, 8),
        "data_source": _find_template_column(activity_ws, DATA_SOURCE_ALIASES, 12),
        "data_source_other": _find_template_column(activity_ws, DATA_SOURCE_OTHER_ALIASES, 13),
        "calculate_transportation_emissions": _find_template_column(activity_ws, CALCULATE_TRANSPORTATION_EMISSIONS_ALIASES, 14),
        "supplier_name": _find_template_column(activity_ws, SUPPLIER_NAME_ALIASES, 15),
        "transport_origin": _find_template_column(activity_ws, TRANSPORT_ORIGIN_ALIASES, 16),
        "transport_destination": _find_template_column(activity_ws, TRANSPORT_DESTINATION_ALIASES, 17),
        "target_product": _find_template_column(activity_ws, PRODUCT_LINK_ALIASES, 18),
        "comment": _find_template_column(activity_ws, COMMENT_ALIASES, 19),
        "material_group": _find_template_column(activity_ws, MATERIAL_GROUP_ALIASES, 20),
        "net_weight": _find_template_optional_column(activity_ws, NET_WEIGHT_ALIASES),
        "gross_weight": _find_template_optional_column(activity_ws, GROSS_WEIGHT_ALIASES),
        "weight_unit": _find_template_optional_column(activity_ws, WEIGHT_UNIT_ALIASES),
    }
    raw_cols = {
        "raw_name": _find_template_column(raw_ws, RAW_MATERIAL_NAME_ALIASES, 1),
        "raw_code": _find_template_column(raw_ws, RAW_MATERIAL_CODE_ALIASES, 2),
        "description": _find_template_column(raw_ws, RAW_MATERIAL_DESC_ALIASES, 6),
    }

    dropdown_cache = DropdownValuesCache.from_workbook(wb)
    document_type_value = _document_type_for_template(wb, dropdown_cache)
    transport_calculation_yes_value = _transport_calculation_yes_for_template(wb, dropdown_cache)
    activity_unit_display_map = dropdown_cache.input_to_display_map("activity_data_unit")
    weight_unit_display_map = dropdown_cache.input_to_display_map("weight_unit")
    data_source_display_map = dropdown_cache.input_to_display_map("data_source")
    supplier_options = _extract_supplier_name_options_from_raw_template(wb)
    expanded, supplier_write_summary = _apply_supplier_mapping_to_exploded(
        exploded,
        supplier_map or {},
        supplier_options,
        tbc_supplier_map=tbc_supplier_map,
    )

    _clear_template_columns(activity_ws, DATA_START_ROW, list(activity_cols.values()))
    _clear_template_columns(raw_ws, DATA_START_ROW, list(raw_cols.values()))

    row_idx = DATA_START_ROW
    for _, r in expanded.iterrows():
        valid_from = r["valid_from"]
        if not isinstance(valid_from, date):
            valid_from = _date_from_value(valid_from)
        usage_value = float(r["usage"]) if not pd.isna(r["usage"]) else 0
        _write_template_value(activity_ws, row_idx, activity_cols["raw_name"], r["raw_material"])
        _write_template_value(activity_ws, row_idx, activity_cols["raw_code"], r["raw_material"])
        _write_template_value(activity_ws, row_idx, activity_cols["start_date"], _year_start(valid_from))
        _write_template_value(activity_ws, row_idx, activity_cols["end_date"], _year_end(valid_from))
        _write_template_value(activity_ws, row_idx, activity_cols["document_type"], document_type_value)
        _write_template_value(activity_ws, row_idx, activity_cols["document_number"], "")
        _write_template_value(activity_ws, row_idx, activity_cols["usage"], usage_value)
        _write_template_value(activity_ws, row_idx, activity_cols["unit"], r["unit"])
        _write_template_value(activity_ws, row_idx, activity_cols["data_source"], "SAP")
        _write_template_value(activity_ws, row_idx, activity_cols["data_source_other"], "")
        _write_template_value(
            activity_ws, row_idx, activity_cols["calculate_transportation_emissions"],
            r.get("calculate_transportation_emissions", "") or transport_calculation_yes_value,
        )
        _write_template_value(activity_ws, row_idx, activity_cols["supplier_name"], r.get("supplier_name", ""))
        _write_template_value(activity_ws, row_idx, activity_cols["transport_origin"], r.get("transport_origin", ""))
        _write_template_value(activity_ws, row_idx, activity_cols["transport_destination"], r.get("transport_destination", ""))
        _write_template_value(activity_ws, row_idx, activity_cols["target_product"], r["target_product"])
        _write_template_value(activity_ws, row_idx, activity_cols["comment"], "")
        _write_template_value(activity_ws, row_idx, activity_cols["material_group"], r["material_group"])
        _write_template_value(activity_ws, row_idx, activity_cols.get("net_weight"), r.get("net_weight", ""))
        _write_template_value(activity_ws, row_idx, activity_cols.get("gross_weight"), r.get("gross_weight", ""))
        _write_template_value(activity_ws, row_idx, activity_cols.get("weight_unit"), r.get("weight_uom", ""))
        activity_ws.cell(row_idx, activity_cols["start_date"]).number_format = "yyyy/mm/dd"
        activity_ws.cell(row_idx, activity_cols["end_date"]).number_format = "yyyy/mm/dd"
        row_idx += 1

    raw_unique = expanded.sort_values(["raw_material"]).drop_duplicates(subset=["raw_material"])[["raw_material", "description"]] if not expanded.empty else pd.DataFrame(columns=["raw_material", "description"])
    row_idx = DATA_START_ROW
    for _, r in raw_unique.iterrows():
        _write_template_value(raw_ws, row_idx, raw_cols["raw_name"], r["raw_material"])
        _write_template_value(raw_ws, row_idx, raw_cols["raw_code"], r["raw_material"])
        _write_template_value(raw_ws, row_idx, raw_cols["description"], r["description"])
        row_idx += 1

    wb.save(output_path)
    result = {
        "output_filename": output_path.name,
        "activity_template_columns": activity_cols,
        "raw_material_template_columns": raw_cols,
        "activity_rows": int(len(expanded)),
        "raw_materials": int(expanded["raw_material"].nunique()) if not expanded.empty else 0,
        "zero_usage_rows_excluded": int(zero_usage_rows_excluded),
        "supplier_name_options": int(len(supplier_options)),
        "site_tbc_supplier_count": int(len({id(v) for v in (tbc_supplier_map or {}).values()})),
        "tbc_fallback_policy": "prefer_uploaded_plant_tbc_then_system_fallback",
    }
    result.update(supplier_write_summary)
    if return_expanded:
        return result, expanded
    return result


def generate_raw_material_bulk_file(
    bom_path: str | Path | list[str | Path] | tuple[str | Path, ...],
    raw_material_template_path: str | Path,
    output_path: str | Path,
    mapping: dict[str, str | None] | None = None,
    supplier_paths: list[str | Path] | tuple[str | Path, ...] | None = None,
    supplier_bulk_template_path: str | Path | None = None,
    supplier_bulk_output_path: str | Path | None = None,
) -> Dict[str, Any]:
    output_path = Path(output_path)
    bom_df, used_columns = _read_boms(bom_path, mapping=mapping)
    supplier_map, tbc_supplier_map, supplier_summary = _read_supplier_files(supplier_paths)
    exploded, summary = _explode_bom(bom_df)
    exploded, zero_usage_rows_excluded = _exclude_zero_usage_rows(exploded)
    exploded, blank_material_group_rows_excluded, blank_material_group_raw_materials, blank_material_group_targets = _exclude_blank_material_group_rows(exploded)
    summary["zero_usage_rows_excluded"] = int(zero_usage_rows_excluded)
    summary["blank_material_group_rows_excluded"] = int(blank_material_group_rows_excluded)
    summary["blank_material_group_raw_materials"] = blank_material_group_raw_materials[:50]
    summary["blank_material_group_raw_material_count"] = int(len(blank_material_group_raw_materials))
    summary["blank_material_group_targets"] = blank_material_group_targets[:50]
    summary["blank_material_group_target_count"] = int(len(blank_material_group_targets))
    summary["blank_material_group_rule"] = "M2B excludes final BOM rows whose Standard BOM Material group is blank; M2A expanded BOM remains unchanged."
    summary["activity_rows"] = int(len(exploded))
    summary["raw_materials"] = int(exploded["raw_material"].nunique()) if not exploded.empty else 0
    write_summary, expanded = _write_raw_material_bulk_from_exploded(
        exploded=exploded,
        raw_material_template_path=raw_material_template_path,
        output_path=output_path,
        supplier_map=supplier_map,
        tbc_supplier_map=tbc_supplier_map,
        return_expanded=True,
    )
    summary.update(write_summary)
    summary.update(supplier_summary)
    if supplier_bulk_template_path and supplier_bulk_output_path:
        summary.update(_write_supplier_bulk_create_file(expanded, supplier_bulk_template_path, supplier_bulk_output_path))
    summary["used_columns"] = used_columns
    summary["bom_files"] = int(used_columns.get("bom_files", 1)) if isinstance(used_columns, dict) else 1
    summary["bom_rows_before_dedup"] = int(used_columns.get("bom_rows_before_dedup", 0)) if isinstance(used_columns, dict) else 0
    summary["bom_rows_after_dedup"] = int(used_columns.get("bom_rows_after_dedup", 0)) if isinstance(used_columns, dict) else 0
    summary["bom_duplicate_rows_removed"] = int(used_columns.get("bom_duplicate_rows_removed", 0)) if isinstance(used_columns, dict) else 0
    return summary


def generate_raw_material_bulk_files_by_site_zip(
    bom_path: str | Path | list[str | Path] | tuple[str | Path, ...],
    raw_material_template_path: str | Path,
    output_dir: str | Path,
    token: str,
    step1_output_path: str | Path,
    mapping: dict[str, str | None] | None = None,
    supplier_paths: list[str | Path] | tuple[str | Path, ...] | None = None,
    supplier_bulk_template_path: str | Path | None = None,
    supplier_bulk_output_path: str | Path | None = None,
) -> Dict[str, Any]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    bom_df, used_columns = _read_boms(bom_path, mapping=mapping)
    supplier_map, tbc_supplier_map, supplier_summary = _read_supplier_files(supplier_paths)
    exploded, base_summary = _explode_bom(bom_df)
    exploded, zero_usage_rows_excluded = _exclude_zero_usage_rows(exploded)
    annual_qty_map, annual_qty_source_summary = _read_step1_annual_quantity_map(step1_output_path)
    # Production mode: do not generate bom_trace_detail_*.xlsx.
    # That file was only for BOM debugging and must not be included in the Module 2 ZIP,
    # because Module 3 consumes every Excel in the package as raw-material bulk input.
    trace_summary: dict[str, Any] = {}
    exploded, annual_usage_summary = _apply_annual_quantity_to_exploded_usage(exploded, annual_qty_map)
    exploded, zero_annual_usage_rows_excluded = _exclude_zero_usage_rows(exploded)
    total_hour_by_target, working_hour_summary = _calculate_total_working_hour_by_target(step1_output_path=step1_output_path, bom_df=bom_df)
    exploded, zero_total_working_hour_rows_excluded = _exclude_zero_total_working_hour_target_rows(exploded=exploded, total_hour_by_material=total_hour_by_target)
    exploded, blank_material_group_rows_excluded, blank_material_group_raw_materials, blank_material_group_targets = _exclude_blank_material_group_rows(exploded)
    base_summary["zero_usage_rows_excluded"] = int(zero_usage_rows_excluded)
    base_summary["zero_annual_usage_rows_excluded"] = int(zero_annual_usage_rows_excluded)
    base_summary["zero_total_working_hour_rows_excluded"] = int(zero_total_working_hour_rows_excluded)
    base_summary["blank_material_group_rows_excluded"] = int(blank_material_group_rows_excluded)
    base_summary["blank_material_group_raw_materials"] = blank_material_group_raw_materials[:50]
    base_summary["blank_material_group_raw_material_count"] = int(len(blank_material_group_raw_materials))
    base_summary["blank_material_group_targets"] = blank_material_group_targets[:50]
    base_summary["blank_material_group_target_count"] = int(len(blank_material_group_targets))
    base_summary["blank_material_group_rule"] = "M2B excludes final BOM rows whose Standard BOM Material group is blank; M2A expanded BOM remains unchanged."
    base_summary.update(annual_qty_source_summary)
    base_summary.update(annual_usage_summary)
    base_summary.update(trace_summary)
    base_summary["activity_rows"] = int(len(exploded))
    base_summary["raw_materials"] = int(exploded["raw_material"].nunique()) if not exploded.empty else 0
    base_summary.update(working_hour_summary)

    site_map, step1_summary = _read_step1_product_master_maps(step1_output_path)
    work = exploded.copy()
    if work.empty:
        work["_production_site"] = ""
    else:
        work["_target_key"] = work["target_product"].apply(_normalize_material_key)
        work["_production_site"] = work["_target_key"].map(site_map).fillna("Unassigned")
        work["_production_site"] = work["_production_site"].apply(lambda x: str(x or "").strip() or "Unassigned")
        # Original official logic: Transportation Destination follows Step1 Production Site.
        work["transport_destination"] = work["_production_site"]
        if "material_group" not in work.columns:
            work["material_group"] = ""

    site_values = sorted({str(x).strip() or "Unassigned" for x in work["_production_site"].tolist()}) if not work.empty else ["Unassigned"]
    generated_files: list[dict[str, Any]] = []
    expanded_all: list[pd.DataFrame] = []
    supplier_matched_total = 0
    supplier_expanded_total = 0
    supplier_name_matched_total = 0
    supplier_name_missing_total = 0
    supplier_usage_split_source_total = 0
    supplier_usage_split_output_total = 0
    supplier_usage_total_before_split = 0.0
    supplier_usage_total_after_split = 0.0
    supplier_options_total = 0
    zip_filename = f"raw_material_activity_data_bulk_by_site_{token}.zip"
    zip_path = output_dir / zip_filename

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for site in site_values:
            site_df = work[work["_production_site"] == site].copy().drop(columns=["_target_key", "_production_site"], errors="ignore")
            safe_site = _sanitize_filename_part(site)
            file_path = output_dir / f"raw_material_activity_data_bulk_{safe_site}_{token}.xlsx"
            write_summary, expanded_site = _write_raw_material_bulk_from_exploded(
                exploded=site_df,
                raw_material_template_path=raw_material_template_path,
                output_path=file_path,
                supplier_map=supplier_map,
                tbc_supplier_map=tbc_supplier_map,
                return_expanded=True,
            )
            expanded_all.append(expanded_site)
            supplier_matched_total += int(write_summary.get("supplier_matched_rows", 0))
            supplier_expanded_total += int(write_summary.get("supplier_expanded_rows", 0))
            supplier_name_matched_total += int(write_summary.get("supplier_name_matched_rows", 0))
            supplier_name_missing_total += int(write_summary.get("supplier_name_missing_rows", 0))
            supplier_usage_split_source_total += int(write_summary.get("supplier_usage_split_source_rows", 0))
            supplier_usage_split_output_total += int(write_summary.get("supplier_usage_split_output_rows", 0))
            supplier_usage_total_before_split += float(write_summary.get("supplier_usage_total_before_split", 0.0) or 0.0)
            supplier_usage_total_after_split += float(write_summary.get("supplier_usage_total_after_split", 0.0) or 0.0)
            supplier_options_total = max(supplier_options_total, int(write_summary.get("supplier_name_options", 0)))
            zf.write(file_path, arcname=file_path.name)
            generated_files.append({
                "production_site": site,
                "filename": file_path.name,
                "activity_rows": int(write_summary.get("activity_rows", 0)),
                "raw_materials": int(write_summary.get("raw_materials", 0)),
            })

    combined_expanded = pd.concat(expanded_all, ignore_index=True) if expanded_all else pd.DataFrame()
    supplier_bulk_summary = {}
    if supplier_bulk_template_path and supplier_bulk_output_path:
        supplier_bulk_summary = _write_supplier_bulk_create_file(combined_expanded, supplier_bulk_template_path, supplier_bulk_output_path)

    unassigned_rows = int((work["_production_site"] == "Unassigned").sum()) if not work.empty else 0
    summary = dict(base_summary)
    summary.update({
        "output_filename": zip_filename,
        "download_url": f"/download/{zip_filename}",
        "split_by_production_site": True,
        "production_site_files": generated_files,
        "production_site_count": int(len(site_values)),
        "unassigned_rows": unassigned_rows,
        "used_columns": used_columns,
        "bom_files": int(used_columns.get("bom_files", 1)) if isinstance(used_columns, dict) else 1,
        "bom_rows_before_dedup": int(used_columns.get("bom_rows_before_dedup", 0)) if isinstance(used_columns, dict) else 0,
        "bom_rows_after_dedup": int(used_columns.get("bom_rows_after_dedup", 0)) if isinstance(used_columns, dict) else 0,
        "bom_duplicate_rows_removed": int(used_columns.get("bom_duplicate_rows_removed", 0)) if isinstance(used_columns, dict) else 0,
        "supplier_matched_rows": int(supplier_matched_total),
        "supplier_expanded_rows": int(supplier_expanded_total),
        "supplier_name_matched_rows": int(supplier_name_matched_total),
        "supplier_name_missing_rows": int(supplier_name_missing_total),
        "supplier_usage_split_source_rows": int(supplier_usage_split_source_total),
        "supplier_usage_split_output_rows": int(supplier_usage_split_output_total),
        "supplier_usage_total_before_split": float(supplier_usage_total_before_split),
        "supplier_usage_total_after_split": float(supplier_usage_total_after_split),
        "supplier_usage_split_rule": "equal_share_original_usage_divided_by_unique_supplier_count",
        "supplier_dropdown_matched_rows": int(supplier_name_matched_total),
        "supplier_dropdown_missing_rows": int(supplier_name_missing_total),
        "supplier_name_options": int(supplier_options_total),
    })
    summary.update(step1_summary)
    summary.update(supplier_summary)
    summary.update(supplier_bulk_summary)
    return summary


# =========================================================
# Module 2A · Standard BOM Total Usage
# Standard BOM -> Standard BOM total usage workbook only.
# This function intentionally does not read or write Raw Material Bulk templates.
# =========================================================
STANDARD_BOM_TOTAL_USAGE_BASE_SHEET_NAME = "標準BOM表總用量"
STANDARD_BOM_TOTAL_USAGE_SUMMARY_SHEET_NAME = "輸出摘要"
STANDARD_BOM_TOTAL_USAGE_SAFE_ROWS_PER_SHEET = 900000


def _excel_safe_sheet_name(name: str, fallback: str = "Sheet") -> str:
    text = re.sub(r"[\\/*?:\[\]]", "_", str(name or "").strip())
    text = text[:31].strip()
    return text or fallback


def _jsonable_text(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if pd.isna(value):
        return ""
    return value


def _standard_bom_total_usage_rows(
    bom_df: pd.DataFrame,
    exploded: pd.DataFrame,
    used_columns: dict[str, Any],
) -> tuple[list[str], pd.DataFrame]:
    """Build the output table for Module 2A while keeping original BOM columns.

    The output table preserves all original Standard BOM columns. Only the key
    BOM fields are rewritten so each row represents final raw-material usage
    under one finished product:
      - Material / Parent Node -> finished product
      - Component -> final raw material
      - CS03 Qty -> rolled-up total usage
      - CS03 UoM -> raw material unit

    Alternative item probability is already applied before BOM expansion, so
    Altitem group and Usage probability% are cleared to avoid double counting in
    downstream stages.
    """
    original_columns = [c for c in bom_df.columns if not str(c).startswith("_")]
    if not original_columns:
        original_columns = ["Material", "Parent Node", "Component", "CS03 Qty", "CS03 UoM"]

    material_col = str(used_columns.get("material_col") or "").strip()
    parent_col = str(used_columns.get("parent_col") or "Parent Node").strip()
    component_col = str(used_columns.get("component_col") or "Component").strip()
    qty_col = str(used_columns.get("qty_col") or "CS03 Qty").strip()
    unit_col = str(used_columns.get("unit_col") or "CS03 UoM").strip()
    description_col = str(used_columns.get("description_col") or "").strip()
    material_group_col = str(used_columns.get("material_group_col") or "").strip()
    valid_from_col = str(used_columns.get("valid_from_col") or "").strip()
    altitem_group_col = str(used_columns.get("altitem_group_col") or "").strip()
    usage_probability_col = str(used_columns.get("usage_probability_col") or "").strip()
    net_weight_col = str(used_columns.get("net_weight_col") or "").strip()
    gross_weight_col = str(used_columns.get("gross_weight_col") or "").strip()
    weight_uom_col = str(used_columns.get("weight_uom_col") or "").strip()

    # Representative original BOM row lookup for preserving all non-key fields.
    by_source_component: dict[tuple[str, str, str], dict[str, Any]] = {}
    by_component: dict[tuple[str, str], dict[str, Any]] = {}
    for _, row_dict in bom_df.iterrows():
        source = _safe_text(row_dict.get("_bom_material"))
        component = _safe_text(row_dict.get("_component"))
        unit = _safe_text(row_dict.get("_uom"))
        if not component:
            continue
        original = {col: row_dict.get(col, "") for col in original_columns}
        by_source_component.setdefault((source, component, unit), original)
        by_component.setdefault((component, unit), original)

    trace_detail = exploded.attrs.get("trace_detail")
    trace_source: dict[tuple[str, Any, str, str], str] = {}
    if isinstance(trace_detail, pd.DataFrame) and not trace_detail.empty:
        for row in trace_detail.itertuples(index=False):
            target = _safe_text(getattr(row, "target_product", ""))
            target_valid_from = getattr(row, "target_valid_from", None)
            raw = _safe_text(getattr(row, "raw_material", ""))
            unit = _safe_text(getattr(row, "unit", ""))
            source = _safe_text(getattr(row, "source_material", ""))
            trace_source.setdefault((target, target_valid_from, raw, unit), source)

    output_rows: list[dict[str, Any]] = []
    if exploded is None or exploded.empty:
        return original_columns, pd.DataFrame(columns=original_columns)

    sort_columns = [c for c in ["target_product", "target_valid_from", "raw_material", "unit"] if c in exploded.columns]
    work = exploded.sort_values(sort_columns, kind="mergesort").reset_index(drop=True)
    for row in work.itertuples(index=False):
        target_product = _safe_text(getattr(row, "target_product", ""))
        target_valid_from = getattr(row, "target_valid_from", None)
        raw_material = _safe_text(getattr(row, "raw_material", ""))
        unit = _safe_text(getattr(row, "unit", ""))
        source = trace_source.get((target_product, target_valid_from, raw_material, unit), "")
        template = by_source_component.get((source, raw_material, unit)) or by_component.get((raw_material, unit)) or {}
        out = {col: template.get(col, "") for col in original_columns}

        for col, value in [
            (material_col, target_product),
            (parent_col, target_product),
            (component_col, raw_material),
            (qty_col, float(getattr(row, "usage", 0.0) or 0.0)),
            (unit_col, unit),
            (description_col, _safe_text(getattr(row, "description", ""))),
            (material_group_col, _safe_text(getattr(row, "material_group", ""))),
            (valid_from_col, getattr(row, "target_valid_from", getattr(row, "valid_from", ""))),
            (net_weight_col, getattr(row, "net_weight", "")),
            (gross_weight_col, getattr(row, "gross_weight", "")),
            (weight_uom_col, getattr(row, "weight_uom", "")),
        ]:
            if col and col in out:
                out[col] = value

        if altitem_group_col and altitem_group_col in out:
            out[altitem_group_col] = ""
        if usage_probability_col and usage_probability_col in out:
            out[usage_probability_col] = ""

        output_rows.append(out)

    return original_columns, pd.DataFrame(output_rows, columns=original_columns)


def _write_standard_bom_total_usage_workbook(
    output_df: pd.DataFrame,
    output_path: str | Path,
    summary: dict[str, Any],
    max_rows_per_sheet: int = STANDARD_BOM_TOTAL_USAGE_SAFE_ROWS_PER_SHEET,
    product_column_name: str | None = None,
    progress_callback=None,
) -> dict[str, Any]:
    """Write Module 2A workbook and split sheets without splitting products."""
    from openpyxl import Workbook

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    max_rows_per_sheet = int(max_rows_per_sheet or STANDARD_BOM_TOTAL_USAGE_SAFE_ROWS_PER_SHEET)
    max_rows_per_sheet = min(max_rows_per_sheet, 1048575)
    if max_rows_per_sheet < 1:
        max_rows_per_sheet = STANDARD_BOM_TOTAL_USAGE_SAFE_ROWS_PER_SHEET

    if output_df is None:
        output_df = pd.DataFrame()

    product_col = product_column_name if product_column_name in output_df.columns else None
    if product_col is None:
        product_col = output_df.columns[0] if len(output_df.columns) else "Material"

    sheet_plan: list[dict[str, Any]] = []
    if output_df.empty:
        sheet_plan.append({"sheet_name": f"{STANDARD_BOM_TOTAL_USAGE_BASE_SHEET_NAME}_001", "products": [], "rows": 0})
    else:
        current_products: list[str] = []
        current_rows = 0
        sheet_idx = 1
        for product, group in output_df.groupby(product_col, sort=True, dropna=False):
            product_text = _safe_text(product) or "Unassigned"
            group_rows = int(len(group))
            if group_rows > max_rows_per_sheet:
                raise ValueError(
                    f"單一成品料號 {product_text} 的資料列數 {group_rows:,} 超過 Excel 單一分頁安全上限 {max_rows_per_sheet:,}，無法在不切斷成品料號的前提下輸出。"
                )
            if current_rows and current_rows + group_rows > max_rows_per_sheet:
                sheet_plan.append({
                    "sheet_name": f"{STANDARD_BOM_TOTAL_USAGE_BASE_SHEET_NAME}_{sheet_idx:03d}",
                    "products": current_products,
                    "rows": current_rows,
                })
                sheet_idx += 1
                current_products = []
                current_rows = 0
            current_products.append(product_text)
            current_rows += group_rows
        sheet_plan.append({
            "sheet_name": f"{STANDARD_BOM_TOTAL_USAGE_BASE_SHEET_NAME}_{sheet_idx:03d}",
            "products": current_products,
            "rows": current_rows,
        })

    if progress_callback:
        progress_callback(step="Writing output workbook", processed=0, total=int(len(output_df)), progress=86)

    wb = Workbook(write_only=True)
    summary_ws = wb.create_sheet(STANDARD_BOM_TOTAL_USAGE_SUMMARY_SHEET_NAME)
    summary_rows = [
        ["項目", "值"],
        ["輸出檔案", output_path.name],
        ["Module 1 Step 1 來源檔案", str(summary.get("source_filename", "") or "")],
        ["來源版次", str(summary.get("bom_version", "") or "")],
        ["來源日期", str(summary.get("bom_date", "") or "")],
        ["來源建立時間", str(summary.get("source_modified_at", "") or "")],
        ["BOM 檔案數", int(summary.get("bom_files", 0) or 0)],
        ["BOM 原始列數", int(summary.get("bom_rows_before_dedup", 0) or 0)],
        ["BOM 去重後列數", int(summary.get("bom_rows_after_dedup", 0) or 0)],
        ["重複列移除", int(summary.get("bom_duplicate_rows_removed", 0) or 0)],
        ["重複 BOM 版本移除", int(summary.get("bom_version_duplicate_groups_removed", 0) or 0)],
        ["跨 Material 半品展開次數", int(summary.get("cross_material_expansions", 0) or 0)],
        ["循環路徑略過", int(summary.get("cycles_skipped", 0) or 0)],
        ["成品數", int(summary.get("products", 0) or 0)],
        ["半品數", int(summary.get("semi_finished", 0) or 0)],
        ["最終原物料列數", int(len(output_df))],
        ["最終原物料數", int(summary.get("raw_materials", 0) or 0)],
        ["最大階層", int(summary.get("max_level", 0) or 0)],
        ["Altitem rows", int(summary.get("altitem_rows", 0) or 0)],
        ["Altitem adjusted rows", int(summary.get("altitem_adjusted_rows", 0) or 0)],
        ["Altitem probability missing rows", int(summary.get("altitem_probability_missing_rows", 0) or 0)],
        ["分頁安全上限", int(max_rows_per_sheet)],
        ["分頁數", int(len(sheet_plan))],
    ]
    for row in summary_rows:
        summary_ws.append(row)
    summary_ws.append([])
    summary_ws.append(["Sheet", "資料列數", "成品數", "第一個成品", "最後一個成品"])
    for plan in sheet_plan:
        products = plan.get("products") or []
        summary_ws.append([plan["sheet_name"], int(plan["rows"]), int(len(products)), products[0] if products else "", products[-1] if products else ""])

    headers = list(output_df.columns)
    processed = 0
    total = int(len(output_df))
    for plan in sheet_plan:
        ws = wb.create_sheet(_excel_safe_sheet_name(plan["sheet_name"]))
        ws.append(headers)
        products = set(plan.get("products") or [])
        if output_df.empty:
            continue
        part = output_df[output_df[product_col].apply(lambda x: _safe_text(x) or "Unassigned").isin(products)]
        for row in part.itertuples(index=False, name=None):
            ws.append([_jsonable_text(v) for v in row])
            processed += 1
            if progress_callback and (processed % 5000 == 0 or processed == total):
                progress_callback(step=f"Writing {plan['sheet_name']}", processed=processed, total=total, progress=min(98, 86 + int((processed / max(total, 1)) * 12)))

    wb.save(output_path)
    return {
        "output_filename": output_path.name,
        "download_url": f"/download/{output_path.name}",
        "sheet_count": int(len(sheet_plan)),
        "sheet_plan": sheet_plan,
        "rows_per_sheet_limit": int(max_rows_per_sheet),
        "standard_bom_total_usage_rows": int(len(output_df)),
    }


def generate_standard_bom_total_usage_file(
    bom_path: str | Path | list[str | Path] | tuple[str | Path, ...],
    output_path: str | Path,
    mapping: dict[str, str | None] | None = None,
    bom_version: str | None = None,
    bom_date: str | None = None,
    source_filename: str | None = None,
    source_modified_at: str | None = None,
    max_rows_per_sheet: int = STANDARD_BOM_TOTAL_USAGE_SAFE_ROWS_PER_SHEET,
    progress_callback=None,
) -> Dict[str, Any]:
    """Generate Module 2A Standard BOM total usage workbook.

    This is intentionally separated from Raw Material Bulk generation. It reads
    only Standard BOM files, explodes semi-finished items to final raw materials,
    aggregates usage back to finished products, and writes an intermediate Excel
    workbook for Module 2B.
    """
    if progress_callback:
        progress_callback(step="Reading Standard BOM", processed=0, total=0, progress=8)
    bom_df, used_columns = _read_boms(bom_path, mapping=mapping)

    if progress_callback:
        progress_callback(step="Expanding BOM", processed=0, total=int(used_columns.get("bom_rows_after_dedup", len(bom_df))), progress=32)
    exploded, summary = _explode_bom(bom_df)
    summary.update(used_columns)
    summary["module2a_rule"] = "Standard BOM -> final raw-material total usage only; Raw Material Bulk Template is not read or written."
    summary["bom_version"] = str(bom_version or "").strip()
    summary["bom_date"] = str(bom_date or "").strip()
    summary["source_filename"] = str(source_filename or "").strip()
    summary["source_modified_at"] = str(source_modified_at or "").strip()

    if progress_callback:
        progress_callback(step="Aggregating final raw material usage", processed=int(len(exploded)), total=int(len(exploded)), progress=72)
    headers, output_df = _standard_bom_total_usage_rows(bom_df=bom_df, exploded=exploded, used_columns=used_columns)
    summary["output_columns"] = headers
    summary["standard_bom_total_usage_rows"] = int(len(output_df))

    product_col = str(used_columns.get("material_col") or "").strip() or str(used_columns.get("parent_col") or "Parent Node").strip()
    write_summary = _write_standard_bom_total_usage_workbook(
        output_df=output_df,
        output_path=output_path,
        summary=summary,
        max_rows_per_sheet=max_rows_per_sheet,
        product_column_name=product_col,
        progress_callback=progress_callback,
    )
    summary.update(write_summary)
    if progress_callback:
        progress_callback(step="Completed", processed=int(len(output_df)), total=int(len(output_df)), progress=100)
    return summary


# =========================================================
# Module 2B · Raw Material Bulk from Module 2A Total Usage
# Standard BOM total usage workbook + Raw Material Bulk Template -> Raw Material Bulk ZIP by Production Site.
# This function intentionally does not re-expand BOM and does not apply Supplier mapping.
# =========================================================
def _read_standard_bom_total_usage_workbook(
    standard_total_usage_path: str | Path,
    mapping: dict[str, str | None] | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Read Module 2A output workbook and convert it to the exploded schema.

    Module 2A output preserves Standard BOM columns but already contains final
    raw-material usage per finished product. Module 2B must not call _explode_bom
    again and must not apply Altitem probability again.
    """
    path = Path(standard_total_usage_path)
    if not path.exists():
        raise FileNotFoundError(f"找不到 Module 2A 標準BOM表總用量檔案：{path}")

    m = _resolve_mapping(mapping)
    try:
        sheets = pd.read_excel(path, sheet_name=None, dtype=object)
    except Exception as exc:
        raise ValueError(f"無法讀取標準BOM表總用量檔案：{exc}") from exc

    parts: list[pd.DataFrame] = []
    sheet_names: list[str] = []
    for sheet_name, df in sheets.items():
        name = str(sheet_name or "").strip()
        if name == STANDARD_BOM_TOTAL_USAGE_SUMMARY_SHEET_NAME:
            continue
        if not name.startswith(STANDARD_BOM_TOTAL_USAGE_BASE_SHEET_NAME):
            continue
        if df is None:
            continue
        part = df.copy()
        if part.empty:
            continue
        part["_module2a_sheet"] = name
        parts.append(part)
        sheet_names.append(name)

    if not parts:
        raise ValueError("標準BOM表總用量檔案中找不到可讀取的『標準BOM表總用量』分頁，請先完成 Module 2A。")

    df = pd.concat(parts, ignore_index=True)
    material_col = _find_optional_column(df, m.get("material_col", "Material"))
    parent_col = _find_optional_column(df, m.get("parent_col", "Parent Node"))
    component_col = _find_column(df, m["component_col"])
    qty_col = _find_column(df, m["qty_col"])
    unit_col = _find_column(df, m["unit_col"])
    description_col = _find_optional_column(df, m.get("description_col", "Component Description"))
    material_group_col = _find_optional_column(df, m.get("material_group_col", "Material group"))
    valid_from_col = _find_optional_column(df, m.get("valid_from_col", "BOM Valid From"))
    net_weight_col = _find_optional_column(df, m.get("net_weight_col", "Net weight"))
    gross_weight_col = _find_optional_column(df, m.get("gross_weight_col", "Gross weight"))
    weight_uom_col = _find_optional_column(df, m.get("weight_uom_col", "Weight UoM"))

    target_col = material_col or parent_col
    if not target_col:
        raise ValueError("標準BOM表總用量缺少成品料號欄位：Material 或 Parent Node")

    work = pd.DataFrame({
        "target_product": df[target_col].apply(_safe_text),
        "source_material": df[target_col].apply(_safe_text),
        "raw_material": df[component_col].apply(_safe_text),
        "usage": df[qty_col].apply(_safe_number),
        "unit": df[unit_col].apply(_safe_text),
        "description": df[description_col].apply(_safe_text) if description_col else "",
        "material_group": df[material_group_col].apply(_safe_text) if material_group_col else "",
        "valid_from": df[valid_from_col].apply(_date_from_value) if valid_from_col else date(datetime.now().year, 1, 1),
        "level": 0,
        "net_weight": df[net_weight_col].apply(_safe_number) if net_weight_col else "",
        "gross_weight": df[gross_weight_col].apply(_safe_number) if gross_weight_col else "",
        "weight_uom": df[weight_uom_col].apply(_safe_text) if weight_uom_col else "",
    })
    before_filter = int(len(work))
    work = work[(work["target_product"] != "") & (work["raw_material"] != "")].copy()
    return work, {
        "module2a_total_usage_source_filename": path.name,
        "module2a_total_usage_source_sheets": sheet_names,
        "module2a_total_usage_rows_read": int(before_filter),
        "module2a_total_usage_valid_rows": int(len(work)),
        "module2a_total_usage_rule": "Read Module 2A total usage as per-PC final raw-material usage; BOM is not re-expanded and Altitem probability is not re-applied.",
        "used_columns": {
            "target_product_col": str(target_col),
            "component_col": str(component_col),
            "qty_col": str(qty_col),
            "unit_col": str(unit_col),
            "description_col": str(description_col or ""),
            "material_group_col": str(material_group_col or ""),
            "valid_from_col": str(valid_from_col or ""),
            "net_weight_col": str(net_weight_col or ""),
            "gross_weight_col": str(gross_weight_col or ""),
            "weight_uom_col": str(weight_uom_col or ""),
        },
    }




_ACTIVITY_VISIBLE_DEFAULT_COLS: dict[str, int] = {
    "raw_name": 1,
    "raw_code": 2,
    "start_date": 3,
    "end_date": 4,
    "document_type": 5,
    "document_number": 6,
    "usage": 7,
    "unit": 8,
    "net_weight": 9,
    "gross_weight": 10,
    "weight_unit": 11,
    "data_source": 12,
    "data_source_other": 13,
    "calculate_transportation_emissions": 14,
    "supplier_name": 15,
    "transport_origin": 16,
    "transport_destination": 17,
    "target_product": 18,
    "comment": 19,
    "material_group": 20,
}

_ACTIVITY_VISIBLE_HEADERS: dict[str, str] = {
    "raw_name": "Raw Material Name",
    "raw_code": "Raw Material Code",
    "start_date": "Doc. Start Date",
    "end_date": "Doc. End Date",
    "document_type": "Document Type",
    "document_number": "Document Number (optional)",
    "usage": "Usage",
    "unit": "Activity Data Unit",
    "net_weight": "Net Weight (optional)",
    "gross_weight": "Gross Weight (optional)",
    "weight_unit": "Weight Unit (optional)",
    "data_source": "Data Source",
    "data_source_other": "Data Source Other",
    "calculate_transportation_emissions": "Calculate Transportation Emissions (Required)",
    "supplier_name": "Supplier Name (optional)",
    "transport_origin": "Transportation Origin",
    "transport_destination": "Transportation Destination",
    "target_product": "Allocated Target Product/Service",
    "comment": "Comment (optional)",
    "material_group": "Material Group",
}

_RAW_VISIBLE_DEFAULT_COLS: dict[str, int] = {
    "raw_name": 1,
    "raw_code": 2,
    "description": 6,
}

_RAW_VISIBLE_HEADERS: dict[str, str] = {
    "raw_name": "Raw Material Name",
    "raw_code": "Raw Material Code",
    "description": "Raw Material Description (Optional)",
}


def _read_bulk_header_rows(ws, width: int | None = None) -> list[list[Any]]:
    """Return the first two Bulk Template header rows, padded to a stable width."""
    max_col = max(1, int(width or 0), int(getattr(ws, "max_column", 1) or 1))
    rows: list[list[Any]] = []
    for row_idx in (1, 2):
        row = [ws.cell(row_idx, col).value or "" for col in range(1, max_col + 1)]
        rows.append(row)
    return rows


def _bulk_visible_header_missing(header_rows: list[list[Any]]) -> bool:
    if len(header_rows) < 2:
        return True
    visible_values = [str(v or "").strip() for v in header_rows[1]]
    return sum(1 for v in visible_values if v) < 3


def _bulk_find_col_from_rows(header_rows: list[list[Any]], aliases: list[str], fallback_col: int | None = None) -> int | None:
    """Find a visible Bulk column by header rows, preferring row 2 labels.

    Legacy M2B/M2C lightweight files accidentally omitted row 2 and still kept
    internal/helper keys such as activity_data_unit in hidden columns AA:AG.
    When row 2 is missing, prefer the known visible fallback positions instead
    of those helper columns so M2C/M3 continue reading H/K/L/N etc. correctly.
    """
    alias_keys = {_normalize_template_header(a) for a in aliases if str(a or "").strip()}
    if not alias_keys:
        return fallback_col
    # Prefer user-visible row 2, then internal row 1 only when row 2 exists.
    row_order = [1, 0] if len(header_rows) >= 2 else [0]
    visible_missing = _bulk_visible_header_missing(header_rows)
    for row_idx in row_order:
        if row_idx >= len(header_rows):
            continue
        if visible_missing and row_idx == 0 and fallback_col:
            # Avoid mapping to hidden helper columns in legacy files with blank row 2.
            return int(fallback_col)
        for col_idx, value in enumerate(header_rows[row_idx], start=1):
            if _normalize_template_header(value) in alias_keys:
                return int(col_idx)
    return int(fallback_col) if fallback_col else None


def _ensure_bulk_visible_header_row(header_rows: list[list[Any]], cols: dict[str, int | None], labels: dict[str, str]) -> list[list[Any]]:
    """Ensure row 2 contains the official visible Bulk headers.

    M2B/M2C outputs are intermediate files but M3 must be able to identify the
    visible template columns.  Row 1 may contain internal keys or hidden helper
    keys; row 2 is the reliable user-facing header row.
    """
    while len(header_rows) < 2:
        header_rows.append([])
    width = max(
        [len(r) for r in header_rows] + [int(c or 0) for c in cols.values()] + [1]
    )
    out: list[list[Any]] = []
    for row in header_rows[:2]:
        new_row = list(row)
        if len(new_row) < width:
            new_row.extend([""] * (width - len(new_row)))
        else:
            new_row = new_row[:width]
        out.append(new_row)
    for key, label in labels.items():
        col_idx = cols.get(key)
        if col_idx and int(col_idx) > 0:
            out[1][int(col_idx) - 1] = out[1][int(col_idx) - 1] or label
    return out


def _read_xlsx_header_rows_and_row3_formulas(
    path: str | Path,
    sheet_name: str,
) -> tuple[list[list[Any]], dict[int, str]]:
    """Read rows 1-2 and row-3 formulas directly from worksheet XML.

    The revised third-party template has a stale worksheet dimension ending at
    AC even though real cells/formulas extend through AI. Reading the XML by
    cell coordinates avoids losing AD~AI in openpyxl read-only mode.
    """
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    sheet_paths = _xlsx_sheet_paths_by_name(path)
    sheet_xml_path = sheet_paths.get(sheet_name)
    if not sheet_xml_path:
        raise ValueError(f"找不到 raw material bulk 分頁：{sheet_name}")

    shared: list[str] = []
    header_maps: dict[int, dict[int, Any]] = {1: {}, 2: {}}
    formula_by_col: dict[int, str] = {}
    max_col = max(M2_ACTIVITY_HELPER_FORMULA_COLS) if sheet_name == ACTIVITY_SHEET_NAME else 1
    with zipfile.ZipFile(path) as zf:
        if "xl/sharedStrings.xml" in zf.namelist():
            for _event, si in ET.iterparse(zf.open("xl/sharedStrings.xml"), events=("end",)):
                if si.tag == f"{ns}si":
                    shared.append("".join(t.text or "" for t in si.iter(f"{ns}t")))
                    si.clear()
        for _event, row_el in ET.iterparse(zf.open(sheet_xml_path), events=("end",)):
            if row_el.tag != f"{ns}row":
                continue
            try:
                row_idx = int(row_el.attrib.get("r", "0") or 0)
            except Exception:
                row_idx = 0
            if row_idx > DATA_START_ROW:
                row_el.clear()
                break
            for cell in row_el.findall(f"{ns}c"):
                col_idx = _xlsx_col_to_index(cell.attrib.get("r", "")) + 1
                max_col = max(max_col, col_idx)
                if row_idx in (1, 2):
                    cell_type = cell.attrib.get("t", "")
                    value: Any = ""
                    if cell_type == "inlineStr":
                        value = "".join(t.text or "" for t in cell.iter(f"{ns}t"))
                    else:
                        v = cell.find(f"{ns}v")
                        if v is not None and v.text is not None:
                            raw = v.text
                            if cell_type == "s":
                                try:
                                    value = shared[int(raw)]
                                except Exception:
                                    value = raw
                            else:
                                value = raw
                    header_maps[row_idx][col_idx] = value
                elif row_idx == DATA_START_ROW and col_idx in M2_ACTIVITY_HELPER_FORMULA_COLS:
                    formula = cell.find(f"{ns}f")
                    if formula is not None and formula.text:
                        formula_by_col[col_idx] = "=" + formula.text
            row_el.clear()

    rows: list[list[Any]] = []
    for row_idx in (1, 2):
        rows.append([header_maps[row_idx].get(col, "") for col in range(1, max_col + 1)])
    return rows, formula_by_col



_BUILTIN_XLSX_NUM_FORMATS = {
    0: "General", 1: "0", 2: "0.00", 3: "#,##0", 4: "#,##0.00",
    9: "0%", 10: "0.00%", 11: "0.00E+00", 14: "mm-dd-yy",
    15: "d-mmm-yy", 16: "d-mmm", 17: "mmm-yy", 18: "h:mm AM/PM",
    19: "h:mm:ss AM/PM", 20: "h:mm", 21: "h:mm:ss",
    22: "m/d/yy h:mm", 49: "@",
}


def _xlsx_row3_number_formats(path: str | Path, sheet_name: str) -> dict[int, str]:
    """Read each row-3 cell's real Excel number format directly from OOXML."""
    sheet_paths = _xlsx_sheet_paths_by_name(path)
    sheet_xml_path = sheet_paths.get(sheet_name)
    if not sheet_xml_path:
        return {}
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with zipfile.ZipFile(path) as zf:
        styles_root = ET.fromstring(zf.read("xl/styles.xml"))
        custom: dict[int, str] = {}
        numfmts = styles_root.find(f"{ns}numFmts")
        if numfmts is not None:
            for item in numfmts.findall(f"{ns}numFmt"):
                custom[int(item.attrib.get("numFmtId", "0"))] = item.attrib.get("formatCode", "General")
        xfs = styles_root.find(f"{ns}cellXfs")
        style_formats: list[str] = []
        if xfs is not None:
            for xf in xfs.findall(f"{ns}xf"):
                num_fmt_id = int(xf.attrib.get("numFmtId", "0"))
                style_formats.append(custom.get(num_fmt_id, _BUILTIN_XLSX_NUM_FORMATS.get(num_fmt_id, "General")))
        result: dict[int, str] = {}
        for _event, row_el in ET.iterparse(zf.open(sheet_xml_path), events=("end",)):
            if row_el.tag != f"{ns}row":
                continue
            row_idx = int(row_el.attrib.get("r", "0") or 0)
            if row_idx == DATA_START_ROW:
                for cell in row_el.findall(f"{ns}c"):
                    col_idx = _xlsx_col_to_index(cell.attrib.get("r", "")) + 1
                    style_idx = int(cell.attrib.get("s", "0") or 0)
                    result[col_idx] = style_formats[style_idx] if style_idx < len(style_formats) else "General"
                break
            if row_idx > DATA_START_ROW:
                break
            row_el.clear()
        return result


def _xlsxwriter_column_formats(workbook, number_formats: dict[int, str]) -> dict[int, Any]:
    cache: dict[str, Any] = {}
    result: dict[int, Any] = {}
    for col_idx, code in (number_formats or {}).items():
        fmt_code = str(code or "General")
        if fmt_code.lower() == "general":
            continue
        if fmt_code not in cache:
            cache[fmt_code] = workbook.add_format({"num_format": fmt_code})
        result[int(col_idx)] = cache[fmt_code]
    return result


def _write_xlsxwriter_typed_cell(ws, zero_based_row: int, zero_based_col: int, value: Any, number_formats: dict[int, str], formats: dict[int, Any]) -> None:
    col_idx = int(zero_based_col) + 1
    fmt_code = str((number_formats or {}).get(col_idx, "General") or "General")
    fmt = (formats or {}).get(col_idx)
    if value in (None, ""):
        if fmt is not None:
            ws.write_blank(zero_based_row, zero_based_col, None, fmt)
        return
    if isinstance(value, datetime):
        ws.write_datetime(zero_based_row, zero_based_col, value, fmt)
        return
    if isinstance(value, date):
        ws.write_datetime(zero_based_row, zero_based_col, datetime(value.year, value.month, value.day), fmt)
        return
    if fmt_code == "@":
        ws.write_string(zero_based_row, zero_based_col, str(value), fmt)
        return
    if isinstance(value, bool):
        ws.write_boolean(zero_based_row, zero_based_col, value, fmt)
        return
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        ws.write_number(zero_based_row, zero_based_col, float(value), fmt)
        return
    ws.write(zero_based_row, zero_based_col, value, fmt)

def _translate_m2_helper_formula(formula: str, target_row: int) -> str:
    """Translate a row-3 helper formula to the requested worksheet row."""
    if not formula or int(target_row) == DATA_START_ROW:
        return formula
    pattern = r'(?<![A-Za-z0-9_])(\$?[A-Z]{1,3})3(?![0-9])'
    return re.sub(pattern, lambda m: m.group(1) + str(int(target_row)), formula)


def _write_m2_helper_formulas_xlsxwriter(ws, zero_based_row: int, formula_by_col: dict[int, str], formats: dict[int, Any] | None = None) -> None:
    excel_row = int(zero_based_row) + 1
    for col_idx in M2_ACTIVITY_HELPER_FORMULA_COLS:
        formula = formula_by_col.get(int(col_idx))
        if formula:
            ws.write_formula(int(zero_based_row), int(col_idx) - 1, _translate_m2_helper_formula(formula, excel_row), (formats or {}).get(int(col_idx)), "")


def _apply_m2_helper_formulas_to_row(row: list[Any], formula_by_col: dict[int, str], excel_row: int) -> list[Any]:
    width = max([len(row), *M2_ACTIVITY_HELPER_FORMULA_COLS])
    if len(row) < width:
        row.extend([""] * (width - len(row)))
    for col_idx in M2_ACTIVITY_HELPER_FORMULA_COLS:
        formula = formula_by_col.get(int(col_idx))
        if formula:
            row[int(col_idx) - 1] = _translate_m2_helper_formula(formula, int(excel_row))
    return row


def _template_headers_for_lightweight_bulk(raw_material_template_path: str | Path) -> tuple[list[list[Any]], list[list[Any]], dict[str, int], dict[str, int], str, str, dict[int, str], dict[int, str], dict[int, str], dict[str, dict[str, str]]]:
    """Read Raw Material Bulk template two-row headers for Module 2B Large Dataset Mode.

    Module 2B intentionally keeps the template sheet names and column layout but
    does not copy workbook styles/dropdowns/formulas to avoid memory spikes.
    The intermediate workbook must still keep both template header rows:
    row 1 = internal keys, row 2 = visible field labels.  M3 relies on row 2 to
    avoid accidentally reading hidden helper columns such as activity_data_unit.
    """
    activity_header_rows, activity_formula_templates = _read_xlsx_header_rows_and_row3_formulas(
        raw_material_template_path,
        ACTIVITY_SHEET_NAME,
    )
    raw_header_rows, _raw_formula_templates = _read_xlsx_header_rows_and_row3_formulas(
        raw_material_template_path,
        RAW_MATERIAL_SHEET_NAME,
    )
    wb = load_workbook(raw_material_template_path, read_only=True, data_only=True)
    try:
        if ACTIVITY_SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"找不到 raw material bulk 分頁：{ACTIVITY_SHEET_NAME}")
        if RAW_MATERIAL_SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"找不到 raw material bulk 分頁：{RAW_MATERIAL_SHEET_NAME}")
        activity_cols = {
            "raw_name": _bulk_find_col_from_rows(activity_header_rows, RAW_MATERIAL_NAME_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["raw_name"]),
            "raw_code": _bulk_find_col_from_rows(activity_header_rows, RAW_MATERIAL_CODE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["raw_code"]),
            "start_date": _bulk_find_col_from_rows(activity_header_rows, DOC_START_DATE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["start_date"]),
            "end_date": _bulk_find_col_from_rows(activity_header_rows, DOC_END_DATE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["end_date"]),
            "document_type": _bulk_find_col_from_rows(activity_header_rows, DOCUMENT_TYPE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["document_type"]),
            "document_number": _bulk_find_col_from_rows(activity_header_rows, DOCUMENT_NUMBER_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["document_number"]),
            "usage": _bulk_find_col_from_rows(activity_header_rows, USAGE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["usage"]),
            "unit": _bulk_find_col_from_rows(activity_header_rows, ACTIVITY_DATA_UNIT_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["unit"]),
            "data_source": _bulk_find_col_from_rows(activity_header_rows, DATA_SOURCE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["data_source"]),
            "data_source_other": _bulk_find_col_from_rows(activity_header_rows, DATA_SOURCE_OTHER_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["data_source_other"]),
            "calculate_transportation_emissions": _bulk_find_col_from_rows(activity_header_rows, CALCULATE_TRANSPORTATION_EMISSIONS_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["calculate_transportation_emissions"]),
            "supplier_name": _bulk_find_col_from_rows(activity_header_rows, SUPPLIER_NAME_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["supplier_name"]),
            "transport_origin": _bulk_find_col_from_rows(activity_header_rows, TRANSPORT_ORIGIN_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["transport_origin"]),
            "transport_destination": _bulk_find_col_from_rows(activity_header_rows, TRANSPORT_DESTINATION_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["transport_destination"]),
            "target_product": _bulk_find_col_from_rows(activity_header_rows, PRODUCT_LINK_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["target_product"]),
            "comment": _bulk_find_col_from_rows(activity_header_rows, COMMENT_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["comment"]),
            "material_group": _bulk_find_col_from_rows(activity_header_rows, MATERIAL_GROUP_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["material_group"]),
            "net_weight": _bulk_find_col_from_rows(activity_header_rows, NET_WEIGHT_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["net_weight"]),
            "gross_weight": _bulk_find_col_from_rows(activity_header_rows, GROSS_WEIGHT_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["gross_weight"]),
            "weight_unit": _bulk_find_col_from_rows(activity_header_rows, WEIGHT_UNIT_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["weight_unit"]),
        }
        raw_cols = {
            "raw_name": _bulk_find_col_from_rows(raw_header_rows, RAW_MATERIAL_NAME_ALIASES, _RAW_VISIBLE_DEFAULT_COLS["raw_name"]),
            "raw_code": _bulk_find_col_from_rows(raw_header_rows, RAW_MATERIAL_CODE_ALIASES, _RAW_VISIBLE_DEFAULT_COLS["raw_code"]),
            "description": _bulk_find_col_from_rows(raw_header_rows, RAW_MATERIAL_DESC_ALIASES, _RAW_VISIBLE_DEFAULT_COLS["description"]),
        }
        activity_header_rows = _ensure_bulk_visible_header_row(activity_header_rows, activity_cols, _ACTIVITY_VISIBLE_HEADERS)
        raw_header_rows = _ensure_bulk_visible_header_row(raw_header_rows, raw_cols, _RAW_VISIBLE_HEADERS)
        dropdown_cache = DropdownValuesCache.from_workbook(wb, required=True)
        document_type_value = _document_type_for_template(wb, dropdown_cache)
        transport_calculation_yes_value = _transport_calculation_yes_for_template(wb, dropdown_cache)
        missing_formula_cols = [col for col in M2_ACTIVITY_HELPER_FORMULA_COLS if col not in activity_formula_templates]
        if missing_formula_cols:
            missing_letters = ", ".join(_xlsx_index_to_col_letter(col) for col in missing_formula_cols)
            raise ValueError(f"Raw Material Bulk Template 第 3 列缺少 AB～AI 公式：{missing_letters}")
        activity_number_formats = _xlsx_row3_number_formats(raw_material_template_path, ACTIVITY_SHEET_NAME)
        raw_number_formats = _xlsx_row3_number_formats(raw_material_template_path, RAW_MATERIAL_SHEET_NAME)
        dropdown_maps = {
            "activity_unit": dropdown_cache.input_to_display_map("activity_data_unit"),
            "weight_unit": dropdown_cache.input_to_display_map("weight_unit"),
            "data_source": dropdown_cache.input_to_display_map("data_source"),
            "country_area": dropdown_cache.input_to_display_map("country_area"),
            "document_type": dropdown_cache.input_to_display_map("document_type"),
            "calculate_transportation_emissions": dropdown_cache.input_to_display_map("calculate_transportation_emissions"),
            "_cache_summary": dropdown_cache.summary(),
        }
        return activity_header_rows, raw_header_rows, activity_cols, raw_cols, document_type_value, transport_calculation_yes_value, activity_formula_templates, activity_number_formats, raw_number_formats, dropdown_maps
    finally:
        wb.close()

def _set_row_value(row: list[Any], col_idx: int | None, value: Any) -> None:
    if col_idx:
        while len(row) < int(col_idx):
            row.append("")
        row[int(col_idx) - 1] = value


def _find_header_index(headers: list[Any], wanted: str | None, aliases: list[str] | None = None) -> int | None:
    candidates = []
    if wanted:
        candidates.append(wanted)
    if aliases:
        candidates.extend(aliases)
    keys = {_normalize_template_header(x) for x in candidates if str(x or "").strip()}
    for idx, header in enumerate(headers):
        if _normalize_template_header(header) in keys:
            return idx
    return None


def _read_module2a_header_indices(headers: list[Any], mapping: dict[str, str | None]) -> dict[str, int | None]:
    material_idx = _find_header_index(headers, mapping.get("material_col", "Material"), ["Material"])
    parent_idx = _find_header_index(headers, mapping.get("parent_col", "Parent Node"), ["Parent Node"])
    component_idx = _find_header_index(headers, mapping.get("component_col", "Component"), ["Component"])
    qty_idx = _find_header_index(headers, mapping.get("qty_col", "CS03 Qty"), ["CS03 Qty", "Qty", "Usage"])
    unit_idx = _find_header_index(headers, mapping.get("unit_col", "CS03 UoM"), ["CS03 UoM", "UoM", "Unit"])
    if material_idx is None and parent_idx is None:
        raise ValueError("標準BOM表總用量缺少成品料號欄位：Material 或 Parent Node")
    missing = []
    if component_idx is None:
        missing.append(str(mapping.get("component_col") or "Component"))
    if qty_idx is None:
        missing.append(str(mapping.get("qty_col") or "CS03 Qty"))
    if unit_idx is None:
        missing.append(str(mapping.get("unit_col") or "CS03 UoM"))
    if missing:
        raise ValueError("標準BOM表總用量缺少必要欄位：" + ", ".join(missing))
    return {
        "target": material_idx if material_idx is not None else parent_idx,
        "component": component_idx,
        "qty": qty_idx,
        "unit": unit_idx,
        "description": _find_header_index(headers, mapping.get("description_col", "Component Description"), ["Component Description", "Description"]),
        "material_group": _find_header_index(headers, mapping.get("material_group_col", "Material group"), ["Material group", "Material Group"]),
        "valid_from": _find_header_index(headers, mapping.get("valid_from_col", "BOM Valid From"), ["BOM Valid From", "Valid From"]),
        "net_weight": _find_header_index(headers, mapping.get("net_weight_col", "Net weight"), ["Net weight", "Net Weight"]),
        "gross_weight": _find_header_index(headers, mapping.get("gross_weight_col", "Gross weight"), ["Gross weight", "Gross Weight"]),
        "weight_uom": _find_header_index(headers, mapping.get("weight_uom_col", "Weight UoM"), ["Weight UoM", "Weight UOM"]),
    }


def _row_get(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    return row[idx]



def _fast_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.upper() in {"", "NAN", "NONE"} else text


def _fast_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        text = str(value).strip().replace(",", "")
        if not text or text.upper() in {"NAN", "NONE"}:
            return 0.0
        return float(text)
    except Exception:
        return 0.0


def _fast_optional_number(value: Any) -> float | str:
    """Return a real number when present, otherwise a blank cell."""
    if value is None or value == "":
        return ""
    text = str(value).strip()
    if not text or text.upper() in {"NAN", "NONE"}:
        return ""
    try:
        return float(text.replace(",", ""))
    except Exception:
        return ""


def _fast_date_iso(value: Any) -> str:
    if isinstance(value, date):
        return value.isoformat()
    text = _fast_text(value)
    if len(text) >= 4 and text[:4].isdigit():
        return f"{int(text[:4]):04d}-01-01" if len(text) == 4 else text[:10]
    return date(datetime.now().year, 1, 1).isoformat()


def _fast_year_bounds(value: Any) -> tuple[date, date]:
    if isinstance(value, date):
        year = value.year
    else:
        text = _fast_text(value)
        year = int(text[:4]) if len(text) >= 4 and text[:4].isdigit() else datetime.now().year
    return date(year, 1, 1), date(year, 12, 31)



def _xlsx_col_to_index(cell_ref: str) -> int:
    letters = re.sub(r"[^A-Z]", "", str(cell_ref or "").upper())
    value = 0
    for ch in letters:
        value = value * 26 + (ord(ch) - ord("A") + 1)
    return max(value - 1, 0)


def _xlsx_index_to_col_letter(col_idx: int) -> str:
    value = int(col_idx)
    out = ""
    while value > 0:
        value, rem = divmod(value - 1, 26)
        out = chr(65 + rem) + out
    return out or "A"


def _xlsx_sheet_paths_by_name(path: str | Path) -> dict[str, str]:
    ns_main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    ns_rel = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    ns_pkg_rel = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    with zipfile.ZipFile(path) as zf:
        workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
        rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rels = {}
        for rel in rel_root.findall(f"{ns_pkg_rel}Relationship"):
            rid = rel.attrib.get("Id")
            target = rel.attrib.get("Target", "")
            if rid:
                rels[rid] = target.lstrip("/") if target.startswith("/xl/") else ("xl/" + target.lstrip("/") if not target.startswith("xl/") else target)
        sheet_paths = {}
        sheets_el = workbook_root.find(f"{ns_main}sheets")
        if sheets_el is not None:
            for sheet in sheets_el.findall(f"{ns_main}sheet"):
                name = sheet.attrib.get("name", "")
                rid = sheet.attrib.get(f"{ns_rel}id")
                if name and rid in rels:
                    sheet_paths[name] = rels[rid]
        return sheet_paths


def _iter_xlsx_sheet_rows_fast(path: str | Path, sheet_xml_path: str):
    """Yield worksheet rows from XLSX XML without openpyxl cell objects.

    Supports inline strings, shared-string indices and numeric values. This is
    substantially faster and lower-memory for Module 2A's large flat output.
    """
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    shared: list[str] = []
    with zipfile.ZipFile(path) as zf:
        if "xl/sharedStrings.xml" in zf.namelist():
            for _event, si in ET.iterparse(zf.open("xl/sharedStrings.xml"), events=("end",)):
                if si.tag == f"{ns}si":
                    texts = [t.text or "" for t in si.iter(f"{ns}t")]
                    shared.append("".join(texts))
                    si.clear()
        context = ET.iterparse(zf.open(sheet_xml_path), events=("end",))
        for _event, row_el in context:
            if row_el.tag != f"{ns}row":
                continue
            values: dict[int, Any] = {}
            max_idx = -1
            for c in row_el.findall(f"{ns}c"):
                idx = _xlsx_col_to_index(c.attrib.get("r", ""))
                max_idx = max(max_idx, idx)
                cell_type = c.attrib.get("t", "")
                value: Any = ""
                if cell_type == "inlineStr":
                    texts = [t.text or "" for t in c.iter(f"{ns}t")]
                    value = "".join(texts)
                else:
                    v = c.find(f"{ns}v")
                    if v is not None and v.text is not None:
                        raw = v.text
                        if cell_type == "s":
                            try:
                                value = shared[int(raw)]
                            except Exception:
                                value = raw
                        else:
                            value = raw
                values[idx] = value
            row = [values.get(i, "") for i in range(max_idx + 1)] if max_idx >= 0 else []
            row_el.clear()
            yield tuple(row)


def _stream_module2b_rows_to_site_csv(
    standard_total_usage_path: str | Path,
    step1_output_path: str | Path,
    output_dir: Path,
    working_hour_rollup_path: str | Path | None = None,
    mapping: dict[str, str | None] | None = None,
    progress_callback=None,
) -> tuple[dict[str, Path], dict[str, int], dict[str, Any]]:
    """Stream Module 2A total-usage workbook into per-site CSV spool files.

    This is the memory-safe core of Module 2B. It does not create a pandas
    DataFrame for the 2A workbook, does not call pd.concat(), and never creates
    site_df copies. Rows are transformed and immediately written to small disk
    spools grouped by Production Site.
    """
    path = Path(standard_total_usage_path)
    if not path.exists():
        raise FileNotFoundError(f"找不到 Module 2A 標準BOM表總用量檔案：{path}")
    m = _resolve_mapping(mapping)
    annual_qty_map, annual_qty_source_summary = _read_step1_annual_quantity_map(step1_output_path)
    site_map, step1_summary = _read_step1_product_master_maps(step1_output_path)
    if working_hour_rollup_path and Path(working_hour_rollup_path).exists():
        product_eligibility_map, product_exclusion_reasons, product_filter_summary = (
            _read_working_hour_rollup_m2b_product_eligibility_map(working_hour_rollup_path)
        )
        product_filter_summary["m2b_product_filter_fallback_used"] = False
    else:
        product_eligibility_map, product_exclusion_reasons, product_filter_summary = (
            _read_step1_m2b_product_eligibility_map(step1_output_path)
        )
        product_filter_summary["m2b_product_filter_fallback_used"] = True
        product_filter_summary["m2b_product_filter_fallback_reason"] = (
            "M2A working_hour_rollup unavailable; fallback to Module 1A direct annual working hour."
        )

    spool_dir = output_dir / f"m2b_spool_{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    spool_dir.mkdir(parents=True, exist_ok=True)
    csv_paths: dict[str, Path] = {}
    csv_files: dict[str, Any] = {}
    csv_writers: dict[str, Any] = {}
    site_counts: dict[str, int] = defaultdict(int)
    missing_annual_targets: set[str] = set()
    matched_annual_rows = 0
    missing_annual_rows = 0
    zero_usage_rows_excluded = 0
    zero_annual_usage_rows_excluded = 0
    blank_material_group_rows_excluded = 0
    blank_material_group_raw_materials: set[str] = set()
    blank_material_group_targets: set[str] = set()
    m2b_product_filter_rows_excluded = 0
    m2b_product_filter_excluded_targets: set[str] = set()
    m2b_product_filter_reason_counts: dict[str, int] = defaultdict(int)
    rows_read = 0
    valid_rows = 0
    source_sheets: list[str] = []
    used_columns: dict[str, str] = {}
    csv_headers = [
        "target_product", "raw_material", "usage", "unit", "description", "material_group",
        "valid_from", "net_weight", "gross_weight", "weight_uom", "transport_destination",
    ]

    def writer_for(site: str):
        safe_site = _sanitize_filename_part(site)
        if site not in csv_writers:
            csv_path = spool_dir / f"{safe_site}.csv"
            fh = open(csv_path, "w", newline="", encoding="utf-8-sig")
            writer = csv.DictWriter(fh, fieldnames=csv_headers)
            writer.writeheader()
            csv_paths[site] = csv_path
            csv_files[site] = fh
            csv_writers[site] = writer
        return csv_writers[site]

    try:
        sheet_paths = _xlsx_sheet_paths_by_name(path)
        candidate_sheets = [name for name in sheet_paths if str(name or "").strip().startswith(STANDARD_BOM_TOTAL_USAGE_BASE_SHEET_NAME)]
        if not candidate_sheets:
            raise ValueError("標準BOM表總用量檔案中找不到可讀取的『標準BOM表總用量』分頁，請先完成 Module 2A。")
        for sheet_name in candidate_sheets:
            source_sheets.append(sheet_name)
            row_iter = _iter_xlsx_sheet_rows_fast(path, sheet_paths[sheet_name])
            try:
                headers = list(next(row_iter))
            except StopIteration:
                continue
            idx = _read_module2a_header_indices(headers, m)
            if idx.get("material_group") is None:
                raise ValueError(
                    "標準BOM表總用量缺少 Material group 欄位；M2B 無法判斷真正原物料，請確認標準 BOM 欄位後重新執行 M2A。"
                )
            if not used_columns:
                def header_name(i):
                    return str(headers[i] or "") if i is not None and i < len(headers) else ""
                used_columns = {
                    "target_product_col": header_name(idx["target"]),
                    "component_col": header_name(idx["component"]),
                    "qty_col": header_name(idx["qty"]),
                    "unit_col": header_name(idx["unit"]),
                    "description_col": header_name(idx["description"]),
                    "material_group_col": header_name(idx["material_group"]),
                    "valid_from_col": header_name(idx["valid_from"]),
                    "net_weight_col": header_name(idx["net_weight"]),
                    "gross_weight_col": header_name(idx["gross_weight"]),
                    "weight_uom_col": header_name(idx["weight_uom"]),
                }
            for row in row_iter:
                rows_read += 1
                target_product = _fast_text(_row_get(row, idx["target"]))
                raw_material = _fast_text(_row_get(row, idx["component"]))
                if not target_product or not raw_material:
                    continue
                material_group = _fast_text(_row_get(row, idx["material_group"]))
                if not material_group:
                    blank_material_group_rows_excluded += 1
                    blank_material_group_raw_materials.add(raw_material)
                    blank_material_group_targets.add(target_product)
                    continue
                usage_per_pc = _fast_number(_row_get(row, idx["qty"]))
                if usage_per_pc == 0:
                    zero_usage_rows_excluded += 1
                    continue
                target_key = _normalize_material_key(target_product)
                if product_eligibility_map and target_key in product_eligibility_map and not product_eligibility_map[target_key]:
                    m2b_product_filter_rows_excluded += 1
                    m2b_product_filter_excluded_targets.add(target_key)
                    reason = product_exclusion_reasons.get(target_key, "excluded_by_m2b_product_filter")
                    m2b_product_filter_reason_counts[reason] += 1
                    continue
                annual_qty = annual_qty_map.get(target_key)
                if annual_qty is None:
                    annual_qty = 1.0
                    missing_annual_rows += 1
                    missing_annual_targets.add(target_key)
                else:
                    matched_annual_rows += 1
                annual_usage = float(usage_per_pc) * float(annual_qty)
                if annual_usage == 0:
                    zero_annual_usage_rows_excluded += 1
                    continue
                site = str(site_map.get(target_key, "Unassigned") or "Unassigned").strip() or "Unassigned"
                writer_for(site).writerow({
                    "target_product": target_product,
                    "raw_material": raw_material,
                    "usage": annual_usage,
                    "unit": _fast_text(_row_get(row, idx["unit"])),
                    "description": _fast_text(_row_get(row, idx["description"])),
                    "material_group": material_group,
                    "valid_from": _fast_date_iso(_row_get(row, idx["valid_from"])),
                    "net_weight": _fast_number(_row_get(row, idx["net_weight"])) if idx.get("net_weight") is not None else "",
                    "gross_weight": _fast_number(_row_get(row, idx["gross_weight"])) if idx.get("gross_weight") is not None else "",
                    "weight_uom": _fast_text(_row_get(row, idx["weight_uom"])),
                    "transport_destination": site,
                })
                site_counts[site] += 1
                valid_rows += 1
                if progress_callback and rows_read % 20000 == 0:
                    progress_callback(
                        step="Streaming Module 2A total usage",
                        processed=rows_read,
                        total=0,
                        progress=10,
                        blank_material_group_rows_excluded=int(blank_material_group_rows_excluded),
                    )
    finally:
        for fh in csv_files.values():
            fh.close()

    summary: dict[str, Any] = {
        "module2a_total_usage_source_filename": path.name,
        "module2a_total_usage_source_sheets": source_sheets,
        "module2a_total_usage_rows_read": int(rows_read),
        "module2a_total_usage_valid_rows": int(valid_rows),
        "module2a_total_usage_rule": "Read Module 2A total usage in streaming mode as per-PC final raw-material usage; BOM is not re-expanded and Altitem probability is not re-applied.",
        "used_columns": used_columns,
        "annual_quantity_applied": True,
        "annual_quantity_matched_rows": int(matched_annual_rows),
        "annual_quantity_missing_rows": int(missing_annual_rows),
        "annual_quantity_missing_targets": sorted(missing_annual_targets)[:50],
        "usage_per_pc_column_added": False,
        "annual_finished_product_qty_column_added": False,
        "zero_usage_rows_excluded": int(zero_usage_rows_excluded),
        "zero_annual_usage_rows_excluded": int(zero_annual_usage_rows_excluded),
        "blank_material_group_rows_excluded": int(blank_material_group_rows_excluded),
        "blank_material_group_raw_materials": sorted(blank_material_group_raw_materials)[:50],
        "blank_material_group_raw_material_count": int(len(blank_material_group_raw_materials)),
        "blank_material_group_targets": sorted(blank_material_group_targets)[:50],
        "blank_material_group_target_count": int(len(blank_material_group_targets)),
        "blank_material_group_rule": "M2B excludes final BOM rows whose Standard BOM Material group is blank; M2A expanded BOM remains unchanged.",
        "m2b_product_filter_rows_excluded": int(m2b_product_filter_rows_excluded),
        "m2b_product_filter_excluded_targets": sorted(m2b_product_filter_excluded_targets)[:50],
        "m2b_product_filter_excluded_target_count": int(len(m2b_product_filter_excluded_targets)),
        "m2b_product_filter_reason_counts": dict(m2b_product_filter_reason_counts),
        "m2b_large_dataset_mode": True,
        "m2b_spool_dir": str(spool_dir),
    }
    summary.update(annual_qty_source_summary)
    summary.update(step1_summary)
    summary.update(product_filter_summary)
    return csv_paths, dict(site_counts), summary


def _write_raw_material_bulk_from_site_csv_streaming(
    csv_path: str | Path,
    raw_material_template_path: str | Path,
    output_path: str | Path,
    progress_callback=None,
    processed_offset: int = 0,
    total_rows: int | None = None,
    current_site: str = "",
    current_site_rows: int = 0,
) -> Dict[str, Any]:
    """Write a lightweight Raw Material Bulk workbook from a per-site CSV spool.

    Uses xlsxwriter constant_memory when available because it is much faster for
    200k+ flat rows than openpyxl write_only. Falls back to openpyxl if the
    package is not installed.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    activity_header_rows, raw_header_rows, activity_cols, raw_cols, document_type_value, transport_calculation_yes_value, activity_formula_templates, activity_number_formats, raw_number_formats, dropdown_maps = _template_headers_for_lightweight_bulk(raw_material_template_path)
    activity_headers = activity_header_rows[0]
    raw_headers = raw_header_rows[0]

    raw_seen: set[str] = set()
    raw_descriptions: dict[str, str] = {}
    written = 0
    actual_data_rows = 0
    total_rows = int(total_rows or current_site_rows or 0)
    progress_every = max(1000, min(10000, max(int(current_site_rows or 0) // 20, 1)))

    def build_activity_row(row_data: dict[str, Any]) -> tuple[list[Any], str]:
        start_date_value, end_date_value = _fast_year_bounds(row_data.get("valid_from"))
        usage_value = _fast_number(row_data.get("usage"))
        raw_material = _fast_text(row_data.get("raw_material"))
        activity_row = ["" for _ in activity_headers]
        _set_row_value(activity_row, activity_cols["raw_name"], raw_material)
        _set_row_value(activity_row, activity_cols["raw_code"], raw_material)
        _set_row_value(activity_row, activity_cols["start_date"], start_date_value)
        _set_row_value(activity_row, activity_cols["end_date"], end_date_value)
        _set_row_value(activity_row, activity_cols["document_type"], row_data.get("document_type", "") or document_type_value)
        _set_row_value(activity_row, activity_cols["document_number"], "")
        _set_row_value(activity_row, activity_cols["usage"], usage_value)
        _set_row_value(activity_row, activity_cols["unit"], _localized_dropdown_display(row_data.get("unit", ""), dropdown_maps.get("activity_unit", {})))
        _set_row_value(activity_row, activity_cols["data_source"], _localized_dropdown_display("SAP", dropdown_maps.get("data_source", {}), "SAP"))
        _set_row_value(activity_row, activity_cols["data_source_other"], "")
        _set_row_value(activity_row, activity_cols["calculate_transportation_emissions"], transport_calculation_yes_value)
        _set_row_value(activity_row, activity_cols["supplier_name"], "")
        _set_row_value(activity_row, activity_cols["transport_origin"], "")
        _set_row_value(activity_row, activity_cols["transport_destination"], row_data.get("transport_destination", current_site))
        _set_row_value(activity_row, activity_cols["target_product"], row_data.get("target_product", ""))
        _set_row_value(activity_row, activity_cols["comment"], "")
        _set_row_value(activity_row, activity_cols["material_group"], row_data.get("material_group", ""))
        _set_row_value(activity_row, activity_cols.get("net_weight"), _fast_optional_number(row_data.get("net_weight", "")))
        _set_row_value(activity_row, activity_cols.get("gross_weight"), _fast_optional_number(row_data.get("gross_weight", "")))
        _set_row_value(activity_row, activity_cols.get("weight_unit"), _localized_dropdown_display(row_data.get("weight_uom", ""), dropdown_maps.get("weight_unit", {})))
        return activity_row, raw_material

    if xlsxwriter is not None:
        workbook = xlsxwriter.Workbook(str(output_path), {"constant_memory": True})
        activity_ws = workbook.add_worksheet(ACTIVITY_SHEET_NAME[:31])
        raw_ws = workbook.add_worksheet(RAW_MATERIAL_SHEET_NAME[:31])
        activity_ws.set_column(27, 34, None, None, {"hidden": True})
        activity_formats = _xlsxwriter_column_formats(workbook, activity_number_formats)
        raw_formats = _xlsxwriter_column_formats(workbook, raw_number_formats)
        for row_idx, header_row in enumerate(activity_header_rows):
            for col, value in enumerate(header_row):
                if value not in (None, ""):
                    activity_ws.write(row_idx, col, value)
        for row_idx, header_row in enumerate(raw_header_rows):
            for col, value in enumerate(header_row):
                if value not in (None, ""):
                    raw_ws.write(row_idx, col, value)
        excel_row_idx = DATA_START_ROW - 1  # zero-based row 2
        with open(csv_path, "r", newline="", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh)
            for row_data in reader:
                activity_row, raw_material = build_activity_row(row_data)
                for col, value in enumerate(activity_row):
                    _write_xlsxwriter_typed_cell(activity_ws, excel_row_idx, col, value, activity_number_formats, activity_formats)
                _write_m2_helper_formulas_xlsxwriter(activity_ws, excel_row_idx, activity_formula_templates, activity_formats)
                if raw_material and raw_material not in raw_seen:
                    raw_seen.add(raw_material)
                    raw_descriptions[raw_material] = row_data.get("description", "") or ""
                written += 1
                actual_data_rows += 1
                excel_row_idx += 1
                if progress_callback and (written == 1 or written % progress_every == 0 or written == current_site_rows):
                    progress_callback(
                        step=f"Writing Raw Material Bulk: {current_site}",
                        processed=processed_offset + written,
                        total=total_rows,
                        progress=min(95, 45 + int((processed_offset + written) / max(total_rows, 1) * 45)),
                        current_site=current_site,
                        current_site_rows=current_site_rows,
                        current_site_written=written,
                    )
        raw_excel_row_idx = DATA_START_ROW - 1
        for raw_material in sorted(raw_seen):
            raw_row = ["" for _ in raw_headers]
            _set_row_value(raw_row, raw_cols["raw_name"], raw_material)
            _set_row_value(raw_row, raw_cols["raw_code"], raw_material)
            _set_row_value(raw_row, raw_cols["description"], raw_descriptions.get(raw_material, ""))
            for col, value in enumerate(raw_row):
                _write_xlsxwriter_typed_cell(raw_ws, raw_excel_row_idx, col, value, raw_number_formats, raw_formats)
            raw_excel_row_idx += 1
        workbook.close()
        writer_name = "xlsxwriter_constant_memory"
    else:
        wb = Workbook(write_only=True)
        activity_ws = wb.create_sheet(ACTIVITY_SHEET_NAME)
        raw_ws = wb.create_sheet(RAW_MATERIAL_SHEET_NAME)
        for col_idx in M2_ACTIVITY_HELPER_FORMULA_COLS:
            activity_ws.column_dimensions[_xlsx_index_to_col_letter(col_idx)].hidden = True
        for header_row in activity_header_rows:
            activity_ws.append(header_row)
        for header_row in raw_header_rows:
            raw_ws.append(header_row)
        with open(csv_path, "r", newline="", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh)
            for row_data in reader:
                activity_row, raw_material = build_activity_row(row_data)
                activity_ws.append(_apply_m2_helper_formulas_to_row(activity_row, activity_formula_templates, DATA_START_ROW + written))
                if raw_material and raw_material not in raw_seen:
                    raw_seen.add(raw_material)
                    raw_descriptions[raw_material] = row_data.get("description", "") or ""
                written += 1
                actual_data_rows += 1
                if progress_callback and (written == 1 or written % progress_every == 0 or written == current_site_rows):
                    progress_callback(
                        step=f"Writing Raw Material Bulk: {current_site}",
                        processed=processed_offset + written,
                        total=total_rows,
                        progress=min(95, 45 + int((processed_offset + written) / max(total_rows, 1) * 45)),
                        current_site=current_site,
                        current_site_rows=current_site_rows,
                        current_site_written=written,
                    )
        for raw_material in sorted(raw_seen):
            raw_row = ["" for _ in raw_headers]
            _set_row_value(raw_row, raw_cols["raw_name"], raw_material)
            _set_row_value(raw_row, raw_cols["raw_code"], raw_material)
            _set_row_value(raw_row, raw_cols["description"], raw_descriptions.get(raw_material, ""))
            raw_ws.append(raw_row)
        wb.save(output_path)
        writer_name = "openpyxl_write_only"

    return {
        "output_filename": output_path.name,
        "activity_template_columns": activity_cols,
        "raw_material_template_columns": raw_cols,
        "activity_rows": int(actual_data_rows),
        "raw_materials": int(len(raw_seen)),
        "zero_usage_rows_excluded": 0,
        "supplier_name_options": 0,
        "site_tbc_supplier_count": 0,
        "supplier_status": "Deferred to Module 2C",
        "m2b_writer": f"large_dataset_{writer_name}_template_number_formats_plus_AB_AI_formulas",
        "calculate_transportation_emissions_value": transport_calculation_yes_value,
        "dropdown_values_cache": dropdown_maps.get("_cache_summary", {}),
        "helper_formula_columns": ["AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI"],
        "helper_formula_rows": int(actual_data_rows),
    }


def generate_raw_material_bulk_from_standard_total_usage_zip(
    standard_total_usage_path: str | Path,
    raw_material_template_path: str | Path,
    output_dir: str | Path,
    token: str,
    step1_output_path: str | Path,
    working_hour_rollup_path: str | Path | None = None,
    mapping: dict[str, str | None] | None = None,
    progress_callback=None,
) -> Dict[str, Any]:
    """Generate Module 2B Raw Material Bulk ZIP by site from Module 2A output.

    Large Dataset Mode rules:
    - Stream Module 2A rows directly into per-site disk spools; no full DataFrame,
      no pd.concat(), and no site_df.copy().
    - Usage = 2A per-PC final raw-material usage × Module 1 Step1 annual quantity.
    - Transportation Destination = Module 1 Step1 Production Site.
    - Output is split by Production Site.
    - M2B keeps template sheet names/header columns only. It intentionally does
      not load/copy template formatting, formulas, dropdowns, validations or
      hidden helper sheets. M2C/M3 can apply formal template formatting later.
    - Supplier mapping is not applied in Module 2B; it is reserved for Module 2C.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    if progress_callback:
        progress_callback(step="Preparing Module 2B Large Dataset Mode", processed=0, total=0, progress=6)

    csv_paths, site_counts, summary = _stream_module2b_rows_to_site_csv(
        standard_total_usage_path=standard_total_usage_path,
        step1_output_path=step1_output_path,
        output_dir=output_dir,
        working_hour_rollup_path=working_hour_rollup_path,
        mapping=mapping,
        progress_callback=progress_callback,
    )

    site_values = sorted(site_counts.keys()) if site_counts else ["Unassigned"]
    generated_files: list[dict[str, Any]] = []
    zip_filename = f"raw_material_activity_data_bulk_by_site_{token}.zip"
    zip_path = output_dir / zip_filename
    processed_rows = 0
    total_rows = int(sum(site_counts.values()))

    if progress_callback:
        progress_callback(step="Writing Raw Material Bulk files", processed=0, total=total_rows, progress=45)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for idx, site in enumerate(site_values, start=1):
            site_row_count = int(site_counts.get(site, 0))
            safe_site = _sanitize_filename_part(site)
            file_path = output_dir / f"raw_material_activity_data_bulk_{safe_site}_{token}.xlsx"
            if progress_callback:
                progress_callback(
                    step=f"Writing Raw Material Bulk: {site}",
                    processed=processed_rows,
                    total=total_rows,
                    progress=min(95, 45 + int((idx - 1) / max(len(site_values), 1) * 45)),
                    current_site=str(site),
                    current_site_rows=site_row_count,
                    current_site_written=0,
                )
            write_summary = _write_raw_material_bulk_from_site_csv_streaming(
                csv_path=csv_paths.get(site, ""),
                raw_material_template_path=raw_material_template_path,
                output_path=file_path,
                progress_callback=progress_callback,
                processed_offset=processed_rows,
                total_rows=total_rows,
                current_site=str(site),
                current_site_rows=site_row_count,
            )
            processed_rows += site_row_count
            zf.write(file_path, arcname=file_path.name)
            generated_files.append({
                "production_site": site,
                "filename": file_path.name,
                "activity_rows": int(write_summary.get("activity_rows", 0)),
                "raw_materials": int(write_summary.get("raw_materials", 0)),
            })
            try:
                file_path.unlink(missing_ok=True)
            except Exception:
                pass

    for csv_path in csv_paths.values():
        try:
            Path(csv_path).unlink(missing_ok=True)
        except Exception:
            pass
    try:
        Path(summary.get("m2b_spool_dir", "")).rmdir()
    except Exception:
        pass

    unassigned_rows = int(site_counts.get("Unassigned", 0))
    raw_material_total = 0
    # raw_materials are counted per generated workbook; keep the previous summary
    # field conservative because cross-site unique counting would require another
    # global in-memory set. Per-site counts are available in production_site_files.
    result = dict(summary)
    result.update({
        "output_filename": zip_filename,
        "download_url": f"/download/{zip_filename}",
        "split_by_production_site": True,
        "production_site_files": generated_files,
        "production_site_count": int(len(site_values)),
        "unassigned_rows": int(unassigned_rows),
        "activity_rows": int(total_rows),
        "raw_materials": int(raw_material_total),
        "supplier_status": "Deferred to Module 2C",
        "supplier_upload_files": 0,
        "supplier_bulk_generated": False,
        "module2b_rule": "2A total usage -> Module 1 annual quantity/site mapping -> lightweight Raw Material Bulk ZIP by site; supplier mapping is not applied in Module 2B.",
        "module2b_template_policy": "M2B/M2C lightweight files preserve row-3 number formats and AB~AI formulas; the full dropdown/validation package is restored by M3A final template application.",
    })
    if progress_callback:
        progress_callback(step="Completed", processed=total_rows, total=total_rows, progress=100)
    return result


def _read_raw_material_bulk_workbook_as_exploded(raw_bulk_path: str | Path) -> tuple[pd.DataFrame, Dict[str, Any]]:
    """Read an existing Raw Material Bulk workbook back into the internal exploded schema.

    Module 2C uses this to apply supplier mapping to Module 2B output without
    re-expanding BOM and without requiring the Raw Material Bulk template again.
    The workbook itself is reused as the output template so dropdowns/hidden
    columns/formatting are preserved.
    """
    path = Path(raw_bulk_path)
    if not path.exists():
        raise FileNotFoundError(f"找不到 Raw Material Bulk 檔案：{path}")

    wb = load_workbook(path, data_only=True)
    if ACTIVITY_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"{path.name} 找不到 raw material bulk 分頁：{ACTIVITY_SHEET_NAME}")
    if RAW_MATERIAL_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"{path.name} 找不到 raw material bulk 分頁：{RAW_MATERIAL_SHEET_NAME}")

    activity_ws = wb[ACTIVITY_SHEET_NAME]
    raw_ws = wb[RAW_MATERIAL_SHEET_NAME]

    activity_cols = {
        "raw_name": _find_template_column(activity_ws, RAW_MATERIAL_NAME_ALIASES, 1),
        "raw_code": _find_template_column(activity_ws, RAW_MATERIAL_CODE_ALIASES, 2),
        "start_date": _find_template_column(activity_ws, DOC_START_DATE_ALIASES, 3),
        "usage": _find_template_column(activity_ws, USAGE_ALIASES, 7),
        "unit": _find_template_column(activity_ws, ACTIVITY_DATA_UNIT_ALIASES, 8),
        "supplier_name": _find_template_column(activity_ws, SUPPLIER_NAME_ALIASES, 14),
        "transport_origin": _find_template_column(activity_ws, TRANSPORT_ORIGIN_ALIASES, 16),
        "transport_destination": _find_template_column(activity_ws, TRANSPORT_DESTINATION_ALIASES, 17),
        "target_product": _find_template_column(activity_ws, PRODUCT_LINK_ALIASES, 18),
        "material_group": _find_template_column(activity_ws, MATERIAL_GROUP_ALIASES, 20),
        "net_weight": _find_template_optional_column(activity_ws, NET_WEIGHT_ALIASES),
        "gross_weight": _find_template_optional_column(activity_ws, GROSS_WEIGHT_ALIASES),
        "weight_unit": _find_template_optional_column(activity_ws, WEIGHT_UNIT_ALIASES),
    }
    raw_cols = {
        "raw_name": _find_template_column(raw_ws, RAW_MATERIAL_NAME_ALIASES, 1),
        "raw_code": _find_template_column(raw_ws, RAW_MATERIAL_CODE_ALIASES, 2),
        "description": _find_template_column(raw_ws, RAW_MATERIAL_DESC_ALIASES, 6),
    }

    description_map: dict[str, str] = {}
    for row_idx in range(DATA_START_ROW, raw_ws.max_row + 1):
        code = _safe_text(raw_ws.cell(row_idx, raw_cols["raw_code"]).value) or _safe_text(raw_ws.cell(row_idx, raw_cols["raw_name"]).value)
        if not code:
            continue
        description_map[_normalize_material_key(code)] = _safe_text(raw_ws.cell(row_idx, raw_cols["description"]).value)

    rows: list[dict[str, Any]] = []
    for row_idx in range(DATA_START_ROW, activity_ws.max_row + 1):
        raw_material = _safe_text(activity_ws.cell(row_idx, activity_cols["raw_code"]).value) or _safe_text(activity_ws.cell(row_idx, activity_cols["raw_name"]).value)
        target_product = _safe_text(activity_ws.cell(row_idx, activity_cols["target_product"]).value)
        unit = _safe_text(activity_ws.cell(row_idx, activity_cols["unit"]).value)
        usage = _safe_number(activity_ws.cell(row_idx, activity_cols["usage"]).value)
        if not raw_material and not target_product and not unit and usage == 0:
            continue
        if not raw_material or not target_product:
            continue
        raw_key = _normalize_material_key(raw_material)
        rows.append({
            "target_product": target_product,
            "source_material": target_product,
            "raw_material": raw_material,
            "usage": usage,
            "unit": unit,
            "description": description_map.get(raw_key, ""),
            "material_group": _safe_text(activity_ws.cell(row_idx, activity_cols["material_group"]).value),
            "valid_from": _date_from_value(activity_ws.cell(row_idx, activity_cols["start_date"]).value),
            "level": 0,
            "transport_destination": _safe_text(activity_ws.cell(row_idx, activity_cols["transport_destination"]).value),
            "transport_origin": _safe_text(activity_ws.cell(row_idx, activity_cols["transport_origin"]).value),
            "calculate_transportation_emissions": _safe_text(activity_ws.cell(row_idx, activity_cols["calculate_transportation_emissions"]).value),
            "supplier_name": _safe_text(activity_ws.cell(row_idx, activity_cols["supplier_name"]).value),
            "net_weight": _safe_number(activity_ws.cell(row_idx, activity_cols["net_weight"]).value) if activity_cols.get("net_weight") else "",
            "gross_weight": _safe_number(activity_ws.cell(row_idx, activity_cols["gross_weight"]).value) if activity_cols.get("gross_weight") else "",
            "weight_uom": _safe_text(activity_ws.cell(row_idx, activity_cols["weight_unit"]).value) if activity_cols.get("weight_unit") else "",
        })

    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=[
            "target_product", "source_material", "raw_material", "usage", "unit", "description",
            "material_group", "valid_from", "level", "transport_destination", "transport_origin",
            "calculate_transportation_emissions", "supplier_name", "net_weight", "gross_weight", "weight_uom",
        ])
    return df, {
        "input_filename": path.name,
        "activity_rows_read": int(len(df)),
        "raw_materials_read": int(df["raw_material"].nunique()) if not df.empty else 0,
        "template_columns": activity_cols,
    }




# =========================================================
# Module 2C · Large Dataset Supplier Mapping
# Module 2B lightweight Raw Material Bulk ZIP -> Supplier-mapped Bulk ZIP.
# This path intentionally avoids pandas DataFrame materialization and avoids
# load_workbook(output_template) for 200k+ row workbooks.
# =========================================================

def _safe_cell_text(value: Any) -> str:
    return _fast_text(value)


def _read_raw_material_descriptions_streaming(path: str | Path) -> dict[str, str]:
    """Read Raw Material sheet into a compact raw_material -> description map."""
    descriptions: dict[str, str] = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if RAW_MATERIAL_SHEET_NAME not in wb.sheetnames:
            return descriptions
        ws = wb[RAW_MATERIAL_SHEET_NAME]
        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()))
        def find(headers: list[Any], aliases: list[str], fallback: int) -> int:
            keys = {_normalize_template_header(a) for a in aliases}
            for i, h in enumerate(headers, start=1):
                if _normalize_template_header(h) in keys:
                    return i
            return fallback
        raw_code_col = find(headers, RAW_MATERIAL_CODE_ALIASES, 2)
        raw_name_col = find(headers, RAW_MATERIAL_NAME_ALIASES, 1)
        desc_col = find(headers, RAW_MATERIAL_DESC_ALIASES, 6)
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            raw = _safe_cell_text(_row_get(row, raw_code_col - 1)) or _safe_cell_text(_row_get(row, raw_name_col - 1))
            if raw and raw not in descriptions:
                desc = _safe_cell_text(_row_get(row, desc_col - 1))
                descriptions[raw] = desc
                descriptions[_normalize_material_key(raw)] = desc
    finally:
        wb.close()
    return descriptions


def _activity_layout_from_bulk_workbook(path: str | Path) -> tuple[list[list[Any]], list[list[Any]], dict[str, int], dict[str, int], str]:
    """Read two-row headers and visible-column indexes from a Raw Material Bulk workbook.

    M2C consumes M2B lightweight workbooks.  Those files must be interpreted by
    the visible Bulk Template columns, not hidden helper columns.  If an older
    file is missing row 2, fall back to the official visible positions.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if ACTIVITY_SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"找不到 raw material bulk 分頁：{ACTIVITY_SHEET_NAME}")
        if RAW_MATERIAL_SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"找不到 raw material bulk 分頁：{RAW_MATERIAL_SHEET_NAME}")
        activity_ws = wb[ACTIVITY_SHEET_NAME]
        raw_ws = wb[RAW_MATERIAL_SHEET_NAME]
        activity_header_rows = _read_bulk_header_rows(activity_ws)
        raw_header_rows = _read_bulk_header_rows(raw_ws)
        activity_cols = {
            "raw_name": _bulk_find_col_from_rows(activity_header_rows, RAW_MATERIAL_NAME_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["raw_name"]),
            "raw_code": _bulk_find_col_from_rows(activity_header_rows, RAW_MATERIAL_CODE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["raw_code"]),
            "start_date": _bulk_find_col_from_rows(activity_header_rows, DOC_START_DATE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["start_date"]),
            "end_date": _bulk_find_col_from_rows(activity_header_rows, DOC_END_DATE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["end_date"]),
            "document_type": _bulk_find_col_from_rows(activity_header_rows, DOCUMENT_TYPE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["document_type"]),
            "document_number": _bulk_find_col_from_rows(activity_header_rows, DOCUMENT_NUMBER_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["document_number"]),
            "usage": _bulk_find_col_from_rows(activity_header_rows, USAGE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["usage"]),
            "unit": _bulk_find_col_from_rows(activity_header_rows, ACTIVITY_DATA_UNIT_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["unit"]),
            "data_source": _bulk_find_col_from_rows(activity_header_rows, DATA_SOURCE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["data_source"]),
            "data_source_other": _bulk_find_col_from_rows(activity_header_rows, DATA_SOURCE_OTHER_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["data_source_other"]),
            "calculate_transportation_emissions": _bulk_find_col_from_rows(activity_header_rows, CALCULATE_TRANSPORTATION_EMISSIONS_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["calculate_transportation_emissions"]),
            "supplier_name": _bulk_find_col_from_rows(activity_header_rows, SUPPLIER_NAME_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["supplier_name"]),
            "transport_origin": _bulk_find_col_from_rows(activity_header_rows, TRANSPORT_ORIGIN_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["transport_origin"]),
            "transport_destination": _bulk_find_col_from_rows(activity_header_rows, TRANSPORT_DESTINATION_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["transport_destination"]),
            "target_product": _bulk_find_col_from_rows(activity_header_rows, PRODUCT_LINK_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["target_product"]),
            "comment": _bulk_find_col_from_rows(activity_header_rows, COMMENT_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["comment"]),
            "material_group": _bulk_find_col_from_rows(activity_header_rows, MATERIAL_GROUP_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["material_group"]),
            "net_weight": _bulk_find_col_from_rows(activity_header_rows, NET_WEIGHT_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["net_weight"]),
            "gross_weight": _bulk_find_col_from_rows(activity_header_rows, GROSS_WEIGHT_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["gross_weight"]),
            "weight_unit": _bulk_find_col_from_rows(activity_header_rows, WEIGHT_UNIT_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["weight_unit"]),
        }
        raw_cols = {
            "raw_name": _bulk_find_col_from_rows(raw_header_rows, RAW_MATERIAL_NAME_ALIASES, _RAW_VISIBLE_DEFAULT_COLS["raw_name"]),
            "raw_code": _bulk_find_col_from_rows(raw_header_rows, RAW_MATERIAL_CODE_ALIASES, _RAW_VISIBLE_DEFAULT_COLS["raw_code"]),
            "description": _bulk_find_col_from_rows(raw_header_rows, RAW_MATERIAL_DESC_ALIASES, _RAW_VISIBLE_DEFAULT_COLS["description"]),
        }
        activity_header_rows = _ensure_bulk_visible_header_row(activity_header_rows, activity_cols, _ACTIVITY_VISIBLE_HEADERS)
        raw_header_rows = _ensure_bulk_visible_header_row(raw_header_rows, raw_cols, _RAW_VISIBLE_HEADERS)
        document_type_value = _document_type_for_template(wb)
        return activity_header_rows, raw_header_rows, activity_cols, raw_cols, document_type_value
    finally:
        wb.close()


def _iter_activity_rows_streaming(path: str | Path, activity_cols: dict[str, int], description_map: dict[str, str]):
    """Yield normalized activity rows from a Raw Material Bulk workbook one by one."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if ACTIVITY_SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"找不到 raw material bulk 分頁：{ACTIVITY_SHEET_NAME}")
        ws = wb[ACTIVITY_SHEET_NAME]
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            raw_material = _safe_cell_text(_row_get(row, activity_cols["raw_code"] - 1)) or _safe_cell_text(_row_get(row, activity_cols["raw_name"] - 1))
            target_product = _safe_cell_text(_row_get(row, activity_cols["target_product"] - 1))
            unit = _safe_cell_text(_row_get(row, activity_cols["unit"] - 1))
            usage = _fast_number(_row_get(row, activity_cols["usage"] - 1))
            if not raw_material and not target_product and not unit and usage == 0:
                continue
            if not raw_material or not target_product:
                continue
            valid_from = _row_get(row, activity_cols["start_date"] - 1)
            raw_key = _normalize_material_key(raw_material)
            yield {
                "target_product": target_product,
                "source_material": target_product,
                "raw_material": raw_material,
                "usage": usage,
                "unit": unit,
                "description": description_map.get(raw_material) or description_map.get(raw_key) or "",
                "material_group": _safe_cell_text(_row_get(row, activity_cols["material_group"] - 1)),
                "valid_from": valid_from if isinstance(valid_from, date) else _fast_date_iso(valid_from),
                "level": 0,
                "transport_destination": _safe_cell_text(_row_get(row, activity_cols["transport_destination"] - 1)),
                "transport_origin": _safe_cell_text(_row_get(row, activity_cols["transport_origin"] - 1)),
                "document_type": _safe_cell_text(_row_get(row, activity_cols["document_type"] - 1)),
                "data_source": _safe_cell_text(_row_get(row, activity_cols["data_source"] - 1)),
                "calculate_transportation_emissions": _safe_cell_text(_row_get(row, activity_cols["calculate_transportation_emissions"] - 1)),
                "supplier_name": _safe_cell_text(_row_get(row, activity_cols["supplier_name"] - 1)),
                "net_weight": _fast_number(_row_get(row, activity_cols["net_weight"] - 1)) if activity_cols.get("net_weight") else "",
                "gross_weight": _fast_number(_row_get(row, activity_cols["gross_weight"] - 1)) if activity_cols.get("gross_weight") else "",
                "weight_uom": _safe_cell_text(_row_get(row, activity_cols["weight_unit"] - 1)) if activity_cols.get("weight_unit") else "",
            }
    finally:
        wb.close()


def _supplier_rows_for_activity_row(row: dict[str, Any], supplier_map: dict[str, list[dict[str, str]]], supplier_options: list[str], tbc_supplier_map: dict[str, dict[str, str]]):
    """Yield one or more supplier-mapped rows for a normalized activity row."""
    raw_key = _normalize_material_key(row.get("raw_material"))
    destination = _safe_text(row.get("transport_destination"))
    suppliers = supplier_map.get(raw_key) or []
    if not suppliers:
        uploaded_tbc = _select_uploaded_tbc_supplier_for_destination(tbc_supplier_map, destination)
        uploaded_address = ""
        uploaded_country = ""
        uploaded_plant = ""
        if uploaded_tbc:
            uploaded_address = uploaded_tbc.get("supplier_address", "") or uploaded_tbc.get("transport_origin", "")
            uploaded_country = uploaded_tbc.get("country_area", "")
            uploaded_plant = uploaded_tbc.get("plant", "")
        out = dict(row)
        out["transport_destination"] = destination
        out["supplier_name"] = _raw_material_supplier_display_name(destination, "TBC", "TBC")
        out["transport_origin"] = uploaded_address or "TBC"
        out["supplier_code"] = "TBC"
        out["supplier_master_name"] = "TBC"
        out["supplier_country_area"] = _country_area_for_unit_name(destination, uploaded_plant, uploaded_country)
        out["supplier_address"] = uploaded_address or "TBC"
        yield out, False, True, True
        return
    supplier_count = len(suppliers)
    usage_per_supplier = _split_usage_evenly_by_supplier_count(row.get("usage"), supplier_count)
    for info in suppliers:
        out = dict(row)
        out["usage"] = usage_per_supplier
        out["transport_destination"] = destination
        supplier_address = info.get("supplier_address", "") or info.get("transport_origin", "")
        supplier_code = info.get("supplier_code", "") or info.get("vendor_code", "")
        if _normalize_vendor_code(supplier_code) == "TBC":
            uploaded_tbc = _select_uploaded_tbc_supplier_for_destination(tbc_supplier_map, destination)
            if uploaded_tbc:
                supplier_address = supplier_address or uploaded_tbc.get("supplier_address", "") or uploaded_tbc.get("transport_origin", "")
        supplier_master_name = info.get("supplier_master_name", "") or _supplier_name_from_option(info.get("supplier_name", ""))
        supplier_name = _raw_material_supplier_display_name(destination, supplier_code, supplier_master_name)
        if not supplier_name:
            supplier_name = _select_supplier_name_option(supplier_options, destination, supplier_code)
        out["transport_origin"] = supplier_address
        out["supplier_code"] = supplier_code
        out["supplier_master_name"] = info.get("supplier_master_name", "") or _supplier_name_from_option(supplier_name)
        if _normalize_vendor_code(supplier_code) == "TBC":
            out["supplier_master_name"] = "TBC"
            out["supplier_country_area"] = _country_area_for_unit_name(destination, info.get("plant", ""), info.get("country_area", ""))
            supplier_name = _raw_material_supplier_display_name(destination, "TBC", "TBC")
        else:
            out["supplier_country_area"] = info.get("country_area", "")
        out["supplier_address"] = supplier_address
        out["supplier_name"] = supplier_name
        yield out, True, bool(supplier_name), False


def _build_supplier_activity_row(row_data: dict[str, Any], activity_headers: list[str], activity_cols: dict[str, int], document_type_value: str) -> tuple[list[Any], str]:
    start_date_value, end_date_value = _fast_year_bounds(row_data.get("valid_from"))
    raw_material = _fast_text(row_data.get("raw_material"))
    activity_row = ["" for _ in activity_headers]
    _set_row_value(activity_row, activity_cols["raw_name"], raw_material)
    _set_row_value(activity_row, activity_cols["raw_code"], raw_material)
    _set_row_value(activity_row, activity_cols["start_date"], start_date_value)
    _set_row_value(activity_row, activity_cols["end_date"], end_date_value)
    _set_row_value(
        activity_row, activity_cols["document_type"],
        row_data.get("document_type", "") or document_type_value,
    )
    _set_row_value(activity_row, activity_cols["document_number"], "")
    _set_row_value(activity_row, activity_cols["usage"], _fast_number(row_data.get("usage")))
    _set_row_value(activity_row, activity_cols["unit"], row_data.get("unit", ""))
    _set_row_value(activity_row, activity_cols["data_source"], row_data.get("data_source", "") or "SAP")
    _set_row_value(activity_row, activity_cols["data_source_other"], "")
    _set_row_value(
        activity_row, activity_cols["calculate_transportation_emissions"],
        row_data.get("calculate_transportation_emissions", "") or "Yes",
    )
    _set_row_value(activity_row, activity_cols["supplier_name"], row_data.get("supplier_name", ""))
    _set_row_value(activity_row, activity_cols["transport_origin"], row_data.get("transport_origin", ""))
    _set_row_value(activity_row, activity_cols["transport_destination"], row_data.get("transport_destination", ""))
    _set_row_value(activity_row, activity_cols["target_product"], row_data.get("target_product", ""))
    _set_row_value(activity_row, activity_cols["comment"], "")
    _set_row_value(activity_row, activity_cols["material_group"], row_data.get("material_group", ""))
    _set_row_value(activity_row, activity_cols.get("net_weight"), row_data.get("net_weight", ""))
    _set_row_value(activity_row, activity_cols.get("gross_weight"), row_data.get("gross_weight", ""))
    _set_row_value(activity_row, activity_cols.get("weight_unit"), row_data.get("weight_uom", ""))
    return activity_row, raw_material



_XLSX_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_XLSX_MAIN_TAG = "{" + _XLSX_MAIN_NS + "}"
_XLSX_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_XLSX_PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def _m2c_sheet_paths_from_zip(zf: zipfile.ZipFile) -> dict[str, str]:
    workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
    rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rels: dict[str, str] = {}
    for rel in rel_root.findall(f"{{{_XLSX_PKG_REL_NS}}}Relationship"):
        rid = rel.attrib.get("Id")
        target = rel.attrib.get("Target", "")
        if not rid:
            continue
        if target.startswith("/xl/"):
            rels[rid] = target.lstrip("/")
        elif target.startswith("xl/"):
            rels[rid] = target
        else:
            rels[rid] = "xl/" + target.lstrip("/")
    result: dict[str, str] = {}
    sheets = workbook_root.find(f"{{{_XLSX_MAIN_NS}}}sheets")
    if sheets is not None:
        for sheet in sheets.findall(f"{{{_XLSX_MAIN_NS}}}sheet"):
            name = sheet.attrib.get("name", "")
            rid = sheet.attrib.get(f"{{{_XLSX_REL_NS}}}id")
            if name and rid in rels:
                result[name] = rels[rid]
    return result


def _m2c_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    values: list[str] = []
    for _event, si in ET.iterparse(zf.open("xl/sharedStrings.xml"), events=("end",)):
        if si.tag == f"{_XLSX_MAIN_TAG}si":
            values.append("".join(t.text or "" for t in si.iter(f"{_XLSX_MAIN_TAG}t")))
            si.clear()
    return values


def _m2c_cell_value(cell: ET.Element, shared: list[str]) -> Any:
    cell_type = cell.attrib.get("t", "")
    if cell_type == "inlineStr":
        return "".join(t.text or "" for t in cell.iter(f"{_XLSX_MAIN_TAG}t"))
    value_el = cell.find(f"{_XLSX_MAIN_TAG}v")
    if value_el is None or value_el.text is None:
        return ""
    raw = value_el.text
    if cell_type == "s":
        try:
            return shared[int(raw)]
        except Exception:
            return raw
    if cell_type == "b":
        return raw == "1"
    return raw


def _m2c_row_cell_map(row_el: ET.Element, shared: list[str], needed_cols: set[int] | None = None) -> dict[int, Any]:
    values: dict[int, Any] = {}
    for cell in row_el.findall(f"{_XLSX_MAIN_TAG}c"):
        col_idx = _xlsx_col_to_index(cell.attrib.get("r", "")) + 1
        if needed_cols is not None and col_idx not in needed_cols:
            continue
        values[col_idx] = _m2c_cell_value(cell, shared)
    return values


def _m2c_clean_xml_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    # XML 1.0 disallows most ASCII control characters.  Supplier masters may
    # contain them, so strip only invalid controls while preserving tabs/newlines.
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)


def _m2c_xml_escape(value: Any) -> bytes:
    text = _m2c_clean_xml_text(value)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .encode("utf-8")
    )


def _m2c_cell_xml(
    row_idx: int,
    col_idx: int,
    value: Any,
    style_id: str | int | None = None,
    formula: str | None = None,
) -> bytes:
    ref = f"{_xlsx_index_to_col_letter(col_idx)}{int(row_idx)}".encode("ascii")
    style_attr = b""
    if style_id is not None and str(style_id).strip() not in {"", "0"}:
        style_attr = f' s="{str(style_id).strip()}"'.encode("ascii")
    if formula:
        formula_text = formula[1:] if formula.startswith("=") else formula
        return (
            b'<c r="' + ref + b'"' + style_attr + b' t="str"><f>'
            + _m2c_xml_escape(formula_text)
            + b'</f><v></v></c>'
        )
    if value is None or value == "":
        return b""
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        serial = (value - date(1899, 12, 30)).days
        return b'<c r="' + ref + b'"' + style_attr + b'><v>' + str(serial).encode("ascii") + b'</v></c>'
    if isinstance(value, bool):
        return b'<c r="' + ref + b'"' + style_attr + f' t="b"><v>{1 if value else 0}</v></c>'.encode("ascii")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            number = float(value)
            if math.isfinite(number):
                raw = str(int(number)) if number.is_integer() else repr(number)
                return b'<c r="' + ref + b'"' + style_attr + b'><v>' + raw.encode("ascii") + b'</v></c>'
        except Exception:
            pass
    return (
        b'<c r="' + ref + b'"' + style_attr
        + b' t="inlineStr"><is><t xml:space="preserve">'
        + _m2c_xml_escape(value)
        + b'</t></is></c>'
    )


def _m2c_row_xml(
    row_idx: int,
    col_values: list[tuple[int, Any]],
    style_by_col: dict[int, str] | None = None,
    formula_by_col: dict[int, str] | None = None,
    width: int = 1,
) -> bytes:
    cells: list[bytes] = []
    styles = style_by_col or {}
    formulas = formula_by_col or {}
    for col_idx, value in col_values:
        formula = formulas.get(int(col_idx))
        cell = _m2c_cell_xml(row_idx, int(col_idx), value, styles.get(int(col_idx), "0"), formula=formula)
        if cell:
            cells.append(cell)
    return (
        f'<row r="{int(row_idx)}" spans="1:{max(1, int(width))}">'.encode("ascii")
        + b"".join(cells)
        + b"</row>"
    )


def _m2c_header_rows_and_styles(
    row_maps: dict[int, dict[int, Any]],
    row_styles: dict[int, dict[int, str]],
    width: int,
) -> tuple[list[list[Any]], dict[int, dict[int, str]]]:
    rows = [
        [row_maps.get(row_idx, {}).get(col_idx, "") for col_idx in range(1, width + 1)]
        for row_idx in (1, 2)
    ]
    return rows, row_styles


def _m2c_excel_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _fast_text(value)
    if not text:
        return None
    try:
        serial = float(text)
        if math.isfinite(serial) and 1 <= serial <= 2958465:
            return (datetime(1899, 12, 30) + timedelta(days=serial)).date()
    except Exception:
        pass
    if len(text) >= 4 and text[:4].isdigit():
        try:
            return date(int(text[:4]), 1, 1)
        except Exception:
            pass
    return None


def _m2c_year_bounds(value: Any) -> tuple[date, date]:
    parsed = _m2c_excel_date(value)
    year = parsed.year if parsed else datetime.now().year
    return date(year, 1, 1), date(year, 12, 31)


def _m2c_supplier_options_from_zip(
    zf: zipfile.ZipFile,
    sheet_paths: dict[str, str],
    shared: list[str],
) -> list[str]:
    path = sheet_paths.get("Dropdown Values")
    if not path:
        return []
    alias_keys = {
        _normalize_template_header(x)
        for x in ("Supplier Name (optional)", "Supplier Name(optional)", "Supplier Name", "supplier_name", "供應商名稱")
    }
    target_col: int | None = None
    values: list[str] = []
    seen: set[str] = set()
    for _event, row_el in ET.iterparse(zf.open(path), events=("end",)):
        if row_el.tag != f"{_XLSX_MAIN_TAG}row":
            continue
        row_idx = int(row_el.attrib.get("r", "0") or 0)
        row_values = _m2c_row_cell_map(row_el, shared)
        if row_idx == 1:
            for col_idx, value in row_values.items():
                if _normalize_template_header(value) in alias_keys:
                    target_col = col_idx
                    break
        elif target_col:
            value = _safe_text(row_values.get(target_col, ""))
            if value and value not in seen:
                seen.add(value)
                values.append(value)
        row_el.clear()
    return values


def _m2c_raw_context_and_descriptions(
    zf: zipfile.ZipFile,
    sheet_path: str,
    shared: list[str],
) -> tuple[list[list[Any]], dict[int, dict[int, str]], dict[int, str], dict[str, int], int, dict[str, str]]:
    row_maps: dict[int, dict[int, Any]] = {1: {}, 2: {}}
    row_styles: dict[int, dict[int, str]] = {1: {}, 2: {}}
    row3_styles: dict[int, str] = {}
    width = max(_RAW_VISIBLE_DEFAULT_COLS.values())
    raw_cols: dict[str, int] | None = None
    descriptions: dict[str, str] = {}
    for _event, row_el in ET.iterparse(zf.open(sheet_path), events=("end",)):
        if row_el.tag != f"{_XLSX_MAIN_TAG}row":
            continue
        row_idx = int(row_el.attrib.get("r", "0") or 0)
        if row_idx in (1, 2):
            for cell in row_el.findall(f"{_XLSX_MAIN_TAG}c"):
                col_idx = _xlsx_col_to_index(cell.attrib.get("r", "")) + 1
                width = max(width, col_idx)
                row_maps[row_idx][col_idx] = _m2c_cell_value(cell, shared)
                row_styles[row_idx][col_idx] = cell.attrib.get("s", "0")
            if row_idx == 2:
                headers, _ = _m2c_header_rows_and_styles(row_maps, row_styles, width)
                raw_cols = {
                    "raw_name": _bulk_find_col_from_rows(headers, RAW_MATERIAL_NAME_ALIASES, _RAW_VISIBLE_DEFAULT_COLS["raw_name"]),
                    "raw_code": _bulk_find_col_from_rows(headers, RAW_MATERIAL_CODE_ALIASES, _RAW_VISIBLE_DEFAULT_COLS["raw_code"]),
                    "description": _bulk_find_col_from_rows(headers, RAW_MATERIAL_DESC_ALIASES, _RAW_VISIBLE_DEFAULT_COLS["description"]),
                }
        elif row_idx >= DATA_START_ROW:
            if raw_cols is None:
                headers, _ = _m2c_header_rows_and_styles(row_maps, row_styles, width)
                raw_cols = {
                    "raw_name": _bulk_find_col_from_rows(headers, RAW_MATERIAL_NAME_ALIASES, _RAW_VISIBLE_DEFAULT_COLS["raw_name"]),
                    "raw_code": _bulk_find_col_from_rows(headers, RAW_MATERIAL_CODE_ALIASES, _RAW_VISIBLE_DEFAULT_COLS["raw_code"]),
                    "description": _bulk_find_col_from_rows(headers, RAW_MATERIAL_DESC_ALIASES, _RAW_VISIBLE_DEFAULT_COLS["description"]),
                }
            needed = {int(v) for v in raw_cols.values() if v}
            values = _m2c_row_cell_map(row_el, shared, needed)
            if row_idx == DATA_START_ROW:
                for cell in row_el.findall(f"{_XLSX_MAIN_TAG}c"):
                    col_idx = _xlsx_col_to_index(cell.attrib.get("r", "")) + 1
                    width = max(width, col_idx)
                    row3_styles[col_idx] = cell.attrib.get("s", "0")
            raw = _safe_text(values.get(raw_cols["raw_code"], "")) or _safe_text(values.get(raw_cols["raw_name"], ""))
            if raw and raw not in descriptions:
                desc = _safe_text(values.get(raw_cols["description"], ""))
                descriptions[raw] = desc
                descriptions[_normalize_material_key(raw)] = desc
        row_el.clear()
    headers, header_styles = _m2c_header_rows_and_styles(row_maps, row_styles, width)
    assert raw_cols is not None
    headers = _ensure_bulk_visible_header_row(headers, raw_cols, _RAW_VISIBLE_HEADERS)
    for col_idx in range(1, width + 1):
        row3_styles.setdefault(col_idx, "0")
    return headers, header_styles, row3_styles, raw_cols, width, descriptions


def _m2c_remove_calc_chain_content_types(data: bytes) -> bytes:
    return re.sub(rb'<Override\b[^>]*PartName="/xl/calcChain\.xml"[^>]*/>', b"", data)


def _m2c_remove_calc_chain_rels(data: bytes) -> bytes:
    return re.sub(rb'<Relationship\b[^>]*Type="[^"]*/calcChain"[^>]*/>', b"", data)


def _m2c_force_full_calc(data: bytes) -> bytes:
    attrs = b' calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"'
    if b"<calcPr" not in data:
        return data.replace(b"</workbook>", b"<calcPr" + attrs + b"/></workbook>", 1)
    def repl(match: re.Match[bytes]) -> bytes:
        tag = match.group(0)
        for attr in (b"calcMode", b"fullCalcOnLoad", b"forceFullCalc"):
            tag = re.sub(attr + rb'="[^"]*"', b"", tag)
        if tag.endswith(b"/>"):
            return tag[:-2].rstrip() + attrs + b"/>"
        return tag[:-1].rstrip() + attrs + b">"
    return re.sub(rb"<calcPr\b[^>]*/?>", repl, data, count=1)


def _m2c_write_sheet_part(
    zout: zipfile.ZipFile,
    arcname: str,
    header_rows: list[list[Any]],
    header_styles: dict[int, dict[int, str]],
    data_spool_path: Path,
    data_count: int,
    width: int,
    hide_helper_columns: bool = False,
) -> None:
    max_row = max(2, DATA_START_ROW - 1 + int(data_count))
    max_col = max(1, int(width))
    dimension = f'A1:{_xlsx_index_to_col_letter(max_col)}{max_row}'
    with zout.open(arcname, "w", force_zip64=True) as out:
        out.write(b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>')
        out.write(f'<worksheet xmlns="{_XLSX_MAIN_NS}">'.encode("ascii"))
        out.write(f'<dimension ref="{dimension}"/>'.encode("ascii"))
        out.write(b'<sheetViews><sheetView workbookViewId="0"/></sheetViews>')
        out.write(b'<sheetFormatPr defaultRowHeight="15"/>')
        if hide_helper_columns:
            out.write(b'<cols><col min="28" max="35" width="0" hidden="1" customWidth="1"/></cols>')
        out.write(b'<sheetData>')
        for row_idx in (1, 2):
            values = header_rows[row_idx - 1] if len(header_rows) >= row_idx else []
            col_values = [(col_idx, values[col_idx - 1]) for col_idx in range(1, min(len(values), max_col) + 1) if values[col_idx - 1] not in (None, "")]
            out.write(_m2c_row_xml(row_idx, col_values, header_styles.get(row_idx, {}), width=max_col))
        if data_spool_path.exists():
            with data_spool_path.open("rb") as src:
                shutil.copyfileobj(src, out, length=M2C_COPY_BUFFER_BYTES)
        out.write(b'</sheetData>')
        out.write(b'<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/>')
        out.write(b'</worksheet>')


def _m2c_copy_static_zip_entry(zin: zipfile.ZipFile, zout: zipfile.ZipFile, item: zipfile.ZipInfo) -> None:
    if item.is_dir():
        zout.writestr(item, b"")
        return
    with zin.open(item, "r") as src, zout.open(item, "w", force_zip64=True) as dst:
        shutil.copyfileobj(src, dst, length=M2C_COPY_BUFFER_BYTES)


def _m2c_build_output_package(
    source_file: Path,
    output_path: Path,
    activity_sheet_path: str,
    raw_sheet_path: str,
    activity_headers: list[list[Any]],
    activity_header_styles: dict[int, dict[int, str]],
    activity_spool: Path,
    activity_count: int,
    activity_width: int,
    raw_headers: list[list[Any]],
    raw_header_styles: dict[int, dict[int, str]],
    raw_spool: Path,
    raw_count: int,
    raw_width: int,
) -> None:
    partial = output_path.with_suffix(output_path.suffix + ".partial")
    partial.unlink(missing_ok=True)
    try:
        with zipfile.ZipFile(source_file, "r") as zin, zipfile.ZipFile(
            partial,
            "w",
            compression=zipfile.ZIP_DEFLATED,
            compresslevel=M2C_XLSX_COMPRESSION_LEVEL,
            allowZip64=True,
        ) as zout:
            replaced = {activity_sheet_path, raw_sheet_path, "xl/calcChain.xml"}
            for item in zin.infolist():
                name = item.filename
                if name in replaced:
                    continue
                if name in {"[Content_Types].xml", "xl/_rels/workbook.xml.rels", "xl/workbook.xml"}:
                    data = zin.read(name)
                    if name == "[Content_Types].xml":
                        data = _m2c_remove_calc_chain_content_types(data)
                    elif name == "xl/_rels/workbook.xml.rels":
                        data = _m2c_remove_calc_chain_rels(data)
                    else:
                        data = _m2c_force_full_calc(data)
                    zout.writestr(item, data)
                else:
                    _m2c_copy_static_zip_entry(zin, zout, item)
            _m2c_write_sheet_part(
                zout, activity_sheet_path, activity_headers, activity_header_styles,
                activity_spool, activity_count, activity_width, hide_helper_columns=True,
            )
            _m2c_write_sheet_part(
                zout, raw_sheet_path, raw_headers, raw_header_styles,
                raw_spool, raw_count, raw_width, hide_helper_columns=False,
            )
        os.replace(partial, output_path)
    except Exception:
        partial.unlink(missing_ok=True)
        raise




def _m2c_compile_formula_template(formula: str) -> str:
    """Compile row-3 references once; per-row expansion becomes a cheap replace."""
    if not formula:
        return formula
    pattern = r'(?<![A-Za-z0-9_])(\$?[A-Z]{1,3})3(?![0-9])'
    return re.sub(pattern, lambda m: m.group(1) + "{ROW}", formula)

def _m2c_supplier_variants(
    base: dict[str, Any],
    suppliers: list[dict[str, str]],
    supplier_options: list[str],
    tbc_supplier_map: dict[str, dict[str, str]],
):
    destination = _safe_text(base.get("transport_destination"))
    if not suppliers:
        uploaded_tbc = _select_uploaded_tbc_supplier_for_destination(tbc_supplier_map, destination)
        uploaded_address = ""
        uploaded_country = ""
        uploaded_plant = ""
        if uploaded_tbc:
            uploaded_address = uploaded_tbc.get("supplier_address", "") or uploaded_tbc.get("transport_origin", "")
            uploaded_country = uploaded_tbc.get("country_area", "")
            uploaded_plant = uploaded_tbc.get("plant", "")
        yield {
            "usage": base.get("usage", 0),
            "supplier_name": _raw_material_supplier_display_name(destination, "TBC", "TBC"),
            "transport_origin": uploaded_address or "TBC",
            "supplier_code": "TBC",
            "supplier_master_name": "TBC",
            "supplier_country_area": _country_area_for_unit_name(destination, uploaded_plant, uploaded_country),
            "supplier_address": uploaded_address or "TBC",
        }, False, True, True
        return
    usage_per_supplier = _split_usage_evenly_by_supplier_count(base.get("usage"), len(suppliers))
    for info in suppliers:
        supplier_address = info.get("supplier_address", "") or info.get("transport_origin", "")
        supplier_code = info.get("supplier_code", "") or info.get("vendor_code", "")
        normalized_code = _normalize_vendor_code(supplier_code)
        if normalized_code == "TBC":
            uploaded_tbc = _select_uploaded_tbc_supplier_for_destination(tbc_supplier_map, destination)
            if uploaded_tbc:
                supplier_address = supplier_address or uploaded_tbc.get("supplier_address", "") or uploaded_tbc.get("transport_origin", "")
        supplier_master_name = info.get("supplier_master_name", "") or _supplier_name_from_option(info.get("supplier_name", ""))
        supplier_name = _raw_material_supplier_display_name(destination, supplier_code, supplier_master_name)
        if not supplier_name:
            supplier_name = _select_supplier_name_option(supplier_options, destination, supplier_code)
        if normalized_code == "TBC":
            supplier_master_name = "TBC"
            supplier_country_area = _country_area_for_unit_name(destination, info.get("plant", ""), info.get("country_area", ""))
            supplier_name = _raw_material_supplier_display_name(destination, "TBC", "TBC")
        else:
            supplier_country_area = info.get("country_area", "")
        yield {
            "usage": usage_per_supplier,
            "supplier_name": supplier_name,
            "transport_origin": supplier_address,
            "supplier_code": supplier_code,
            "supplier_master_name": supplier_master_name,
            "supplier_country_area": supplier_country_area,
            "supplier_address": supplier_address,
        }, True, bool(supplier_name), False


def _write_supplier_mapped_bulk_streaming(
    source_file: str | Path,
    output_path: str | Path,
    supplier_map: dict[str, list[dict[str, str]]],
    tbc_supplier_map: dict[str, dict[str, str]] | None = None,
    progress_callback=None,
    current_file: str = "",
) -> tuple[Dict[str, Any], set[tuple[str, str, str, str, str]]]:
    """Complete Streaming OpenXML M2C pipeline.

    The source XLSX is read directly from worksheet XML, supplier mapping is
    applied row-by-row, row XML is spooled to disk, and a new XLSX package is
    assembled by replacing only the Activity Data and Raw Material worksheet
    parts.  No pandas DataFrame, openpyxl workbook or XlsxWriter cell loop is
    used for the large activity dataset.
    """
    started = time.perf_counter()
    source_file = Path(source_file)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tbc_supplier_map = tbc_supplier_map or {}

    input_rows = 0
    actual_mapped_rows = 0
    matched_source_rows = 0
    supplier_expanded_rows = 0
    supplier_name_matched = 0
    supplier_name_missing = 0
    tbc_fallback_rows = 0
    supplier_usage_split_source_rows = 0
    supplier_usage_split_output_rows = 0
    supplier_usage_total_before_split = 0.0
    supplier_usage_total_after_split = 0.0
    raw_seen: set[str] = set()
    raw_descriptions: dict[str, str] = {}
    supplier_unique: set[tuple[str, str, str, str, str]] = set()
    progress_every = 5000

    with tempfile.TemporaryDirectory(prefix="dip_m2c_openxml_") as tmp:
        tmpdir = Path(tmp)
        activity_spool = tmpdir / "activity_rows.xml"
        raw_spool = tmpdir / "raw_rows.xml"

        with zipfile.ZipFile(source_file, "r") as zf:
            sheet_paths = _m2c_sheet_paths_from_zip(zf)
            activity_sheet_path = sheet_paths.get(ACTIVITY_SHEET_NAME)
            raw_sheet_path = sheet_paths.get(RAW_MATERIAL_SHEET_NAME)
            if not activity_sheet_path:
                raise ValueError(f"找不到 raw material bulk 分頁：{ACTIVITY_SHEET_NAME}")
            if not raw_sheet_path:
                raise ValueError(f"找不到 raw material bulk 分頁：{RAW_MATERIAL_SHEET_NAME}")
            shared = _m2c_shared_strings(zf)
            supplier_options = _m2c_supplier_options_from_zip(zf, sheet_paths, shared)
            (
                raw_header_rows,
                raw_header_styles,
                raw_row3_styles,
                raw_cols,
                raw_width,
                description_map,
            ) = _m2c_raw_context_and_descriptions(zf, raw_sheet_path, shared)

            row_maps: dict[int, dict[int, Any]] = {1: {}, 2: {}}
            header_styles: dict[int, dict[int, str]] = {1: {}, 2: {}}
            activity_row3_styles: dict[int, str] = {}
            activity_formula_templates: dict[int, str] = {}
            activity_formula_row_templates: dict[int, str] = {}
            activity_cols: dict[str, int] | None = None
            activity_header_rows: list[list[Any]] = []
            activity_width = max(max(M2_ACTIVITY_HELPER_FORMULA_COLS), max(_ACTIVITY_VISIBLE_DEFAULT_COLS.values()))
            needed_cols: set[int] | None = None
            write_plan: list[tuple[int, int]] = []

            with activity_spool.open("wb") as activity_out:
                for _event, row_el in ET.iterparse(zf.open(activity_sheet_path), events=("end",)):
                    if row_el.tag != f"{_XLSX_MAIN_TAG}row":
                        continue
                    row_idx = int(row_el.attrib.get("r", "0") or 0)
                    if row_idx in (1, 2):
                        for cell in row_el.findall(f"{_XLSX_MAIN_TAG}c"):
                            col_idx = _xlsx_col_to_index(cell.attrib.get("r", "")) + 1
                            activity_width = max(activity_width, col_idx)
                            row_maps[row_idx][col_idx] = _m2c_cell_value(cell, shared)
                            header_styles[row_idx][col_idx] = cell.attrib.get("s", "0")
                        if row_idx == 2:
                            activity_header_rows, _ = _m2c_header_rows_and_styles(row_maps, header_styles, activity_width)
                            activity_cols = {
                                "raw_name": _bulk_find_col_from_rows(activity_header_rows, RAW_MATERIAL_NAME_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["raw_name"]),
                                "raw_code": _bulk_find_col_from_rows(activity_header_rows, RAW_MATERIAL_CODE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["raw_code"]),
                                "start_date": _bulk_find_col_from_rows(activity_header_rows, DOC_START_DATE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["start_date"]),
                                "end_date": _bulk_find_col_from_rows(activity_header_rows, DOC_END_DATE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["end_date"]),
                                "document_type": _bulk_find_col_from_rows(activity_header_rows, DOCUMENT_TYPE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["document_type"]),
                                "document_number": _bulk_find_col_from_rows(activity_header_rows, DOCUMENT_NUMBER_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["document_number"]),
                                "usage": _bulk_find_col_from_rows(activity_header_rows, USAGE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["usage"]),
                                "unit": _bulk_find_col_from_rows(activity_header_rows, ACTIVITY_DATA_UNIT_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["unit"]),
                                "data_source": _bulk_find_col_from_rows(activity_header_rows, DATA_SOURCE_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["data_source"]),
                                "data_source_other": _bulk_find_col_from_rows(activity_header_rows, DATA_SOURCE_OTHER_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["data_source_other"]),
                                "calculate_transportation_emissions": _bulk_find_col_from_rows(activity_header_rows, CALCULATE_TRANSPORTATION_EMISSIONS_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["calculate_transportation_emissions"]),
                                "supplier_name": _bulk_find_col_from_rows(activity_header_rows, SUPPLIER_NAME_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["supplier_name"]),
                                "transport_origin": _bulk_find_col_from_rows(activity_header_rows, TRANSPORT_ORIGIN_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["transport_origin"]),
                                "transport_destination": _bulk_find_col_from_rows(activity_header_rows, TRANSPORT_DESTINATION_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["transport_destination"]),
                                "target_product": _bulk_find_col_from_rows(activity_header_rows, PRODUCT_LINK_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["target_product"]),
                                "comment": _bulk_find_col_from_rows(activity_header_rows, COMMENT_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["comment"]),
                                "material_group": _bulk_find_col_from_rows(activity_header_rows, MATERIAL_GROUP_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["material_group"]),
                                "net_weight": _bulk_find_col_from_rows(activity_header_rows, NET_WEIGHT_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["net_weight"]),
                                "gross_weight": _bulk_find_col_from_rows(activity_header_rows, GROSS_WEIGHT_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["gross_weight"]),
                                "weight_unit": _bulk_find_col_from_rows(activity_header_rows, WEIGHT_UNIT_ALIASES, _ACTIVITY_VISIBLE_DEFAULT_COLS["weight_unit"]),
                            }
                            activity_header_rows = _ensure_bulk_visible_header_row(activity_header_rows, activity_cols, _ACTIVITY_VISIBLE_HEADERS)
                            needed_cols = {int(v) for v in activity_cols.values() if v}
                            # Tuple positions generated below. Sort once by real template column.
                            field_keys = [
                                "raw_name", "raw_code", "start_date", "end_date", "document_type", "document_number",
                                "usage", "unit", "data_source", "data_source_other", "calculate_transportation_emissions",
                                "supplier_name", "transport_origin", "transport_destination", "target_product", "comment",
                                "material_group", "net_weight", "gross_weight", "weight_unit",
                            ]
                            write_plan = sorted((int(activity_cols[key]), idx) for idx, key in enumerate(field_keys) if activity_cols.get(key))
                    elif row_idx >= DATA_START_ROW:
                        if activity_cols is None or needed_cols is None:
                            raise ValueError("M2C 無法解析 Activity Data 表頭。")
                        if row_idx == DATA_START_ROW:
                            for cell in row_el.findall(f"{_XLSX_MAIN_TAG}c"):
                                col_idx = _xlsx_col_to_index(cell.attrib.get("r", "")) + 1
                                activity_width = max(activity_width, col_idx)
                                activity_row3_styles[col_idx] = cell.attrib.get("s", "0")
                                if col_idx in M2_ACTIVITY_HELPER_FORMULA_COLS:
                                    formula_el = cell.find(f"{_XLSX_MAIN_TAG}f")
                                    if formula_el is not None and formula_el.text:
                                        activity_formula_templates[col_idx] = "=" + formula_el.text
                            missing = [col for col in M2_ACTIVITY_HELPER_FORMULA_COLS if col not in activity_formula_templates]
                            if missing:
                                missing_letters = ", ".join(_xlsx_index_to_col_letter(col) for col in missing)
                                raise ValueError(f"M2C 輸入的 Raw Material Bulk 第 3 列缺少 AB～AI 公式：{missing_letters}")
                            activity_formula_row_templates = {
                                col_idx: _m2c_compile_formula_template(formula)
                                for col_idx, formula in activity_formula_templates.items()
                            }
                            for col_idx in range(1, activity_width + 1):
                                activity_row3_styles.setdefault(col_idx, "0")

                        values = _m2c_row_cell_map(row_el, shared, needed_cols)
                        raw_material = _safe_text(values.get(activity_cols["raw_code"], "")) or _safe_text(values.get(activity_cols["raw_name"], ""))
                        target_product = _safe_text(values.get(activity_cols["target_product"], ""))
                        unit = _safe_text(values.get(activity_cols["unit"], ""))
                        usage = _fast_number(values.get(activity_cols["usage"], ""))
                        if not raw_material and not target_product and not unit and usage == 0:
                            row_el.clear()
                            continue
                        if not raw_material or not target_product:
                            row_el.clear()
                            continue

                        input_rows += 1
                        raw_key = _normalize_material_key(raw_material)
                        destination = _safe_text(values.get(activity_cols["transport_destination"], ""))
                        base = {
                            "raw_material": raw_material,
                            "target_product": target_product,
                            "usage": usage,
                            "unit": unit,
                            "description": description_map.get(raw_material) or description_map.get(raw_key) or "",
                            "material_group": _safe_text(values.get(activity_cols["material_group"], "")),
                            "valid_from": values.get(activity_cols["start_date"], ""),
                            "transport_destination": destination,
                            "document_type": _safe_text(values.get(activity_cols["document_type"], "")),
                            "data_source": _safe_text(values.get(activity_cols["data_source"], "")),
                            "calculate_transportation_emissions": _safe_text(values.get(activity_cols["calculate_transportation_emissions"], "")),
                            "net_weight": _fast_number(values.get(activity_cols["net_weight"], "")) if activity_cols.get("net_weight") else "",
                            "gross_weight": _fast_number(values.get(activity_cols["gross_weight"], "")) if activity_cols.get("gross_weight") else "",
                            "weight_uom": _safe_text(values.get(activity_cols["weight_unit"], "")) if activity_cols.get("weight_unit") else "",
                        }
                        suppliers = supplier_map.get(raw_key) or []
                        source_supplier_count = len(suppliers)
                        if source_supplier_count > 1:
                            split_usage = _split_usage_evenly_by_supplier_count(usage, source_supplier_count)
                            supplier_usage_split_source_rows += 1
                            supplier_usage_split_output_rows += source_supplier_count
                            supplier_usage_total_before_split += usage
                            supplier_usage_total_after_split += split_usage * source_supplier_count

                        start_date_value, end_date_value = _m2c_year_bounds(base["valid_from"])
                        for variant, matched, name_ok, used_tbc in _m2c_supplier_variants(base, suppliers, supplier_options, tbc_supplier_map):
                            if matched:
                                matched_source_rows += 1
                                supplier_expanded_rows += 1
                            if name_ok:
                                supplier_name_matched += 1
                            else:
                                supplier_name_missing += 1
                            if used_tbc:
                                tbc_fallback_rows += 1

                            field_values = (
                                raw_material,
                                raw_material,
                                start_date_value,
                                end_date_value,
                                base["document_type"] or "Bill of Materials (BOM)",
                                "",
                                _fast_number(variant["usage"]),
                                unit,
                                base["data_source"] or "SAP",
                                "",
                                base["calculate_transportation_emissions"] or "Yes",
                                variant["supplier_name"],
                                variant["transport_origin"],
                                destination,
                                target_product,
                                "",
                                base["material_group"],
                                base["net_weight"],
                                base["gross_weight"],
                                base["weight_uom"],
                            )
                            excel_row = DATA_START_ROW + actual_mapped_rows
                            col_values = [(col_idx, field_values[field_idx]) for col_idx, field_idx in write_plan]
                            row_text = str(excel_row)
                            formula_map = {
                                col_idx: activity_formula_row_templates[col_idx].replace("{ROW}", row_text)
                                for col_idx in M2_ACTIVITY_HELPER_FORMULA_COLS
                            }
                            # Visible columns are before AB; helper cells can be appended without sorting.
                            col_values.extend((col_idx, "") for col_idx in M2_ACTIVITY_HELPER_FORMULA_COLS)
                            activity_out.write(_m2c_row_xml(
                                excel_row,
                                col_values,
                                activity_row3_styles,
                                formula_map,
                                width=activity_width,
                            ))

                            if raw_material not in raw_seen:
                                raw_seen.add(raw_material)
                                raw_descriptions[raw_material] = base["description"]
                            supplier_code = _normalize_vendor_code(variant["supplier_code"])
                            if supplier_code:
                                supplier_unique.add((
                                    _supplier_bulk_name_only(
                                        variant["supplier_master_name"],
                                        variant["supplier_name"],
                                        supplier_code,
                                        destination,
                                    ),
                                    supplier_code,
                                    _safe_text(variant["supplier_country_area"]),
                                    _safe_text(variant["supplier_address"]) or _safe_text(variant["transport_origin"]),
                                    destination,
                                ))
                            actual_mapped_rows += 1

                        if progress_callback and (input_rows == 1 or input_rows % progress_every == 0):
                            progress_callback(
                                step=f"Streaming OpenXML Supplier mapping: {current_file or source_file.name}",
                                processed=actual_mapped_rows,
                                total=0,
                                progress=30,
                                current_file=current_file or source_file.name,
                                input_rows=input_rows,
                                output_rows=actual_mapped_rows,
                            )
                    row_el.clear()

        if not activity_header_rows or activity_cols is None:
            raise ValueError("M2C 無法解析 Activity Data 表頭。")

        with raw_spool.open("wb") as raw_out:
            for offset, raw_material in enumerate(sorted(raw_seen)):
                excel_row = DATA_START_ROW + offset
                field_values = {
                    int(raw_cols["raw_name"]): raw_material,
                    int(raw_cols["raw_code"]): raw_material,
                    int(raw_cols["description"]): raw_descriptions.get(raw_material, ""),
                }
                raw_out.write(_m2c_row_xml(
                    excel_row,
                    sorted(field_values.items()),
                    raw_row3_styles,
                    width=raw_width,
                ))

        if progress_callback:
            progress_callback(
                step=f"Packaging Streaming OpenXML: {current_file or source_file.name}",
                processed=actual_mapped_rows,
                total=actual_mapped_rows,
                progress=82,
                current_file=current_file or source_file.name,
            )
        package_started = time.perf_counter()
        _m2c_build_output_package(
            source_file=source_file,
            output_path=output_path,
            activity_sheet_path=activity_sheet_path,
            raw_sheet_path=raw_sheet_path,
            activity_headers=activity_header_rows,
            activity_header_styles=header_styles,
            activity_spool=activity_spool,
            activity_count=actual_mapped_rows,
            activity_width=activity_width,
            raw_headers=raw_header_rows,
            raw_header_styles=raw_header_styles,
            raw_spool=raw_spool,
            raw_count=len(raw_seen),
            raw_width=raw_width,
        )
        package_seconds = time.perf_counter() - package_started

    total_seconds = time.perf_counter() - started
    return {
        "input_filename": source_file.name,
        "activity_rows_read": int(input_rows),
        "activity_rows": int(actual_mapped_rows),
        "raw_materials": int(len(raw_seen)),
        "supplier_matched_rows": int(matched_source_rows),
        "supplier_expanded_rows": int(supplier_expanded_rows),
        "supplier_name_matched_rows": int(supplier_name_matched),
        "supplier_name_missing_rows": int(supplier_name_missing),
        "supplier_dropdown_matched_rows": int(supplier_name_matched),
        "supplier_dropdown_missing_rows": int(supplier_name_missing),
        "tbc_fallback_rows": int(tbc_fallback_rows),
        "supplier_usage_split_source_rows": int(supplier_usage_split_source_rows),
        "supplier_usage_split_output_rows": int(supplier_usage_split_output_rows),
        "supplier_usage_total_before_split": float(supplier_usage_total_before_split),
        "supplier_usage_total_after_split": float(supplier_usage_total_after_split),
        "supplier_usage_split_rule": "equal_share_original_usage_divided_by_unique_supplier_count",
        "supplier_name_options": int(len(supplier_options)),
        "site_tbc_supplier_count": int(len({id(v) for v in tbc_supplier_map.values()})),
        "tbc_fallback_policy": "prefer_uploaded_plant_tbc_then_system_fallback",
        "m2c_large_dataset_mode": True,
        "m2c_writer": "complete_streaming_openxml_package",
        "m2c_xlsx_compression_level": int(M2C_XLSX_COMPRESSION_LEVEL),
        "m2c_total_seconds": round(total_seconds, 3),
        "m2c_package_seconds": round(package_seconds, 3),
        "m2c_template_policy": "Row-2 headers and row-3 styles plus hidden AB~AI formulas are preserved in a lightweight Streaming OpenXML workbook; full dropdown/validation package is restored by M3A.",
        "calculate_transportation_emissions_value": "Preserved from M2B template (YES semantic)",
        "helper_formula_columns": ["AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI"],
        "helper_formula_rows": int(actual_mapped_rows),
    }, supplier_unique


def _write_supplier_bulk_create_file_from_unique_rows(supplier_rows: set[tuple[str, str, str, str, str]], supplier_bulk_template_path: str | Path, output_path: str | Path) -> Dict[str, Any]:
    """Write Supplier Bulk Create from a compact unique supplier row set."""
    if not supplier_rows:
        return {"supplier_bulk_rows": 0, "supplier_bulk_filename": "", "supplier_bulk_download_url": ""}
    supplier_bulk_template_path = Path(supplier_bulk_template_path)
    output_path = Path(output_path)
    if not supplier_bulk_template_path.exists():
        return {"supplier_bulk_rows": 0, "supplier_bulk_filename": "", "supplier_bulk_download_url": "", "supplier_bulk_error": f"找不到 Supplier Bulk Template：{supplier_bulk_template_path.name}"}
    # Supplier list is much smaller than activity data. Preserve the existing supplier bulk template here.
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_meta = load_workbook(supplier_bulk_template_path, read_only=True, data_only=False)
    ws = wb_meta[SUPPLIER_BULK_SHEET_NAME] if SUPPLIER_BULK_SHEET_NAME in wb_meta.sheetnames else wb_meta[wb_meta.sheetnames[0]]
    cols = {
        "supplier_name": _find_template_column(ws, SUPPLIER_BULK_NAME_ALIASES, 1),
        "supplier_code": _find_template_column(ws, SUPPLIER_BULK_CODE_ALIASES, 2),
        "country_area": _find_template_column(ws, SUPPLIER_BULK_COUNTRY_ALIASES, 3),
        "supplier_address": _find_template_column(ws, SUPPLIER_BULK_ADDRESS_ALIASES, 4),
        "unit_name": _find_template_column(ws, SUPPLIER_BULK_UNIT_ALIASES, 5),
        "country_hidden": _find_template_column(ws, ["country"], 15),
        "industry_hidden": _find_template_column(ws, ["industry_category"], 16),
        "tier_hidden": _find_template_column(ws, ["tier"], 17),
    }
    wb_meta.close()
    normalized_rows = _normalize_supplier_bulk_rows(supplier_rows)
    _write_supplier_bulk_openxml(normalized_rows, supplier_bulk_template_path, output_path, cols)
    return {
        "supplier_bulk_rows": int(len(normalized_rows)),
        "supplier_bulk_filename": output_path.name,
        "supplier_bulk_download_url": f"/download/{output_path.name}",
        "supplier_bulk_template_columns": cols,
    }


def generate_supplier_mapped_raw_material_bulk_from_zip(
    raw_material_bulk_zip_path: str | Path,
    supplier_paths: list[str | Path] | tuple[str | Path, ...],
    output_dir: str | Path,
    token: str,
    supplier_bulk_template_path: str | Path | None = None,
    supplier_bulk_output_path: str | Path | None = None,
    progress_callback=None,
) -> Dict[str, Any]:
    """Apply Module 2C supplier mapping with complete Streaming OpenXML.

    Each Module 2B XLSX member is copied from the source ZIP one at a time,
    transformed without openpyxl/XlsxWriter activity-cell objects, and added to
    the result ZIP using ZIP_STORED because XLSX is already compressed.
    """
    started = time.perf_counter()
    source_zip = Path(raw_material_bulk_zip_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    if not source_zip.exists():
        raise FileNotFoundError(f"找不到 Module 2B Raw Material Bulk ZIP：{source_zip}")
    if source_zip.suffix.lower() != ".zip":
        raise ValueError("Module 2C 目前需要讀取 Module 2B 產出的 ZIP 檔案。")

    if progress_callback:
        progress_callback(step="Reading Supplier files", processed=0, total=0, progress=6)
    supplier_map, tbc_supplier_map, supplier_summary = _read_supplier_files(supplier_paths)

    zip_filename = f"supplier_mapped_raw_material_bulk_by_site_{token}.zip"
    zip_path = output_dir / zip_filename
    partial_zip_path = zip_path.with_suffix(zip_path.suffix + ".partial")
    partial_zip_path.unlink(missing_ok=True)
    generated_files: list[dict[str, Any]] = []
    input_files = 0
    total_input_rows = 0
    total_output_rows = 0
    supplier_matched_total = 0
    supplier_expanded_total = 0
    supplier_name_matched_total = 0
    supplier_name_missing_total = 0
    supplier_usage_split_source_total = 0
    supplier_usage_split_output_total = 0
    supplier_usage_total_before_split = 0.0
    supplier_usage_total_after_split = 0.0
    supplier_options_total = 0
    m2c_writer_seconds_total = 0.0
    m2c_package_seconds_total = 0.0
    combined_supplier_rows: set[tuple[str, str, str, str, str]] = set()

    try:
        with tempfile.TemporaryDirectory(prefix="dip_module2c_openxml_") as tmp:
            tmpdir = Path(tmp)
            with zipfile.ZipFile(source_zip, "r") as source_bundle:
                members = [
                    m for m in source_bundle.namelist()
                    if m.lower().endswith((".xlsx", ".xlsm")) and not Path(m).name.startswith("~$")
                ]
                if not members:
                    raise ValueError("Module 2B ZIP 中找不到 Raw Material Bulk Excel 檔案。")

                with zipfile.ZipFile(partial_zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=1, allowZip64=True) as out_zip:
                    for idx, member in enumerate(members, start=1):
                        current_name = Path(member).name
                        source_file = tmpdir / f"{idx:03d}_{current_name}"
                        with source_bundle.open(member, "r") as src, source_file.open("wb") as dst:
                            shutil.copyfileobj(src, dst, length=M2C_COPY_BUFFER_BYTES)

                        if progress_callback:
                            progress_callback(
                                step=f"Streaming OpenXML Supplier mapping: {current_name}",
                                processed=0,
                                total=0,
                                progress=min(88, 12 + int((idx - 1) / max(len(members), 1) * 72)),
                                current_file=current_name,
                            )
                        safe_name = _sanitize_filename_part(Path(member).stem)
                        output_path = tmpdir / f"supplier_mapped_{safe_name}_{token}.xlsx"
                        write_summary, supplier_unique = _write_supplier_mapped_bulk_streaming(
                            source_file=source_file,
                            output_path=output_path,
                            supplier_map=supplier_map,
                            tbc_supplier_map=tbc_supplier_map,
                            progress_callback=progress_callback,
                            current_file=current_name,
                        )
                        combined_supplier_rows.update(supplier_unique)
                        input_files += 1
                        input_rows = int(write_summary.get("activity_rows_read", 0))
                        output_rows = int(write_summary.get("activity_rows", 0))
                        total_input_rows += input_rows
                        total_output_rows += output_rows
                        supplier_matched_total += int(write_summary.get("supplier_matched_rows", 0))
                        supplier_expanded_total += int(write_summary.get("supplier_expanded_rows", 0))
                        supplier_name_matched_total += int(write_summary.get("supplier_name_matched_rows", 0))
                        supplier_name_missing_total += int(write_summary.get("supplier_name_missing_rows", 0))
                        supplier_usage_split_source_total += int(write_summary.get("supplier_usage_split_source_rows", 0))
                        supplier_usage_split_output_total += int(write_summary.get("supplier_usage_split_output_rows", 0))
                        supplier_usage_total_before_split += float(write_summary.get("supplier_usage_total_before_split", 0.0) or 0.0)
                        supplier_usage_total_after_split += float(write_summary.get("supplier_usage_total_after_split", 0.0) or 0.0)
                        supplier_options_total = max(supplier_options_total, int(write_summary.get("supplier_name_options", 0)))
                        m2c_writer_seconds_total += float(write_summary.get("m2c_total_seconds", 0.0) or 0.0)
                        m2c_package_seconds_total += float(write_summary.get("m2c_package_seconds", 0.0) or 0.0)
                        out_zip.write(output_path, arcname=output_path.name, compress_type=zipfile.ZIP_DEFLATED)
                        generated_files.append({
                            "source_filename": current_name,
                            "filename": output_path.name,
                            "input_rows": input_rows,
                            "activity_rows": output_rows,
                            "raw_materials": int(write_summary.get("raw_materials", 0)),
                            "supplier_matched_rows": int(write_summary.get("supplier_matched_rows", 0)),
                            "supplier_expanded_rows": int(write_summary.get("supplier_expanded_rows", 0)),
                            "supplier_name_missing_rows": int(write_summary.get("supplier_name_missing_rows", 0)),
                            "supplier_usage_split_source_rows": int(write_summary.get("supplier_usage_split_source_rows", 0)),
                            "supplier_usage_split_output_rows": int(write_summary.get("supplier_usage_split_output_rows", 0)),
                            "supplier_usage_total_before_split": float(write_summary.get("supplier_usage_total_before_split", 0.0) or 0.0),
                            "supplier_usage_total_after_split": float(write_summary.get("supplier_usage_total_after_split", 0.0) or 0.0),
                            "m2c_writer": write_summary.get("m2c_writer", "complete_streaming_openxml_package"),
                            "m2c_total_seconds": float(write_summary.get("m2c_total_seconds", 0.0) or 0.0),
                            "m2c_package_seconds": float(write_summary.get("m2c_package_seconds", 0.0) or 0.0),
                        })
                        output_path.unlink(missing_ok=True)
                        source_file.unlink(missing_ok=True)

                    supplier_bulk_summary: Dict[str, Any] = {}
                    if supplier_bulk_template_path and supplier_bulk_output_path:
                        if progress_callback:
                            progress_callback(
                                step="Writing Supplier Bulk Create file",
                                processed=total_output_rows,
                                total=total_output_rows,
                                progress=94,
                            )
                        supplier_bulk_summary = _write_supplier_bulk_create_file_from_unique_rows(
                            combined_supplier_rows,
                            supplier_bulk_template_path,
                            supplier_bulk_output_path,
                        )
                        supplier_bulk_filename = str(supplier_bulk_summary.get("supplier_bulk_filename") or "")
                        if supplier_bulk_filename:
                            supplier_bulk_path = Path(supplier_bulk_output_path)
                            if supplier_bulk_path.exists():
                                out_zip.write(supplier_bulk_path, arcname=supplier_bulk_path.name, compress_type=zipfile.ZIP_DEFLATED)
        os.replace(partial_zip_path, zip_path)
    except Exception:
        partial_zip_path.unlink(missing_ok=True)
        raise

    result: Dict[str, Any] = {
        "output_filename": zip_filename,
        "download_url": f"/download/{zip_filename}",
        "module2b_raw_bulk_source_filename": source_zip.name,
        "module2c_rule": "Read Module 2B ZIP -> direct worksheet XML streaming -> supplier expansion -> direct OpenXML package output; original usage is divided equally by supplier count.",
        "module2c_large_dataset_mode": True,
        "module2c_writer": "complete_streaming_openxml_package",
        "module2c_outer_zip": "ZIP_DEFLATED_level_1",
        "module2c_xlsx_compression_level": int(M2C_XLSX_COMPRESSION_LEVEL),
        "module2c_template_policy": "Row-2 headers and row-3 styles plus hidden AB~AI formulas are preserved; full dropdown/validation package is restored by M3A.",
        "input_files": int(input_files),
        "input_rows": int(total_input_rows),
        "activity_rows": int(total_output_rows),
        "generated_files": generated_files,
        "supplier_matched_rows": int(supplier_matched_total),
        "supplier_expanded_rows": int(supplier_expanded_total),
        "supplier_name_matched_rows": int(supplier_name_matched_total),
        "supplier_name_missing_rows": int(supplier_name_missing_total),
        "supplier_usage_split_source_rows": int(supplier_usage_split_source_total),
        "supplier_usage_split_output_rows": int(supplier_usage_split_output_total),
        "supplier_usage_total_before_split": float(supplier_usage_total_before_split),
        "supplier_usage_total_after_split": float(supplier_usage_total_after_split),
        "supplier_usage_split_rule": "equal_share_original_usage_divided_by_unique_supplier_count",
        "supplier_dropdown_matched_rows": int(supplier_name_matched_total),
        "supplier_dropdown_missing_rows": int(supplier_name_missing_total),
        "supplier_name_options": int(supplier_options_total),
        "supplier_status": "Generated",
        "supplier_bulk_generated": False,
        "m2c_writer_seconds": round(m2c_writer_seconds_total, 3),
        "m2c_package_seconds": round(m2c_package_seconds_total, 3),
        "m2c_total_seconds": round(time.perf_counter() - started, 3),
    }
    result.update(supplier_summary)
    if 'supplier_bulk_summary' in locals():
        result.update(supplier_bulk_summary)
        result["supplier_bulk_generated"] = bool(supplier_bulk_summary.get("supplier_bulk_filename"))
    if progress_callback:
        progress_callback(step="Completed", processed=total_output_rows, total=total_output_rows, progress=100)
    return result


BOM_FORMATTER_VERSION = "DIP_V29_M2C_STREAMING_OPENXML_20260724"
