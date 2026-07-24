from __future__ import annotations

import json
import re
import os
import shutil
import time
import traceback
import uuid
from contextvars import ContextVar
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path
from threading import Lock
from typing import Any, Dict, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from bulk_formatter import generate_product_activity_bulk_file, generate_product_activity_bulk_files_by_site, generate_product_activity_bulk_files_by_site_zip
from bom_formatter import BOM_FORMATTER_VERSION, generate_raw_material_bulk_file, generate_raw_material_bulk_files_by_site_zip, export_bom_structure_file, generate_working_hour_rollup_file, generate_working_hour_rollup_file_from_standard_bom, generate_standard_bom_total_usage_file, generate_raw_material_bulk_from_standard_total_usage_zip, generate_supplier_mapped_raw_material_bulk_from_zip
from factor_selector import FACTOR_SELECTOR_VERSION, apply_ccl_factors_to_raw_material_bulk, apply_ccl_factors_to_raw_material_bulk_package, apply_final_template_to_factor_filled_package, collect_factor_library_geographies, preload_factor_libraries, search_factor_library

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"
DATA_DIR = BASE_DIR / "data"
RULE_LIBRARY_DIR = BASE_DIR / "rule_library"
FACTOR_LIBRARY_DIR = DATA_DIR / "factor_library"
LATEST_BOM_STRUCTURE_PATH = OUTPUT_DIR / "bom_structure_latest.xlsx"
LATEST_WORKING_HOUR_ROLLUP_PATH = OUTPUT_DIR / "working_hour_rollup_latest.xlsx"
MODULE2_STANDARD_BOM_TOTAL_USAGE_PATH = OUTPUT_DIR / "standard_bom_total_usage_latest.xlsx"
MODULE2_RAW_MATERIAL_BULK_PATH: Optional[Path] = None
MODULE2B_RAW_MATERIAL_BULK_ZIP_LATEST_PATH = OUTPUT_DIR / "module2b_raw_material_bulk_latest.zip"
MODULE2C_SUPPLIER_MAPPED_BULK_ZIP_LATEST_PATH = OUTPUT_DIR / "module2c_supplier_mapped_raw_material_bulk_latest.zip"
RAW_MATERIAL_BULK_TEMPLATE_LATEST_PATH = OUTPUT_DIR / "raw_material_bulk_template_latest.xlsx"
MODULE1B_PRODUCT_ACTIVITY_BULK_LATEST_PATH = OUTPUT_DIR / "module1b_product_activity_bulk_latest.zip"
MODULE3_CCL_FILLED_LATEST_PATH = OUTPUT_DIR / "module3_ccl_factor_filled_latest.zip"
MODULE3A_FINAL_BULK_LATEST_PATH = OUTPUT_DIR / "module3a_final_raw_material_bulk_latest.zip"


RULE_SET_MAP = {
    "IPS": "LiteOn_IPS",
    "AE": "LiteOn_AE",
    "PC_CE": "LiteOn_PC_CE",
}

DEFAULT_RULE_SET = "IPS"

UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
DATA_DIR.mkdir(exist_ok=True)
RULE_LIBRARY_DIR.mkdir(exist_ok=True)
FACTOR_LIBRARY_DIR.mkdir(exist_ok=True)

app = FastAPI(title="Annual Output Platform v6", version="6.0.0")
print("===== CMP MAIN VERSION: CMP_V15_5_M2B_ROLLUP_TOTAL_HOUR_FILTER =====")
print(f"===== BOM FORMATTER VERSION: {BOM_FORMATTER_VERSION} =====")

MODULE3_CCL_EXECUTOR = ThreadPoolExecutor(max_workers=2)
MODULE3_CCL_JOBS: Dict[str, Dict[str, Any]] = {}
MODULE3A_EXECUTOR = ThreadPoolExecutor(max_workers=1)
MODULE3A_JOBS: Dict[str, Dict[str, Any]] = {}
MODULE1A_EXECUTOR = ThreadPoolExecutor(max_workers=1)
MODULE1A_JOBS: Dict[str, Dict[str, Any]] = {}
MODULE1A_JOB_DIR = OUTPUT_DIR / "module1a_jobs"
MODULE1A_JOB_DIR.mkdir(parents=True, exist_ok=True)
MODULE1B_EXECUTOR = ThreadPoolExecutor(max_workers=1)
MODULE1B_JOBS: Dict[str, Dict[str, Any]] = {}
MODULE1B_JOB_DIR = OUTPUT_DIR / "module1b_jobs"
MODULE1B_JOB_DIR.mkdir(parents=True, exist_ok=True)
MODULE2A_EXECUTOR = ThreadPoolExecutor(max_workers=1)
MODULE2B_EXECUTOR = ThreadPoolExecutor(max_workers=1)
MODULE2C_EXECUTOR = ThreadPoolExecutor(max_workers=1)
MODULE2A_JOBS: Dict[str, Dict[str, Any]] = {}
MODULE2A_JOB_DIR = OUTPUT_DIR / "module2a_jobs"
MODULE2A_JOB_DIR.mkdir(parents=True, exist_ok=True)

CMP_TIMEZONE = ZoneInfo("Asia/Taipei")

# Every full page load receives a server-side timestamp. All AJAX requests from
# that page send the timestamp in X-CMP-Workspace-Started-At. Auto-fetch and
# progress APIs only recognize files created after that timestamp, so F5 starts
# a clean logical workspace without deleting shared output files.
CMP_WORKSPACE_CUTOFF: ContextVar[float] = ContextVar("cmp_workspace_cutoff", default=0.0)
CMP_WORKSPACE_ID: ContextVar[str] = ContextVar("cmp_workspace_id", default="")
CMP_WORKSPACE_OUTPUT_OWNERS: Dict[str, str] = {}
CMP_WORKSPACE_OUTPUT_LOCK = Lock()


def _workspace_cutoff() -> float:
    try:
        return max(0.0, float(CMP_WORKSPACE_CUTOFF.get() or 0.0))
    except (TypeError, ValueError):
        return 0.0


def _workspace_id() -> str:
    return str(CMP_WORKSPACE_ID.get() or "").strip()


def _workspace_path_key(path: Path) -> str:
    try:
        return str(path.resolve())
    except OSError:
        return str(path.absolute())


def _register_workspace_output(path: Path | None, workspace_id: str = "") -> None:
    if not path or not path.exists():
        return
    owner = str(workspace_id or _workspace_id()).strip()
    if not owner:
        return
    with CMP_WORKSPACE_OUTPUT_LOCK:
        CMP_WORKSPACE_OUTPUT_OWNERS[_workspace_path_key(path)] = owner


def _is_workspace_fresh(path: Path | None) -> bool:
    if not path or not path.exists():
        return False
    workspace_id = _workspace_id()
    if workspace_id:
        with CMP_WORKSPACE_OUTPUT_LOCK:
            return CMP_WORKSPACE_OUTPUT_OWNERS.get(_workspace_path_key(path)) == workspace_id
    cutoff = _workspace_cutoff()
    if cutoff <= 0:
        return True
    try:
        return path.stat().st_mtime >= cutoff
    except OSError:
        return False


def _fresh_path(path: Path | None) -> Path | None:
    return path if _is_workspace_fresh(path) else None


def _freshest_path(paths: list[Path]) -> Path | None:
    candidates = [path for path in paths if _is_workspace_fresh(path)]
    return max(candidates, key=lambda p: p.stat().st_mtime) if candidates else None


@app.middleware("http")
async def cmp_workspace_scope(request: Request, call_next):
    raw_cutoff = request.headers.get("x-cmp-workspace-started-at", "")
    workspace_id = str(request.headers.get("x-cmp-workspace-id", "") or "").strip()
    try:
        cutoff = max(0.0, float(raw_cutoff)) if raw_cutoff else 0.0
    except (TypeError, ValueError):
        cutoff = 0.0
    cutoff_token = CMP_WORKSPACE_CUTOFF.set(cutoff)
    workspace_token = CMP_WORKSPACE_ID.set(workspace_id)
    try:
        response = await call_next(request)
        response.headers["X-CMP-Workspace-Started-At"] = str(cutoff or "")
        response.headers["X-CMP-Workspace-ID"] = workspace_id
        return response
    finally:
        CMP_WORKSPACE_ID.reset(workspace_token)
        CMP_WORKSPACE_CUTOFF.reset(cutoff_token)

def _cmp_now_iso() -> str:
    return datetime.now(CMP_TIMEZONE).isoformat(timespec="seconds")

def _cmp_mtime_iso(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime, CMP_TIMEZONE).isoformat(timespec="seconds")

def _extract_source_version_date(filename: str) -> Dict[str, str]:
    text = filename or ""
    version = ""
    date = ""
    version_match = re.search(r"(?:^|[_\-\s])(?:v|V)(\d+(?:\.\d+)*)", text)
    if version_match:
        version = f"v{version_match.group(1)}"
    date_match = re.search(r"(20\d{6})", text)
    if date_match:
        raw = date_match.group(1)
        date = f"{raw[:4]}-{raw[4:6]}-{raw[6:8]}"
    else:
        date_match = re.search(r"(20\d{2})[-_/\.](\d{1,2})[-_/\.](\d{1,2})", text)
        if date_match:
            date = f"{int(date_match.group(1)):04d}-{int(date_match.group(2)):02d}-{int(date_match.group(3)):02d}"
    return {"source_version": version, "source_date": date, "timezone": "Asia/Taipei"}

def _set_module3_ccl_job(job_id: str, **updates: Any) -> None:
    job = MODULE3_CCL_JOBS.setdefault(job_id, {})
    job.update(updates)
    job["updated_at"] = _cmp_now_iso()

def _set_module3a_job(job_id: str, **updates: Any) -> None:
    job = MODULE3A_JOBS.setdefault(job_id, {})
    job.update(updates)
    job["updated_at"] = _cmp_now_iso()


def _module1a_job_path(job_id: str) -> Path:
    return MODULE1A_JOB_DIR / f"{job_id}.json"


def _json_safe_job_value(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, dict):
        return {str(k): _json_safe_job_value(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe_job_value(v) for v in value]
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    return value


def _set_module1a_job(job_id: str, **updates: Any) -> None:
    job = MODULE1A_JOBS.setdefault(job_id, {})
    job.update(updates)
    job["job_id"] = job_id
    job["last_heartbeat"] = _cmp_now_iso()
    try:
        _module1a_job_path(job_id).write_text(
            json.dumps(_json_safe_job_value(job), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception:
        traceback.print_exc()


def _get_module1a_job(job_id: str) -> Dict[str, Any] | None:
    job = MODULE1A_JOBS.get(job_id)
    if job:
        return job
    path = _module1a_job_path(job_id)
    if path.exists():
        try:
            loaded = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                MODULE1A_JOBS[job_id] = loaded
                return loaded
        except Exception:
            traceback.print_exc()
    return None


def _module1b_job_path(job_id: str) -> Path:
    return MODULE1B_JOB_DIR / f"{job_id}.json"


def _set_module1b_job(job_id: str, **updates: Any) -> None:
    job = MODULE1B_JOBS.setdefault(job_id, {})
    job.update(updates)
    job["job_id"] = job_id
    job["last_heartbeat"] = _cmp_now_iso()
    try:
        _module1b_job_path(job_id).write_text(
            json.dumps(_json_safe_job_value(job), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception:
        traceback.print_exc()


def _get_module1b_job(job_id: str) -> Dict[str, Any] | None:
    job = MODULE1B_JOBS.get(job_id)
    if job:
        return job
    path = _module1b_job_path(job_id)
    if path.exists():
        try:
            loaded = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                MODULE1B_JOBS[job_id] = loaded
                return loaded
        except Exception:
            traceback.print_exc()
    return None


def _module2a_job_path(job_id: str) -> Path:
    return MODULE2A_JOB_DIR / f"{job_id}.json"


def _json_safe_module2a(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, dict):
        return {str(k): _json_safe_module2a(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe_module2a(v) for v in value]
    return value


def _set_module2a_job(job_id: str, **updates: Any) -> None:
    job = MODULE2A_JOBS.setdefault(job_id, {})
    job.update(updates)
    job["job_id"] = job_id
    job["last_heartbeat"] = _cmp_now_iso()
    try:
        _module2a_job_path(job_id).write_text(
            json.dumps(_json_safe_module2a(job), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception:
        traceback.print_exc()


def _get_module2a_job(job_id: str) -> Dict[str, Any] | None:
    job = MODULE2A_JOBS.get(job_id)
    if job:
        return job
    path = _module2a_job_path(job_id)
    if path.exists():
        try:
            loaded = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                MODULE2A_JOBS[job_id] = loaded
                return loaded
        except Exception:
            traceback.print_exc()
    return None


def _run_module2a_total_usage_job(
    job_id: str,
    bom_paths: list[Path],
    output_path: Path,
    bom_version: str,
    bom_date: str,
    step1_source_filename: str = "",
    step1_source_modified_at: str = "",
    step1_source_path: Path | None = None,
    workspace_id: str = "",
) -> None:
    started_at = _cmp_now_iso()

    def progress_callback(step: str, processed: int = 0, total: int = 0, progress: int = 0, **extra: Any) -> None:
        _set_module2a_job(
            job_id,
            status="running",
            step=step,
            processed_rows=int(processed or 0),
            total_rows=int(total or 0),
            progress=int(max(0, min(100, progress or 0))),
            **extra,
        )

    def standard_progress_callback(step: str, processed: int = 0, total: int = 0, progress: int = 0, **extra: Any) -> None:
        # Keep room after the total-usage workbook for BOM Structure + Working Hour Roll-up.
        scaled_progress = 2 + int(max(0, min(100, progress or 0)) * 0.83)
        progress_callback(step, processed=processed, total=total, progress=min(85, scaled_progress), **extra)

    try:
        _set_module2a_job(
            job_id,
            status="running",
            step="Queued",
            progress=1,
            processed_rows=0,
            total_rows=0,
            started_at=started_at,
            output_filename=output_path.name,
        )
        summary = generate_standard_bom_total_usage_file(
            bom_path=bom_paths,
            output_path=output_path,
            bom_version=bom_version,
            bom_date=bom_date,
            source_filename=step1_source_filename,
            source_modified_at=step1_source_modified_at,
            progress_callback=standard_progress_callback,
        )
        if output_path.exists():
            _register_workspace_output(output_path, workspace_id)
            shutil.copy2(output_path, MODULE2_STANDARD_BOM_TOTAL_USAGE_PATH)
            _register_workspace_output(MODULE2_STANDARD_BOM_TOTAL_USAGE_PATH, workspace_id)

        summary["module2a_working_hour_rollup_policy"] = "Module 2A creates working_hour_rollup_latest.xlsx for Module 1B when Module 1A annual output/classification is available. Module 1B requires this file only when Working Hour Source = Include Semi-finished Working Hour."
        summary["working_hour_rollup_required_by_step2"] = False
        summary["working_hour_rollup_status"] = "skipped"
        summary["working_hour_rollup_message"] = "未找到 Module 1A 年度產品產量與分類結果，因此僅產出標準BOM表總用量；Module 1B 若選擇 Direct Working Hour 不需要此檔。"

        resolved_step1_path = Path(step1_source_path) if step1_source_path else None
        if resolved_step1_path and resolved_step1_path.exists():
            try:
                working_hour_rollup_output_path = OUTPUT_DIR / f"working_hour_rollup_{job_id}.xlsx"

                # Memory fix: do not export a large BOM Structure workbook in M2A.
                # Module 1B only needs the Summary sheet in working_hour_rollup_latest.xlsx,
                # so the roll-up is generated directly from Standard BOM with a streaming writer.
                summary["bom_structure_status"] = "skipped_streaming_rollup"
                summary["bom_structure_filename"] = ""
                summary["bom_structure_download_url"] = ""
                summary["bom_structure_latest"] = LATEST_BOM_STRUCTURE_PATH.name if LATEST_BOM_STRUCTURE_PATH.exists() else ""
                summary["bom_structure_latest_download_url"] = f"/download/{LATEST_BOM_STRUCTURE_PATH.name}" if LATEST_BOM_STRUCTURE_PATH.exists() else ""
                summary["bom_structure_rows"] = 0
                summary["bom_structure_message"] = "M2A 已改為直接串流產出 Working Hour Roll-up，不再輸出大型 BOM Structure Excel，避免記憶體 crash。"

                progress_callback(
                    "Generating Working Hour Roll-up (streaming)",
                    processed=0,
                    total=int(summary.get("standard_bom_total_usage_rows", 0) or 0),
                    progress=88,
                )
                rollup_summary = generate_working_hour_rollup_file_from_standard_bom(
                    step1_output_path=resolved_step1_path,
                    bom_path=bom_paths,
                    output_path=working_hour_rollup_output_path,
                    mapping=None,
                    progress_callback=progress_callback,
                )
                if working_hour_rollup_output_path.exists():
                    _register_workspace_output(working_hour_rollup_output_path, workspace_id)
                    shutil.copy2(working_hour_rollup_output_path, LATEST_WORKING_HOUR_ROLLUP_PATH)
                    _register_workspace_output(LATEST_WORKING_HOUR_ROLLUP_PATH, workspace_id)
                summary["working_hour_rollup_status"] = "success"
                summary["working_hour_rollup_required_by_step2"] = True
                summary["working_hour_rollup_message"] = "M2A 已以低記憶體串流方式產出 working hour rollup；Module 1B 選擇包含半品工時時會自動引用。"
                summary["working_hour_rollup_filename"] = working_hour_rollup_output_path.name
                summary["working_hour_rollup_download_url"] = f"/download/{working_hour_rollup_output_path.name}"
                summary["working_hour_rollup_latest"] = LATEST_WORKING_HOUR_ROLLUP_PATH.name if LATEST_WORKING_HOUR_ROLLUP_PATH.exists() else ""
                summary["working_hour_rollup_latest_download_url"] = f"/download/{LATEST_WORKING_HOUR_ROLLUP_PATH.name}" if LATEST_WORKING_HOUR_ROLLUP_PATH.exists() else ""
                summary["working_hour_rollup_rows"] = int(rollup_summary.get("summary_rows", 0) or 0)
                summary["working_hour_rollup_detail_rows"] = int(rollup_summary.get("detail_rows", 0) or 0)
                summary["working_hour_rollup_total_direct_hours"] = float(rollup_summary.get("total_direct_hours", 0) or 0)
                summary["working_hour_rollup_total_semi_hours"] = float(rollup_summary.get("total_semi_hours", 0) or 0)
                summary["working_hour_rollup_total_hours"] = float(rollup_summary.get("total_hours", 0) or 0)
                summary["working_hour_rollup_step1_source_filename"] = resolved_step1_path.name
            except Exception as rollup_exc:
                traceback.print_exc()
                summary["working_hour_rollup_status"] = "failed"
                summary["working_hour_rollup_required_by_step2"] = True
                summary["working_hour_rollup_message"] = f"M2A 標準BOM表總用量已完成，但 working hour rollup 產出失敗：{rollup_exc}"
                summary["working_hour_rollup_error"] = str(rollup_exc)
                summary["working_hour_rollup_rows"] = 0
        else:
            summary["working_hour_rollup_filename"] = ""
            summary["working_hour_rollup_download_url"] = ""
            summary["working_hour_rollup_rows"] = 0

        _set_module2a_job(
            job_id,
            status="success",
            step="Completed",
            progress=100,
            processed_rows=int(summary.get("standard_bom_total_usage_rows", 0) or 0),
            total_rows=int(summary.get("standard_bom_total_usage_rows", 0) or 0),
            completed_at=_cmp_now_iso(),
            output_filename=output_path.name,
            download_url=f"/download/{output_path.name}",
            latest_filename=MODULE2_STANDARD_BOM_TOTAL_USAGE_PATH.name if MODULE2_STANDARD_BOM_TOTAL_USAGE_PATH.exists() else "",
            working_hour_rollup_status=summary.get("working_hour_rollup_status", ""),
            working_hour_rollup_filename=summary.get("working_hour_rollup_filename", ""),
            working_hour_rollup_download_url=summary.get("working_hour_rollup_download_url", ""),
            summary=summary,
        )
    except Exception as exc:
        traceback.print_exc()
        _set_module2a_job(
            job_id,
            status="failed",
            step="Failed",
            progress=100,
            error=str(exc),
            traceback=traceback.format_exc(),
            completed_at=_cmp_now_iso(),
        )


def _run_module2b_raw_bulk_job(
    job_id: str,
    total_usage_path: Path,
    template_path: Path,
    output_dir: Path,
    token: str,
    step1_path: Path,
    working_hour_rollup_path: Path,
    step1_source: Dict[str, str],
    workspace_id: str = "",
) -> None:
    global MODULE2_RAW_MATERIAL_BULK_PATH
    started_at = _cmp_now_iso()

    def progress_callback(step: str, processed: int = 0, total: int = 0, progress: int = 0, **extra: Any) -> None:
        _set_module2a_job(
            job_id,
            status="running",
            module="2B",
            step=step,
            processed_rows=int(processed or 0),
            total_rows=int(total or 0),
            progress=int(max(0, min(100, progress or 0))),
            **extra,
        )

    try:
        _set_module2a_job(
            job_id,
            status="running",
            module="2B",
            step="Queued",
            progress=1,
            processed_rows=0,
            total_rows=0,
            started_at=started_at,
            module1_step1_source=step1_source,
            module2a_total_usage_filename=total_usage_path.name,
        )
        summary = generate_raw_material_bulk_from_standard_total_usage_zip(
            standard_total_usage_path=total_usage_path,
            raw_material_template_path=template_path,
            output_dir=output_dir,
            token=token,
            step1_output_path=step1_path,
            working_hour_rollup_path=working_hour_rollup_path,
            progress_callback=progress_callback,
        )
        output_path = output_dir / str(summary.get("output_filename", f"raw_material_activity_data_bulk_by_site_{token}.zip"))
        if output_path.exists():
            MODULE2_RAW_MATERIAL_BULK_PATH = output_path
            _register_workspace_output(output_path, workspace_id)
            shutil.copy2(output_path, MODULE2B_RAW_MATERIAL_BULK_ZIP_LATEST_PATH)
            _register_workspace_output(MODULE2B_RAW_MATERIAL_BULK_ZIP_LATEST_PATH, workspace_id)
        if template_path.exists():
            _register_workspace_output(template_path, workspace_id)
            shutil.copy2(template_path, RAW_MATERIAL_BULK_TEMPLATE_LATEST_PATH)
            _register_workspace_output(RAW_MATERIAL_BULK_TEMPLATE_LATEST_PATH, workspace_id)
            summary["final_raw_material_template_latest"] = RAW_MATERIAL_BULK_TEMPLATE_LATEST_PATH.name
            summary["final_raw_material_template_latest_download_url"] = f"/download/{RAW_MATERIAL_BULK_TEMPLATE_LATEST_PATH.name}"
        summary["module1_step1_source_filename"] = step1_path.name
        summary["module2a_working_hour_rollup_filename"] = working_hour_rollup_path.name
        summary["module2a_working_hour_rollup_download_url"] = f"/download/{working_hour_rollup_path.name}"
        summary["module1_step1_source_download_url"] = f"/download/{step1_path.name}"
        summary["module1_step1_source"] = step1_source
        _set_module2a_job(
            job_id,
            status="success",
            module="2B",
            step="Completed",
            progress=100,
            processed_rows=int(summary.get("activity_rows", 0) or 0),
            total_rows=int(summary.get("activity_rows", 0) or 0),
            completed_at=_cmp_now_iso(),
            output_filename=output_path.name,
            download_url=f"/download/{output_path.name}",
            summary=summary,
        )
    except Exception as exc:
        traceback.print_exc()
        _set_module2a_job(
            job_id,
            status="failed",
            module="2B",
            step="Failed",
            progress=100,
            error=str(exc),
            traceback=traceback.format_exc(),
            completed_at=_cmp_now_iso(),
        )


def _run_module2c_supplier_mapping_job(
    job_id: str,
    raw_bulk_zip_path: Path,
    supplier_paths: list[Path],
    supplier_bulk_template_path: Path,
    output_dir: Path,
    token: str,
    step1_source: Dict[str, str],
    raw_bulk_source: Dict[str, str],
    workspace_id: str = "",
) -> None:
    global MODULE2_RAW_MATERIAL_BULK_PATH
    started_at = _cmp_now_iso()

    def progress_callback(step: str, processed: int = 0, total: int = 0, progress: int = 0, **extra: Any) -> None:
        _set_module2a_job(
            job_id,
            status="running",
            module="2C",
            step=step,
            processed_rows=int(processed or 0),
            total_rows=int(total or 0),
            progress=int(max(0, min(100, progress or 0))),
            **extra,
        )

    try:
        _set_module2a_job(
            job_id,
            status="running",
            module="2C",
            step="Queued",
            progress=1,
            processed_rows=0,
            total_rows=0,
            started_at=started_at,
            module1_step1_source=step1_source,
            module2b_raw_bulk_source=raw_bulk_source,
            supplier_upload_files=len(supplier_paths),
            supplier_bulk_template_filename=supplier_bulk_template_path.name,
        )
        if not supplier_bulk_template_path.exists():
            raise FileNotFoundError(f"找不到已上傳的 Supplier Bulk Template：{supplier_bulk_template_path.name}")
        supplier_bulk_output_path = output_dir / f"supplier_bulk_create_{token}.xlsx"
        summary = generate_supplier_mapped_raw_material_bulk_from_zip(
            raw_material_bulk_zip_path=raw_bulk_zip_path,
            supplier_paths=supplier_paths,
            output_dir=output_dir,
            token=token,
            supplier_bulk_template_path=supplier_bulk_template_path,
            supplier_bulk_output_path=supplier_bulk_output_path,
            progress_callback=progress_callback,
        )
        output_path = output_dir / str(summary.get("output_filename", f"supplier_mapped_raw_material_bulk_by_site_{token}.zip"))
        if output_path.exists():
            MODULE2_RAW_MATERIAL_BULK_PATH = output_path
            _register_workspace_output(output_path, workspace_id)
            summary["latest_alias_mode"] = _refresh_latest_output_alias(output_path, MODULE2C_SUPPLIER_MAPPED_BULK_ZIP_LATEST_PATH)
            if MODULE2C_SUPPLIER_MAPPED_BULK_ZIP_LATEST_PATH.exists() or MODULE2C_SUPPLIER_MAPPED_BULK_ZIP_LATEST_PATH.is_symlink():
                _register_workspace_output(MODULE2C_SUPPLIER_MAPPED_BULK_ZIP_LATEST_PATH, workspace_id)
        if supplier_bulk_output_path.exists():
            _register_workspace_output(supplier_bulk_output_path, workspace_id)
        summary["module1_step1_source"] = step1_source
        summary["module2b_raw_bulk_source"] = raw_bulk_source
        _set_module2a_job(
            job_id,
            status="success",
            module="2C",
            step="Completed",
            progress=100,
            processed_rows=int(summary.get("activity_rows", 0) or 0),
            total_rows=int(summary.get("activity_rows", 0) or 0),
            completed_at=_cmp_now_iso(),
            output_filename=output_path.name,
            download_url=f"/download/{output_path.name}",
            summary=summary,
        )
    except Exception as exc:
        traceback.print_exc()
        _set_module2a_job(
            job_id,
            status="failed",
            module="2C",
            step="Failed",
            progress=100,
            error=str(exc),
            traceback=traceback.format_exc(),
            completed_at=_cmp_now_iso(),
        )

def _run_module3_ccl_job(job_id: str, raw_path: Path, ccl_path: Path, output_path: Path, raw_template_path: Path | None = None, workspace_id: str = "") -> None:
    def report(
        progress: int,
        step: str,
        remaining_seconds: int | None = None,
        *,
        processed_rows: int | None = None,
        total_rows: int | None = None,
    ) -> None:
        payload = {
            "status": "running",
            "progress": max(0, min(100, int(progress))),
            "step": step,
        }
        if remaining_seconds is not None:
            payload["remaining_seconds"] = max(0, int(remaining_seconds))
        if processed_rows is not None:
            payload["processed_rows"] = max(0, int(processed_rows))
        if total_rows is not None:
            payload["total_rows"] = max(0, int(total_rows))
        _set_module3_ccl_job(job_id, **payload)

    try:
        report(1, "建立 CCL 係數對應工作", 45)
        summary = apply_ccl_factors_to_raw_material_bulk_package(
            raw_path, ccl_path, output_path, progress_callback=report, raw_material_template_path=None
        )
        summary["app_version"] = "DIP_MODULE3_LIGHTWEIGHT_INTERMEDIATE_PERF_V5"
        if output_path.exists():
            _register_workspace_output(output_path, workspace_id)
            shutil.copy2(output_path, MODULE3_CCL_FILLED_LATEST_PATH)
            _register_workspace_output(MODULE3_CCL_FILLED_LATEST_PATH, workspace_id)
            summary["module3_latest_filename"] = MODULE3_CCL_FILLED_LATEST_PATH.name
            try:
                summary["output_file_size_bytes"] = output_path.stat().st_size
                summary["output_file_size_mb"] = round(output_path.stat().st_size / 1024 / 1024, 2)
            except OSError:
                pass
        _set_module3_ccl_job(
            job_id,
            status="success",
            progress=100,
            step="M3 CCL 係數對應與輕量中繼檔完成",
            message="M3 已完成 CCL 係數對應；請於 M3A 上傳正式原物料 Template 產出最終 Bulk。",
            remaining_seconds=0,
            summary=summary,
            final_template_filename="",
            final_template_modified_at="",
            stage="M3",
            awaiting_official_template=True,
            download_url=summary.get("download_url", f"/download/{output_path.name}"),
        )
    except Exception as exc:
        traceback.print_exc()
        _set_module3_ccl_job(
            job_id,
            status="error",
            progress=100,
            step="CCL 係數對應失敗",
            message=str(exc),
        )


def _refresh_latest_output_alias(source_path: Path, alias_path: Path) -> str:
    """Refresh a latest-file alias without duplicating a large ZIP on disk."""
    source_path = Path(source_path)
    alias_path = Path(alias_path)
    alias_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        alias_path.unlink(missing_ok=True)
    except OSError:
        pass
    try:
        os.link(source_path, alias_path)
        return "hardlink"
    except OSError:
        try:
            alias_path.symlink_to(source_path.name)
            return "symlink"
        except OSError:
            return "direct-output-only"


def _run_module3a_template_job(
    job_id: str,
    m3_output_path: Path,
    official_template_path: Path,
    output_path: Path,
    workspace_id: str = "",
) -> None:
    def report(
        progress: int,
        step: str,
        remaining_seconds: int | None = None,
        *,
        processed_rows: int | None = None,
        total_rows: int | None = None,
    ) -> None:
        payload: Dict[str, Any] = {
            "status": "running",
            "progress": max(0, min(100, int(progress))),
            "step": step,
        }
        if remaining_seconds is not None:
            payload["remaining_seconds"] = max(0, int(remaining_seconds))
        if processed_rows is not None:
            payload["processed_rows"] = max(0, int(processed_rows))
        if total_rows is not None:
            payload["total_rows"] = max(0, int(total_rows))
        _set_module3a_job(job_id, **payload)

    try:
        report(1, "建立 M3A 正式原物料批次檔套版工作", 60)
        summary = apply_final_template_to_factor_filled_package(
            factor_filled_bulk_path=m3_output_path,
            raw_material_template_path=official_template_path,
            output_path=output_path,
            progress_callback=report,
        )
        summary["app_version"] = "DIP_MODULE3A_STREAM_PACKAGE_PERF_V3"
        if output_path.exists():
            _register_workspace_output(output_path, workspace_id)
            alias_mode = _refresh_latest_output_alias(output_path, MODULE3A_FINAL_BULK_LATEST_PATH)
            summary["latest_output_alias_mode"] = alias_mode
            if alias_mode != "direct-output-only":
                _register_workspace_output(MODULE3A_FINAL_BULK_LATEST_PATH, workspace_id)
            try:
                summary["output_file_size_bytes"] = output_path.stat().st_size
                summary["output_file_size_mb"] = round(output_path.stat().st_size / 1024 / 1024, 2)
            except OSError:
                pass
        _set_module3a_job(
            job_id,
            status="success",
            progress=100,
            step="M3A 正式原物料批次檔已完成",
            message="M3A 正式原物料批次檔已完成。",
            remaining_seconds=0,
            summary=summary,
            source_filename=m3_output_path.name,
            official_template_filename=official_template_path.name,
            download_url=summary.get("download_url", f"/download/{output_path.name}"),
        )
    except Exception as exc:
        traceback.print_exc()
        _set_module3a_job(
            job_id,
            status="error",
            progress=100,
            step="M3A 正式套版失敗",
            message=str(exc),
        )

app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

# =========================================================
# v6 分類邏輯：
# 1. rule_master.csv 依 Priority 由小到大判斷
#    - Material Number Exact
#    - Material Number Prefix
#    - Description Contains
#    - Series Prefix
#    - Plant Exact / Plant Prefix
#    - Default
# 2. 若 rule_master 未命中，再查 product_series_master.csv
# 3. Strict Site 未命中則排除；非 Strict Site 未命中才寫 WIP
# =========================================================

GENERIC_WORDS = {
    "FG", "ASSY", "ASSEMBLY", "MODULE", "MOD", "BL", "BLANK", "NEW", "EURO",
    "NOEURO", "NO", "PCBA", "KB", "KEYBOARD", "TOUCH", "PAD", "CH",
}
SERIES_RE = re.compile(r"([A-Z]{2}[A-Z0-9]*\d[A-Z0-9]*)")

# Product Series 抽取用：遇到這些描述詞時要截斷，避免把 Assy/Touchpad/Module 等描述文字吃進系列名稱
SERIES_STOP_WORDS = [
    "ASSY", "ASSEMBLY", "MODULE", "TOUCHPAD", "TOUCH", "HAPTIC", "FORCE",
    "KEYBOARD", "MOUSE", "TRIBUTO", "LIGHTGRAY", "NOKEYCAP", "NOKEYCA",
    "NOKEY", "BLANK", "NEW", "EURO", "NOEURO",
    "MITS", "US", "UK", "JP", "JIS", "105K", "106K", "110K",
]

# Product series extraction noise words:
# these words describe process/material/assembly context and must not become Product series.
SERIES_NOISE_WORDS = [
    "SMTFOR", "SMT FOR", "SMT", "FOR", "ASSY", "ASSEMBLY", "MODULE",
    "PCB", "ELECTRON", "ELECTRONIC", "PCBA", "BL", "PC", "PET", "PCPET",
    "CH",
]

REQUIRED_ALIASES = {
    "order": ["Order"],
    "plant": ["Plant", "Plant code"],
    "material_number": ["Material Number", "Material", "Product Material Number"],
    "material_description": ["Material description", "Material Description"],
    "delivered_quantity": ["Delivered quantity (GMEIN)", "Delivered quantity", "Delivered Quantity"],
    "finish_date": ["Actual finish date", "Actual Finish Date", "Finish date", "Actual finish"],
}


LABOR_ALIASES = {
    "order": ["Order Number", "Order", "Production Order", "Process order"],
    "plant": ["Plant", "Plant code"],
    "material_number": ["Material Number", "Material", "Product Material Number"],
    "labor_hr": ["Labor HR.Act", "Labor HR Act", "Labor HR.Act.", "Labor HR", "Labor Hour"],
    "foh_others": ["FOH-Others.Act", "FOH Others.Act", "FOH-Others Act", "FOH Others Act", "FOH-Others.Act.", "FOH Others"],
}

VALID_LABOR_MODES = {"labor_hr", "foh", "both"}

FINISHED_PRODUCT_PREFIXES: list[str] = []

def normalize_order_key(value: object) -> str:
    """Normalize SAP order number for matching quantity orders with working-hour orders.

    Rules:
    - Remove trailing .0 generated by Excel numeric conversion.
    - Remove spaces.
    - Keep only letters and numbers.
    - If the order number is numeric, remove leading zeros.
      Example: 000307544123 -> 307544123
    """
    if pd.isna(value):
        return ""

    text = str(value).strip().upper()
    text = re.sub(r"\.0+$", "", text)
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[^A-Z0-9]", "", text)

    if text.isdigit():
        text = text.lstrip("0") or "0"

    return text


def normalize_labor_mode(value: object) -> str:
    """Normalize working-hour source selection.

    Canonical values:
    - both: Labor HR.Act + FOH-Others.Act
    - labor_hr: Labor HR.Act only
    - foh: FOH-Others.Act only
    """
    text = str(value or "both").strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)

    if text in {"labor_hr", "laborhr", "labor-hr", "labor", "hr"}:
        return "labor_hr"
    if text in {"foh", "foh_others", "fohothers", "foh-others"}:
        return "foh"
    if text in {"both", "all", "labor+foh", "laborhr.act+foh-others.act"}:
        return "both"

    if "人員+設備" in text or "人員設備" in text:
        return "both"
    if "人員工時" in text and "設備工時" in text:
        return "both"
    if "人員工時" in text:
        return "labor_hr"
    if "設備工時" in text:
        return "foh"

    if "only" in text:
        if "foh" in text:
            return "foh"
        if "labor" in text:
            return "labor_hr"

    if "foh" in text and "labor" not in text:
        return "foh"
    if "labor" in text and "foh" not in text:
        return "labor_hr"

    return "both"



RULE_COLUMNS = [
    "Priority", "Rule Type", "Key", "Product Type", "Product Line",
    "Production Site", "Is_WIP", "Enabled", "Prefix Length",
]

DEFAULT_RULE_MASTER = (
    "Priority,Rule Type,Key,Product Type,Product Line,Production Site,Is_WIP,Enabled,Prefix Length\n"
    "1,Material Number Prefix,851-,WIP,,,Y,Y,\n"
    "2,Material Number Prefix,852-,WIP,,,Y,Y,\n"
    "10,Series Prefix,SN,NB,,,,Y,\n"
    "11,Series Prefix,FU,NB,,,,Y,\n"
    "12,Series Prefix,SP,TP,,,,Y,\n"
    "13,Series Prefix,SM,DT Mouse,DT,,,Y,\n"
    "14,Series Prefix,SA,DT Accessory,DT,,,Y,\n"
    "15,Description Contains,RECEIVER,DT Dongle,DT,,,Y,\n"
    "16,Series Prefix,SK,DT Keyboard,DT,,,Y,\n"
    "17,Series Prefix,SB,DT Keyboard+Mouse,DT,,,Y,\n"
    "18,Series Prefix,ST,DT Tablet Keyboard,DT,,,Y,\n"
    "19,Description Contains,TOUCH PAD MODULE,TP,,,,Y,\n"
    "20,Description Contains,TOUCHPAD MODULE,TP,,,,Y,\n"
    "21,Series Prefix,SCMC,WIP,,,Y,Y,\n"
    "90,Description Contains,ASSY,WIP,,,Y,Y,\n"
    "999,Default,*,WIP,,,Y,Y,\n"
)

def normalize_rule_set(value: object) -> str:
    text = str(value or DEFAULT_RULE_SET).strip().upper()
    text = text.replace("&", "_").replace("-", "_").replace(" ", "_")
    if text in {"PC&CE", "PC_CE", "PCCE", "PC CE"}:
        text = "PC_CE"
    if text not in RULE_SET_MAP:
        text = DEFAULT_RULE_SET
    return text


def get_rule_set_dir(rule_set: object = DEFAULT_RULE_SET) -> Path:
    normalized = normalize_rule_set(rule_set)
    return RULE_LIBRARY_DIR / RULE_SET_MAP[normalized]


def ensure_master_files() -> None:
    rule_path = DATA_DIR / "rule_master.csv"
    series_path = DATA_DIR / "product_series_master.csv"
    if not rule_path.exists():
        rule_path.write_text(DEFAULT_RULE_MASTER, encoding="utf-8-sig")
    if not series_path.exists():
        series_path.write_text("Plant,Product series,產品類型\n", encoding="utf-8-sig")

    for rule_set_key in RULE_SET_MAP:
        rule_dir = get_rule_set_dir(rule_set_key)
        rule_dir.mkdir(parents=True, exist_ok=True)
        bu_rule_path = rule_dir / "rule_master.csv"
        bu_series_path = rule_dir / "product_series_master.csv"

        if not bu_rule_path.exists():
            if rule_path.exists():
                bu_rule_path.write_bytes(rule_path.read_bytes())
            else:
                bu_rule_path.write_text(DEFAULT_RULE_MASTER, encoding="utf-8-sig")

        if not bu_series_path.exists():
            if series_path.exists():
                bu_series_path.write_bytes(series_path.read_bytes())
            else:
                bu_series_path.write_text("Plant,Product series,產品類型\n", encoding="utf-8-sig")


ensure_master_files()



def production_site_from_line(product_line: object) -> str:
    """Rule Master only mode.

    Production Site must come from rule_master.csv.
    This function is kept only for backward compatibility and never infers a site.
    """
    return ""

def resolve_production_site(product_line: object, production_site: object = "") -> str:
    """Rule Master only mode.

    Production Site is controlled by rule_master.csv only.
    If rule_master.csv leaves Production Site blank, keep it blank.
    """
    return str(production_site or "").strip()

def read_csv_flexible(path: Path, columns: Optional[list[str]] = None) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(columns=columns or [])
    try:
        df = pd.read_csv(path, dtype=str, encoding="utf-8-sig").fillna("")
    except UnicodeDecodeError:
        df = pd.read_csv(path, dtype=str, encoding="big5").fillna("")
    if columns is not None:
        for col in columns:
            if col not in df.columns:
                df[col] = ""
        df = df[columns].copy()
    return df


def load_rule_master(rule_set: object = DEFAULT_RULE_SET) -> pd.DataFrame:
    df = read_csv_flexible(get_rule_set_dir(rule_set) / "rule_master.csv", RULE_COLUMNS)
    for c in RULE_COLUMNS:
        df[c] = df[c].astype(str).str.strip()
    df["Enabled"] = df["Enabled"].str.upper().replace("", "Y")
    df["Rule Type"] = df["Rule Type"].str.strip()
    df["Key"] = df["Key"].str.upper().str.strip()
    df["Priority_num"] = pd.to_numeric(df["Priority"], errors="coerce").fillna(9999)
    df = df[df["Enabled"].isin(["Y", "YES", "TRUE", "1"])]
    return df.sort_values(["Priority_num", "Rule Type", "Key"], kind="stable").reset_index(drop=True)


def load_product_series_master(rule_set: object = DEFAULT_RULE_SET) -> pd.DataFrame:
    path = get_rule_set_dir(rule_set) / "product_series_master.csv"
    df = read_csv_flexible(path)
    for col in ["Plant", "Product series", "產品類型"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].astype(str).str.strip()
    df["Plant"] = df["Plant"].str.replace(r"\.0$", "", regex=True)
    df["Product series"] = df["Product series"].str.upper().str.replace(r"\s+", "", regex=True)
    return df[["Plant", "Product series", "產品類型"]].copy()


def build_masters(rule_set: object = DEFAULT_RULE_SET) -> dict:
    rule_set = normalize_rule_set(rule_set)
    rule_master = load_rule_master(rule_set)
    series_master = load_product_series_master(rule_set)

    by_plant_series: dict[tuple[str, str], pd.Series] = {}
    by_series: dict[str, pd.Series] = {}
    for _, row in series_master.iterrows():
        series = str(row.get("Product series", "") or "").upper().strip()
        plant = str(row.get("Plant", "") or "").replace(".0", "").strip()
        if series:
            by_plant_series[(plant, series)] = row
            if series not in by_series:
                by_series[series] = row

    return {
        "rule_set": rule_set,
        "rules": rule_master,
        "by_plant_series": by_plant_series,
        "by_series": by_series,
    }


def find_col(df: pd.DataFrame, aliases: list[str]) -> Optional[str]:
    normalized = {str(c).strip().lower(): c for c in df.columns}
    for alias in aliases:
        key = alias.strip().lower()
        if key in normalized:
            return normalized[key]
    return None


def get_series_prefixes(masters: Optional[dict] = None) -> list[str]:
    """從 rule_master.csv 動態讀取 Series Prefix，讓新增 SN/SP/SM/SK/FU... 不用再改 Python。"""
    prefixes: list[str] = []
    try:
        rules = masters.get("rules") if masters else load_rule_master()
        if isinstance(rules, pd.DataFrame):
            for _, row in rules.iterrows():
                rule_type = normalize_rule_type(row.get("Rule Type", ""))
                key = str(row.get("Key", "") or "").upper().strip()
                if rule_type in ["series prefix", "product series prefix", "產品系列前綴", "系列前綴"] and key and key not in ["*", "DEFAULT"]:
                    prefixes.append(re.escape(key))
    except Exception:
        prefixes = []

    # fallback：避免 rule_master 空白時完全抓不到
    if not prefixes:
        prefixes = ["SN", "SP", "SM", "SK", "SB", "ST", "SA", "FU"]

    return sorted(set(prefixes), key=len, reverse=True)


def _strip_leading_noise(text: str) -> str:
    """Remove process/material words before Product series extraction."""
    value = str(text or "").upper()
    value = value.replace("&", " ").replace("+", " ")
    value = re.sub(r"SMT\s*FOR", " ", value)
    for word in SERIES_NOISE_WORDS:
        value = re.sub(rf"\b{re.escape(word)}\b", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _extract_series_core(candidate: str) -> str:
    """Extract the core product series and remove trailing region/layout descriptors.

    Examples:
    - SN3103B02US105KBLANKNOKEYCANEW -> SN3103B02
    - SP7D01B02MITS -> SP7D01B02
    - SL6D00B00US -> SL6D00B00
    - SCMC50B11XXX -> SCMC50B11
    """
    value = str(candidate or "").upper()
    if not value:
        return ""

    # First cut at known description words.
    cut_positions = [pos for word in SERIES_STOP_WORDS if (pos := value.find(word)) > 0]

    # Country/layout descriptors often follow the real series immediately,
    # e.g. SN3103B02US105K -> SN3103B02, SL6D00B00US -> SL6D00B00.
    for token in ["US", "UK", "JP", "JIS"]:
        pos = value.find(token)
        if pos > 0:
            cut_positions.append(pos)

    if cut_positions:
        value = value[:min(cut_positions)]

    # Prefer the longest core ending with Letter + 2 digits.
    # This covers common Lite-On series such as SN3103B02, SP7D01B02, SL6D00B00, SCMC50B11.
    core_match = re.match(r"^([A-Z]{2,5}[A-Z0-9]*\d[A-Z]\d{2})", value)
    if core_match:
        return core_match.group(1)

    return value


def trim_series_candidate(candidate: str) -> str:
    """Clean a candidate product series.

    Avoid treating process/material words such as SMTFOR, ASSY, MODULE, PC+PET as part of Product series.
    """
    candidate = str(candidate or "").upper()
    candidate = re.sub(r"[^A-Z0-9].*$", "", candidate)
    candidate = _extract_series_core(candidate)
    return candidate


def _series_candidate_from_text(text: str, pattern: re.Pattern) -> Optional[str]:
    """在一段文字中尋找第一個符合 Series Prefix 的候選值，並套用描述詞截斷。"""
    if not text:
        return None

    # Remove process/material words first, so "SMTfor SP7D01B02" becomes "SP7D01B02".
    cleaned_text = _strip_leading_noise(str(text).upper())

    # 移除空白與非英數符號，讓 CHSN4396BL1 / AssyCH SN5372BL 都能被抓到。
    compact_text = re.sub(r"[^A-Z0-9]+", "", cleaned_text)

    for match in pattern.finditer(compact_text):
        candidate = trim_series_candidate(match.group(0))
        if not candidate or candidate in GENERIC_WORDS:
            continue
        if not re.search(r"\d", candidate):
            continue
        if len(candidate) < 5:
            continue
        return candidate

    return None


def parse_product_series(description: object, masters: Optional[dict] = None) -> tuple[str, str]:
    """
    v6 Product Series 抽取邏輯：
    1. 保留標點符號判斷：先依逗號 / 分號 / 底線切段，在每個區段內搜尋 Series。
    2. 若標點切段找不到，再用整段 Material description 做 Regex Prefix Search。
    3. Prefix 由 rule_master.csv 的 Series Prefix 動態讀取，不用寫死在 Python。
    4. 遇到 ASSY / TOUCHPAD / HAPTIC / MODULE / BLANK / NEW 等描述詞自動截斷。

    可處理：
    - Assy,PCB&Electron, SMTfor SP7D01B02,Mits -> SP7D01B02
    - Assy,Module,BL,PC+PET,SL6D00B00,US -> SL6D00B00
    - Assy,CH,SN3103B02US105KBlanknokeycaNEW -> SN3103B02
    - Assy,CHSN4396BL1,UK,106KBlank,nokeycaNEW -> SN4396BL1
    - SP2B20XF0AssyHapticForce Touchpad module -> SP2B20XF0
    """
    if pd.isna(description):
        return "", "Material description 空白"

    raw_text = str(description).upper()
    pattern = None
    if isinstance(masters, dict):
        pattern = masters.get("_series_prefix_pattern")
    if pattern is None:
        prefixes = get_series_prefixes(masters)
        prefix_pattern = "|".join(prefixes)
        pattern = re.compile(rf"({prefix_pattern})[A-Z0-9]{{3,40}}")
        if isinstance(masters, dict):
            masters["_series_prefix_pattern"] = pattern

    # 第一階段：保留標點符號作為重要邊界，先分段判斷。
    # 這可以避免 SN5372BL,110K 被抓成 SN5372BL110K。
    segments = [seg for seg in re.split(r"[,;_]+", raw_text) if str(seg).strip()]
    for idx, segment in enumerate(segments, start=1):
        candidate = _series_candidate_from_text(segment, pattern)
        if candidate:
            if idx == 1:
                return candidate, "標點切段解析產品系列"
            return candidate, f"標點切段解析產品系列：取第{idx}段"

    # 第二階段：若資料完全沒有標點，改用整段搜尋。
    candidate = _series_candidate_from_text(raw_text, pattern)
    if candidate:
        return candidate, "全文 Regex Prefix Search 解析產品系列"

    # fallback：若 rule_master 沒有涵蓋 prefix，仍保留舊邏輯，但同樣套用截斷
    parts = raw_text.replace("_", ",").replace(";", ",").split(",")
    for idx, part in enumerate(parts, start=1):
        compact = re.sub(r"\s+", "", part)
        compact = trim_series_candidate(compact)
        if not compact or compact in GENERIC_WORDS:
            continue
        match = SERIES_RE.match(compact)
        if match:
            series = trim_series_candidate(match.group(0))
            if series and re.search(r"\d", series) and len(series) >= 5:
                return series, "fallback：逗號分段解析產品系列" if idx == 1 else f"fallback：前段非系列，取第{idx}段"

    return "", "未找到符合產品系列格式的英數碼"



def normalize_rule_type(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def rule_matches(rule_type: str, key: str, material_number: str, description: str, series: str, plant: str = "") -> bool:
    rt = normalize_rule_type(rule_type)
    key_u = str(key or "").upper().strip()
    plant_u = str(plant or "").upper().strip().replace(".0", "")
    if not key_u:
        return False

    # 支援英文與常見中文寫法
    if rt in ["plant exact", "plant", "plant code exact", "廠別", "廠別完全符合", "廠區", "廠區完全符合"]:
        return plant_u == key_u
    if rt in ["plant prefix", "plant code prefix", "廠別前綴", "廠區前綴"]:
        return plant_u.startswith(key_u)
    if rt in ["material number exact", "material exact", "成品料號", "成品料號完全符合"]:
        return material_number == key_u
    if rt in ["material number prefix", "material prefix", "成品料號前綴", "料號前綴"]:
        return material_number.startswith(key_u)
    if rt in ["description contains", "material description contains", "描述包含", "品名包含"]:
        return key_u in description
    if rt in ["series prefix", "product series prefix", "產品系列前綴", "系列前綴"]:
        return series.startswith(key_u)
    if rt in ["series exact", "product series exact", "產品系列", "產品系列完全符合"]:
        return series == key_u
    if rt in ["default", "預設"]:
        return key_u in ["*", "DEFAULT", ""]
    return False


def resolve_plant_production_site_from_rule_master(plant: object, masters: dict) -> tuple[str, str]:
    """Resolve Production Site by Plant rules in rule_master.csv.

    Supported Rule Type:
    - Plant Exact
    - Plant
    - Plant Prefix

    This function only resolves Production Site. It does not change Product Type or Product Line.
    """
    plant_u = str(plant or "").upper().strip().replace(".0", "")
    if not plant_u:
        return "", ""

    rules: pd.DataFrame = masters["rules"]
    for _, row in rules.iterrows():
        rule_type = str(row.get("Rule Type", "") or "").strip()
        key = str(row.get("Key", "") or "").strip()
        rt = normalize_rule_type(rule_type)

        if rt not in [
            "plant exact", "plant", "plant code exact", "廠別", "廠別完全符合", "廠區", "廠區完全符合",
            "plant prefix", "plant code prefix", "廠別前綴", "廠區前綴",
        ]:
            continue

        if not rule_matches(rule_type, key, "", "", "", plant_u):
            continue

        production_site = str(row.get("Production Site", "") or "").strip()
        if production_site:
            return production_site, f"{rule_type}={key}"

    return "", ""




def _normalize_production_site_key(value: object) -> str:
    """Normalize Production Site text for strict-site comparison."""
    return re.sub(r"\s+", "", str(value or "").strip().upper())


def get_strict_production_sites(masters: dict) -> set[str]:
    """Read strict Production Site rules from rule_master.csv.

    Rule purpose:
    - Shared sites such as 越南海防廠-IPS may contain products from other BUs.
    - If no normal classification rule/Product Series Master rule matches, rows from
      these sites must be excluded instead of falling back to Default WIP.

    Supported Rule Type values:
    - Production Site Strict Rule Match
    - Production Site Strict
    - Strict Production Site
    - Default Exclude
    - Exclude If No Match

    Site value can be maintained in either Key or Production Site.
    """
    strict_sites: set[str] = set()
    rules: pd.DataFrame = masters.get("rules", pd.DataFrame())
    for _, row in rules.iterrows():
        rt = normalize_rule_type(row.get("Rule Type", ""))
        if rt not in [
            "production site strict rule match",
            "production site strict",
            "strict production site",
            "default exclude",
            "exclude if no match",
            "廠區嚴格規則",
            "生產廠區嚴格規則",
            "未命中排除",
        ]:
            continue

        for col in ["Production Site", "Key"]:
            site = str(row.get(col, "") or "").strip()
            if site and site not in ["*", "DEFAULT"]:
                strict_sites.add(_normalize_production_site_key(site))
    return strict_sites


def should_exclude_unmatched_by_strict_site(plant: object, masters: dict, current_production_site: object = "") -> tuple[bool, str, str]:
    """Return whether an unmatched row should be excluded by strict-site rule."""
    strict_sites = get_strict_production_sites(masters)
    if not strict_sites:
        return False, "", ""

    production_site = str(current_production_site or "").strip()
    plant_rule_hit = ""
    if not production_site:
        production_site, plant_rule_hit = resolve_plant_production_site_from_rule_master(plant, masters)

    site_key = _normalize_production_site_key(production_site)
    if site_key and site_key in strict_sites:
        return True, production_site, plant_rule_hit
    return False, production_site, plant_rule_hit


def _rule_site_is_compatible(rule_site: object, current_production_site: object) -> bool:
    """Return whether a classification rule may apply to the current plant's Production Site.

    Blank Production Site on a classification rule means the rule is generic and may be used
    by multiple sites. When a rule has a Production Site value, it may only apply to rows
    whose Plant has been resolved to the same Production Site. This prevents a site-specific
    DT/NB/TP rule from assigning the wrong site to another plant while still allowing shared
    generic product-line rules such as DT Mouse/DT Keyboard.
    """
    rule_site_key = _normalize_production_site_key(rule_site)
    if not rule_site_key:
        return True
    current_site_key = _normalize_production_site_key(current_production_site)
    if not current_site_key:
        return True
    return rule_site_key == current_site_key



def _material_number_prefix(value: object, length: int) -> str:
    """Return the first N non-space characters of a material number."""
    text = str(value or "").upper().strip()
    text = re.sub(r"\s+", "", text)
    if length <= 0:
        return text
    return text[:length]


def _as_int_or_default(value: object, default: int = 2) -> int:
    try:
        return int(float(str(value or "").strip()))
    except Exception:
        return default


def get_site_material_prefix_whitelist(masters: dict, current_production_site: object = "") -> list[dict]:
    """Read optional Production Site material-prefix whitelist rules from rule_master.csv.

    Supported Rule Type values:
    - Production Site Material Prefix Whitelist
    - Site Material Prefix Whitelist
    - Material Number Prefix Whitelist
    - Site Prefix Whitelist

    If any whitelist rows exist for the current Production Site, a row must match one
    of the maintained prefixes before normal classification is allowed to continue.
    This is generic and BU-neutral: each BU/site controls its own whitelist in Rule Master.

    Optional Prefix Length values:
    - PREFIX2, FIRST2, 2 => compare first 2 characters.
    - PREFIX3, FIRST3, 3 => compare first 3 characters.
    - FULL, EXACT => compare the full Key length.
    Blank defaults to first 2 characters for backward compatibility with current IPS rules.
    """
    site_key = _normalize_production_site_key(current_production_site)
    if not site_key:
        return []

    rows: list[dict] = []
    rules: pd.DataFrame = masters.get("rules", pd.DataFrame())
    for _, row in rules.iterrows():
        rt = normalize_rule_type(row.get("Rule Type", ""))
        if rt not in [
            "production site material prefix whitelist",
            "site material prefix whitelist",
            "material number prefix whitelist",
            "site prefix whitelist",
            "廠區料號前綴白名單",
            "料號前綴白名單",
        ]:
            continue

        rule_site = str(row.get("Production Site", "") or "").strip()
        if not _rule_site_is_compatible(rule_site, current_production_site):
            continue

        key = str(row.get("Key", "") or "").upper().strip()
        if not key or key in ["*", "DEFAULT"]:
            continue

        logic = str(row.get("Prefix Length", "") or "").strip().upper()
        if logic in ["PREFIX3", "FIRST3", "3"]:
            prefix_len = 3
        elif logic in ["PREFIX4", "FIRST4", "4"]:
            prefix_len = 4
        elif logic in ["FULL", "EXACT"]:
            prefix_len = len(key)
        else:
            prefix_len = _as_int_or_default(logic, 2)

        rows.append({
            "key": key,
            "prefix_len": prefix_len,
            "production_site": rule_site,
            "rule_type": str(row.get("Rule Type", "") or "").strip(),
        })

    return rows


def enforce_site_material_prefix_whitelist(material_number: object, masters: dict, current_production_site: object = "") -> dict:
    """Apply Rule Master-driven site material-prefix whitelist.

    When a Production Site has whitelist rows, only allowed material-number prefixes
    may continue into the normal classifier. Non-whitelisted rows are excluded before
    generic Series Prefix / Description Contains rules can accidentally classify products
    from other BUs at shared factories.
    """
    whitelist = get_site_material_prefix_whitelist(masters, current_production_site)
    if not whitelist:
        return {}

    material_u = str(material_number or "").upper().strip()
    material_compact = re.sub(r"\s+", "", material_u)
    for item in whitelist:
        prefix = _material_number_prefix(material_compact, item["prefix_len"])
        if prefix == item["key"]:
            return {
                "_site_prefix_whitelist_allowed": True,
                "_site_prefix_whitelist_rule": f"{item['rule_type']}={item['key']}",
            }

    prefixes = sorted({item["key"] for item in whitelist})
    return {
        "產品類型": "",
        "Product Line": "",
        "Production Site": str(current_production_site or "").strip(),
        "判斷來源": "Rule Master",
        "規則判定結果": "Excluded",
        "命中規則": "Material Number prefix not in Production Site whitelist: "
                  + (str(material_number or "").strip() or "(blank)")
                  + " not in "
                  + "/".join(prefixes),
        "Is_WIP": "N",
        "_exclude": True,
        "_site_prefix_whitelist_excluded": True,
    }


def classify_by_rule_master(material_number: object, description: object, series: str, masters: dict, current_production_site: object = "") -> dict:
    material_number_u = str(material_number or "").upper().strip()
    description_u = str(description or "").upper().strip()
    series_u = str(series or "").upper().strip()

    rules: pd.DataFrame = masters["rules"]
    for _, row in rules.iterrows():
        rule_type = str(row.get("Rule Type", "") or "").strip()
        key = str(row.get("Key", "") or "").strip()

        # Default 規則只作為文件與下載範本保留；實際 Default WIP 放在 Product Series Master 之後。
        if normalize_rule_type(rule_type) in ["default", "預設", "default exclude", "exclude if no match", "未命中排除"]:
            continue

        if not rule_matches(rule_type, key, material_number_u, description_u, series_u):
            continue

        product_type = str(row.get("Product Type", "") or "").strip()
        is_wip = str(row.get("Is_WIP", "") or "").upper().strip()
        if not is_wip:
            is_wip = "Y" if product_type.upper() == "WIP" else "N"

        # Rules with blank Product Type are markers/metadata, not final classification.
        # Example: 越南海防廠-IPS prefix 90 is only a finished-product whitelist;
        # it must continue through the normal Rule Master / Product Series Master flow.
        if not product_type:
            continue

        # Rule Master only mode:
        # Product Line and Production Site must come from rule_master.csv.
        # Do not auto-fill Product Line from Product Type.
        product_line = str(row.get("Product Line", "") or "").strip()
        production_site = str(row.get("Production Site", "") or "").strip()
        if not _rule_site_is_compatible(production_site, current_production_site):
            continue

        return {
            "產品類型": product_type,
            "Product Line": product_line,
            "Production Site": production_site,
            "判斷來源": "Rule Master",
            "規則判定結果": "符合" if product_type else "待補產品分類",
            "命中規則": f"{rule_type}={key}",
            "Is_WIP": "Y" if is_wip in ["Y", "YES", "TRUE", "1"] else "N",
        }
    return {}


def classify_by_series_master(plant: object, series: str, masters: dict) -> dict:
    plant_str = str(plant or "").strip().replace(".0", "")
    series_u = str(series or "").upper().strip()
    if not series_u:
        return {}

    # 注意：不能使用 row_a or row_b，Pandas Series 不能直接做布林判斷
    row = masters["by_plant_series"].get((plant_str, series_u))
    if row is None:
        row = masters["by_series"].get(series_u)
    if row is None:
        return {}

    product_type = str(row.get("產品類型", "") or "").strip()
    # Rule Master only mode:
    # Product Series Master may classify Product Type only.
    # Product Line / Production Site are not inferred here.
    product_line = ""
    return {
        "產品類型": product_type,
        "Product Line": product_line,
        "Production Site": "",
        "判斷來源": "Product Series Master",
        "規則判定結果": "符合" if product_type else "待補產品分類",
        "命中規則": f"Product series={series_u}",
        "Is_WIP": "Y" if product_type.upper() == "WIP" else "N",
    }


def classify(material_number: object, description: object, series: str, plant: object, masters: dict, current_production_site: object = "") -> dict:
    whitelist_result = enforce_site_material_prefix_whitelist(material_number, masters, current_production_site)
    if whitelist_result.get("_exclude"):
        return whitelist_result

    result = classify_by_rule_master(material_number, description, series, masters, current_production_site)
    if result.get("產品類型"):
        return result

    result = classify_by_series_master(plant, series, masters)
    if result.get("產品類型"):
        return result

    exclude, strict_site, plant_rule_hit = should_exclude_unmatched_by_strict_site(plant, masters, current_production_site)
    if exclude:
        return {
            "產品類型": "",
            "Product Line": "",
            "Production Site": strict_site,
            "判斷來源": "Rule Master",
            "規則判定結果": "Excluded",
            "命中規則": f"No rule matched → Excluded by strict Production Site ({strict_site})" + (f" | {plant_rule_hit}" if plant_rule_hit else ""),
            "Is_WIP": "N",
            "_exclude": True,
        }

    return {
        "產品類型": "WIP",
        "Product Line": "",
        "Production Site": "",
        "判斷來源": "Default WIP",
        "規則判定結果": "WIP",
        "命中規則": "No rule matched → WIP",
        "Is_WIP": "Y",
        "_exclude": False,
    }





def is_finished_product_whitelist(material_number: object) -> bool:
    """Deprecated no-op. Finished product handling is controlled by rule_master.csv."""
    return False


def is_wip_by_rule_master(material_number: object, description: object, series: str, masters: dict) -> bool:
    """Deprecated no-op. Rule Master priority is now the single source of truth."""
    return False


def infer_product_type_line_site_from_series_rules(description: object, series: str, masters: dict) -> tuple[str, str, str]:
    """Infer finished-product classification from rule_master.csv only.

    SG- only means non-WIP. Product Type / Product Line / Production Site
    must still come from matched rule_master.csv fields.
    No fallback from Product Type to Product Line, and no fallback from Product Line to Production Site.
    """
    description_u = str(description or "").upper().strip()
    series_u = str(series or "").upper().strip()
    rules: pd.DataFrame = masters["rules"]

    for _, row in rules.iterrows():
        rule_type = str(row.get("Rule Type", "") or "").strip()
        key = str(row.get("Key", "") or "").strip()
        rt = normalize_rule_type(rule_type)

        if rt not in ["series prefix", "product series prefix", "產品系列前綴", "系列前綴",
                      "series exact", "product series exact", "產品系列", "產品系列完全符合",
                      "description contains", "material description contains", "描述包含", "品名包含"]:
            continue

        if not rule_matches(rule_type, key, "", description_u, series_u):
            continue

        product_type = str(row.get("Product Type", "") or "").strip()
        if product_type.upper() == "WIP":
            continue

        product_line = str(row.get("Product Line", "") or "").strip()
        production_site = str(row.get("Production Site", "") or "").strip()

        if product_type or product_line or production_site:
            return product_type, product_line, production_site

    return "", "", ""

def infer_product_line_site_from_rules(description: object, series: str, masters: dict) -> tuple[str, str]:
    """Infer Product Line / Production Site for WIP from rule_master.csv only.

    WIP Product Type remains WIP. Product Line and Production Site may be filled
    only when the matched rule_master.csv row explicitly provides them.
    """
    description_u = str(description or "").upper().strip()
    series_u = str(series or "").upper().strip()
    rules: pd.DataFrame = masters["rules"]

    for _, row in rules.iterrows():
        rule_type = str(row.get("Rule Type", "") or "").strip()
        key = str(row.get("Key", "") or "").strip()
        rt = normalize_rule_type(rule_type)

        if rt not in [
            "series prefix", "product series prefix", "產品系列前綴", "系列前綴",
            "series exact", "product series exact", "產品系列", "產品系列完全符合",
            "description contains", "material description contains", "描述包含", "品名包含",
        ]:
            continue

        if not rule_matches(rule_type, key, "", description_u, series_u):
            continue

        product_line = str(row.get("Product Line", "") or "").strip()
        production_site = str(row.get("Production Site", "") or "").strip()

        if product_line or production_site:
            return product_line, production_site

    return "", ""

def load_production_dataframe(paths: list[Path]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    errors: list[str] = []

    for path in paths:
        df = pd.read_excel(path, dtype=str)
        cols = {key: find_col(df, aliases) for key, aliases in REQUIRED_ALIASES.items()}
        missing = [key for key, col in cols.items() if col is None]
        if missing:
            errors.append(f"{path.name} 缺少必要欄位：{', '.join(missing)}")
            continue

        part = pd.DataFrame(index=df.index)
        part["Source file"] = path.name
        part["Order"] = df[cols["order"]]
        part["Order Merge Key"] = part["Order"].apply(normalize_order_key)
        part["Plant"] = df[cols["plant"]].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        part["Material Number"] = df[cols["material_number"]].astype(str).str.strip()
        part["Material description"] = df[cols["material_description"]]
        part["Delivered quantity"] = pd.to_numeric(df[cols["delivered_quantity"]], errors="coerce").fillna(0)
        part["Actual finish date"] = pd.to_datetime(df[cols["finish_date"]], errors="coerce")
        part["Year"] = part["Actual finish date"].dt.year
        frames.append(part)

    if errors:
        raise ValueError("；".join(errors))
    if not frames:
        raise ValueError("沒有可處理的工單資料")

    return pd.concat(frames, ignore_index=True)



def load_labor_dataframe(paths: list[Path], labor_mode: str = "both") -> pd.DataFrame:
    """Load production labor work order files and aggregate labor hours by Order first,
    with Plant + Material Number retained as fallback merge keys.
    """
    if not paths:
        return pd.DataFrame(
            columns=[
                "Order", "Order Merge Key", "Plant", "Material Number",
                "Labor HR.Act", "FOH-Others.Act", "Selected Hours", "Labor Source files"
            ]
        )

    mode = str(labor_mode or "both").strip().lower()
    if mode not in VALID_LABOR_MODES:
        mode = "both"

    frames: list[pd.DataFrame] = []
    errors: list[str] = []

    for path in paths:
        df = pd.read_excel(path, dtype=str)
        cols = {key: find_col(df, aliases) for key, aliases in LABOR_ALIASES.items()}

        missing = [key for key in ["order", "labor_hr", "foh_others"] if cols.get(key) is None]
        if missing:
            errors.append(f"{path.name} 缺少必要工時欄位：{', '.join(missing)}")
            continue

        part = pd.DataFrame(index=df.index)
        part["Labor Source file"] = path.name
        part["Order"] = df[cols["order"]].astype(str).str.strip()
        part["Order Merge Key"] = part["Order"].apply(normalize_order_key)

        if cols.get("plant") is not None:
            part["Plant"] = df[cols["plant"]].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        else:
            part["Plant"] = ""

        if cols.get("material_number") is not None:
            part["Material Number"] = df[cols["material_number"]].astype(str).str.strip()
        else:
            part["Material Number"] = ""

        part["Labor HR.Act"] = pd.to_numeric(df[cols["labor_hr"]], errors="coerce").fillna(0)
        part["FOH-Others.Act"] = pd.to_numeric(df[cols["foh_others"]], errors="coerce").fillna(0)

        if mode == "labor_hr":
            part["Selected Hours"] = part["Labor HR.Act"]
        elif mode == "foh":
            part["Selected Hours"] = part["FOH-Others.Act"]
        else:
            part["Selected Hours"] = part["Labor HR.Act"] + part["FOH-Others.Act"]

        frames.append(part)

    if errors:
        raise ValueError("；".join(errors))

    if not frames:
        return pd.DataFrame(
            columns=[
                "Order", "Order Merge Key", "Plant", "Material Number",
                "Labor HR.Act", "FOH-Others.Act", "Selected Hours", "Labor Source files"
            ]
        )

    labor = pd.concat(frames, ignore_index=True)
    labor = labor[labor["Order Merge Key"].astype(str).str.strip() != ""].copy()

    return (
        labor.groupby(["Order Merge Key"], dropna=False, as_index=False)
        .agg({
            "Order": lambda s: "; ".join(sorted(set(str(x) for x in s if str(x).strip()))),
            "Plant": lambda s: "; ".join(sorted(set(str(x) for x in s if str(x).strip()))),
            "Material Number": lambda s: "; ".join(sorted(set(str(x) for x in s if str(x).strip()))),
            "Labor HR.Act": "sum",
            "FOH-Others.Act": "sum",
            "Selected Hours": "sum",
            "Labor Source file": lambda s: "; ".join(sorted(set(str(x) for x in s if str(x).strip())))
        })
        .rename(columns={"Labor Source file": "Labor Source files"})
    )


def attach_labor_hours(out: pd.DataFrame, labor: pd.DataFrame) -> pd.DataFrame:
    """Attach labor hours to production output.

    Primary key only:
    - Production quantity file Order
    - Working-hour file Order Number / Order

    Both sides are matched by normalized Order Merge Key.
    No Plant + Material Number fallback is used.
    """
    out = out.copy()
    if "Order Merge Key" not in out.columns:
        out["Order Merge Key"] = out["Order"].apply(normalize_order_key)

    for col in ["Labor HR.Act", "FOH-Others.Act", "Selected Hours", "Labor Source files"]:
        if col not in out.columns:
            out[col] = 0 if col != "Labor Source files" else ""

    if labor is None or labor.empty:
        return out

    if "Order Merge Key" not in labor.columns:
        labor = labor.copy()
        labor["Order Merge Key"] = labor["Order"].apply(normalize_order_key)

    order_labor = (
        labor[labor["Order Merge Key"].astype(str).str.strip() != ""]
        .groupby(["Order Merge Key"], dropna=False, as_index=False)
        .agg({
            "Order": lambda s: "; ".join(sorted(set(str(x) for x in s if str(x).strip()))),
            "Labor HR.Act": "sum",
            "FOH-Others.Act": "sum",
            "Selected Hours": "sum",
            "Labor Source files": lambda s: "; ".join(sorted(set(str(x) for x in s if str(x).strip())))
        })
        .rename(columns={"Order": "Matched Labor Order Number"})
    )

    if not order_labor.empty:
        out = out.merge(order_labor, on="Order Merge Key", how="left", suffixes=("", "_labor"))
        for col in ["Labor HR.Act", "FOH-Others.Act", "Selected Hours", "Labor Source files"]:
            labor_col = f"{col}_labor"
            if labor_col in out.columns:
                if col == "Labor Source files":
                    out[col] = out[labor_col].fillna("").astype(str)
                else:
                    out[col] = pd.to_numeric(out[labor_col], errors="coerce")
                out = out.drop(columns=[labor_col])

    for col in ["Labor HR.Act", "FOH-Others.Act", "Selected Hours"]:
        out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0)

    for col in ["Labor HR.Act", "FOH-Others.Act", "Selected Hours"]:
        out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0)
    out["Labor Source files"] = out["Labor Source files"].fillna("").astype(str)

    if "Matched Labor Order Number" not in out.columns:
        out["Matched Labor Order Number"] = ""
    out["Matched Labor Order Number"] = out["Matched Labor Order Number"].fillna("").astype(str)
    out["Labor Match Status"] = out["Matched Labor Order Number"].apply(
        lambda x: "Matched" if str(x).strip() else "Not matched"
    )

    return out


def _safe_job_progress_callback(callback: Any, step: str, processed: int = 0, total: int = 0, progress: int = 0, **extra: Any) -> None:
    """Report Module 1A progress without allowing UI callbacks to break processing."""
    if callback is None:
        return
    try:
        callback(
            step=step,
            processed=int(processed or 0),
            total=int(total or 0),
            progress=int(max(0, min(100, progress or 0))),
            **extra,
        )
    except Exception:
        traceback.print_exc()


def _excel_cell_value(value: Any) -> Any:
    """Convert pandas/numpy values to values openpyxl can write like pandas.to_excel would."""
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if hasattr(value, "to_pydatetime"):
        try:
            return value.to_pydatetime()
        except Exception:
            return value
    return value


def _m1a_excel_number_format(column_name: str, series: Optional[pd.Series] = None) -> Optional[str]:
    """Return an Excel display format for M1A streaming exports.

    Numeric values are kept as the original Python/Excel numbers; only the cell
    display format is changed so users see two decimals and Excel does not fall
    back to scientific notation. Identifier/text columns are intentionally left
    as text values.
    """
    name = str(column_name or "").strip()
    if not name:
        return None

    normalized_name = name.lower().replace(" ", "")
    if name in {"Year", "年度"}:
        return "0"
    if name in {"筆數", "工時Order數"} or normalized_name.endswith("count") or normalized_name.endswith("rows"):
        return "0"

    if series is not None:
        try:
            if not pd.api.types.is_numeric_dtype(series):
                return None
        except Exception:
            return None

    numeric_keywords = [
        "quantity", "qty", "hours", "hour", "hr.act", "foh-others.act",
        "年度生產量", "年度人員工時", "年度設備工時", "年度總工時",
        "生產數量占比", "生產工時占比", "工時", "合計", "生產量",
    ]
    if any(keyword in normalized_name for keyword in numeric_keywords):
        return "#,##0.00"

    if series is not None:
        try:
            if pd.api.types.is_numeric_dtype(series):
                return "#,##0.00"
        except Exception:
            pass

    return None


def _format_stream_cell(ws: Any, value: Any, number_format: Optional[str]) -> Any:
    if number_format is None or value is None:
        return value
    cell = WriteOnlyCell(ws, value=value)
    cell.number_format = number_format
    return cell


def _stream_dataframe_to_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    sheet_name: str,
    progress_callback: Any = None,
    progress_start: int = 85,
    progress_end: int = 98,
) -> None:
    """Write a DataFrame to Excel with openpyxl write_only mode.

    This keeps the current M1A workbook structure (sheet names, columns and freeze panes)
    but avoids pandas.to_excel's high memory / slow full-workbook writer path.
    Column widths follow the previous rule: header plus the first 999 data rows, capped at 45.
    """
    ws = wb.create_sheet(title=sheet_name)
    ws.freeze_panes = "A2"

    headers = [str(c) for c in df.columns]
    ws.append(headers)

    # Match the old auto-width behavior without scanning an already-written worksheet.
    widths = [max(12, len(str(h or "")) + 2) for h in headers]
    number_formats = [
        _m1a_excel_number_format(header, df.iloc[:, idx] if idx < len(df.columns) else None)
        for idx, header in enumerate(headers)
    ]
    sample_limit = 999
    total_rows = int(len(df))
    denom = max(1, total_rows)
    progress_span = max(0, progress_end - progress_start)

    for row_index, row in enumerate(df.itertuples(index=False, name=None), start=1):
        values = [_excel_cell_value(v) for v in row]
        excel_row = [
            _format_stream_cell(ws, value, number_formats[idx])
            for idx, value in enumerate(values)
        ]
        ws.append(excel_row)

        if row_index <= sample_limit:
            for i, value in enumerate(values):
                widths[i] = max(widths[i], len(str(value or "")) + 2)

        if progress_callback is not None and (row_index == total_rows or row_index % 5000 == 0):
            progress = progress_start + int(progress_span * row_index / denom)
            _safe_job_progress_callback(
                progress_callback,
                f"寫入 {sheet_name}",
                processed=row_index,
                total=total_rows,
                progress=progress,
            )

    for idx, width in enumerate(widths, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = min(width, 45)


def _export_module1a_workbook_streaming(
    output_path: Path,
    out_export: pd.DataFrame,
    annual_export: pd.DataFrame,
    type_summary: pd.DataFrame,
    progress_callback: Any = None,
) -> None:
    """Export the same M1A workbook sheets using streaming write mode."""
    wb = Workbook(write_only=True)
    _stream_dataframe_to_sheet(wb, out_export, "工單明細_已分類", progress_callback, 82, 92)
    _stream_dataframe_to_sheet(wb, annual_export, "Plant_Material年度產量", progress_callback, 92, 96)
    _stream_dataframe_to_sheet(wb, type_summary, "Plant_產品類型年度產量", progress_callback, 96, 98)
    _safe_job_progress_callback(progress_callback, "儲存 Module 1A Excel", progress=99)
    wb.save(output_path)

def process_files(
    paths: list[Path],
    year: Optional[int],
    labor_paths: Optional[list[Path]] = None,
    labor_mode: str = "both",
    rule_set: str = DEFAULT_RULE_SET,
    progress_callback: Any = None,
) -> tuple[Path, dict]:
    rule_set = normalize_rule_set(rule_set)
    _safe_job_progress_callback(progress_callback, "載入 Rule Master", progress=2, total=len(paths or []))
    masters = build_masters(rule_set)
    labor_mode = normalize_labor_mode(labor_mode)

    _safe_job_progress_callback(progress_callback, "讀取生產工單 Excel", progress=6, total=len(paths or []))
    out = load_production_dataframe(paths)
    _safe_job_progress_callback(progress_callback, "讀取工時工單 Excel", progress=12, processed=len(out), total=len(out))
    labor = load_labor_dataframe(labor_paths or [], labor_mode)
    _safe_job_progress_callback(progress_callback, "合併工單與工時資料", progress=18, processed=len(out), total=len(out))
    out = attach_labor_hours(out, labor)

    # Final safety guard: total working hours must always follow selected working-hour source.
    for col in ["Labor HR.Act", "FOH-Others.Act", "Selected Hours"]:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0)

    if labor_mode == "labor_hr":
        out["Selected Hours"] = out["Labor HR.Act"]
    elif labor_mode == "foh":
        out["Selected Hours"] = out["FOH-Others.Act"]
    else:
        out["Selected Hours"] = out["Labor HR.Act"] + out["FOH-Others.Act"]

    out["Working Hour Source"] = labor_mode

    if year:
        _safe_job_progress_callback(progress_callback, "篩選申報年度", progress=24, processed=len(out), total=len(out))
        out = out[out["Year"] == int(year)].copy()

    _safe_job_progress_callback(progress_callback, "解析 Product series（已啟用快取）", progress=30, processed=0, total=len(out))
    series_cache: dict[str, tuple[str, str]] = {}

    def cached_parse_product_series(desc: object) -> tuple[str, str]:
        key = "" if pd.isna(desc) else str(desc)
        cached = series_cache.get(key)
        if cached is None:
            cached = parse_product_series(desc, masters)
            series_cache[key] = cached
        return cached

    parsed_values = [cached_parse_product_series(desc) for desc in out["Material description"].tolist()]
    out["Product series"] = [x[0] for x in parsed_values]
    out["解析說明"] = [x[1] for x in parsed_values]

    # Resolve Plant -> Production Site before classification.
    # This makes Rule Master classification plant-aware:
    # - site-specific rules only apply to the same Production Site
    # - blank-site rules remain generic and can be used by both 越南海防廠-IPS and 中國石碣廠-IPS
    _safe_job_progress_callback(progress_callback, "解析 Plant 對應生產廠區", progress=38, processed=0, total=len(out))
    plant_site_cache: dict[str, tuple[str, str]] = {}

    def cached_plant_site(plant: object) -> tuple[str, str]:
        key = "" if pd.isna(plant) else str(plant)
        cached = plant_site_cache.get(key)
        if cached is None:
            cached = resolve_plant_production_site_from_rule_master(plant, masters)
            plant_site_cache[key] = cached
        return cached

    initial_plant_site_values = [cached_plant_site(p) for p in out["Plant"].tolist()]
    out["_Plant Production Site"] = [x[0] for x in initial_plant_site_values]
    out["_Plant Production Site Rule"] = [x[1] for x in initial_plant_site_values]

    _safe_job_progress_callback(progress_callback, "套用 Rule Master 與產品分類（已啟用快取）", progress=46, processed=0, total=len(out))
    classify_cache: dict[tuple[str, str, str, str, str], dict] = {}
    classified_values: list[dict] = []
    total_classify_rows = int(len(out))
    for idx, row in enumerate(out[["Material Number", "Material description", "Product series", "Plant", "_Plant Production Site"]].itertuples(index=False, name=None), start=1):
        material_number, description, series, plant, current_site = row
        key = (
            "" if pd.isna(material_number) else str(material_number),
            "" if pd.isna(description) else str(description),
            "" if pd.isna(series) else str(series),
            "" if pd.isna(plant) else str(plant),
            "" if pd.isna(current_site) else str(current_site),
        )
        cached = classify_cache.get(key)
        if cached is None:
            cached = classify(material_number, description, series, plant, masters, current_site)
            classify_cache[key] = cached
        classified_values.append(cached)
        if progress_callback is not None and (idx == total_classify_rows or idx % 10000 == 0):
            _safe_job_progress_callback(
                progress_callback,
                "套用 Rule Master 與產品分類（已啟用快取）",
                processed=idx,
                total=total_classify_rows,
                progress=46 + int(12 * idx / max(1, total_classify_rows)),
            )

    classified = pd.Series(classified_values, index=out.index)
    out["產品類型"] = [x.get("產品類型", "") for x in classified_values]
    out["Product Line"] = [x.get("Product Line", "") for x in classified_values]
    out["Production Site"] = [x.get("Production Site", "") for x in classified_values]
    out["判斷來源"] = [x.get("判斷來源", "") for x in classified_values]
    out["規則判定結果"] = [x.get("規則判定結果", "") for x in classified_values]
    out["命中規則"] = [x.get("命中規則", "") for x in classified_values]
    out["Is_WIP"] = [x.get("Is_WIP", "N") for x in classified_values]
    # Generic rule-engine behavior:
    # If Rule Master classifies a row as WIP and does not explicitly provide Product Line,
    # keep Product Line and Product Series blank. This prevents downstream inference from
    # re-attaching product-family attributes to WIP/SFG rows.
    wip_without_line_mask = (
        out["Is_WIP"].astype(str).str.upper().isin(["Y", "YES", "TRUE", "1"])
        & out["Product Line"].astype(str).str.strip().eq("")
    )
    if wip_without_line_mask.any():
        out.loc[wip_without_line_mask, "Product series"] = ""
        out.loc[wip_without_line_mask, "Product Line"] = ""
        out.loc[wip_without_line_mask, "解析說明"] = "Rule Master：WIP且未指定Product Line，不使用Product Series"

    # If a generic product classification rule is used, keep its Product Type/Product Line
    # and fill Production Site from Plant Exact/Prefix rules. This prevents shared DT rules
    # from forcing all outputs to 越南海防廠-IPS when 石碣廠 uses the same product types.
    blank_site_mask = out["Production Site"].astype(str).str.strip().eq("")
    plant_site_available_mask = out["_Plant Production Site"].astype(str).str.strip().ne("")
    fill_site_mask = blank_site_mask & plant_site_available_mask
    if fill_site_mask.any():
        out.loc[fill_site_mask, "Production Site"] = out.loc[fill_site_mask, "_Plant Production Site"].to_numpy()

    strict_site_excluded_rows = int(classified.apply(lambda x: bool(x.get("_exclude", False))).sum())
    if strict_site_excluded_rows:
        keep_mask = ~classified.apply(lambda x: bool(x.get("_exclude", False)))
        out = out.loc[keep_mask].copy().reset_index(drop=True)

    # Rule Master only mode:
    # Missing Product Line / Production Site may be filled only from explicit rule_master.csv rows.
    missing_line_mask = out["Product Line"].astype(str).str.strip().eq("")
    missing_line_mask = missing_line_mask & ~out["Is_WIP"].astype(str).str.upper().isin(["Y", "YES", "TRUE", "1"])
    if missing_line_mask.any():
        infer_cache: dict[tuple[str, str], tuple[str, str]] = {}
        inferred_values: list[tuple[str, str]] = []
        for description, series in out.loc[missing_line_mask, ["Material description", "Product series"]].itertuples(index=False, name=None):
            key = (
                "" if pd.isna(description) else str(description),
                "" if pd.isna(series) else str(series),
            )
            cached = infer_cache.get(key)
            if cached is None:
                cached = infer_product_line_site_from_rules(description, series, masters)
                infer_cache[key] = cached
            inferred_values.append(cached)
        out.loc[missing_line_mask, "Product Line"] = [x[0] for x in inferred_values]
        out.loc[missing_line_mask, "Production Site"] = [x[1] for x in inferred_values]

    _safe_job_progress_callback(progress_callback, "完成產品分類與生產廠區判斷", progress=62, processed=len(out), total=len(out))
    out["Production Site"] = [
        resolve_production_site(product_line, production_site)
        for product_line, production_site in out[["Product Line", "Production Site"]].itertuples(index=False, name=None)
    ]

    # Final safety guard:
    # Product Line / Production Site can be inferred from series rules, but Product Type must remain WIP
    # if WIP rules such as 850-/851-/852-/H50-/SFG/ASSY/SCMC match.
    wip_cache: dict[tuple[str, str, str], bool] = {}

    def cached_is_wip(material_number: object, description: object, series: object) -> bool:
        key = (
            "" if pd.isna(material_number) else str(material_number),
            "" if pd.isna(description) else str(description),
            "" if pd.isna(series) else str(series),
        )
        cached = wip_cache.get(key)
        if cached is None:
            cached = is_wip_by_rule_master(material_number, description, series, masters)
            wip_cache[key] = cached
        return cached

    wip_rule_mask = pd.Series(
        [cached_is_wip(material_number, description, series) for material_number, description, series in out[["Material Number", "Material description", "Product series"]].itertuples(index=False, name=None)],
        index=out.index,
    )
    if wip_rule_mask.any():
        out.loc[wip_rule_mask, "產品類型"] = "WIP"
        out.loc[wip_rule_mask, "Is_WIP"] = "Y"
        out.loc[wip_rule_mask, "規則判定結果"] = "WIP"

    # Finished-product whitelist:
    # SG- means non-WIP only. Product Type / Product Line / Production Site still comes from Product Series.
    finished_product_mask = out["Material Number"].apply(is_finished_product_whitelist)
    if finished_product_mask.any():
        finished_cache: dict[tuple[str, str], tuple[str, str, str]] = {}
        inferred_finished_values: list[tuple[str, str, str]] = []
        for description, series in out.loc[finished_product_mask, ["Material description", "Product series"]].itertuples(index=False, name=None):
            key = (
                "" if pd.isna(description) else str(description),
                "" if pd.isna(series) else str(series),
            )
            cached = finished_cache.get(key)
            if cached is None:
                cached = infer_product_type_line_site_from_series_rules(description, series, masters)
                finished_cache[key] = cached
            inferred_finished_values.append(cached)

        inferred_product_type = [x[0] for x in inferred_finished_values]
        inferred_product_line = [x[1] for x in inferred_finished_values]
        inferred_production_site = [x[2] for x in inferred_finished_values]

        idx = out.index[finished_product_mask]
        for pos, row_idx in enumerate(idx):
            if inferred_product_type[pos]:
                out.at[row_idx, "產品類型"] = inferred_product_type[pos]
            if inferred_product_line[pos]:
                out.at[row_idx, "Product Line"] = inferred_product_line[pos]
            if inferred_production_site[pos]:
                out.at[row_idx, "Production Site"] = inferred_production_site[pos]
            out.at[row_idx, "Is_WIP"] = "N"

        out.loc[finished_product_mask, "Production Site"] = [
            resolve_production_site(product_line, production_site)
            for product_line, production_site in out.loc[finished_product_mask, ["Product Line", "Production Site"]].itertuples(index=False, name=None)
        ]

    # Plant Rule Master override:
    # Production Site may be controlled directly by Plant Exact / Plant Prefix rules in rule_master.csv.
    # This only updates Production Site and does not change Product Type / Product Line / WIP status.
    plant_site_values = [cached_plant_site(p) for p in out["Plant"].tolist()]
    plant_sites = pd.Series([x[0] for x in plant_site_values], index=out.index)
    plant_rule_hits = pd.Series([x[1] for x in plant_site_values], index=out.index)

    plant_site_mask = plant_sites.astype(str).str.strip().ne("")
    if plant_site_mask.any():
        out.loc[plant_site_mask, "Production Site"] = plant_sites.loc[plant_site_mask].to_numpy()
        out.loc[plant_site_mask, "命中規則"] = out.loc[plant_site_mask, "命中規則"].astype(str) + " | " + plant_rule_hits.loc[plant_site_mask].astype(str)
        out.loc[plant_site_mask, "判斷來源"] = out.loc[plant_site_mask, "判斷來源"].astype(str).where(
            out.loc[plant_site_mask, "判斷來源"].astype(str).str.strip().ne(""),
            "Rule Master"
        )

    out = out.drop(columns=["_Plant Production Site", "_Plant Production Site Rule"], errors="ignore")

    _safe_job_progress_callback(progress_callback, "彙總年度產量與工時", progress=68, processed=len(out), total=len(out))

    group_cols = [
        "Year", "Plant", "Production Site", "Product Line", "Material Number", "Material description", "Product series",
        "產品類型", "判斷來源", "Is_WIP"
    ]
    annual = (
        out.groupby(group_cols, dropna=False, as_index=False)
        .agg({
            "Delivered quantity": "sum",
            "Labor HR.Act": "sum",
            "FOH-Others.Act": "sum",
            "Selected Hours": "sum",
        })
        .rename(columns={
            "Delivered quantity": "年度生產量",
            "Labor HR.Act": "年度人員工時",
            "FOH-Others.Act": "年度設備工時",
            "Selected Hours": "年度總工時",
        })
        .sort_values(["Plant", "Material Number"])
    )

    plant_qty_total = annual.groupby(["Year", "Plant"], dropna=False)["年度生產量"].transform("sum")
    plant_hour_total = annual.groupby(["Year", "Plant"], dropna=False)["年度總工時"].transform("sum")
    annual["生產數量占比(%)"] = 0.0
    annual["生產工時占比(%)"] = 0.0

    qty_mask = plant_qty_total.ne(0)
    hour_mask = plant_hour_total.ne(0)

    annual.loc[qty_mask, "生產數量占比(%)"] = (
        annual.loc[qty_mask, "年度生產量"] / plant_qty_total.loc[qty_mask] * 100
    )
    annual.loc[hour_mask, "生產工時占比(%)"] = (
        annual.loc[hour_mask, "年度總工時"] / plant_hour_total.loc[hour_mask] * 100
    )

    type_summary = (
        out.groupby(["Year", "Plant", "產品類型", "Is_WIP"], dropna=False, as_index=False)["Delivered quantity"]
        .sum()
        .rename(columns={"Delivered quantity": "年度生產量"})
        .sort_values(["Plant", "產品類型"])
    )


    source_summary = (
        out.groupby(["判斷來源", "規則判定結果", "命中規則"], dropna=False, as_index=False)["Delivered quantity"]
        .agg(筆數="count", 生產量="sum")
    )

    file_summary = (
        out.groupby(["Source file"], dropna=False, as_index=False)["Delivered quantity"]
        .agg(筆數="count", 生產量="sum")
        .sort_values(["Source file"])
    )

    # Labor source file summary: helps verify whether each labor file was loaded and how many hours were read.
    if labor is not None and not labor.empty:
        labor_source_summary = (
            labor.copy()
            .assign(**{
                "Labor HR.Act": pd.to_numeric(labor["Labor HR.Act"], errors="coerce").fillna(0),
                "FOH-Others.Act": pd.to_numeric(labor["FOH-Others.Act"], errors="coerce").fillna(0),
                "Selected Hours": pd.to_numeric(labor["Selected Hours"], errors="coerce").fillna(0),
            })
            .groupby(["Labor Source files"], dropna=False, as_index=False)
            .agg({
                "Order Merge Key": "count",
                "Labor HR.Act": "sum",
                "FOH-Others.Act": "sum",
                "Selected Hours": "sum",
            })
            .rename(columns={
                "Labor Source files": "Labor Source file",
                "Order Merge Key": "工時Order數",
                "Labor HR.Act": "Labor HR.Act合計",
                "FOH-Others.Act": "FOH-Others.Act合計",
                "Selected Hours": "Selected Hours合計",
            })
        )
    else:
        labor_source_summary = pd.DataFrame(columns=[
            "Labor Source file", "工時Order數", "Labor HR.Act合計", "FOH-Others.Act合計", "Selected Hours合計"
        ])

    # Labor matching diagnostics: shows how production Orders matched labor Order Numbers.
    diagnostic_cols = [
        "Source file", "Order", "Order Merge Key", "Matched Labor Order Number",
        "Labor Match Status", "Material Number", "Material description",
        "Labor HR.Act", "FOH-Others.Act", "Selected Hours", "Labor Source files"
    ]
    for c in diagnostic_cols:
        if c not in out.columns:
            out[c] = ""
    labor_match_diagnostics = out[diagnostic_cols].copy()
    for c in ["Labor HR.Act", "FOH-Others.Act", "Selected Hours"]:
        labor_match_diagnostics[c] = pd.to_numeric(labor_match_diagnostics[c], errors="coerce").fillna(0)
    labor_match_diagnostics = labor_match_diagnostics.rename(columns={
        "Selected Hours": "Total working hours"
    })

    labor_match_summary = (
        labor_match_diagnostics.groupby(["Labor Match Status"], dropna=False, as_index=False)["Order"]
        .count()
        .rename(columns={"Order": "筆數"})
    )

    wip = out[out["Is_WIP"] == "Y"].copy()

    file_id = uuid.uuid4().hex[:10]
    output_path = OUTPUT_DIR / f"年度產品產量與分類結果_v6_{year or 'ALL'}_{file_id}.xlsx"

    # Export view: only show the selected working-hour source in Excel.
    out_export = out.copy()
    annual_export = annual.copy()

    # Rename Excel output header only; keep internal calculation column as Selected Hours.
    out_export = out_export.rename(columns={"Selected Hours": "Total working hours"})

    if labor_mode == "labor_hr":
        out_export = out_export.drop(columns=["FOH-Others.Act"], errors="ignore")
        annual_export = annual_export.drop(columns=["年度設備工時"], errors="ignore")
    elif labor_mode == "foh":
        out_export = out_export.drop(columns=["Labor HR.Act"], errors="ignore")
        annual_export = annual_export.drop(columns=["年度人員工時"], errors="ignore")

    _safe_job_progress_callback(progress_callback, "寫入 Module 1A Excel（streaming）", progress=82, processed=0, total=len(out_export))
    _export_module1a_workbook_streaming(output_path, out_export, annual_export, type_summary, progress_callback)

    summary = {
        "files": int(len(paths)),
        "labor_files": int(len(labor_paths or [])),
        "labor_mode": labor_mode,
        "rule_set": rule_set,
        "rows": int(len(out)),
        "annual_rows": int(len(annual)),
        "total_qty": float(out["Delivered quantity"].sum()),
        "total_hours": float(out["Selected Hours"].sum()) if "Selected Hours" in out.columns else 0.0,
        "wip_rows": int(len(wip)),
        "strict_site_excluded_rows": int(strict_site_excluded_rows),
        "output_filename": output_path.name,
        "year": year or "ALL",
    }
    return output_path, summary

def process_file(path: Path, year: Optional[int]) -> tuple[Path, dict]:
    """Backward-compatible wrapper for single-file processing."""
    return process_files([path], year, None, "both", DEFAULT_RULE_SET)


def normalize_rule_upload(df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {}
    for c in df.columns:
        key = str(c).strip().lower()
        if key in ["priority", "優先順序", "排序"]:
            rename_map[c] = "Priority"
        elif key in ["rule type", "規則類型", "判斷類型"]:
            rename_map[c] = "Rule Type"
        elif key in ["key", "規則值", "關鍵字", "prefix", "前綴"]:
            rename_map[c] = "Key"
        elif key in ["product type", "產品分類", "產品類型"]:
            rename_map[c] = "Product Type"
        elif key in ["prefix length", "prefix_len", "match length", "match_length", "前綴長度", "比對長度"]:
            rename_map[c] = "Prefix Length"
        elif key in ["is_wip", "is wip", "wip", "半品"]:
            rename_map[c] = "Is_WIP"
        elif key in ["product line", "產品線", "歸屬類型", "production site type"]:
            rename_map[c] = "Product Line"
        elif key in ["production site", "生產廠區", "廠區", "廠別"]:
            rename_map[c] = "Production Site"
        elif key in ["enabled", "啟用"]:
            rename_map[c] = "Enabled"
    df = df.rename(columns=rename_map).fillna("")
    for col in RULE_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[RULE_COLUMNS].copy()
    for c in RULE_COLUMNS:
        df[c] = df[c].astype(str).str.strip()
    df["Enabled"] = df["Enabled"].replace("", "Y")
    df["Is_WIP"] = df["Is_WIP"].replace("", "N")
    # Product Type may be blank for rules that only maintain Production Site,
    # e.g. Plant Exact, 3760 -> Production Site.
    df = df[(df["Rule Type"] != "") & (df["Key"] != "")]
    return df


def save_uploaded_rule(file_path: Path, rule_set: object = DEFAULT_RULE_SET) -> int:
    if file_path.suffix.lower() in [".xlsx", ".xlsm", ".xls"]:
        df = pd.read_excel(file_path, dtype=str).fillna("")
    else:
        df = pd.read_csv(file_path, dtype=str, encoding="utf-8-sig").fillna("")
    df = normalize_rule_upload(df)
    target_dir = get_rule_set_dir(rule_set)
    target_dir.mkdir(parents=True, exist_ok=True)
    df.to_csv(target_dir / "rule_master.csv", index=False, encoding="utf-8-sig")
    return len(df)


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    workspace_started_at = time.time()
    workspace_id = uuid.uuid4().hex
    response = templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "workspace_started_at": workspace_started_at,
            "workspace_id": workspace_id,
        },
    )
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response




@app.get("/debug-version")
def debug_version():
    return {
        "ok": True,
        "app": "Carbon Management Platform",
        "version": "PROCESS_MANUAL_FORM_V6",
        "process_endpoint": "manual form compatible",
        "supports": ["files multi-upload", "file single-upload", "Module 2 multi-BOM upload", "blank year", "BU rule library"],
    }


def _run_module1a_process_job(
    job_id: str,
    saved_paths: list[Path],
    year_value: Optional[int],
    saved_labor_paths: list[Path],
    labor_mode: str,
    rule_set: str,
    workspace_id: str = "",
) -> None:
    started_at = datetime.now(CMP_TIMEZONE)

    def report(step: str, processed: int = 0, total: int = 0, progress: int = 0, **extra: Any) -> None:
        elapsed_seconds = max(1, int((datetime.now(CMP_TIMEZONE) - started_at).total_seconds()))
        current_progress = max(0, min(100, int(progress or 0)))
        remaining_seconds = None
        if 1 <= current_progress < 100:
            remaining_seconds = int(elapsed_seconds * max(0, 100 - current_progress) / current_progress)
        _set_module1a_job(
            job_id,
            status="running",
            step=step,
            progress=current_progress,
            processed_rows=int(processed or 0),
            total_rows=int(total or 0),
            elapsed_seconds=elapsed_seconds,
            remaining_seconds=remaining_seconds,
            **extra,
        )

    try:
        _set_module1a_job(
            job_id,
            status="running",
            step="準備處理 Module 1A 工單",
            progress=1,
            processed_rows=0,
            total_rows=0,
            started_at=_cmp_now_iso(),
        )
        output_path, summary = process_files(
            saved_paths,
            year_value,
            saved_labor_paths,
            labor_mode,
            rule_set,
            progress_callback=report,
        )
        _register_workspace_output(output_path, workspace_id)
        elapsed_seconds = max(1, int((datetime.now(CMP_TIMEZONE) - started_at).total_seconds()))
        _set_module1a_job(
            job_id,
            status="done",
            step="Module 1A 年度產品產量與分類結果已產出",
            progress=100,
            processed_rows=int(summary.get("rows") or 0),
            total_rows=int(summary.get("rows") or 0),
            elapsed_seconds=elapsed_seconds,
            remaining_seconds=0,
            summary=summary,
            output_filename=output_path.name,
            download_url=f"/download/{output_path.name}",
            completed_at=_cmp_now_iso(),
        )
    except Exception as exc:
        traceback.print_exc()
        _set_module1a_job(
            job_id,
            status="error",
            step="Module 1A 處理失敗",
            progress=100,
            message=str(exc),
            completed_at=_cmp_now_iso(),
        )


@app.post("/module1a/process-job")
async def module1a_process_job(request: Request):
    """Module 1A background processing endpoint with real progress status."""
    try:
        form = await request.form()
    except Exception as exc:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": f"無法讀取上傳表單：{exc}"}, status_code=400)

    available_keys = list(form.keys())

    def is_upload_file_like(item) -> bool:
        return bool(getattr(item, "filename", None)) and hasattr(item, "read")

    upload_files = []
    labor_uploads = []

    for item in form.getlist("files") + form.getlist("file"):
        if is_upload_file_like(item):
            upload_files.append(item)

    for item in form.getlist("labor_files") + form.getlist("labor_file"):
        if is_upload_file_like(item):
            labor_uploads.append(item)

    seen_ids = {id(x) for x in upload_files + labor_uploads}
    for key in available_keys:
        for item in form.getlist(key):
            if not is_upload_file_like(item) or id(item) in seen_ids:
                continue
            key_lower = str(key).lower()
            if "labor" in key_lower or "hour" in key_lower or "worktime" in key_lower:
                labor_uploads.append(item)
            else:
                upload_files.append(item)
            seen_ids.add(id(item))

    if not upload_files:
        return JSONResponse({"ok": False, "message": "請至少上傳一個 Excel 生產數量工單檔案"}, status_code=400)

    labor_mode = normalize_labor_mode(form.get("labor_mode") or "both")
    rule_set = normalize_rule_set(form.get("rule_set") or DEFAULT_RULE_SET)
    year = form.get("year")

    try:
        year_value: Optional[int] = None
        if year is not None and str(year).strip() != "":
            year_value = int(str(year).strip())
    except Exception:
        return JSONResponse({"ok": False, "message": "Reporting Year 請輸入有效年份，或留空代表全部年度。"}, status_code=400)

    job_id = uuid.uuid4().hex[:10]
    workspace_id = _workspace_id()
    saved_paths: list[Path] = []
    for idx, upload in enumerate(upload_files, start=1):
        filename = str(getattr(upload, "filename", "") or f"production_order_{idx}.xlsx")
        if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            return JSONResponse({"ok": False, "message": f"{filename} 不是 Excel 檔案"}, status_code=400)
        saved = UPLOAD_DIR / f"module1a_workorder_{job_id}_{idx}_{Path(filename).name}"
        saved.write_bytes(await upload.read())
        saved_paths.append(saved)

    saved_labor_paths: list[Path] = []
    for idx, upload in enumerate(labor_uploads, start=1):
        filename = str(getattr(upload, "filename", "") or f"labor_order_{idx}.xlsx")
        if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            return JSONResponse({"ok": False, "message": f"{filename} 不是 Excel 工時檔案"}, status_code=400)
        saved = UPLOAD_DIR / f"module1a_labor_{job_id}_{idx}_{Path(filename).name}"
        saved.write_bytes(await upload.read())
        saved_labor_paths.append(saved)

    _set_module1a_job(
        job_id,
        status="queued",
        step="Module 1A 工作已建立，等待背景處理",
        progress=0,
        processed_rows=0,
        total_rows=0,
        created_at=_cmp_now_iso(),
        files=len(saved_paths),
        labor_files=len(saved_labor_paths),
        labor_mode=labor_mode,
        rule_set=rule_set,
        year=year_value if year_value is not None else "ALL",
        workspace_id=workspace_id,
    )
    MODULE1A_EXECUTOR.submit(
        _run_module1a_process_job,
        job_id,
        saved_paths,
        year_value,
        saved_labor_paths,
        labor_mode,
        rule_set,
        workspace_id,
    )
    return {"ok": True, "job_id": job_id, "status_url": f"/module1a/process-job/{job_id}"}


@app.get("/module1a/process-job/{job_id}")
def module1a_get_process_job(job_id: str):
    job = _get_module1a_job(job_id)
    if not job:
        return JSONResponse({"ok": False, "message": "找不到 Module 1A job，可能已被清除或服務重啟前未寫入狀態。"}, status_code=404)
    response = dict(job)
    response["ok"] = True
    response["job"] = dict(job)
    return response

@app.post("/process")
async def process(request: Request):
    """Step 1 processing endpoint.

    Manual multipart reader:
    - accepts production quantity files from files/file
    - accepts working-hour files from labor_files/labor_file
    - avoids FastAPI 422 validation when optional fields are omitted
    """
    try:
        form = await request.form()
    except Exception as exc:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": f"無法讀取上傳表單：{exc}"}, status_code=400)

    available_keys = list(form.keys())

    def is_upload_file_like(item) -> bool:
        return bool(getattr(item, "filename", None)) and hasattr(item, "read")

    upload_files = []
    labor_uploads = []

    for item in form.getlist("files") + form.getlist("file"):
        if is_upload_file_like(item):
            upload_files.append(item)

    for item in form.getlist("labor_files") + form.getlist("labor_file"):
        if is_upload_file_like(item):
            labor_uploads.append(item)

    seen_ids = {id(x) for x in upload_files + labor_uploads}
    for key in available_keys:
        for item in form.getlist(key):
            if not is_upload_file_like(item) or id(item) in seen_ids:
                continue

            key_lower = str(key).lower()
            if "labor" in key_lower or "hour" in key_lower or "worktime" in key_lower:
                labor_uploads.append(item)
            else:
                upload_files.append(item)
            seen_ids.add(id(item))

    if not upload_files:
        return JSONResponse({"ok": False, "message": "請至少上傳一個 Excel 生產數量工單檔案"}, status_code=400)

    labor_mode = normalize_labor_mode(form.get("labor_mode") or "both")
    rule_set = normalize_rule_set(form.get("rule_set") or DEFAULT_RULE_SET)
    year = form.get("year")

    saved_paths: list[Path] = []
    for upload in upload_files:
        filename = str(getattr(upload, "filename", "") or "")
        if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            return JSONResponse({"ok": False, "message": f"{filename} 不是 Excel 檔案"}, status_code=400)
        saved = UPLOAD_DIR / f"{uuid.uuid4().hex}_{Path(filename).name}"
        saved.write_bytes(await upload.read())
        saved_paths.append(saved)

    saved_labor_paths: list[Path] = []
    for upload in labor_uploads:
        filename = str(getattr(upload, "filename", "") or "")
        if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            return JSONResponse({"ok": False, "message": f"{filename} 不是 Excel 工時檔案"}, status_code=400)
        saved = UPLOAD_DIR / f"labor_{uuid.uuid4().hex}_{Path(filename).name}"
        saved.write_bytes(await upload.read())
        saved_labor_paths.append(saved)

    try:
        year_value: Optional[int] = None
        if year is not None and str(year).strip() != "":
            year_value = int(str(year).strip())

        output_path, summary = process_files(saved_paths, year_value, saved_labor_paths, labor_mode, rule_set)
        _register_workspace_output(output_path)
    except Exception as exc:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=400)

    return {"ok": True, "summary": summary, "download_url": f"/download/{output_path.name}"}


# =========================================================
# Step 2 · Batch Data Formatting
# Module 1A annual output/classification + Bulk Template -> Formatted Product Activity Bulk
# =========================================================
def _run_module1b_bulk_job(
    job_id: str,
    step1_path: Path,
    template_path: Path,
    working_hour_source: str,
    bom_structure_path: Path | None,
    working_hour_rollup_path: Path | None,
    step1_source_mode: str,
    workspace_id: str = "",
) -> None:
    started_at = datetime.now(CMP_TIMEZONE)

    def report(step: str, processed: int = 0, total: int = 0, progress: int = 0, **extra: Any) -> None:
        elapsed_seconds = max(1, int((datetime.now(CMP_TIMEZONE) - started_at).total_seconds()))
        current_progress = max(0, min(100, int(progress or 0)))
        remaining_seconds = None
        if 1 <= current_progress < 100:
            remaining_seconds = int(elapsed_seconds * max(0, 100 - current_progress) / current_progress)
        _set_module1b_job(
            job_id,
            status="running",
            step=step,
            progress=current_progress,
            processed_rows=int(processed or 0),
            total_rows=int(total or 0),
            elapsed_seconds=elapsed_seconds,
            remaining_seconds=remaining_seconds,
            **extra,
        )

    try:
        _set_module1b_job(
            job_id,
            status="running",
            step="準備產出 Product Activity Bulk",
            progress=1,
            processed_rows=0,
            total_rows=0,
            started_at=_cmp_now_iso(),
            source_filename=step1_path.name,
            template_filename=template_path.name,
            working_hour_source=working_hour_source,
        )
        summary = generate_product_activity_bulk_files_by_site_zip(
            step1_output_path=step1_path,
            bulk_template_path=template_path,
            output_dir=OUTPUT_DIR,
            token=job_id,
            working_hour_source=working_hour_source,
            bom_structure_path=bom_structure_path,
            working_hour_rollup_path=working_hour_rollup_path,
            progress_callback=report,
        )
        summary["module1_step1_source_mode"] = step1_source_mode
        summary["module1_step1_source_filename"] = step1_path.name if step1_path else ""
        summary["module1_step1_source_download_url"] = f"/download/{step1_path.name}" if step1_path else ""
        summary["module2a_working_hour_rollup_used"] = bool(working_hour_rollup_path and Path(working_hour_rollup_path).exists())
        summary["module2a_working_hour_rollup_filename"] = Path(working_hour_rollup_path).name if working_hour_rollup_path else ""
        summary["working_hour_source_rule"] = "Module 2A working_hour_rollup is required only when Working Hour Source = Include Semi-finished Working Hour."

        output_filename = str(summary.get("output_filename") or "").strip()
        generated_output_path = OUTPUT_DIR / output_filename if output_filename else None
        if generated_output_path and generated_output_path.exists() and generated_output_path.suffix.lower() == ".zip":
            _register_workspace_output(generated_output_path, workspace_id)
            shutil.copy2(generated_output_path, MODULE1B_PRODUCT_ACTIVITY_BULK_LATEST_PATH)
            _register_workspace_output(MODULE1B_PRODUCT_ACTIVITY_BULK_LATEST_PATH, workspace_id)
            summary["module1b_product_activity_bulk_latest"] = MODULE1B_PRODUCT_ACTIVITY_BULK_LATEST_PATH.name
            summary["module1b_product_activity_bulk_latest_download_url"] = f"/download/{MODULE1B_PRODUCT_ACTIVITY_BULK_LATEST_PATH.name}"

        activity_rows = int(summary.get("activity_rows", 0) or 0)
        elapsed_seconds = max(1, int((datetime.now(CMP_TIMEZONE) - started_at).total_seconds()))
        _set_module1b_job(
            job_id,
            status="success",
            step="Product Activity Bulk 已完成",
            progress=100,
            processed_rows=activity_rows,
            total_rows=activity_rows,
            elapsed_seconds=elapsed_seconds,
            remaining_seconds=0,
            completed_at=_cmp_now_iso(),
            output_filename=output_filename,
            download_url=summary.get("download_url", f"/download/{output_filename}" if output_filename else ""),
            summary=summary,
        )
    except Exception as exc:
        traceback.print_exc()
        _set_module1b_job(
            job_id,
            status="error",
            step="Product Activity Bulk 產出失敗",
            progress=100,
            message=str(exc),
            error=str(exc),
            completed_at=_cmp_now_iso(),
        )


@app.post("/module1b/generate-bulk-file-job")
async def module1b_generate_bulk_file_job(request: Request):
    try:
        form = await request.form()
    except Exception as exc:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": f"無法讀取上傳表單：{exc}"}, status_code=400)

    template_file = form.get("template_file")
    step1_file = form.get("step1_file")
    working_hour_source = str(form.get("working_hour_source") or "direct").strip()

    if not template_file or not getattr(template_file, "filename", None):
        return JSONResponse({"ok": False, "message": "Bulk Template 請上傳 Excel 檔案"}, status_code=400)
    template_filename = str(template_file.filename or "")
    if not template_filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        return JSONResponse({"ok": False, "message": "Bulk Template 請上傳 Excel 檔案"}, status_code=400)

    job_id = uuid.uuid4().hex[:10]
    workspace_id = _workspace_id()
    uploaded_step1_filename = str(getattr(step1_file, "filename", "") or "").strip() if step1_file else ""
    if uploaded_step1_filename:
        if not uploaded_step1_filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            return JSONResponse({"ok": False, "message": "Module 1A Output 請上傳 Excel 檔案"}, status_code=400)
        step1_path = UPLOAD_DIR / f"step1_output_{job_id}_{Path(uploaded_step1_filename).name}"
        step1_path.write_bytes(await step1_file.read())
        step1_source_mode = "uploaded"
    else:
        step1_path = _find_latest_module1_step1_output()
        step1_source_mode = "latest"
        if step1_path is None or not step1_path.exists():
            return JSONResponse({"ok": False, "message": "尚未找到 Module 1A 年度產品產量與分類結果。請先完成 Module 1A。"}, status_code=400)

    template_path = UPLOAD_DIR / f"bulk_template_{job_id}_{Path(template_filename).name}"
    template_path.write_bytes(await template_file.read())

    bom_structure_path = None
    working_hour_rollup_path = None
    if _is_include_semi_working_hour_source(working_hour_source):
        if _find_latest_working_hour_rollup() is None:
            return JSONResponse({
                "ok": False,
                "message": "選擇『包含半品工時』時需要 Module 2A working_hour_rollup。請先完成 Module 2A；若只使用直接工時，請將 Working Hour Source 改為 Direct Working Hour。",
            }, status_code=400)
        working_hour_rollup_path = _find_latest_working_hour_rollup()
        bom_structure_path = _fresh_path(LATEST_BOM_STRUCTURE_PATH)

    _set_module1b_job(
        job_id,
        status="queued",
        step="Product Activity Bulk 工作已建立，等待背景處理",
        progress=0,
        processed_rows=0,
        total_rows=0,
        created_at=_cmp_now_iso(),
        source_filename=step1_path.name,
        template_filename=template_path.name,
        working_hour_source=working_hour_source,
        workspace_id=workspace_id,
    )
    MODULE1B_EXECUTOR.submit(
        _run_module1b_bulk_job,
        job_id,
        step1_path,
        template_path,
        working_hour_source,
        bom_structure_path,
        working_hour_rollup_path,
        step1_source_mode,
        workspace_id,
    )
    return {"ok": True, "job_id": job_id, "status_url": f"/module1b/generate-bulk-file-job/{job_id}"}


@app.get("/module1b/generate-bulk-file-job/{job_id}")
def module1b_get_generate_bulk_file_job(job_id: str):
    job = _get_module1b_job(job_id)
    if not job:
        return JSONResponse({"ok": False, "message": "找不到 Module 1B job，可能已被清除或服務重啟前未寫入狀態。"}, status_code=404)
    response = dict(job)
    response["ok"] = True
    response["job"] = dict(job)
    return response


@app.post("/generate-bulk-file")
async def generate_bulk_file(
    template_file: UploadFile = File(...),
    step1_file: Optional[UploadFile] = File(None),
    working_hour_source: str = Form("direct"),
):
    if not template_file or not str(template_file.filename or "").lower().endswith((".xlsx", ".xlsm", ".xls")):
        return JSONResponse({"ok": False, "message": "Bulk Template 請上傳 Excel 檔案"}, status_code=400)

    token = uuid.uuid4().hex[:10]
    workspace_id = _workspace_id()

    uploaded_step1_filename = str(getattr(step1_file, "filename", "") or "").strip() if step1_file else ""
    if uploaded_step1_filename:
        if not uploaded_step1_filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            return JSONResponse({"ok": False, "message": "Module 1A Output 請上傳 Excel 檔案"}, status_code=400)
        step1_path = UPLOAD_DIR / f"step1_output_{token}_{Path(uploaded_step1_filename).name}"
        step1_path.write_bytes(await step1_file.read())
        step1_source_mode = "uploaded"
    else:
        step1_path = _find_latest_module1_step1_output()
        step1_source_mode = "latest"
        if step1_path is None or not step1_path.exists():
            return JSONResponse({
                "ok": False,
                "message": "尚未找到 Module 1A 年度產品產量與分類結果。請先完成 Module 1A。",
            }, status_code=400)

    template_path = UPLOAD_DIR / f"bulk_template_{token}_{Path(template_file.filename).name}"
    template_path.write_bytes(await template_file.read())

    working_hour_source = str(working_hour_source or "direct").strip()
    bom_structure_path = None
    working_hour_rollup_path = None
    if _is_include_semi_working_hour_source(working_hour_source):
        if _find_latest_working_hour_rollup() is None:
            return JSONResponse({
                "ok": False,
                "message": "選擇『包含半品工時』時需要 Module 2A working_hour_rollup。請先完成 Module 2A；若只使用直接工時，請將 Working Hour Source 改為 Direct Working Hour。"
            }, status_code=400)
        working_hour_rollup_path = _find_latest_working_hour_rollup()
        bom_structure_path = _fresh_path(LATEST_BOM_STRUCTURE_PATH)

    try:
        summary = generate_product_activity_bulk_files_by_site_zip(
            step1_output_path=step1_path,
            bulk_template_path=template_path,
            output_dir=OUTPUT_DIR,
            token=token,
            working_hour_source=working_hour_source,
            bom_structure_path=bom_structure_path,
            working_hour_rollup_path=working_hour_rollup_path,
        )
        summary["module1_step1_source_mode"] = step1_source_mode
        summary["module1_step1_source_filename"] = step1_path.name if step1_path else ""
        summary["module1_step1_source_download_url"] = f"/download/{step1_path.name}" if step1_path else ""
        summary["module2a_working_hour_rollup_used"] = bool(working_hour_rollup_path and Path(working_hour_rollup_path).exists())
        summary["module2a_working_hour_rollup_filename"] = Path(working_hour_rollup_path).name if working_hour_rollup_path else ""
        summary["working_hour_source_rule"] = "Module 2A working_hour_rollup is required only when Working Hour Source = Include Semi-finished Working Hour."
        output_filename = str(summary.get("output_filename") or "").strip()
        if output_filename:
            generated_output_path = OUTPUT_DIR / output_filename
            if generated_output_path.exists() and generated_output_path.suffix.lower() == ".zip":
                _register_workspace_output(generated_output_path, workspace_id)
                shutil.copy2(generated_output_path, MODULE1B_PRODUCT_ACTIVITY_BULK_LATEST_PATH)
                _register_workspace_output(MODULE1B_PRODUCT_ACTIVITY_BULK_LATEST_PATH, workspace_id)
                summary["module1b_product_activity_bulk_latest"] = MODULE1B_PRODUCT_ACTIVITY_BULK_LATEST_PATH.name
                summary["module1b_product_activity_bulk_latest_download_url"] = f"/download/{MODULE1B_PRODUCT_ACTIVITY_BULK_LATEST_PATH.name}"
    except Exception as exc:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=400)

    return {
        "ok": True,
        "message": "Bulk file generated successfully.",
        "summary": summary,
        "download_url": summary.get("download_url", ""),
    }


@app.post("/upload-rule-master")
async def upload_rule_master(request: Request):
    try:
        form = await request.form()
    except Exception as exc:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": f"無法讀取上傳表單：{exc}"}, status_code=400)

    file = form.get("file")
    rule_set = normalize_rule_set(form.get("rule_set") or DEFAULT_RULE_SET)

    if not file or not getattr(file, "filename", None):
        return JSONResponse({"ok": False, "message": "請上傳 Rule Master 檔案"}, status_code=400)

    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xls", ".csv")):
        return JSONResponse({"ok": False, "message": "請上傳 Excel 或 CSV 檔案"}, status_code=400)

    saved = UPLOAD_DIR / f"{uuid.uuid4().hex}_{file.filename}"
    saved.write_bytes(await file.read())
    try:
        count = save_uploaded_rule(saved, rule_set)
    except Exception as exc:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=400)
    return {"ok": True, "count": count, "rule_set": rule_set}


@app.get("/download/{filename}")
def download(filename: str):
    path = OUTPUT_DIR / filename
    if not path.exists():
        return JSONResponse({"ok": False, "message": "檔案不存在"}, status_code=404)
    media_type = "application/zip" if path.suffix.lower() == ".zip" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return FileResponse(path, filename=filename, media_type=media_type)


@app.get("/download-rule-master")
def download_rule_master(rule_set: str = DEFAULT_RULE_SET):
    ensure_master_files()
    rule_set = normalize_rule_set(rule_set)
    path = get_rule_set_dir(rule_set) / "rule_master.csv"
    return FileResponse(path, filename=f"rule_master_{rule_set}.csv", media_type="text/csv")


@app.get("/download-product-series-master")
def download_product_series_master(rule_set: str = DEFAULT_RULE_SET):
    ensure_master_files()
    rule_set = normalize_rule_set(rule_set)
    path = get_rule_set_dir(rule_set) / "product_series_master.csv"
    return FileResponse(path, filename=f"product_series_master_{rule_set}.csv", media_type="text/csv")





def _find_latest_module1_step1_output() -> Path | None:
    """Return the most recent Module 1A output file for Module 2 roll-up."""
    candidates: list[Path] = []
    seen: set[Path] = set()
    for pattern in ["年度產品產量與分類結果_v6_*.xlsx", "年度產品產量與分類結果*.xlsx"]:
        for path in OUTPUT_DIR.glob(pattern):
            if path in seen:
                continue
            seen.add(path)
            if path.name.startswith("~$"):
                continue
            if path.suffix.lower() in [".xlsx", ".xlsm", ".xls"]:
                candidates.append(path)
    return _freshest_path(candidates)


def _find_latest_module2a_total_usage() -> Path | None:
    """Return this workspace's latest Module 2A Standard BOM Total Usage."""
    if _is_workspace_fresh(MODULE2_STANDARD_BOM_TOTAL_USAGE_PATH):
        return MODULE2_STANDARD_BOM_TOTAL_USAGE_PATH
    candidates = [
        path
        for path in OUTPUT_DIR.glob("standard_bom_total_usage_*.xlsx")
        if path != MODULE2_STANDARD_BOM_TOTAL_USAGE_PATH and not path.name.startswith("~$")
    ]
    return _freshest_path(candidates)


def _find_latest_working_hour_rollup() -> Path | None:
    """Return this workspace's latest Module 2A Working Hour Roll-up."""
    if _is_workspace_fresh(LATEST_WORKING_HOUR_ROLLUP_PATH):
        return LATEST_WORKING_HOUR_ROLLUP_PATH
    candidates = [
        path
        for path in OUTPUT_DIR.glob("working_hour_rollup_*.xlsx")
        if path != LATEST_WORKING_HOUR_ROLLUP_PATH and not path.name.startswith("~$")
    ]
    return _freshest_path(candidates)


def _find_latest_module1b_product_activity_bulk() -> Path | None:
    """Return the latest Module 1B Product Activity Bulk output for progress display."""
    if _is_workspace_fresh(MODULE1B_PRODUCT_ACTIVITY_BULK_LATEST_PATH):
        return MODULE1B_PRODUCT_ACTIVITY_BULK_LATEST_PATH
    candidates: list[Path] = []
    for pattern in (
        "product_activity_data_bulk_by_production_site_*.zip",
        "product_activity_data_bulk_create_*.xlsx",
        # Backward compatibility for outputs generated before the M1B filename update.
        "formatted_product_activity_data_bulk_by_production_site_*.zip",
        "formatted_product_activity_data_bulk_create_*.xlsx",
    ):
        for path in OUTPUT_DIR.glob(pattern):
            if path.name.startswith("~$"):
                continue
            if path.suffix.lower() in [".zip", ".xlsx", ".xlsm", ".xls"]:
                candidates.append(path)
    return _freshest_path(candidates)


@app.get("/module1/progress-status")
def module1_progress_status():
    """Module 1 overview progress for the entry page."""
    step1_path = _find_latest_module1_step1_output()
    step2_path = _find_latest_module1b_product_activity_bulk()
    ready_module1a = bool(step1_path)
    ready_module1b = bool(step2_path)

    if ready_module1b:
        progress_percent = 100
        title = "MODULE 1 已完成"
        message = "Module 1A 年度產品產量與分類結果與 Module 1B Product Activity Bulk 都已完成，可進入 Module 2。"
        status_label = "Completed"
        next_step = "Module 2A"
    elif ready_module1a:
        progress_percent = 50
        title = "MODULE 1A 已完成，等待 Module 1B"
        message = "已找到 Module 1A 年度產品產量與分類結果；下一步請執行 Module 1B Batch Data Formatting。"
        status_label = "50%"
        next_step = "Module 1B"
    else:
        progress_percent = 0
        title = "MODULE 1 尚未開始"
        message = "尚未找到 Module 1A 年度產品產量與分類結果；請先執行 Module 1A Work Order Processing。"
        status_label = "0%"
        next_step = "Module 1A"

    return {
        "ok": True,
        "ready_module1a": ready_module1a,
        "ready_module1b": ready_module1b,
        "progress_percent": progress_percent,
        "title": title,
        "message": message,
        "status_label": status_label,
        "next_step": next_step,
        "module1a_output": _source_info_for_existing_path(
            step1_path,
            "Module 1A 年度產品產量與分類結果",
            "尚未找到 Module 1A 年度產品產量與分類結果。請先完成 Module 1A。",
        ),
        "module1b_output": _source_info_for_existing_path(
            step2_path,
            "Module 1B Product Activity Bulk",
            "尚未找到 Module 1B Product Activity Bulk。請完成 Module 1B Batch Data Formatting。",
        ),
    }


@app.get("/module2/step1-output-source")
def module2_step1_output_source():
    step1_path = _find_latest_module1_step1_output()
    if not step1_path:
        return {
            "ok": False,
            "message": "尚未找到 Module 1A 年度產品產量與分類結果，請先完成 Module 1A。",
        }
    stat = step1_path.stat()
    return {
        "ok": True,
        "filename": step1_path.name,
        "size_bytes": stat.st_size,
        "modified_at": _cmp_mtime_iso(step1_path),
        "download_url": f"/download/{step1_path.name}",
        **_source_meta_for_path(step1_path, "Module 1A"),
    }


def _source_meta_for_path(path: Path, default_version: str = "") -> Dict[str, str]:
    meta = _extract_source_version_date(path.name)
    if not meta.get("source_date"):
        meta["source_date"] = datetime.fromtimestamp(path.stat().st_mtime, CMP_TIMEZONE).date().isoformat()
    if default_version and not meta.get("source_version"):
        meta["source_version"] = default_version
    return meta


def _is_include_semi_working_hour_source(value: str | None) -> bool:
    return str(value or "").strip().lower() in {"include_semi", "semi", "semi_finished", "rollup", "rolled_up", "total"}


def _source_info_for_existing_path(path: Path | None, label: str, missing_message: str) -> Dict[str, Any]:
    if _is_workspace_fresh(path):
        stat = path.stat()
        return {
            "ok": True,
            "filename": path.name,
            "size_bytes": stat.st_size,
            "modified_at": _cmp_mtime_iso(path),
            "download_url": f"/download/{path.name}",
            **_source_meta_for_path(path, label),
        }
    return {"ok": False, "message": missing_message}


@app.get("/module1/step2/source-info")
def module1_step2_source_info():
    """Source status for Module 1B auto-fetch.

    Module 1B always needs Module 1A annual output/classification. Module 2A working_hour_rollup is
    required only when users select Include Semi-finished Working Hour.
    """
    step1_path = _find_latest_module1_step1_output()
    total_usage_path = _find_latest_module2a_total_usage()
    rollup_path = _find_latest_working_hour_rollup()
    return {
        "ok": True,
        "ready_direct": bool(step1_path),
        "ready_include_semi": bool(step1_path and rollup_path),
        "module1_step1": _source_info_for_existing_path(
            step1_path,
            "Module 1A 年度產品產量與分類結果",
            "尚未找到 Module 1A 年度產品產量與分類結果。請先完成 Module 1A。",
        ),
        "module2a_standard_bom_total_usage": _source_info_for_existing_path(
            total_usage_path,
            "Module 2A Standard BOM Total Usage",
            "尚未找到 Module 2A 標準BOM表總用量。這不影響 Direct Working Hour；只有後續 M2B 或半品工時流程才需要。",
        ),
        "module2a_working_hour_rollup": _source_info_for_existing_path(
            rollup_path,
            "Module 2A working_hour_rollup",
            "尚未找到 Module 2A working_hour_rollup。Module 1B 選擇『包含半品工時』時才需要；Direct Working Hour 不需要此檔。",
        ),
        "rule": "Module 1B always references Module 1A annual output/classification; Module 2A working_hour_rollup is required only when Working Hour Source = Include Semi-finished Working Hour.",
    }

def _find_latest_module2c_supplier_mapped_raw_material_bulk_zip() -> Path | None:
    """Return the latest Module 2C supplier-mapped Raw Material Bulk ZIP for Module 3."""
    if _is_workspace_fresh(MODULE2C_SUPPLIER_MAPPED_BULK_ZIP_LATEST_PATH):
        return MODULE2C_SUPPLIER_MAPPED_BULK_ZIP_LATEST_PATH
    candidates: list[Path] = []
    for pattern in ("supplier_mapped_raw_material_bulk_by_site_*.zip", "supplier_mapped_raw_material_activity_data_bulk_by_site_*.zip"):
        for path in OUTPUT_DIR.glob(pattern):
            if path.name.startswith("~$"):
                continue
            candidates.append(path)
    return _freshest_path(candidates)


def _module2_raw_bulk_source_label(path: Path) -> str:
    name = path.name.lower()
    if path == MODULE2C_SUPPLIER_MAPPED_BULK_ZIP_LATEST_PATH or "supplier_mapped" in name:
        return "Module 2C Raw Material Bulk"
    if path == MODULE2B_RAW_MATERIAL_BULK_ZIP_LATEST_PATH or name.startswith("module2b_") or "raw_material_activity_data_bulk_by_site" in name:
        return "Module 2B Raw Material Bulk ZIP"
    return "Module 2 Raw Material Bulk"


def _module2_raw_bulk_source_stage(path: Path) -> str:
    name = path.name.lower()
    if path == MODULE2C_SUPPLIER_MAPPED_BULK_ZIP_LATEST_PATH or "supplier_mapped" in name:
        return "module2c"
    if path == MODULE2B_RAW_MATERIAL_BULK_ZIP_LATEST_PATH or name.startswith("module2b_"):
        return "module2b"
    return "module2"


def _find_latest_module2_raw_material_bulk() -> Path | None:
    """Return the latest Module 2 raw material bulk package/file for Module 3.

    Preference order:
    1. Module 2C Raw Material Bulk ZIP
    2. Module 2B Raw Material Bulk ZIP
    3. Legacy Module 2 raw material bulk outputs

    M3 should always prefer Module 2C even if an older Module 2B path is cached
    in memory from a previous source-status refresh.
    """
    global MODULE2_RAW_MATERIAL_BULK_PATH
    module2c_zip = _find_latest_module2c_supplier_mapped_raw_material_bulk_zip()
    if module2c_zip:
        MODULE2_RAW_MATERIAL_BULK_PATH = module2c_zip
        return module2c_zip

    if MODULE2_RAW_MATERIAL_BULK_PATH and _is_workspace_fresh(MODULE2_RAW_MATERIAL_BULK_PATH):
        cached_stage = _module2_raw_bulk_source_stage(MODULE2_RAW_MATERIAL_BULK_PATH)
        if cached_stage in {"module2b", "module2"}:
            return MODULE2_RAW_MATERIAL_BULK_PATH

    module2b_zip = _find_latest_module2b_raw_material_bulk_zip()
    if module2b_zip:
        MODULE2_RAW_MATERIAL_BULK_PATH = module2b_zip
        return module2b_zip

    candidates: list[Path] = []
    for pattern in ("raw_material_activity_data_bulk_by_site_*.zip", "raw_material_activity_data_bulk_*.zip", "raw_material_activity_data_bulk_*.xlsx"):
        for path in OUTPUT_DIR.glob(pattern):
            name = path.name.lower()
            if name.endswith("_latest.xlsx") or path.name.startswith("~$"):
                continue
            candidates.append(path)
    latest = _freshest_path(candidates)
    if latest is None:
        return None
    MODULE2_RAW_MATERIAL_BULK_PATH = latest
    return latest


def _find_latest_raw_material_bulk_template() -> Path | None:
    """Return the latest user-uploaded Raw Material Bulk Template for M3 final export.

    M2B/M2C/M3 now use lightweight intermediate workbooks for large datasets.
    The final third-party-uploadable template is reapplied only at Module 3,
    using the original template uploaded at Module 2B / legacy Module 2.
    """
    candidates: list[Path] = []
    if _is_workspace_fresh(RAW_MATERIAL_BULK_TEMPLATE_LATEST_PATH):
        candidates.append(RAW_MATERIAL_BULK_TEMPLATE_LATEST_PATH)
    for pattern in (
        "raw_material_bulk_template_latest.xlsx",
        "module2b_raw_material_template_*.xlsx",
        "module2b_raw_material_template_*.xlsm",
        "raw_material_template_*.xlsx",
        "raw_material_template_*.xlsm",
    ):
        for root in (OUTPUT_DIR, UPLOAD_DIR):
            for path in root.glob(pattern):
                if path.name.startswith("~$"):
                    continue
                candidates.append(path)
    # Prefer the explicit latest template when present; it is copied from the latest M2B upload.
    if _is_workspace_fresh(RAW_MATERIAL_BULK_TEMPLATE_LATEST_PATH):
        return RAW_MATERIAL_BULK_TEMPLATE_LATEST_PATH
    return _freshest_path(candidates)


def _find_latest_module3_ccl_filled_output() -> Path | None:
    """Return this workspace's latest completed M3 factor-filled package."""
    if _is_workspace_fresh(MODULE3_CCL_FILLED_LATEST_PATH):
        return MODULE3_CCL_FILLED_LATEST_PATH
    candidates = [
        path for path in OUTPUT_DIR.glob("module3_ccl_factor_filled_*.zip")
        if not path.name.startswith("~$") and path.is_file()
    ]
    return _freshest_path(candidates)


def _find_latest_module2b_raw_material_bulk_zip() -> Path | None:
    """Return the latest Module 2B Raw Material Bulk ZIP for Module 2C."""
    if _is_workspace_fresh(MODULE2B_RAW_MATERIAL_BULK_ZIP_LATEST_PATH):
        return MODULE2B_RAW_MATERIAL_BULK_ZIP_LATEST_PATH
    candidates: list[Path] = []
    for path in OUTPUT_DIR.glob("raw_material_activity_data_bulk_by_site_*.zip"):
        name = path.name.lower()
        if name.startswith("supplier_mapped_") or "supplier_mapped" in name:
            continue
        if path.name.startswith("~$"):
            continue
        candidates.append(path)
    return _freshest_path(candidates)


def _find_latest_module2c_supplier_bulk() -> Path | None:
    """Return the latest final Supplier Bulk created by Module 2C."""
    candidates = [
        path
        for path in OUTPUT_DIR.glob("supplier_bulk_create_*.xlsx")
        if not path.name.startswith("~$") and path.is_file()
    ]
    return _freshest_path(candidates)


@app.get("/module2/progress-status")
def module2_progress_status():
    """Module 2 overview progress and auto-fetched source status for the entry page."""
    total_usage_path = _find_latest_module2a_total_usage()
    rollup_path = _find_latest_working_hour_rollup()
    raw_bulk_path = _find_latest_module2b_raw_material_bulk_zip()
    mapped_bulk_path = _find_latest_module2c_supplier_mapped_raw_material_bulk_zip()
    supplier_bulk_path = _find_latest_module2c_supplier_bulk()

    ready_module2a = bool(total_usage_path and rollup_path)
    ready_module2b = bool(raw_bulk_path)
    ready_module2c = bool(mapped_bulk_path and supplier_bulk_path)

    if ready_module2c:
        progress_percent = 100
        title = "MODULE 2 已完成"
        message = "Module 2A、Module 2B 與 Module 2C 都已完成，可進入 Module 3。"
        status_label = "Completed"
        next_step = "Module 3"
    elif ready_module2b:
        progress_percent = 67
        title = "MODULE 2B 已完成，等待 Module 2C"
        message = "已找到 Raw Material Bulk；下一步請執行 Module 2C 供應商資料對應。"
        status_label = "67%"
        next_step = "Module 2C"
    elif ready_module2a:
        progress_percent = 34
        title = "MODULE 2A 已完成，等待 Module 2B"
        message = "已找到標準 BOM 總用量與 working_hour_rollup；下一步請執行 Module 2B。"
        status_label = "34%"
        next_step = "Module 2B"
    else:
        progress_percent = 0
        title = "MODULE 2 尚未開始"
        message = "尚未找到 Module 2A 標準 BOM 總用量與 working_hour_rollup；請先執行 Module 2A。"
        status_label = "0%"
        next_step = "Module 2A"

    return {
        "ok": True,
        "ready_module2a": ready_module2a,
        "ready_module2b": ready_module2b,
        "ready_module2c": ready_module2c,
        "progress_percent": progress_percent,
        "title": title,
        "message": message,
        "status_label": status_label,
        "next_step": next_step,
        "module2a_total_usage": _source_info_for_existing_path(
            total_usage_path,
            "Module 2A Standard BOM Total Usage",
            "尚未找到 Module 2A 標準 BOM 總用量。請先完成 Module 2A。",
        ),
        "module2a_working_hour_rollup": _source_info_for_existing_path(
            rollup_path,
            "Module 2A working_hour_rollup",
            "尚未找到 Module 2A working_hour_rollup。請先完成 Module 2A。",
        ),
        "module2b_raw_material_bulk": _source_info_for_existing_path(
            raw_bulk_path,
            "Module 2B Raw Material Bulk",
            "尚未找到 Module 2B Raw Material Bulk。請先完成 Module 2B。",
        ),
        "module2c_supplier_mapped_bulk": _source_info_for_existing_path(
            mapped_bulk_path,
            "Module 2C Supplier-mapped Raw Material Bulk",
            "尚未找到 Module 2C 供應商對應 Raw Material Bulk。請先完成 Module 2C。",
        ),
        "module2c_supplier_bulk": _source_info_for_existing_path(
            supplier_bulk_path,
            "Module 2C Supplier Bulk",
            "尚未找到 Module 2C Supplier Bulk。請先完成 Module 2C。",
        ),
    }


@app.get("/module3/raw-material-bulk-source")
def module3_raw_material_bulk_source():
    raw_path = _find_latest_module2_raw_material_bulk()
    if not raw_path:
        return {
            "ok": False,
            "message": "尚未找到 Module 2C Raw Material Bulk，請先完成 Module 2C。",
        }
    stat = raw_path.stat()
    source_label = _module2_raw_bulk_source_label(raw_path)
    meta = _source_meta_for_path(raw_path, source_label)
    return {
        "ok": True,
        "filename": raw_path.name,
        "source_type": "zip_package" if raw_path.suffix.lower() == ".zip" else "excel_file",
        "source_stage": _module2_raw_bulk_source_stage(raw_path),
        "source_label": source_label,
        "size_bytes": stat.st_size,
        "modified_at": _cmp_mtime_iso(raw_path),
        "download_url": f"/download/{raw_path.name}",
        **meta,
    }

# =========================================================
# Module 3 · Carbon Emission Factor Selection
# CCL Mapping + Factor Library Search
# =========================================================
@app.post("/module3/apply-ccl-factors-job")
async def module3_apply_ccl_factors_job(
    ccl_mapping_file: UploadFile = File(...),
):
    token = uuid.uuid4().hex[:10]
    job_id = token
    workspace_id = _workspace_id()

    raw_path = _find_latest_module2_raw_material_bulk()
    if not raw_path:
        return JSONResponse(
            {"ok": False, "message": "尚未找到 Module 2C Raw Material Bulk，請先完成 Module 2C。"},
            status_code=400,
        )
    filename = str(getattr(ccl_mapping_file, "filename", "") or "")
    if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        return JSONResponse({"ok": False, "message": "CCL 係數組配表 請上傳 Excel 檔案"}, status_code=400)

    ccl_path = UPLOAD_DIR / f"module3_ccl_mapping_{token}_{Path(ccl_mapping_file.filename).name}"
    output_path = OUTPUT_DIR / f"module3_ccl_factor_filled_{token}.zip"
    ccl_path.write_bytes(await ccl_mapping_file.read())

    _set_module3_ccl_job(
        job_id,
        status="queued",
        progress=0,
        step="工作已建立，等待背景處理",
        message="CCL 係數對應已開始。",
        source_filename=raw_path.name,
        source_meta=_source_meta_for_path(raw_path, _module2_raw_bulk_source_label(raw_path)),
        stage="M3",
        awaiting_official_template=True,
        remaining_seconds=30,
        created_at=_cmp_now_iso(),
        workspace_id=workspace_id,
    )
    MODULE3_CCL_EXECUTOR.submit(_run_module3_ccl_job, job_id, raw_path, ccl_path, output_path, None, workspace_id)
    return {
        "ok": True,
        "job_id": job_id,
        "message": "CCL 係數對應已開始。",
        "source_filename": raw_path.name,
        "source_label": _module2_raw_bulk_source_label(raw_path),
        "source_stage": _module2_raw_bulk_source_stage(raw_path),
        **_source_meta_for_path(raw_path, _module2_raw_bulk_source_label(raw_path)),
    }


@app.get("/module3/ccl-job/{job_id}")
def module3_get_ccl_job(job_id: str):
    job = MODULE3_CCL_JOBS.get(job_id)
    if not job:
        return JSONResponse({"ok": False, "message": "找不到 CCL 對應工作，請重新執行。"}, status_code=404)
    return {"ok": True, "job": job}



@app.get("/module3a/source-info")
def module3a_source_info():
    m3_path = _find_latest_module3_ccl_filled_output()
    return {
        "ok": bool(m3_path),
        "message": "已找到 M3 係數對應完成檔，可上傳正式原物料批次檔執行 M3A。" if m3_path else "尚未找到 M3 係數對應完成檔，請先完成 M3。",
        "module3_output": _source_info_for_existing_path(
            m3_path,
            "Module 3 CCL factor-filled intermediate",
            "尚未找到 M3 係數對應完成檔，請先完成 M3。",
        ),
    }


@app.post("/module3a/apply-official-template-job")
async def module3a_apply_official_template_job(
    raw_material_template_file: UploadFile = File(...),
):
    token = uuid.uuid4().hex[:10]
    job_id = token
    workspace_id = _workspace_id()
    m3_path = _find_latest_module3_ccl_filled_output()
    if not m3_path:
        return JSONResponse(
            {"ok": False, "message": "尚未找到本工作階段的 M3 係數對應完成檔，請先完成 M3；若已按 F5，請重新執行流程。"},
            status_code=400,
        )
    filename = str(getattr(raw_material_template_file, "filename", "") or "")
    if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        return JSONResponse({"ok": False, "message": "正式原物料批次檔請上傳 Excel 檔案。"}, status_code=400)
    template_path = UPLOAD_DIR / f"module3a_official_raw_material_template_{token}_{Path(filename).name}"
    output_path = OUTPUT_DIR / f"module3a_final_raw_material_bulk_{token}.zip"
    template_path.write_bytes(await raw_material_template_file.read())
    _register_workspace_output(template_path, workspace_id)
    _set_module3a_job(
        job_id,
        status="queued",
        progress=0,
        step="工作已建立，等待背景處理",
        message="M3A 正式原物料批次檔套版已開始。",
        source_filename=m3_path.name,
        official_template_filename=template_path.name,
        remaining_seconds=60,
        created_at=_cmp_now_iso(),
        workspace_id=workspace_id,
    )
    MODULE3A_EXECUTOR.submit(_run_module3a_template_job, job_id, m3_path, template_path, output_path, workspace_id)
    return {
        "ok": True,
        "job_id": job_id,
        "message": "M3A 正式原物料批次檔套版已開始。",
        "source_filename": m3_path.name,
        "official_template_filename": template_path.name,
    }


@app.get("/module3a/job/{job_id}")
def module3a_get_job(job_id: str):
    job = MODULE3A_JOBS.get(job_id)
    if not job:
        return JSONResponse({"ok": False, "message": "找不到 M3A 工作，請重新執行。"}, status_code=404)
    return {"ok": True, "job": job}


@app.post("/module3/apply-ccl-factors")
async def module3_apply_ccl_factors(
    ccl_mapping_file: UploadFile = File(...),
):
    token = uuid.uuid4().hex[:10]

    raw_path = _find_latest_module2_raw_material_bulk()
    if not raw_path:
        return JSONResponse(
            {"ok": False, "message": "尚未找到 Module 2C Raw Material Bulk，請先完成 Module 2C。"},
            status_code=400,
        )
    filename = str(getattr(ccl_mapping_file, "filename", "") or "")
    if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        return JSONResponse({"ok": False, "message": "CCL 係數組配表 請上傳 Excel 檔案"}, status_code=400)

    ccl_path = UPLOAD_DIR / f"module3_ccl_mapping_{token}_{Path(ccl_mapping_file.filename).name}"
    output_path = OUTPUT_DIR / f"module3_ccl_factor_filled_{token}.zip"
    ccl_path.write_bytes(await ccl_mapping_file.read())

    try:
        summary = apply_ccl_factors_to_raw_material_bulk_package(raw_path, ccl_path, output_path, raw_material_template_path=None)
        summary["app_version"] = "DIP_MODULE3_LIGHTWEIGHT_INTERMEDIATE_PERF_V5"
        if output_path.exists():
            _register_workspace_output(output_path)
            shutil.copy2(output_path, MODULE3_CCL_FILLED_LATEST_PATH)
            _register_workspace_output(MODULE3_CCL_FILLED_LATEST_PATH)
            try:
                summary["output_file_size_bytes"] = output_path.stat().st_size
                summary["output_file_size_mb"] = round(output_path.stat().st_size / 1024 / 1024, 2)
            except OSError:
                pass
        summary["source_filename"] = raw_path.name
        summary["source_label"] = _module2_raw_bulk_source_label(raw_path)
        summary["source_stage"] = _module2_raw_bulk_source_stage(raw_path)
        summary.update(_source_meta_for_path(raw_path, _module2_raw_bulk_source_label(raw_path)))
    except Exception as exc:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=400)

    return {
        "ok": True,
        "message": "CCL 係數對應完成。",
        "summary": summary,
        "download_url": summary.get("download_url", f"/download/{output_path.name}"),
    }


def _module3_factor_library_paths():
    return (
        FACTOR_LIBRARY_DIR / "APOS Cumulative LCIA v3.12(顧問).xlsx",
        FACTOR_LIBRARY_DIR / "Cut-off Cumulative LCIA v3.12(顧問).xlsx",
    )


@app.on_event("startup")
def module3_preload_factor_libraries_on_startup():
    """Preload APOS / Cut-off factor libraries into memory for faster Module 3 searches."""
    try:
        apos_path, cutoff_path = _module3_factor_library_paths()
        summary = preload_factor_libraries(apos_path, cutoff_path)
        print(f"===== MODULE3 FACTOR LIBRARY PRELOAD: {summary} =====")
    except Exception as exc:
        print(f"===== MODULE3 FACTOR LIBRARY PRELOAD FAILED: {exc} =====")


@app.get("/module3/factor-library-filters")
def module3_factor_library_filters():
    apos_path, cutoff_path = _module3_factor_library_paths()
    try:
        geographies = ["GLO", "RoW", "RER"]
        return {
            "ok": True,
            "geographies": geographies,
            "sources": ["APOS", "Cut-off"],
            "process_types": ["production", "market_for"],
            "app_version": "CMP_MODULE3_STAGE2_V21",
        }
    except Exception as exc:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=400)


@app.get("/module3/search-factor-library")
def module3_search_factor_library(
    keyword: str = "",
    activity_name_keyword: str = "",
    reference_product_keyword: str = "",
    limit: int = 10,
    source: str = "all",
    geography: str = "all",
    process_type: str = "all",
    page: int = 1,
    page_size: int = 10,
):
    apos_path, cutoff_path = _module3_factor_library_paths()
    try:
        result = search_factor_library(
            keyword,
            apos_path,
            cutoff_path,
            limit=limit,
            source=source,
            geography=geography,
            process_type=process_type,
            page=page,
            page_size=page_size,
            activity_name_keyword=activity_name_keyword,
            reference_product_keyword=reference_product_keyword,
        )
        result["ok"] = True
        result["app_version"] = "CMP_MODULE3_STAGE2_V20"
        return result
    except Exception as exc:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": str(exc)}, status_code=400)




# =========================================================
# Module 2A · Standard BOM Total Usage
# Standard BOM -> Standard BOM total usage workbook only.
# Does not read/write Raw Material Bulk Template.
# =========================================================
@app.post("/module2a/standard-bom-total-usage-job")
async def module2a_standard_bom_total_usage_job(request: Request):
    try:
        form = await request.form()
    except Exception as exc:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": f"無法讀取上傳表單：{exc}"}, status_code=400)

    def is_upload_file_like(item) -> bool:
        return bool(getattr(item, "filename", None)) and hasattr(item, "read")

    bom_uploads = []
    for item in form.getlist("module2a_bom_files") + form.getlist("bom_files") + form.getlist("bom_file"):
        if is_upload_file_like(item) and id(item) not in {id(x) for x in bom_uploads}:
            bom_uploads.append(item)

    if not bom_uploads:
        return JSONResponse({"ok": False, "message": "請至少上傳一個標準 BOM Excel 檔案"}, status_code=400)

    for bom_file in bom_uploads:
        filename = str(getattr(bom_file, "filename", "") or "")
        if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            return JSONResponse({"ok": False, "message": f"{filename} 不是標準 BOM Excel 檔案"}, status_code=400)

    job_id = uuid.uuid4().hex[:10]
    workspace_id = _workspace_id()

    # MODULE 2A uses the latest Module 1A output only as source metadata.
    # It does not read Step 1 rows and does not require users to type version/date.
    step1_source_filename = ""
    step1_source_modified_at = ""
    bom_version = ""
    bom_date = ""
    step1_source = _find_latest_module1_step1_output()
    if step1_source:
        step1_source_filename = step1_source.name
        step1_source_modified_at = _cmp_mtime_iso(step1_source)
        step1_meta = _source_meta_for_path(step1_source, "")
        bom_version = str(step1_meta.get("source_version") or "").strip()
        bom_date = str(step1_meta.get("source_date") or "").strip()

    bom_paths: list[Path] = []
    for idx, bom_file in enumerate(bom_uploads, start=1):
        filename = str(getattr(bom_file, "filename", "") or f"standard_bom_{idx}.xlsx")
        saved = UPLOAD_DIR / f"module2a_standard_bom_{job_id}_{idx}_{Path(filename).name}"
        saved.write_bytes(await bom_file.read())
        bom_paths.append(saved)

    output_path = OUTPUT_DIR / f"standard_bom_total_usage_{job_id}.xlsx"
    _set_module2a_job(
        job_id,
        status="queued",
        step="Queued",
        progress=0,
        processed_rows=0,
        total_rows=0,
        created_at=_cmp_now_iso(),
        bom_files=len(bom_paths),
        bom_version=bom_version,
        bom_date=bom_date,
        step1_source_filename=step1_source_filename,
        step1_source_modified_at=step1_source_modified_at,
        output_filename=output_path.name,
        workspace_id=workspace_id,
    )
    MODULE2A_EXECUTOR.submit(
        _run_module2a_total_usage_job,
        job_id,
        bom_paths,
        output_path,
        bom_version,
        bom_date,
        step1_source_filename,
        step1_source_modified_at,
        step1_source,
        workspace_id,
    )
    return {"ok": True, "job_id": job_id, "status_url": f"/module2/bom-job/{job_id}"}


@app.get("/module2/bom-job/{job_id}")
def module2_get_bom_job(job_id: str):
    job = _get_module2a_job(job_id)
    if not job:
        return JSONResponse({"ok": False, "message": "找不到 Module 2A job，可能已被清除或服務重啟前未寫入狀態。"}, status_code=404)
    response = dict(job)
    response["ok"] = True
    return response


@app.get("/module2/standard-bom-total-usage-source")
def module2_standard_bom_total_usage_source():
    path = _find_latest_module2a_total_usage()
    if path is None:
        return JSONResponse({"ok": False, "message": "尚未產出標準BOM表總用量，請先完成 Module 2A。"}, status_code=404)
    meta = _extract_source_version_date(path.name)
    data = {
        "ok": True,
        "filename": path.name,
        "download_url": f"/download/{path.name}",
        "mtime": _cmp_mtime_iso(path),
        **meta,
    }
    rollup_path = _find_latest_working_hour_rollup()
    if rollup_path:
        data["working_hour_rollup"] = {
            "ok": True,
            "filename": rollup_path.name,
            "download_url": f"/download/{rollup_path.name}",
            "mtime": _cmp_mtime_iso(rollup_path),
        }
    return data


@app.get("/module2b/source-info")
def module2b_source_info():
    step1_path = _find_latest_module1_step1_output()
    total_usage_path = _find_latest_module2a_total_usage()

    if step1_path:
        step1_info = {
            "ok": True,
            "filename": step1_path.name,
            "modified_at": _cmp_mtime_iso(step1_path),
            "download_url": f"/download/{step1_path.name}",
            **_source_meta_for_path(step1_path, "Module 1A"),
        }
    else:
        step1_info = {
            "ok": False,
            "message": "尚未找到 Module 1A 年度產品產量與分類結果，請先完成 Module 1A。",
        }

    if total_usage_path is not None:
        total_usage_info = {
            "ok": True,
            "filename": total_usage_path.name,
            "modified_at": _cmp_mtime_iso(total_usage_path),
            "download_url": f"/download/{total_usage_path.name}",
        }
    else:
        total_usage_info = {
            "ok": False,
            "message": "尚未找到 Module 2A 標準BOM表總用量，請先完成 Module 2A。",
        }

    return {
        "ok": True,
        "ready": bool(step1_path and total_usage_path is not None),
        "module1_step1": step1_info,
        "module2a_total_usage": total_usage_info,
    }


@app.post("/module2b/raw-material-bulk-job")
async def module2b_raw_material_bulk_job(request: Request):
    try:
        form = await request.form()
    except Exception as exc:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": f"無法讀取上傳表單：{exc}"}, status_code=400)

    template_file = form.get("module2b_template_file") or form.get("template_file")
    if not template_file or not getattr(template_file, "filename", None):
        return JSONResponse({"ok": False, "message": "請上傳 Raw Material Bulk Template。"}, status_code=400)
    filename = str(getattr(template_file, "filename", "") or "")
    if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        return JSONResponse({"ok": False, "message": "Raw Material Bulk Template 請上傳 Excel 檔案。"}, status_code=400)

    step1_path = _find_latest_module1_step1_output()
    if step1_path is None:
        return JSONResponse({"ok": False, "message": "尚未找到 Module 1A 年度產品產量與分類結果，請先完成 Module 1A。"}, status_code=400)
    total_usage_path = _find_latest_module2a_total_usage()
    if total_usage_path is None:
        return JSONResponse({"ok": False, "message": "尚未找到 Module 2A 標準BOM表總用量，請先完成 Module 2A。"}, status_code=400)
    working_hour_rollup_path = _find_latest_working_hour_rollup()
    if working_hour_rollup_path is None:
        return JSONResponse({
            "ok": False,
            "message": "尚未找到 Module 2A working_hour_rollup，請重新完成 Module 2A 後再執行 Module 2B。",
        }, status_code=400)

    token = uuid.uuid4().hex[:10]
    job_id = token
    workspace_id = _workspace_id()
    template_path = UPLOAD_DIR / f"module2b_raw_material_template_{token}_{Path(filename).name}"
    template_path.write_bytes(await template_file.read())
    step1_source = {
        "filename": step1_path.name,
        "modified_at": _cmp_mtime_iso(step1_path),
        "download_url": f"/download/{step1_path.name}",
        **_source_meta_for_path(step1_path, "Module 1A"),
    }
    _set_module2a_job(
        job_id,
        status="queued",
        module="2B",
        step="Queued",
        progress=1,
        processed_rows=0,
        total_rows=0,
        module1_step1_source=step1_source,
        module2a_total_usage_filename=total_usage_path.name,
        module2a_working_hour_rollup_filename=working_hour_rollup_path.name,
        workspace_id=workspace_id,
    )
    MODULE2B_EXECUTOR.submit(
        _run_module2b_raw_bulk_job,
        job_id,
        total_usage_path,
        template_path,
        OUTPUT_DIR,
        token,
        step1_path,
        working_hour_rollup_path,
        step1_source,
        workspace_id,
    )
    return {"ok": True, "job_id": job_id, "status_url": f"/module2/bom-job/{job_id}"}




@app.get("/module2c/source-info")
def module2c_source_info():
    step1_path = _find_latest_module1_step1_output()
    raw_bulk_zip = _find_latest_module2b_raw_material_bulk_zip()

    if step1_path:
        step1_info = {
            "ok": True,
            "filename": step1_path.name,
            "modified_at": _cmp_mtime_iso(step1_path),
            "download_url": f"/download/{step1_path.name}",
            **_source_meta_for_path(step1_path, "Module 1A"),
        }
    else:
        step1_info = {
            "ok": False,
            "message": "尚未找到 Module 1A 年度產品產量與分類結果，請先完成 Module 1A。",
        }

    if raw_bulk_zip:
        raw_bulk_info = {
            "ok": True,
            "filename": raw_bulk_zip.name,
            "modified_at": _cmp_mtime_iso(raw_bulk_zip),
            "download_url": f"/download/{raw_bulk_zip.name}",
            "source_type": "zip_package",
            **_source_meta_for_path(raw_bulk_zip, "Module 2B ZIP"),
        }
    else:
        raw_bulk_info = {
            "ok": False,
            "message": "尚未找到 Module 2B Raw Material Bulk ZIP，請先完成 Module 2B。",
        }

    return {
        "ok": True,
        "ready": bool(step1_path and raw_bulk_zip),
        "module1_step1": step1_info,
        "module2b_raw_bulk": raw_bulk_info,
    }


@app.post("/module2c/supplier-mapping-bulk-job")
async def module2c_supplier_mapping_bulk_job(request: Request):
    try:
        form = await request.form()
    except Exception as exc:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": f"無法讀取上傳表單：{exc}"}, status_code=400)

    def is_upload_file_like(item) -> bool:
        return bool(getattr(item, "filename", None)) and hasattr(item, "read")

    supplier_uploads = []
    for item in form.getlist("module2c_supplier_files") + form.getlist("supplier_files") + form.getlist("supplier_file"):
        if is_upload_file_like(item):
            supplier_uploads.append(item)

    if not supplier_uploads:
        return JSONResponse({"ok": False, "message": "請上傳供應商資料 Excel 檔案。"}, status_code=400)

    supplier_bulk_template_upload = form.get("module2c_supplier_bulk_template") or form.get("supplier_bulk_template_file")
    if not is_upload_file_like(supplier_bulk_template_upload):
        return JSONResponse({"ok": False, "message": "請上傳正式 Supplier Bulk Template。"}, status_code=400)
    supplier_bulk_template_filename = str(getattr(supplier_bulk_template_upload, "filename", "") or "supplier_bulk_template.xlsx")
    if not supplier_bulk_template_filename.lower().endswith((".xlsx", ".xlsm")):
        return JSONResponse({"ok": False, "message": f"{supplier_bulk_template_filename} 不是有效的 Supplier Bulk Template（僅支援 .xlsx/.xlsm）。"}, status_code=400)

    for supplier_file in supplier_uploads:
        filename = str(getattr(supplier_file, "filename", "") or "")
        if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            return JSONResponse({"ok": False, "message": f"{filename} 不是供應商 Excel 檔案。"}, status_code=400)

    step1_path = _find_latest_module1_step1_output()
    if step1_path is None:
        return JSONResponse({"ok": False, "message": "尚未找到 Module 1A 年度產品產量與分類結果，請先完成 Module 1A。"}, status_code=400)
    raw_bulk_zip = _find_latest_module2b_raw_material_bulk_zip()
    if raw_bulk_zip is None:
        return JSONResponse({"ok": False, "message": "尚未找到 Module 2B Raw Material Bulk ZIP，請先完成 Module 2B。"}, status_code=400)

    token = uuid.uuid4().hex[:10]
    job_id = token
    workspace_id = _workspace_id()
    supplier_paths: list[Path] = []
    for idx, supplier_file in enumerate(supplier_uploads, start=1):
        filename = str(getattr(supplier_file, "filename", "") or f"supplier_{idx}.xlsx")
        supplier_path = UPLOAD_DIR / f"module2c_supplier_{token}_{idx}_{Path(filename).name}"
        supplier_path.write_bytes(await supplier_file.read())
        supplier_paths.append(supplier_path)

    supplier_bulk_template_path = UPLOAD_DIR / f"module2c_supplier_bulk_template_{token}_{Path(supplier_bulk_template_filename).name}"
    supplier_bulk_template_path.write_bytes(await supplier_bulk_template_upload.read())

    step1_source = {
        "filename": step1_path.name,
        "modified_at": _cmp_mtime_iso(step1_path),
        "download_url": f"/download/{step1_path.name}",
        **_source_meta_for_path(step1_path, "Module 1A"),
    }
    raw_bulk_source = {
        "filename": raw_bulk_zip.name,
        "modified_at": _cmp_mtime_iso(raw_bulk_zip),
        "download_url": f"/download/{raw_bulk_zip.name}",
        "source_type": "zip_package",
        **_source_meta_for_path(raw_bulk_zip, "Module 2B ZIP"),
    }
    _set_module2a_job(
        job_id,
        status="queued",
        module="2C",
        step="Queued",
        progress=1,
        processed_rows=0,
        total_rows=0,
        module1_step1_source=step1_source,
        module2b_raw_bulk_source=raw_bulk_source,
        supplier_upload_files=len(supplier_paths),
        supplier_bulk_template_filename=supplier_bulk_template_path.name,
        supplier_bulk_template_original_filename=supplier_bulk_template_filename,
        workspace_id=workspace_id,
    )
    MODULE2C_EXECUTOR.submit(
        _run_module2c_supplier_mapping_job,
        job_id,
        raw_bulk_zip,
        supplier_paths,
        supplier_bulk_template_path,
        OUTPUT_DIR,
        token,
        step1_source,
        raw_bulk_source,
        workspace_id,
    )
    return {"ok": True, "job_id": job_id, "status_url": f"/module2/bom-job/{job_id}"}

# =========================================================
# Module 2 · BOM Expansion
# Standard BOM + Raw Material Bulk Template -> Raw Material Bulk
# =========================================================
@app.post("/process-bom-expansion")
async def process_bom_expansion(request: Request):
    global MODULE2_RAW_MATERIAL_BULK_PATH
    """Module 2 BOM Expansion.

    V13 supports multiple Standard BOM Excel files.
    Accepted file field names:
    - bom_files: new multi-upload field
    - bom_file: backward-compatible single-upload field
    """
    try:
        form = await request.form()
    except Exception as exc:
        traceback.print_exc()
        return JSONResponse({"ok": False, "message": f"無法讀取上傳表單：{exc}"}, status_code=400)

    def is_upload_file_like(item) -> bool:
        return bool(getattr(item, "filename", None)) and hasattr(item, "read")

    bom_uploads = []
    for item in form.getlist("bom_files") + form.getlist("bom_file"):
        if is_upload_file_like(item):
            bom_uploads.append(item)

    template_file = form.get("template_file")
    step1_file = form.get("step1_file")

    supplier_uploads = []
    for item in form.getlist("supplier_files") + form.getlist("supplier_file"):
        if is_upload_file_like(item):
            supplier_uploads.append(item)

    parent_col = str(form.get("parent_col") or "")
    component_col = str(form.get("component_col") or "")
    qty_col = str(form.get("qty_col") or "")
    unit_col = str(form.get("unit_col") or "")
    description_col = str(form.get("description_col") or "")
    material_group_col = str(form.get("material_group_col") or "")
    valid_from_col = str(form.get("valid_from_col") or "")

    if not bom_uploads:
        return JSONResponse(
            {"ok": False, "message": "請至少上傳一個 Standard BOM Excel 檔案"},
            status_code=400,
        )

    for bom_file in bom_uploads:
        filename = str(getattr(bom_file, "filename", "") or "")
        if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            return JSONResponse(
                {"ok": False, "message": f"{filename} 不是 Standard BOM Excel 檔案"},
                status_code=400,
            )

    if not template_file or not getattr(template_file, "filename", None):
        return JSONResponse(
            {"ok": False, "message": "請上傳 Raw Material Bulk Template"},
            status_code=400,
        )

    if not template_file.filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        return JSONResponse(
            {"ok": False, "message": "Raw Material Bulk Template 請上傳 Excel 檔案"},
            status_code=400,
        )


    for supplier_file in supplier_uploads:
        filename = str(getattr(supplier_file, "filename", "") or "")
        if filename and not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            return JSONResponse(
                {"ok": False, "message": f"{filename} 不是 Supplier Excel 檔案"},
                status_code=400,
            )

    token = uuid.uuid4().hex[:10]
    workspace_id = _workspace_id()

    bom_paths: list[Path] = []
    for idx, bom_file in enumerate(bom_uploads, start=1):
        filename = str(getattr(bom_file, "filename", "") or f"bom_{idx}.xlsx")
        saved_bom = UPLOAD_DIR / f"standard_bom_{token}_{idx}_{Path(filename).name}"
        saved_bom.write_bytes(await bom_file.read())
        bom_paths.append(saved_bom)

    template_path = UPLOAD_DIR / f"raw_material_template_{token}_{Path(template_file.filename).name}"
    output_path = OUTPUT_DIR / f"raw_material_activity_data_bulk_{token}.xlsx"
    working_hour_rollup_output_path = OUTPUT_DIR / f"working_hour_rollup_{token}.xlsx"
    supplier_bulk_template_path = BASE_DIR / "templates" / "supplier_bulk_create_template_v1.xlsx"
    supplier_bulk_output_path = OUTPUT_DIR / f"supplier_bulk_create_{token}.xlsx"
    step1_path = _find_latest_module1_step1_output()
    supplier_paths: list[Path] = []

    template_path.write_bytes(await template_file.read())
    if template_path.exists():
        _register_workspace_output(template_path, workspace_id)
        shutil.copy2(template_path, RAW_MATERIAL_BULK_TEMPLATE_LATEST_PATH)
        _register_workspace_output(RAW_MATERIAL_BULK_TEMPLATE_LATEST_PATH, workspace_id)
    if step1_path is None:
        return JSONResponse(
            {"ok": False, "message": "尚未找到 Module 1A 年度產品產量與分類結果，請先完成 Module 1A。"},
            status_code=400,
        )

    for idx, supplier_file in enumerate(supplier_uploads, start=1):
        filename = str(getattr(supplier_file, "filename", "") or f"supplier_{idx}.xlsx")
        supplier_path = UPLOAD_DIR / f"supplier_{token}_{idx}_{Path(filename).name}"
        supplier_path.write_bytes(await supplier_file.read())
        supplier_paths.append(supplier_path)

    mapping = {
        "parent_col": parent_col,
        "component_col": component_col,
        "qty_col": qty_col,
        "unit_col": unit_col,
        "description_col": description_col,
        "material_group_col": material_group_col,
        "valid_from_col": valid_from_col,
    }

    try:
        if step1_path is not None:
            summary = generate_raw_material_bulk_files_by_site_zip(
                bom_path=bom_paths,
                raw_material_template_path=template_path,
                output_dir=OUTPUT_DIR,
                token=token,
                step1_output_path=step1_path,
                mapping=mapping,
                supplier_paths=supplier_paths,
                supplier_bulk_template_path=supplier_bulk_template_path if supplier_paths else None,
                supplier_bulk_output_path=supplier_bulk_output_path if supplier_paths else None,
            )
            output_path = OUTPUT_DIR / str(summary.get("output_filename", f"raw_material_activity_data_bulk_by_site_{token}.zip"))
            _register_workspace_output(output_path, workspace_id)
            summary["module1_step1_source_filename"] = step1_path.name
            summary["module1_step1_source_download_url"] = f"/download/{step1_path.name}"
        else:
            summary = generate_raw_material_bulk_file(
                bom_path=bom_paths,
                raw_material_template_path=template_path,
                output_path=output_path,
                mapping=mapping,
                supplier_paths=supplier_paths,
                supplier_bulk_template_path=supplier_bulk_template_path if supplier_paths else None,
                supplier_bulk_output_path=supplier_bulk_output_path if supplier_paths else None,
            )
        bom_structure_summary = export_bom_structure_file(
            bom_path=bom_paths,
            output_path=LATEST_BOM_STRUCTURE_PATH,
            mapping=mapping,
        )
        _register_workspace_output(LATEST_BOM_STRUCTURE_PATH, workspace_id)
        summary["bom_structure_latest"] = LATEST_BOM_STRUCTURE_PATH.name
        summary["bom_structure_rows"] = int(bom_structure_summary.get("structure_rows", 0))
        summary["bom_structure_download_url"] = f"/download/{LATEST_BOM_STRUCTURE_PATH.name}"
        summary["bom_files"] = int(bom_structure_summary.get("bom_files", summary.get("bom_files", len(bom_paths))))
        summary["bom_rows_before_dedup"] = int(bom_structure_summary.get("bom_rows_before_dedup", summary.get("bom_rows_before_dedup", 0)))
        summary["bom_rows_after_dedup"] = int(bom_structure_summary.get("bom_rows_after_dedup", summary.get("bom_rows_after_dedup", 0)))
        summary["bom_duplicate_rows_removed"] = int(bom_structure_summary.get("bom_duplicate_rows_removed", summary.get("bom_duplicate_rows_removed", 0)))

        if step1_path is not None:
            rollup_summary = generate_working_hour_rollup_file(
                step1_output_path=step1_path,
                bom_structure_path=LATEST_BOM_STRUCTURE_PATH,
                output_path=working_hour_rollup_output_path,
            )
            _register_workspace_output(working_hour_rollup_output_path, workspace_id)
            LATEST_WORKING_HOUR_ROLLUP_PATH.write_bytes(working_hour_rollup_output_path.read_bytes())
            _register_workspace_output(LATEST_WORKING_HOUR_ROLLUP_PATH, workspace_id)
            summary["working_hour_rollup_filename"] = working_hour_rollup_output_path.name
            summary["working_hour_rollup_download_url"] = f"/download/{working_hour_rollup_output_path.name}"
            summary["working_hour_rollup_latest"] = LATEST_WORKING_HOUR_ROLLUP_PATH.name
            summary["working_hour_rollup_latest_download_url"] = f"/download/{LATEST_WORKING_HOUR_ROLLUP_PATH.name}"
            summary["working_hour_rollup_rows"] = int(rollup_summary.get("summary_rows", 0))
            summary["working_hour_rollup_detail_rows"] = int(rollup_summary.get("detail_rows", 0))
            summary["working_hour_rollup_total_direct_hours"] = float(rollup_summary.get("total_direct_hours", 0))
            summary["working_hour_rollup_total_semi_hours"] = float(rollup_summary.get("total_semi_hours", 0))
            summary["working_hour_rollup_total_hours"] = float(rollup_summary.get("total_hours", 0))
        else:
            summary["working_hour_rollup_filename"] = ""
            summary["working_hour_rollup_download_url"] = ""
            summary["working_hour_rollup_rows"] = 0

        if output_path.exists():
            _register_workspace_output(output_path, workspace_id)
        if output_path.suffix.lower() == ".xlsx" and output_path.exists():
            MODULE2_RAW_MATERIAL_BULK_PATH = output_path
            summary["raw_material_bulk_filename"] = output_path.name
            summary["raw_material_bulk_download_url"] = f"/download/{output_path.name}"

        summary["supplier_upload_files"] = len(supplier_paths)
        if not supplier_paths:
            summary["supplier_bulk_filename"] = ""
            summary["supplier_bulk_download_url"] = ""
            summary["supplier_bulk_rows"] = 0
            summary["supplier_bulk_generated"] = False
            summary["supplier_status"] = "Not Uploaded"
        else:
            if supplier_bulk_output_path.exists():
                _register_workspace_output(supplier_bulk_output_path, workspace_id)
            summary["supplier_bulk_generated"] = bool(summary.get("supplier_bulk_download_url"))
            summary["supplier_status"] = "Generated" if summary.get("supplier_bulk_download_url") else "Not Generated"
        summary["app_version"] = "CMP_V16_2_SUPPLIER_BULK_OPTIONAL"
        summary["bom_formatter_version"] = BOM_FORMATTER_VERSION
    except Exception as exc:
        traceback.print_exc()
        return JSONResponse(
            {"ok": False, "message": str(exc)},
            status_code=400,
        )

    return {
        "ok": True,
        "message": "BOM Expansion completed successfully.",
        "app_version": "CMP_V16_2_SUPPLIER_BULK_OPTIONAL",
        "bom_formatter_version": BOM_FORMATTER_VERSION,
        "summary": summary,
        "download_url": summary.get("download_url", f"/download/{output_path.name}"),
    }
