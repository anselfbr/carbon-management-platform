from __future__ import annotations

import math
import shutil
import tempfile
import zipfile
from collections import defaultdict
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Callable, Iterable, Iterator

from openpyxl import load_workbook

ProgressCallback = Callable[[dict[str, Any]], None]

SHEET_ALIASES = {"input sheet activity data", "raw material activity data", "activity data"}
ALIASES = {
    "product": ["Allocated Target Product/Service", "Product Name", "Target Product", "target_product", "product_name", "產品代碼", "產品名稱"],
    "material": ["Raw Material Code", "Raw Material Number", "Material", "raw_material_code", "料號", "原物料代碼"],
    "material_name": ["Raw Material Name", "raw_material_name", "原物料名稱"],
    "usage": ["Usage", "Activity Data", "activity_data", "使用量", "用量", "生產數量", "年度生產量", "Annual Quantity", "Delivered quantity", "Production/ Service Quantity", "Production / Service Quantity", "Production/Service Quantity", "Production Service Quantity", "Production Quantity", "Service Quantity", "生產／服務數量", "生產/服務數量"],
    "unit": ["Activity Data Unit", "Unit", "activity_data_unit", "單位"],
    "net_weight": ["Net Weight (optional)", "Net weight", "Net Weight", "net_weight", "淨重"],
    "weight_unit": ["Weight Unit (optional)", "Weight Unit", "weight_unit", "重量單位"],
    "factor": ["Emission Factor", "Carbon Factor", "emission_factor", "碳係數", "排放係數"],
    "supplier": ["Supplier Name (optional)", "Supplier Name", "supplier_name", "供應商名稱"],
    "plant": ["Transportation Destination", "Production Site", "Unit Name", "transportation_destination", "production_site", "廠區", "製造場所"],
    "country": ["Country/Area", "Country Area", "Country", "country_area", "國家地區"],
}

WEIGHT_ACTIVITY_UNITS = {
    "kg", "kilogram", "kilograms", "公斤", "g", "gram", "grams", "公克",
    "mg", "milligram", "milligrams", "毫克", "t", "ton", "tonne", "tonnes", "公噸",
}


def _emit(callback: ProgressCallback | None, **payload: Any) -> None:
    if callback:
        callback(payload)


def _norm(v: Any) -> str:
    return " ".join(str(v or "").replace("\n", " ").strip().lower().split())


def _num(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    try:
        x = float(str(v).replace(",", "").strip())
        return x if math.isfinite(x) else 0.0
    except Exception:
        return 0.0


def _header_map(rows: list[tuple[Any, ...]]) -> dict[str, int]:
    width = max((len(r) for r in rows), default=0)
    candidates: list[list[str]] = []
    for i in range(width):
        values = []
        for row in rows:
            if i < len(row) and str(row[i] or "").strip():
                values.append(str(row[i]).strip())
        candidates.append(values)
    result: dict[str, int] = {}
    for key, names in ALIASES.items():
        wanted = {_norm(name) for name in names}
        for idx, values in enumerate(candidates):
            if any(_norm(value) in wanted for value in values):
                result[key] = idx
                break
    return result


def _weight_to_kg(value: float, unit: str) -> float:
    normalized = _norm(unit)
    if normalized in {"g", "gram", "grams", "公克"}:
        return value / 1000
    if normalized in {"mg", "milligram", "milligrams", "毫克"}:
        return value / 1_000_000
    if normalized in {"t", "ton", "tonne", "tonnes", "metric ton", "公噸"}:
        return value * 1000
    return value


def _cell(row: tuple[Any, ...], header_map: dict[str, int], key: str, default: str = "") -> str:
    idx = header_map.get(key, -1)
    return str(row[idx] or "").strip() if 0 <= idx < len(row) else default


class WorkbookRef:
    __slots__ = ("container", "member", "name")

    def __init__(self, container: Path, member: str | None, name: str):
        self.container = container
        self.member = member
        self.name = name



def _workbook_refs(paths: Iterable[Path]) -> list[WorkbookRef]:
    refs: list[WorkbookRef] = []
    for path in paths:
        if path.suffix.lower() == ".zip":
            with zipfile.ZipFile(path) as archive:
                for member in archive.namelist():
                    lower = member.lower()
                    if lower.endswith((".xlsx", ".xlsm")) and not Path(member).name.startswith("~$"):
                        refs.append(WorkbookRef(path, member, Path(member).name))
        else:
            refs.append(WorkbookRef(path, None, path.name))
    return refs


@contextmanager
def _local_workbook(ref: WorkbookRef) -> Iterator[Path]:
    if ref.member is None:
        yield ref.container
        return
    suffix = Path(ref.member).suffix.lower() or ".xlsx"
    temp_path: Path | None = None
    try:
        with zipfile.ZipFile(ref.container) as archive, archive.open(ref.member) as source:
            with tempfile.NamedTemporaryFile(prefix="m5_", suffix=suffix, delete=False) as target:
                shutil.copyfileobj(source, target, length=1024 * 1024)
                temp_path = Path(target.name)
        yield temp_path
    finally:
        if temp_path:
            temp_path.unlink(missing_ok=True)


def _open_activity_sheet(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    worksheet = None
    for sheet_name in workbook.sheetnames:
        if _norm(sheet_name) in SHEET_ALIASES:
            worksheet = workbook[sheet_name]
            break
    return workbook, worksheet


def _classify(ref: WorkbookRef) -> tuple[str, dict[str, int], int]:
    with _local_workbook(ref) as path:
        workbook, worksheet = _open_activity_sheet(path)
        try:
            if worksheet is None:
                return "skip", {}, 0
            rows = worksheet.iter_rows(values_only=True)
            header1 = next(rows, ())
            header2 = next(rows, ())
            header_map = _header_map([header1, header2])
            estimated_rows = max(int(worksheet.max_row or 0) - 2, 0)
            if {"product", "material", "usage", "factor"}.issubset(header_map):
                return "raw", header_map, estimated_rows
            if {"product", "usage"}.issubset(header_map) and "material" not in header_map and "factor" not in header_map:
                return "product", header_map, estimated_rows
            return "skip", header_map, estimated_rows
        finally:
            workbook.close()


def _add_metric(store: dict[str, dict[str, float]], key: str, emission: float, total_emission: float) -> None:
    item = store.setdefault(key, {"emission": 0.0, "total_emission": 0.0})
    item["emission"] += emission
    item["total_emission"] += total_emission


def _ranked(store: dict[str, dict[str, float]]) -> list[dict[str, Any]]:
    return [{"name": key, **value} for key, value in sorted(store.items(), key=lambda item: item[1]["emission"], reverse=True)]


def analyze_bulk_many(paths: Iterable[Path], progress_callback: ProgressCallback | None = None) -> dict[str, Any]:
    refs = _workbook_refs(paths)
    if not refs:
        raise ValueError("找不到可讀取的 Excel 工作簿。")

    _emit(progress_callback, phase="scan", message="正在掃描工作簿與欄位…", file_current=0, file_total=len(refs), rows_current=0, rows_total=0)
    classified: list[tuple[WorkbookRef, str, dict[str, int], int]] = []
    product_refs = 0
    raw_refs = 0
    skipped_files = 0
    estimated_total_rows = 0
    for index, ref in enumerate(refs, 1):
        kind, header_map, estimated_rows = _classify(ref)
        classified.append((ref, kind, header_map, estimated_rows))
        if kind == "product":
            product_refs += 1
        elif kind == "raw":
            raw_refs += 1
            estimated_total_rows += estimated_rows
        else:
            skipped_files += 1
        _emit(progress_callback, phase="scan", message=f"已掃描 {ref.name}", file_name=ref.name, file_current=index, file_total=len(refs), rows_current=0, rows_total=estimated_total_rows)

    if raw_refs == 0:
        raise ValueError("找不到可分析的原物料 Activity Data；請確認檔案含 Product、Raw Material、Usage 與 Emission Factor 欄位。")
    if product_refs == 0:
        raise ValueError("缺少成品生產數量。請將 M1B 產出的成品 Bulk 與 M3A 原物料 Bulk 一起上傳，系統才能換算每 1 PC 成品的碳排放與原物料用量。")

    # Phase 1: M1B is small. Build O(1) production-quantity indexes before reading large raw-material files.
    qty_by_product: dict[str, float] = defaultdict(float)
    qty_by_product_plant: dict[tuple[str, str], float] = defaultdict(float)
    plants_by_product: dict[str, set[str]] = defaultdict(set)
    product_quantity_rows = 0
    product_file_index = 0
    for ref, kind, _, _ in classified:
        if kind != "product":
            continue
        product_file_index += 1
        _emit(progress_callback, phase="quantity", message=f"正在建立生產數量索引：{ref.name}", file_name=ref.name, file_current=product_file_index, file_total=product_refs, rows_current=0, rows_total=0)
        with _local_workbook(ref) as path:
            workbook, worksheet = _open_activity_sheet(path)
            try:
                if worksheet is None:
                    continue
                rows = worksheet.iter_rows(values_only=True)
                header_map = _header_map([next(rows, ()), next(rows, ())])
                for row_index, row in enumerate(rows, 1):
                    product = _cell(row, header_map, "product")
                    usage_idx = header_map.get("usage", -1)
                    quantity = _num(row[usage_idx]) if 0 <= usage_idx < len(row) else 0.0
                    if not product or quantity <= 0:
                        continue
                    plant = _cell(row, header_map, "plant")
                    product_key = _norm(product)
                    plant_key = _norm(plant)
                    qty_by_product[product_key] += quantity
                    if plant_key:
                        qty_by_product_plant[(product_key, plant_key)] += quantity
                        plants_by_product[product_key].add(plant_key)
                    product_quantity_rows += 1
                    if row_index % 5000 == 0:
                        _emit(progress_callback, phase="quantity", message=f"已讀取 {row_index:,} 筆成品資料", file_name=ref.name, file_current=product_file_index, file_total=product_refs, rows_current=row_index, rows_total=max(int(worksheet.max_row or 0) - 2, 0))
            finally:
                workbook.close()

    if not qty_by_product:
        raise ValueError("成品 Bulk 中找不到有效生產數量。請確認 Production/ Service Quantity 欄位有大於 0 的數值。")

    # Phase 2: stream each raw-material workbook and aggregate immediately.
    products: dict[str, dict[str, float]] = {}
    materials: dict[str, dict[str, float]] = {}
    suppliers: dict[str, dict[str, float]] = {}
    plants: dict[str, dict[str, float]] = {}
    by_product: dict[str, dict[str, Any]] = {}
    material_detail: dict[str, dict[str, Any]] = {}
    product_material_detail: dict[str, dict[str, Any]] = {}
    product_qty: dict[str, float] = {}
    calculation_methods: set[str] = set()
    missing: set[str] = set()
    ambiguous: set[str] = set()
    record_count = 0
    source_rows = 0
    invalid_rows = 0
    absolute_total = 0.0
    raw_file_index = 0

    for ref, kind, _, estimated_rows in classified:
        if kind != "raw":
            continue
        raw_file_index += 1
        _emit(progress_callback, phase="raw", message=f"正在串流計算：{ref.name}", file_name=ref.name, file_current=raw_file_index, file_total=raw_refs, rows_current=source_rows, rows_total=estimated_total_rows)
        with _local_workbook(ref) as path:
            workbook, worksheet = _open_activity_sheet(path)
            try:
                if worksheet is None:
                    continue
                rows = worksheet.iter_rows(values_only=True)
                header_map = _header_map([next(rows, ()), next(rows, ())])
                for file_row_index, row in enumerate(rows, 1):
                    source_rows += 1
                    product = _cell(row, header_map, "product")
                    material = _cell(row, header_map, "material")
                    usage_idx = header_map.get("usage", -1)
                    factor_idx = header_map.get("factor", -1)
                    usage = _num(row[usage_idx]) if 0 <= usage_idx < len(row) else 0.0
                    factor = _num(row[factor_idx]) if 0 <= factor_idx < len(row) else 0.0
                    if not product or not material or usage == 0 or factor == 0:
                        invalid_rows += 1
                        continue

                    plant = _cell(row, header_map, "plant", "Unassigned") or "Unassigned"
                    product_key = _norm(product)
                    plant_key = _norm(plant)
                    quantity = qty_by_product_plant.get((product_key, plant_key), 0.0) if plant_key else 0.0
                    if quantity <= 0:
                        known_plants = plants_by_product.get(product_key, set())
                        if len(known_plants) > 1 and plant_key and (product_key, plant_key) not in qty_by_product_plant:
                            ambiguous.add(f"{product}（原物料廠區：{plant}）")
                            continue
                        quantity = qty_by_product.get(product_key, 0.0)
                    if quantity <= 0:
                        missing.add(product)
                        continue

                    unit = _cell(row, header_map, "unit")
                    net_weight_idx = header_map.get("net_weight", -1)
                    net_weight = _num(row[net_weight_idx]) if 0 <= net_weight_idx < len(row) else 0.0
                    weight_unit = _cell(row, header_map, "weight_unit")
                    if _norm(unit) in WEIGHT_ACTIVITY_UNITS:
                        activity_kg_total = _weight_to_kg(usage, unit)
                        method = "Usage(weight) × EF ÷ Production Quantity"
                    elif net_weight > 0:
                        activity_kg_total = usage * _weight_to_kg(net_weight, weight_unit)
                        method = "Usage × Net Weight × EF ÷ Production Quantity"
                    else:
                        activity_kg_total = usage
                        method = "Usage × EF ÷ Production Quantity"

                    total_emission = activity_kg_total * factor
                    activity_kg = activity_kg_total / quantity
                    emission = total_emission / quantity
                    supplier = _cell(row, header_map, "supplier", "Unassigned") or "Unassigned"
                    material_name = _cell(row, header_map, "material_name", material)

                    _add_metric(products, product, emission, total_emission)
                    _add_metric(materials, material, emission, total_emission)
                    _add_metric(suppliers, supplier, emission, total_emission)
                    _add_metric(plants, plant, emission, total_emission)

                    product_item = by_product.setdefault(product, {"materials": {}, "suppliers": {}, "production_quantity": quantity, "total_emission": 0.0})
                    product_item["total_emission"] += total_emission
                    _add_metric(product_item["materials"], material, emission, total_emission)
                    _add_metric(product_item["suppliers"], supplier, emission, total_emission)

                    material_item = material_detail.setdefault(material, {
                        "name": material_name, "emission": 0.0, "total_emission": 0.0,
                        "activity_kg": 0.0, "total_activity_kg": 0.0,
                        "suppliers": {}, "products": {}, "plants": {}, "records": 0,
                    })
                    material_item["emission"] += emission
                    material_item["total_emission"] += total_emission
                    material_item["activity_kg"] += activity_kg
                    material_item["total_activity_kg"] += activity_kg_total
                    material_item["records"] += 1
                    _add_metric(material_item["suppliers"], supplier, emission, total_emission)
                    _add_metric(material_item["products"], product, emission, total_emission)
                    _add_metric(material_item["plants"], plant, emission, total_emission)

                    detail_key = f"{product}|||{material}"
                    product_material_item = product_material_detail.setdefault(detail_key, {
                        "product": product, "material": material, "name": material_name,
                        "emission": 0.0, "total_emission": 0.0,
                        "activity_kg": 0.0, "total_activity_kg": 0.0,
                        "production_quantity": quantity, "suppliers": {}, "plants": {}, "records": 0,
                    })
                    product_material_item["emission"] += emission
                    product_material_item["total_emission"] += total_emission
                    product_material_item["activity_kg"] += activity_kg
                    product_material_item["total_activity_kg"] += activity_kg_total
                    product_material_item["records"] += 1
                    _add_metric(product_material_item["suppliers"], supplier, emission, total_emission)
                    _add_metric(product_material_item["plants"], plant, emission, total_emission)

                    product_qty[product] = max(product_qty.get(product, 0.0), quantity)
                    calculation_methods.add(method)
                    record_count += 1
                    absolute_total += total_emission

                    if file_row_index % 5000 == 0:
                        _emit(
                            progress_callback,
                            phase="raw",
                            message=f"{ref.name}：已計算 {file_row_index:,} 筆",
                            file_name=ref.name,
                            file_current=raw_file_index,
                            file_total=raw_refs,
                            rows_current=source_rows,
                            rows_total=estimated_total_rows,
                            valid_rows=record_count,
                            invalid_rows=invalid_rows,
                        )
            finally:
                workbook.close()

    if ambiguous:
        sample = "、".join(sorted(ambiguous)[:8])
        raise ValueError(f"部分成品有多個生產廠區，但原物料廠區無法與成品 Bulk 對應：{sample}。請確認 M1B Production Site 與 M3A Transportation Destination 名稱一致。")
    if missing:
        sample = "、".join(sorted(missing)[:12])
        raise ValueError(f"下列成品找不到生產數量，無法換算每 PC：{sample}。請一併上傳包含這些成品的 M1B 成品 Bulk。")
    if record_count == 0:
        raise ValueError("沒有可完成每 PC 換算的資料。")

    _emit(progress_callback, phase="finalize", message="正在建立 Dashboard 彙總資料…", file_current=raw_refs, file_total=raw_refs, rows_current=source_rows, rows_total=estimated_total_rows, valid_rows=record_count, invalid_rows=invalid_rows)

    product_list = _ranked(products)
    material_list = _ranked(materials)
    supplier_list = _ranked(suppliers)
    plant_list = _ranked(plants)
    total_unit = sum(item["emission"] for item in product_list)

    drilldown = {
        product: {
            "materials": _ranked(values["materials"]),
            "suppliers": _ranked(values["suppliers"]),
            "production_quantity": values["production_quantity"],
            "total_emission": values["total_emission"],
        }
        for product, values in by_product.items()
    }
    material_details = {
        key: {
            **values,
            "suppliers": _ranked(values["suppliers"]),
            "products": _ranked(values["products"]),
            "plants": _ranked(values["plants"]),
        }
        for key, values in material_detail.items()
    }
    product_material_details = {
        key: {
            **values,
            "suppliers": _ranked(values["suppliers"]),
            "plants": _ranked(values["plants"]),
        }
        for key, values in product_material_detail.items()
    }

    result = {
        "ok": True,
        "summary": {
            "total_emission": total_unit,
            "absolute_total_emission": absolute_total,
            "product_count": len(product_list),
            "material_count": len(material_list),
            "supplier_count": len(supplier_list),
            "plant_count": len(plant_list),
            "record_count": record_count,
            "source_row_count": source_rows,
            "invalid_row_count": invalid_rows,
            "file_count": len(refs),
            "skipped_files": skipped_files,
            "product_bulk_files": product_refs,
            "raw_material_files": raw_refs,
            "production_quantity_rows": product_quantity_rows,
            "processing_mode": "streaming_aggregate",
        },
        "products": product_list,
        "materials": material_list,
        "suppliers": supplier_list,
        "plants": plant_list,
        "drilldown": drilldown,
        "material_details": material_details,
        "product_material_details": product_material_details,
        "product_quantities": product_qty,
        "calculation_methods": sorted(calculation_methods),
        "analysis_basis": "per_pc",
    }
    _emit(progress_callback, phase="done", message="分析完成", file_current=raw_refs, file_total=raw_refs, rows_current=source_rows, rows_total=source_rows, valid_rows=record_count, invalid_rows=invalid_rows)
    return result


def analyze_bulk(path: Path) -> dict[str, Any]:
    return analyze_bulk_many([path])
